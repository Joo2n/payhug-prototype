# -*- coding: utf-8 -*-
"""정산주기_정리 xlsx 생성기.

입력  /Users/semi/Downloads/정산주기.xlsx            (조현준 작성 · 읽기 전용)
      payhug-spec/_pipeline/investor_admin/ceo_definitions.md   (할인율 0.11%)
출력  /Users/semi/Downloads/payhug_정산주기_정리/정산주기_정리_<날짜>.xlsx

원본 한 시트에 좌우로 붙어 있는 두 블록을 시트로 가른다.
  좌측 B8:I41   플랫폼별 정산주기 요약   → `플랫폼 요약`
  우측 K1:Q372  2025년 365일 일자별 금융일수 → `일자별 365일`
  `비중` D2:H6  MAU 원자료 · 배달앱 비중    → `구성비`

숫자는 이 파일에 적지 않는다. 전부 원본 셀에서 읽는다.
집계는 값이 아니라 수식으로 넣는다 — COUNTIF · SUMPRODUCT · AVERAGE · 뺄셈.

금융일수 정의 (원본 정산주기!N4 · 좌측 블록 E:G 3열)
    선정산일   = 결제일 + 1일
    락계좌 입금일 = 결제일 + N영업일   (N: 요기요 5 · 쿠팡이츠 4 · 배민 3 · 카드 2)
    금융일수   = 락계좌 입금일 - 선정산일
영업일은 토·일과 공휴일을 건너뛴다. 공휴일 집합은 원본 카드 열(Q6:Q370)이 요일 기준값에서
어긋나는 지점을 되짚어 뽑는다 — 이 파일에 날짜를 적어 두지 않는다.
"""
import datetime as dt
import io
import os
import re
import sys
from collections import Counter

import openpyxl
import openpyxl.cell._writer as _cw
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# openpyxl 3.1 의 셀 직렬화는 float 를 "%.16g" 로 적어 유효숫자 17번째를 버린다.
# 원본 캐시값(정산주기!H40 = 2.0164383561643837)이 1ulp 어긋나 `검산` 차가 0 이 안 된다.
# 저장할 때만 왕복 보장 표기(repr)로 바꾼다 — 이 프로세스 안에서만 유효하다.
_safe_string_orig = _cw.safe_string


def _safe_string_roundtrip(value):
    if isinstance(value, float) and value == value and value not in (float('inf'), float('-inf')):
        return repr(value)
    return _safe_string_orig(value)


_cw.safe_string = _safe_string_roundtrip

SRC = '/Users/semi/Downloads/정산주기.xlsx'
BASE = os.path.dirname(os.path.abspath(__file__))
DEFS = os.path.join(BASE, 'ceo_definitions.md')


def _mark_text():
    """가중 기준 갈라적기 표식.

    이 워크북의 W 2.7504068548610725 · Ty 14.60% 는 MAU(이용자 수) 비중에서 나온 값이다.
    화면·원장이 쓰는 값은 금액(Ai) 가중이라 다르다. 숫자는 ledger_facts.json 에서 읽는다.
    """
    import json
    f = json.load(io.open(os.path.join(BASE, 'ledger_facts.json'), encoding='utf-8'))
    return ('MAU(이용자 수) 비중 기준 시장 평균 참고값 — 화면 계산은 금액(Ai) 기준 '
            'W %s일(표기 %s) · Ty %s%%' % (f['wRaw'], f['w'], f['ty']))


MARK_SRC = '용어 정의.docx [1번 이미지] 4~6번 문단 · ledger_facts.json'
OUTDIR = '/Users/semi/Downloads/payhug_정산주기_정리'
STAMP = os.environ.get('CYCLE_XLSX_STAMP', '20260830')
OUT = os.path.join(OUTDIR, '정산주기_정리_%s.xlsx' % STAMP)

WEEK = '월화수목금토일'

# ── 색 ────────────────────────────────────────────────────────────
C_HEAD = '1F3864'
C_ORIG = 'DDEBF7'      # 원본 값 그대로
C_CALC = 'E2EFDA'      # 우리가 계산해 붙인 값
C_HOL = 'FCE4D6'       # 공휴일 당일
C_HOLX = 'FFF2CC'      # 공휴일로 늘어난 금융일수

F_HEAD = PatternFill('solid', fgColor=C_HEAD)
F_ORIG = PatternFill('solid', fgColor=C_ORIG)
F_CALC = PatternFill('solid', fgColor=C_CALC)
F_HOL = PatternFill('solid', fgColor=C_HOL)
F_HOLX = PatternFill('solid', fgColor=C_HOLX)

FONT_HEAD = Font(bold=True, color='FFFFFF', size=10)
FONT = Font(size=10)
FONT_B = Font(bold=True, size=10)

_s = Side(style='thin', color='BFBFBF')
BORDER = Border(left=_s, right=_s, top=_s, bottom=_s)

AL_L = Alignment(horizontal='left', vertical='center')
AL_C = Alignment(horizontal='center', vertical='center')
AL_R = Alignment(horizontal='right', vertical='center')

NF_D10 = '0.0000000000'
NF_D15 = '0.000000000000000'
NF_INT = '#,##0'
NF_DATE = 'yyyy-mm-dd'
NF_PCT = '0.00%'


def put(ws, r, c, v, fill=None, nf=None, al=AL_L, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = FONT_B if bold else FONT
    cell.border = BORDER
    cell.alignment = al
    if fill is not None:
        cell.fill = fill
    if nf:
        cell.number_format = nf
    return cell


def head(ws, r, labels, start=1):
    for i, t in enumerate(labels):
        cell = ws.cell(row=r, column=start + i, value=t)
        cell.font = FONT_HEAD
        cell.fill = F_HEAD
        cell.border = BORDER
        cell.alignment = AL_C


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ── 1. 원본 읽기 ──────────────────────────────────────────────────
def read_source():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sc, sw = wb['정산주기'], wb['비중']

    def v(ws, addr):
        return ws[addr].value

    memo = v(sc, 'C1')
    m = re.search(r'(\d{4})\s*년', memo or '')
    if not m:
        sys.exit('정산주기!C1 에서 연도를 못 읽었다: %r' % memo)
    year = int(m.group(1))

    # 우측 블록 — 머리글 K5:Q5 에서 열 위치를 잡는다
    hdr = {sc.cell(5, c).value: c for c in range(11, 18)}
    col_m, col_d, col_w = hdr['월'], hdr['일'], hdr['요']
    plat_cols = [(k, hdr[k]) for k in
                 ('요기요 금융일수', '쿠팡이츠 금융일수', '배민 금융일수', '카드사 금융일수')]
    src_cols = [get_column_letter(c) for _k, c in plat_cols]

    rows = []
    r = 6
    while sc.cell(r, col_m).value is not None:
        rows.append((r,
                     int(sc.cell(r, col_m).value),
                     int(sc.cell(r, col_d).value),
                     sc.cell(r, col_w).value,
                     [int(sc.cell(r, c).value) for _, c in plat_cols]))
        r += 1
    last = r - 1

    for _, mm, dd, wd, _vals in rows:
        assert WEEK[dt.date(year, mm, dd).weekday()] == wd, (mm, dd, wd)

    # 좌측 블록 — 8행 머리글, 9행부터 8행 단위 4묶음
    lhdr = [sc.cell(8, c).value for c in range(2, 10)]
    blocks = []
    for i, top in enumerate((9, 17, 25, 33)):
        blocks.append({
            'top': top,
            'group': next(sc.cell(rr, 2).value for rr in range(top, 8, -1)
                          if sc.cell(rr, 2).value),
            'name': sc.cell(top, 3).value,
            'cycle_txt': sc.cell(top, 4).value,
            'cycle_n': int(sc.cell(2, 14 + i).value),
            'label': sc.cell(1, 14 + i).value,
            'days': [(sc.cell(top + k, 5).value,
                      sc.cell(top + k, 6).value,
                      sc.cell(top + k, 7).value) for k in range(7)],
            'avg_addr': 'H%d' % (top + 7),
            'avg': sc.cell(top + 7, 8).value,
            'mix_addr': 'I%d' % (top + 7),
            'mix': sc.cell(top + 7, 9).value,
        })
    for b in blocks:
        assert str(b['cycle_n']) in str(b['cycle_txt']), b

    src = {
        'year': year, 'memo': memo, 'memo_addr': 'C1', 'src_cols': src_cols,
        'note_c3': v(sc, 'C3'),
        'defn': v(sc, 'N4'), 'defn_addr': 'N4',
        'defn2': v(sc, 'N371'), 'defn2_addr': 'N371',
        'sales_label': v(sc, 'K1'),
        'hdr_right': [sc.cell(5, c).value for c in range(11, 18)],
        'due_labels': [sc.cell(1, c).value for c in range(14, 18)],
        'hdr_left': lhdr,
        'blocks': blocks,
        'rows': rows, 'first_row': 6, 'last_row': last,
        'avg_row': last + 2,
        'avg_right': [sc.cell(last + 2, c).value for c in range(14, 18)],
        'wavg': v(sc, 'H41'), 'wavg_addr': 'H41',
        'mau_labels': [sw.cell(3, c).value for c in range(4, 8)],
        'mau': [sw.cell(4, c).value for c in range(4, 7)],
        'mau_sum': v(sw, 'G4'),
        'mau_period': v(sw, 'B4'),
        'mau_share': [sw.cell(5, c).value for c in range(4, 7)],
        'deliv_label': v(sw, 'H2'), 'deliv': v(sw, 'H5'),
        'mix_row': [sw.cell(6, c).value for c in range(4, 7)],
        'mau_head': v(sw, 'D2'),
        'share_head': v(sw, 'C5'),
        'colmap': [(sc.cell(1, c).value, sc.cell(5, c).value) for c in range(19, 23)],
    }
    wb.close()
    return src


def read_rate():
    txt = open(DEFS, encoding='utf-8').read()
    m = re.search(r'유동화투자자의 할인율 = ([\d.]+)%', txt)
    if not m:
        sys.exit('ceo_definitions.md 에서 할인율을 못 읽었다')
    line = txt[:m.start()].count('\n') + 1
    body = txt.splitlines()[line - 1].lstrip('- ').strip()
    return float(m.group(1)) / 100.0, (os.path.basename(DEFS), line, body)


# ── 2. 공휴일 역산 (원본 금융일수 4열에서) ────────────────────────
def bd_add(d, n, hol):
    c = d
    k = 0
    while k < n:
        c += dt.timedelta(days=1)
        if c.weekday() < 5 and c not in hol:
            k += 1
    return c


def derive_holidays(src):
    year = src['year']
    ns = [b['cycle_n'] for b in src['blocks']]

    def bad(hol):
        n = 0
        for _r, mm, dd, _wd, vals in src['rows']:
            date = dt.date(year, mm, dd)
            pre = date + dt.timedelta(days=1)
            for i, nd in enumerate(ns):
                if (bd_add(date, nd, hol) - pre).days != vals[i]:
                    n += 1
        return n

    start = dt.date(year, 1, 1)
    cands = [start + dt.timedelta(days=i) for i in range(400)]
    cands = [c for c in cands if c.weekday() < 5]
    hol, cur = set(), bad(set())
    while cur:
        best, bestv = None, cur
        for c in cands:
            if c in hol:
                continue
            n = bad(hol | {c})
            if n < bestv:
                best, bestv = c, n
        if best is None:
            break
        hol.add(best)
        cur = bestv
    return hol, cur


def weekday_baseline(src):
    """좌측 블록 결제일→선정산일→락계좌 입금일 3열에서 요일별 기준 금융일수."""
    base = []
    for b in src['blocks']:
        m = {}
        for pay, pre, dep in b['days']:
            i_pre = WEEK.index(str(pre).strip()[-1])
            i_dep = WEEK.index(str(dep).strip()[-1])
            m[str(pay).strip()] = (i_dep - i_pre - 1) % 7 + 1
        base.append(m)
    return base


# ── 3. 시트 ───────────────────────────────────────────────────────
SH_READ, SH_PLAT, SH_DAY = '읽는 법', '플랫폼 요약', '일자별 365일'
SH_FREQ, SH_MIX, SH_WAVG, SH_CHK = '도수분포', '구성비', '가중평균', '검산'
Q_PLAT, Q_DAY = "'%s'" % SH_PLAT, "'%s'" % SH_DAY

# 일자별 시트 좌표
D_TOP = 2
D_COLS = ['E', 'F', 'G', 'H']
# 플랫폼 요약 좌표
P_TOP = 2
P_STEP = 8


def sheet_read(wb, src, rate_src, n_days, avg_row, plat_avg_rows):
    ws = wb.create_sheet(SH_READ)
    widths(ws, {'A': 17, 'B': 26, 'C': 58, 'D': 15, 'E': 24, 'F': 62})

    ws.merge_cells('A1:D1')
    put(ws, 1, 1, src['defn'], F_ORIG, al=AL_L, bold=True)
    ws.row_dimensions[1].height = 22
    put(ws, 1, 5, '정산주기!%s' % src['defn_addr'], F_ORIG, al=AL_C)

    put(ws, 3, 1, '원본 값', F_ORIG, al=AL_C)
    put(ws, 3, 2, '계산 값', F_CALC, al=AL_C)
    put(ws, 3, 3, '공휴일', F_HOL, al=AL_C)
    put(ws, 3, 4, '공휴일 영향', F_HOLX, al=AL_C)

    head(ws, 5, ['원본 시트', '원본 셀', '원본 표기', '옮긴 시트', '옮긴 범위', '비고'])
    NOTE = {src['wavg_addr']: _mark_text()}

    lr, lastr = src['last_row'], src['last_row']
    b = src['blocks']
    last_day = D_TOP + n_days - 1
    ent = [
        ('정산주기', src['memo_addr'], src['memo'], '', ''),
        ('정산주기', 'C3', src['note_c3'], '', ''),
        ('정산주기', 'N4', src['defn'], SH_READ, 'A1'),
        ('정산주기', 'N371', src['defn2'], SH_READ, 'A1'),
        ('정산주기', 'B8:I8', ' / '.join(str(x) for x in src['hdr_left']), SH_PLAT, 'A1:G1'),
        ('정산주기', 'B9:B32', b[0]['group'], SH_PLAT, 'A2:A25'),
        ('정산주기', 'B33:B40', b[3]['group'], SH_PLAT, 'A26:A33'),
        ('정산주기', 'C9,C17,C25,C33',
         ' / '.join(x['name'] for x in b), SH_PLAT, 'B2:B33'),
        ('정산주기', 'D9,D17,D25,D33',
         ' / '.join(str(x['cycle_txt']) for x in b), SH_PLAT, 'C2:C33'),
        ('정산주기', 'N2:Q2', ' / '.join(str(x['cycle_n']) for x in b), SH_PLAT, 'C2:C33'),
        ('정산주기', 'E9:G15,E17:G23,E25:G31,E33:G39',
         ' / '.join(str(x) for x in src['hdr_left'][3:6]), SH_PLAT, 'D2:F33'),
        ('정산주기', ','.join(x['avg_addr'] for x in b),
         ' / '.join(repr(x['avg']) for x in b), SH_PLAT,
         ','.join('G%d' % r for r in plat_avg_rows)),
        ('정산주기', ','.join(x['mix_addr'] for x in b),
         ' / '.join(repr(x['mix']) for x in b), SH_MIX, 'B7:B10'),
        ('정산주기', src['wavg_addr'], repr(src['wavg']), SH_WAVG, 'B7'),
        ('정산주기', 'K1', src['sales_label'], SH_DAY, 'A1:D1'),
        ('정산주기', 'N1:Q1', ' / '.join(str(x) for x in src['due_labels']), SH_DAY, 'E1:H1'),
        ('정산주기', 'K5:Q5', ' / '.join(str(x) for x in src['hdr_right']), SH_DAY, 'A1:H1'),
        ('정산주기', 'K6:Q%d' % lr, '%s일' % n_days, SH_DAY, 'A%d:H%d' % (D_TOP, last_day)),
        ('정산주기', 'N%d:Q%d' % (src['avg_row'], src['avg_row']),
         ' / '.join(repr(x) for x in src['avg_right']), SH_DAY, 'E%d:H%d' % (avg_row, avg_row)),
        ('정산주기', 'S1:V1,S5:V5',
         ' / '.join('%s=%s' % (a, c) for a, c in src['colmap']), SH_DAY, 'E1:H1'),
        ('비중', 'D2', src['mau_head'], SH_MIX, 'A2'),
        ('비중', 'B4', str(src['mau_period']), SH_MIX, 'A2'),
        ('비중', 'D3:G3', ' / '.join(str(x) for x in src['mau_labels']), SH_MIX, 'B1:E1'),
        ('비중', 'D4:F4', ' / '.join(str(x) for x in src['mau']), SH_MIX, 'B2:D2'),
        ('비중', 'G4', str(src['mau_sum']), SH_MIX, 'E2'),
        ('비중', 'D5:F5', ' / '.join(repr(x) for x in src['mau_share']), SH_MIX, 'B3:D3'),
        ('비중', 'H2', src['deliv_label'], SH_MIX, 'A4'),
        ('비중', 'H5', str(src['deliv']), SH_MIX, 'B4'),
        ('비중', 'D6:F6', ' / '.join(repr(x) for x in src['mix_row']), SH_MIX, 'B7:B9'),
        (rate_src[0], '%d행' % rate_src[1], rate_src[2], SH_WAVG, 'B9'),
    ]
    r = 6
    for row in ent:
        for i, val in enumerate(row):
            cell = put(ws, r, i + 1, val, F_ORIG if i < 3 else F_CALC,
                       al=AL_C if i in (0, 1, 3, 4) else AL_L)
            if i == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center',
                                           wrap_text=True)
        if row[1] in NOTE:
            cell = put(ws, r, 6, NOTE[row[1]], F_ORIG, al=AL_L)
            cell.alignment = Alignment(horizontal='left', vertical='center',
                                       wrap_text=True)
        r += 1
    ws.freeze_panes = 'A6'
    return r - 1


def sheet_plat(wb, src, n_days, avg_row):
    ws = wb.create_sheet(SH_PLAT)
    widths(ws, {'A': 10, 'B': 12, 'C': 16, 'D': 11, 'E': 11, 'F': 17, 'G': 20})
    head(ws, 1, ['구분', '플랫폼', '정산주기(영업일)', '결제 요일', '선정산일',
                 '락계좌 입금일', '금융일수'])
    avg_rows = []
    r = P_TOP
    for i, b in enumerate(src['blocks']):
        for pay, pre, dep in b['days']:
            put(ws, r, 1, b['group'], F_ORIG, al=AL_C)
            put(ws, r, 2, b['name'], F_ORIG, al=AL_C)
            put(ws, r, 3, b['cycle_n'], F_ORIG, al=AL_C)
            put(ws, r, 4, pay, F_ORIG, al=AL_C)
            put(ws, r, 5, str(pre).strip(), F_ORIG, al=AL_C)
            put(ws, r, 6, str(dep).strip(), F_ORIG, al=AL_C)
            put(ws, r, 7,
                '=MOD(FIND(RIGHT(F{0},1),"{1}")-FIND(E{0},"{1}")-1,7)+1'.format(r, WEEK),
                F_CALC, al=AL_C)
            r += 1
        put(ws, r, 1, b['group'], F_ORIG, al=AL_C)
        put(ws, r, 2, b['name'], F_ORIG, al=AL_C)
        put(ws, r, 4, '평균', F_CALC, al=AL_C, bold=True)
        put(ws, r, 7, '=AVERAGE({0}!{1}{2}:{1}{3})'.format(
            Q_DAY, D_COLS[i], D_TOP, D_TOP + n_days - 1),
            F_CALC, nf=NF_D10, al=AL_R, bold=True)
        avg_rows.append(r)
        r += 1
    put(ws, r, 4, '가중평균', F_CALC, al=AL_C, bold=True)
    put(ws, r, 7, '=%s!B7' % SH_WAVG, F_CALC, nf=NF_D10, al=AL_R, bold=True)
    ws.freeze_panes = 'A2'
    return avg_rows, r


def sheet_day(wb, src, hol, base):
    ws = wb.create_sheet(SH_DAY)
    widths(ws, {'A': 6, 'B': 6, 'C': 7, 'D': 13, 'E': 11, 'F': 12, 'G': 10, 'H': 10})
    names = [c for _a, c in src['colmap']]
    head(ws, 1, ['월', '일', '요일', '날짜'] + names)
    year = src['year']
    r = D_TOP
    n_hol_rows = n_hol_cells = 0
    for _sr, mm, dd, wd, vals in src['rows']:
        date = dt.date(year, mm, dd)
        is_hol = date in hol
        f = F_HOL if is_hol else F_ORIG
        n_hol_rows += 1 if is_hol else 0
        put(ws, r, 1, mm, f, nf=NF_INT, al=AL_C)
        put(ws, r, 2, dd, f, nf=NF_INT, al=AL_C)
        put(ws, r, 3, wd, f, al=AL_C)
        put(ws, r, 4, '=DATE(%d,A%d,B%d)' % (year, r, r), F_HOL if is_hol else F_CALC,
            nf=NF_DATE, al=AL_C)
        for i, v in enumerate(vals):
            off = base[i].get(wd) != v
            n_hol_cells += 1 if off else 0
            put(ws, r, 5 + i, v, F_HOLX if off else F_ORIG, nf=NF_INT, al=AL_C)
        r += 1
    avg_row = r
    for c in range(1, 4):
        put(ws, r, c, None, F_CALC, al=AL_C)
    put(ws, r, 4, '평균', F_CALC, al=AL_C, bold=True)
    for i in range(4):
        put(ws, r, 5 + i, '=AVERAGE({0}{1}:{0}{2})'.format(D_COLS[i], D_TOP, r - 1),
            F_CALC, nf=NF_D10, al=AL_R, bold=True)
    ws.freeze_panes = 'A2'
    return avg_row, n_hol_rows, n_hol_cells


def sheet_freq(wb, src, n_days):
    ws = wb.create_sheet(SH_FREQ)
    widths(ws, {'A': 24, 'B': 12, 'C': 12, 'D': 12, 'E': 12})
    names = [c for _a, c in src['colmap']]
    head(ws, 1, ['금융일수'] + names)
    allv = [v for _r, _m, _d, _w, vals in src['rows'] for v in vals]
    lo, hi = min(allv), max(allv)
    r = 2
    for d in range(lo, hi + 1):
        put(ws, r, 1, d, F_CALC, nf=NF_INT, al=AL_C)
        for i in range(4):
            put(ws, r, 2 + i,
                '=COUNTIF({0}!{1}${2}:{1}${3},$A{4})'.format(
                    Q_DAY, D_COLS[i], D_TOP, D_TOP + n_days - 1, r),
                F_CALC, nf=NF_INT, al=AL_R)
        r += 1
    first, last = 2, r - 1
    put(ws, r, 1, '관측 일수 합계', F_CALC, al=AL_L, bold=True)
    for i in range(4):
        c = get_column_letter(2 + i)
        put(ws, r, 2 + i, '=SUM({0}{1}:{0}{2})'.format(c, first, last),
            F_CALC, nf=NF_INT, al=AL_R, bold=True)
    sum_row = r
    r += 1
    put(ws, r, 1, '금융일수 x 도수 합계', F_CALC, al=AL_L, bold=True)
    for i in range(4):
        c = get_column_letter(2 + i)
        put(ws, r, 2 + i, '=SUMPRODUCT($A${1}:$A${2},{0}{1}:{0}{2})'.format(c, first, last),
            F_CALC, nf=NF_INT, al=AL_R, bold=True)
    prod_row = r
    r += 1
    put(ws, r, 1, '평균 금융일수', F_CALC, al=AL_L, bold=True)
    for i in range(4):
        c = get_column_letter(2 + i)
        put(ws, r, 2 + i, '={0}{1}/{0}{2}'.format(c, prod_row, sum_row),
            F_CALC, nf=NF_D10, al=AL_R, bold=True)
    avg_row = r
    ws.freeze_panes = 'A2'
    return {'first': first, 'last': last, 'sum': sum_row, 'prod': prod_row, 'avg': avg_row,
            'lo': lo, 'hi': hi}


def sheet_mix(wb, src):
    ws = wb.create_sheet(SH_MIX)
    widths(ws, {'A': 24, 'B': 16, 'C': 16, 'D': 16, 'E': 16})
    labels = [str(x) for x in src['mau_labels']]          # 배민 / 쿠팡이츠 / 요기요 / 계
    head(ws, 1, ['항목'] + labels)
    put(ws, 2, 1, '%s %s' % (src['mau_head'], src['mau_period']), F_ORIG, al=AL_L)
    for i, v in enumerate(src['mau']):
        put(ws, 2, 2 + i, v, F_ORIG, nf=NF_INT, al=AL_R)
    put(ws, 2, 5, '=SUM(B2:D2)', F_CALC, nf=NF_INT, al=AL_R)
    put(ws, 3, 1, src['share_head'], F_ORIG, al=AL_L)
    for i in range(3):
        c = get_column_letter(2 + i)
        put(ws, 3, 2 + i, '=%s2/$E$2' % c, F_CALC, nf=NF_D10, al=AL_R)
    put(ws, 3, 5, '=SUM(B3:D3)', F_CALC, nf=NF_D10, al=AL_R)
    put(ws, 4, 1, src['deliv_label'], F_ORIG, al=AL_L)
    put(ws, 4, 2, src['deliv'], F_ORIG, nf=NF_D10, al=AL_R)

    head(ws, 6, ['플랫폼', '구성비'])
    names = [b['name'] for b in src['blocks']]            # 요기요 / 쿠팡이츠 / 배민 / 신용카드
    col_of = {labels[i]: get_column_letter(2 + i) for i in range(3)}
    r = 7
    for n in names:
        put(ws, r, 1, n, F_ORIG, al=AL_C)
        if n in col_of:
            put(ws, r, 2, '=%s3*$B$4' % col_of[n], F_CALC, nf=NF_D10, al=AL_R)
        else:
            put(ws, r, 2, '=1-$B$4', F_CALC, nf=NF_D10, al=AL_R)
        r += 1
    put(ws, r, 1, '계', F_CALC, al=AL_C, bold=True)
    put(ws, r, 2, '=SUM(B7:B%d)' % (r - 1), F_CALC, nf=NF_D10, al=AL_R, bold=True)
    return {'first': 7, 'last': r - 1, 'total': r,
            'card_row': 7 + names.index('신용카드') if '신용카드' in names else r - 1,
            'names': names}


def sheet_wavg(wb, src, plat_avg_rows, mix, rate, rate_src):
    ws = wb.create_sheet(SH_WAVG)
    widths(ws, {'A': 26, 'B': 20, 'C': 20, 'D': 22, 'E': 30})
    head(ws, 1, ['플랫폼', '평균 금융일수', '구성비', '평균 금융일수 x 구성비', '출처'])
    r = 2
    for i, b in enumerate(src['blocks']):
        put(ws, r, 1, b['name'], F_ORIG, al=AL_C)
        put(ws, r, 2, '=%s!G%d' % (Q_PLAT, plat_avg_rows[i]), F_CALC, nf=NF_D10, al=AL_R)
        put(ws, r, 3, '=%s!B%d' % (SH_MIX, mix['first'] + i), F_CALC, nf=NF_D10, al=AL_R)
        put(ws, r, 4, '=B%d*C%d' % (r, r), F_CALC, nf=NF_D10, al=AL_R)
        r += 1
    last = r - 1
    put(ws, 7, 1, 'W 가중평균 금융일수', F_CALC, al=AL_L, bold=True)
    put(ws, 7, 2, '=SUMPRODUCT(B2:B%d,C2:C%d)' % (last, last),
        F_CALC, nf=NF_D10, al=AL_R, bold=True)
    put(ws, 7, 5, '%s!%s' % ('정산주기', src['wavg_addr']), F_ORIG, al=AL_C)
    put(ws, 8, 1, '구성비 합', F_CALC, al=AL_L)
    put(ws, 8, 2, '=SUM(C2:C%d)' % last, F_CALC, nf=NF_D10, al=AL_R)
    put(ws, 9, 1, '할인율', F_ORIG, al=AL_L)
    put(ws, 9, 2, rate, F_ORIG, nf=NF_PCT, al=AL_R)
    put(ws, 9, 5, '%s:%d' % (rate_src[0], rate_src[1]), F_ORIG, al=AL_C)
    put(ws, 10, 1, '연일수', F_CALC, al=AL_L)
    put(ws, 10, 2, len(src['rows']), F_CALC, nf=NF_INT, al=AL_R)
    put(ws, 10, 5, '정산주기!K6:K%d' % src['last_row'], F_ORIG, al=AL_C)
    put(ws, 11, 1, 'Ty수익율', F_CALC, al=AL_L, bold=True)
    put(ws, 11, 2, '=B9*B10/B7', F_CALC, nf=NF_PCT, al=AL_R, bold=True)
    #   B7·B11 이 어느 기준의 값인지 갈라 적는다. 이 워크북만 보면 화면 값과 어긋나 보인다.
    put(ws, 13, 1, '가중 기준', F_ORIG, al=AL_L, bold=True)
    c = put(ws, 13, 2, _mark_text(), F_ORIG, al=AL_L)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells('B13:D13')
    put(ws, 13, 5, MARK_SRC, F_ORIG, al=AL_C)
    ws.row_dimensions[13].height = 30
    return {'w': 'B7', 'ty': 'B11', 'mixsum': 'B8', 'mark': 13}


def sheet_check(wb, src, plat_avg_rows, day_avg_row, freq, mix, wav, n_days):
    ws = wb.create_sheet(SH_CHK)
    widths(ws, {'A': 34, 'B': 26, 'C': 24, 'D': 24, 'E': 22, 'F': 62})
    head(ws, 1, ['항목', '원본 셀', '원본 값', '정리 값', '차', '비고'])
    NOTE = {'W 가중평균 금융일수': _mark_text()}
    rows = []
    for i, b in enumerate(src['blocks']):
        rows.append(('%s 평균 금융일수' % b['name'], '정산주기!%s' % b['avg_addr'], b['avg'],
                     '=%s!G%d' % (Q_PLAT, plat_avg_rows[i]), NF_D10))
        rows.append(('%s 평균 금융일수 (도수)' % b['name'],
                     '정산주기!%s' % b['avg_addr'], b['avg'],
                     '=%s!%s%d' % (SH_FREQ, get_column_letter(2 + i), freq['avg']), NF_D10))
        rows.append(('%s 금융일수 합계' % b['name'],
                     '정산주기!%s6:%s%d' % (src['src_cols'][i], src['src_cols'][i],
                                        src['last_row']),
                     sum(v[i] for _r, _m, _d, _w, v in src['rows']),
                     '=SUM({0}!{1}{2}:{1}{3})'.format(Q_DAY, D_COLS[i], D_TOP,
                                                      D_TOP + n_days - 1), NF_INT))
        rows.append(('%s 관측 일수' % b['name'],
                     '정산주기!%s6:%s%d' % (src['src_cols'][i], src['src_cols'][i],
                                        src['last_row']), n_days,
                     '=%s!%s%d' % (SH_FREQ, get_column_letter(2 + i), freq['sum']), NF_INT))
    for i, b in enumerate(src['blocks']):
        rows.append(('%s 구성비' % b['name'], '정산주기!%s' % b['mix_addr'], b['mix'],
                     '=%s!B%d' % (SH_MIX, mix['first'] + i), NF_D10))
    rows += [
        ('MAU 계', '비중!G4', src['mau_sum'], '=%s!E2' % SH_MIX, NF_INT),
        ('%s' % src['deliv_label'], '비중!H5', src['deliv'], '=%s!B4' % SH_MIX, NF_D10),
        ('구성비 합', '정산주기!I16+I24+I32+I40',
         sum(b['mix'] for b in src['blocks']), '=%s!%s' % (SH_WAVG, wav['mixsum']), NF_D10),
        ('W 가중평균 금융일수', '정산주기!%s' % src['wavg_addr'], src['wavg'],
         '=%s!%s' % (SH_WAVG, wav['w']), NF_D10),
        ('일자별 행 수', '정산주기!K6:K%d' % src['last_row'], n_days,
         '=COUNT({0}!A{1}:A{2})'.format(Q_DAY, D_TOP, D_TOP + n_days - 1), NF_INT),
    ]
    for i, b in enumerate(src['blocks']):
        rows.append(('%s 일자별 평균' % b['name'],
                     '정산주기!%s%d' % (src['src_cols'][i], src['avg_row']),
                     src['avg_right'][i],
                     '=%s!%s%d' % (Q_DAY, D_COLS[i], day_avg_row), NF_D10))
    r = 2
    for name, addr, val, ref, nf in rows:
        put(ws, r, 1, name, F_CALC, al=AL_L)
        put(ws, r, 2, addr, F_ORIG, al=AL_C)
        put(ws, r, 3, val, F_ORIG, nf=nf, al=AL_R)
        put(ws, r, 4, ref, F_CALC, nf=nf, al=AL_R)
        put(ws, r, 5, '=D%d-C%d' % (r, r), F_CALC, nf=NF_D15, al=AL_R, bold=True)
        if name in NOTE:
            c = put(ws, r, 6, NOTE[name], F_ORIG, al=AL_L)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        r += 1
    put(ws, r, 1, '차 절대값 합계', F_CALC, al=AL_L, bold=True)
    for c in range(2, 5):
        put(ws, r, c, None, F_CALC, al=AL_C)
    put(ws, r, 5, '=SUMPRODUCT(ABS(E2:E%d))' % (r - 1), F_CALC, nf=NF_D15,
        al=AL_R, bold=True)
    ws.freeze_panes = 'A2'
    return r


def main():
    src = read_source()
    rate, rate_line = read_rate()
    n_days = len(src['rows'])
    hol, resid = derive_holidays(src)
    base = weekday_baseline(src)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_read = wb.create_sheet('__tmp__')          # 자리만 잡고 마지막에 채운다
    plat_avg_rows, plat_last = sheet_plat(wb, src, n_days, 0)
    day_avg_row, n_hol_rows, n_hol_cells = sheet_day(wb, src, hol, base)
    freq = sheet_freq(wb, src, n_days)
    mix = sheet_mix(wb, src)
    wav = sheet_wavg(wb, src, plat_avg_rows, mix, rate, rate_line)
    chk_last = sheet_check(wb, src, plat_avg_rows, day_avg_row, freq, mix, wav, n_days)

    wb.remove(ws_read)
    sheet_read(wb, src, rate_line, n_days, day_avg_row, plat_avg_rows)
    wb.move_sheet(SH_READ, offset=-(len(wb.sheetnames) - 1))

    os.makedirs(OUTDIR, exist_ok=True)
    wb.save(OUT)

    print('원본        %s' % SRC)
    print('출력        %s' % OUT)
    print('일자 행수   %d' % n_days)
    print('도수 구간   %d ~ %d' % (freq['lo'], freq['hi']))
    print('할인율      %s  (%s:%d)' % (rate, rate_line[0], rate_line[1]))
    print('공휴일 도출 %d건  (원본 메모: %s)' % (len(hol), src['memo']))
    for h in sorted(hol):
        print('            %s %s' % (h, WEEK[h.weekday()]))
    print('색칠 행     공휴일 %d행 / 공휴일 영향 셀 %d개' % (n_hol_rows, n_hol_cells))
    print('역산 잔차   %d셀' % resid)
    print('시트        %s' % ' · '.join(wb.sheetnames))


if __name__ == '__main__':
    main()
