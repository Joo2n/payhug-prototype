# -*- coding: utf-8 -*-
"""검산 통합문서 생성기 — 2026-09-01 대표 검산 미팅용.

산출: 검산_투자자어드민_20260901.xlsx

모든 셀은 엑셀 수식이다. 파이썬이 낸 값을 박아 넣는 자리는 셋뿐이다.
  ① 입력 시트의 가정값 (근거 자료에서 읽는다)
  ② 채권 시트의 금융일수 Di — 정산주기.xlsx N6:Q370 (2025년 365일 실측)의 40일 슬라이스
  ③ 화면대조 시트의 화면 값 — ledger_facts.json 에서 읽는다

화면 값의 출처 키 — 원장 재생성 뒤 갈아 끼우는 자리는 SEED 하나뿐이다.
  SEED 딕셔너리가 '이 통합문서가 쓰는 숫자 = ledger_facts.json 의 어느 키' 를 1:1로 적는다.
  W · Ty · 곳수 같은 값을 코드 본문에 손으로 적지 않는다.

w금융일수의 가중치는 금액(Ai)이다 — 2026-08-31 사용자 결정.
  대표 워드 `용어 정의.docx` [1번 이미지] 4~6번 문단이 Σ Ai x Di / Σ Ai 로 못 박았다.
  그래서 플랫폼 구성비의 시드값이 금액 실측(Figma 2782:5879)이고, 대표 엑셀의 MAU 비중은
  `가중치 대조` 시트에서 참고값으로만 나란히 둔다.

근거 자료 (전부 읽기 전용)
  ~/Downloads/정산주기.xlsx     금융일수 도수 · 플랫폼 평균 금융일수 · MAU · 배달앱/전체 0.35
  ~/Downloads/용어 정의.docx    산식 원문 (ceo_definitions.md 로 옮겨 적은 것)
  ledger_facts.json            화면 값 전량 — 이 통합문서가 대조하는 유일한 기준
  platform_duration.py         금액 실측 MEASURED · BOOK_MIX · 참고 MAU_MIX · 미지급/과지급률
  daily_ledger.py              로스터 BOOKROWS · 규모 BOOK/CASH · 기준일 ASOF
  roster16_model.py            비중 최대잉여법 규칙 · 서명 대기 큐 SIGN_PENDING
  build_xlsx.py                엑셀 내려받기 프리셋 6종의 기간
"""
import json, os, re, shutil, sys
from datetime import date, timedelta
from decimal import Decimal as D, ROUND_HALF_UP, ROUND_FLOOR

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

BASE = os.path.dirname(os.path.abspath(__file__))
CYCLE_XLSX = os.path.expanduser('~/Downloads/정산주기.xlsx')
FNAME = '검산_투자자어드민_20260901.xlsx'
OUT = os.path.join(BASE, FNAME)                              # 정본
OUT2 = os.path.expanduser('~/Downloads/payhug_검산엑셀/' + FNAME)   # 사용자가 여는 것
ORDER = ('card', 'bm', 'cpe', 'yo')
LABEL = {'card': '카드사', 'bm': '배달의민족', 'cpe': '쿠팡이츠', 'yo': '요기요'}
WINDOW = 40                       # 선정산일 축 길이(일)

# ── 화면 값 ↔ ledger_facts.json 키 대응 ─────────────────────────────
#   왼쪽이 이 통합문서 안의 이름, 오른쪽이 ledger_facts.json 의 키다.
#   원장을 다시 만들면 JSON 만 바뀌고 이 파일은 그대로 둔다.
SEED = {
    '투자실행액':        'exec',            '순현금':            'cash',
    '투자자산':          'total',           'W금융일수 raw':     'wRaw',
    'W금융일수 표기':    'w',               'Ty수익율':          'ty',
    'S입금부족율 raw':   'sRaw',            'S입금부족율 표기':  's',
    '할인율(%)':         'rate',            '하루 평균 투자실행금': 'dayAvg',
    '원장 일수':         'ledgerDays',      '원장 구간':         'ledgerSpan',
    '채권 건수':         'receivables',     '미회수 채권 건수':  'openReceivables',
    '채권 Di 범위':      'diRange',         '일별 W 범위':       'wRange',
    '주간 PSA':          'weekExec',        '주간 PSM':          'weekProfit',
    '주간 상환액':       'weekRepay',       '주간 일수':         'weekDays',
    '주간 PSD 표기':     'weekW',           '주간 PSD raw':      'weekWRaw',
    '주간 ④':            'weekTy',          '주간 PSC':          'weekPsc',
    '주간 ⑤':            'weekTyAsset',     '전 구간 PSA':       'fullExec',
    '전 구간 PSM':       'fullProfit',      '전 구간 PSD 표기':  'fullW',
    '전 구간 ④':         'fullTy',          '전 구간 PSC':       'fullPsc',
    '전 구간 ⑤':         'fullTyAsset',     '가맹점별 값':       'merchants',
    '일자별 값':         'tyByDate',
}
#   merchants 원소 = [상호, 투자금액, W, S, Ty, 규모구간, flow]
MCOL = dict(name=0, amount=1, w=2, s=3, ty=4, tier=5, flow=6)
#   tyByDate 원소 = [W, ty(%), 투자실행금, 투자수익, 상환액, 채권매입수수료, 부족액 차감]
DCOL = dict(w=0, ty=1, exec=2, profit=3, repay=4, fee=5, ded=6)


# ══════════════════════════════════════════════════════════════════
# 1. 근거 자료 읽기
# ══════════════════════════════════════════════════════════════════
def read_cycle():
    wb = openpyxl.load_workbook(CYCLE_XLSX, data_only=True)
    ws = wb['정산주기']
    col = {'yo': 'N', 'cpe': 'O', 'bm': 'P', 'card': 'Q'}
    avg_cell = {'yo': 'H16', 'cpe': 'H24', 'bm': 'H32', 'card': 'H40'}
    mix_cell = {'yo': 'I16', 'cpe': 'I24', 'bm': 'I32', 'card': 'I40'}
    freq, avg, mix, raw = {}, {}, {}, {}
    for k, c in col.items():
        vals = [ws['%s%d' % (c, r)].value for r in range(6, 371)]
        assert len(vals) == 365 and all(isinstance(v, int) for v in vals), k
        raw[k] = vals
        f = {}
        for v in vals:
            f[v] = f.get(v, 0) + 1
        freq[k] = dict(sorted(f.items()))
        avg[k] = ws[avg_cell[k]].value
        mix[k] = ws[mix_cell[k]].value
    wsb = wb['비중']
    mau = [('배민', wsb['D4'].value), ('쿠팡이츠', wsb['E4'].value), ('요기요', wsb['F4'].value)]
    return dict(freq=freq, avg=avg, mix=mix, mau=mau, wavg=ws['H41'].value,
                obs_days=365, raw=raw, src_xlsx=os.path.basename(CYCLE_XLSX))


def read_book():
    """로스터 · 규모 · 기준일 — daily_ledger.py 의 상수 그대로 읽는다(읽기 전용)."""
    sys.path.insert(0, BASE)
    import daily_ledger as L
    rows = [dict(mid=m[0], name=m[1], bizno=m[2], ceo=m[3], biz=m[4], item=m[5],
                 contract=m[6], tier=m[7], flow=int(m[8]), b=m[9]) for m in L.BOOKROWS]
    return rows, L.ASOF, int(L.BOOK), int(L.CASH)


def read_plat():
    """플랫폼 상수 — platform_duration.py 가 유일한 출처(읽기 전용).

      MEASURED  금액 실측 하루치 순지급액 (Figma 2782:5879) — 로스터 구성비의 원천
      BOOK_MIX  금액 실측 구성비        MEASURED_W  그 구성의 w금융일수
      MAU_MIX   대표 엑셀 MAU 구성비    MAU_W       참고값(엑셀 H41)
    """
    sys.path.insert(0, BASE)
    import platform_duration as pd
    return dict(
        measured={k: int(v) for k, v in pd.MEASURED.items()},
        measured_src=pd.MEASURED_SRC,
        measured_sum=int(pd.MEASURED_SUM),
        deliv={k: int(v) for k, v in pd.DELIV.items()},
        deliv_sum=int(pd.DELIV_SUM),
        b_bar=float(pd.B_BAR),
        book_mix=dict(zip(ORDER, [float(x) for x in pd.BOOK_MIX])),
        mau_mix=dict(zip(ORDER, [float(x) for x in pd.MAU_MIX])),
        mau={k: int(v) for k, v in pd.MAU.items()},
        card_share=float(pd.CARD_SHARE),
        measured_w=float(pd.MEASURED_W), mau_w=float(pd.MAU_W),
        dur={k: float(v) for k, v in pd.DURATION.items()},
        unpaid={k: float(v) for k, v in pd.UNPAID.items()},
        overpaid={k: float(v) for k, v in pd.OVERPAID.items()},
        di=(pd.DI_MIN, pd.DI_MAX))



def read_rate_pct():
    txt = open(os.path.join(BASE, 'ceo_definitions.md'), encoding='utf-8').read()
    m = re.search(r'유동화투자자의 할인율 = ([\d.]+)%', txt)
    return float(m.group(1))


def read_definitions():
    txt = open(os.path.join(BASE, 'ceo_definitions.md'), encoding='utf-8').read()
    out, sec = [], ''
    for ln in txt.splitlines():
        s = ln.strip()
        if s.startswith('## '):
            sec = s[3:].strip()
        elif s.startswith('- '):
            out.append((sec, s[2:].strip()))
    return out


def read_word_lines(n=(4, 5, 6)):
    """용어 정의.docx [1번 이미지] 의 지정 문단 원문 — 가중치 논점의 인용 근거."""
    txt = open(os.path.join(BASE, 'ceo_definitions.md'), encoding='utf-8').read()
    blk = txt.split('## [1번 이미지]', 1)[1].split('## [2번 이미지]', 1)[0]
    items = [s[2:].strip() for s in blk.splitlines() if s.strip().startswith('- ')]
    return [(i, items[i - 1]) for i in n]


# ── 화면 값 — ledger_facts.json 이 유일한 출처 ──────────────────────
def read_facts():
    """원장이 낸 사실값. 여기 없는 화면 숫자는 이 통합문서에 쓰지 않는다."""
    p = os.path.join(BASE, 'ledger_facts.json')
    f = json.load(open(p, encoding='utf-8'))
    for key in SEED.values():
        assert key in f, 'ledger_facts.json 에 %s 키가 없다' % key
    return f


def S(f, name):
    """SEED 이름으로 화면 값을 꺼낸다 — 숫자를 손으로 적지 않게 하는 통로."""
    return f[SEED[name]]


def fsrc(name):
    return 'ledger_facts.json %s' % SEED[name]


def facts_rollup(f, frm, to):
    """tyByDate 를 기간으로 접는다 — 주간·프리셋 화면 값의 재료."""
    ks = sorted(k for k in S(f, '일자별 값') if frm <= k <= to)
    g = S(f, '일자별 값')
    ex = sum(g[k][DCOL['exec']] for k in ks)
    wx = sum(float(g[k][DCOL['w']]) * g[k][DCOL['exec']] for k in ks)
    return dict(days=len(ks), span=(ks[0], ks[-1]),
                exec_=ex, profit=sum(g[k][DCOL['profit']] for k in ks),
                repay=sum(g[k][DCOL['repay']] for k in ks),
                fee=sum(g[k][DCOL['fee']] for k in ks),
                ded=sum(g[k][DCOL['ded']] for k in ks),
                wraw=(wx / ex if ex else 0.0))


def ratios(amounts, base):
    """비중 소수1자리 — 최대잉여법. roster16_model.ratios 와 같은 규칙(합 정확히 100.0)."""
    n = len(amounts)
    if not n or not base:
        return [D(0)] * n
    unit, frac = [], []
    for i, a in enumerate(amounts):
        raw = D(a) * 1000 / D(base)
        fl = int(raw.to_integral_value(rounding=ROUND_FLOOR))
        unit.append(fl)
        frac.append((raw - fl, D(a), -i))
    rest = 1000 - sum(unit)
    for i in sorted(range(n), key=lambda i: frac[i], reverse=True)[:max(0, rest)]:
        unit[i] += 1
    return [D(u) / D(10) for u in unit]


def read_sign_pending():
    """서명 대기 큐 — roster16_model.SIGN_PENDING 원문."""
    txt = open(os.path.join(BASE, 'roster16_model.py'), encoding='utf-8').read()
    blk = txt.split('SIGN_PENDING = (', 1)[1].split(')\n', 1)[0]
    return dict(re.findall(r"\('(M2026-\d{4})',\s*'(\d{4}-\d{2}-\d{2})'\)", blk))


def read_presets():
    """엑셀 내려받기 프리셋 6종 — build_xlsx.PRESETS 그대로."""
    sys.path.insert(0, BASE)
    import build_xlsx as bx
    return [(k, bx.GRAN_NAME[g], lab, frm, to) for k, g, lab, frm, to in bx.PRESETS]


# ══════════════════════════════════════════════════════════════════
# 2. 채권 원장 배정 — 엑셀과 같은 산술로 파이썬에서 먼저 푼다
# ══════════════════════════════════════════════════════════════════
def xr(x, n=0):
    return float(D(repr(float(x))).quantize(D(1).scaleb(-n), rounding=ROUND_HALF_UP))


def xrd(x, n=0):
    return float(D(repr(float(x))).quantize(D(1).scaleb(-n), rounding=ROUND_FLOOR))


def base_model(cyc, roster, plat, rate):
    """배달 3사 내부 배분은 금액 실측 순지급액 비다 — MAU 가 아니다.

    대표 워드 4~6번 문단이 w금융일수의 가중치를 금액(Ai)으로 못 박았다(2026-08-31 결정).
    platform_duration.DELIV = Figma 2782:5879 하루치 순지급액.
    """
    dur = {p: sum(d * n for d, n in cyc['freq'][p].items()) / cyc['obs_days'] for p in ORDER}
    ed2 = {p: sum(d * d * n for d, n in cyc['freq'][p].items()) / cyc['obs_days'] for p in ORDER}
    dsh = {k: plat['deliv'][k] / plat['deliv_sum'] for k in ('bm', 'cpe', 'yo')}
    M = len(roster)
    b = [float(r['b']) for r in roster]
    flow = [r['flow'] for r in roster]
    mix = [[(1 - b[m]) if p == 'card' else b[m] * dsh[p] for p in ORDER] for m in range(M)]
    durm = [sum(mix[m][i] * dur[ORDER[i]] for i in range(4)) for m in range(M)]
    w_flow = sum(flow[m] * durm[m] for m in range(M)) / sum(flow)
    return dict(dur=dur, ed2=ed2, dsh=dsh, mix=mix, durm=durm, w_flow=w_flow, M=M, flow=flow)


def money(mo, exec_target, rate):
    M = mo['M']
    tot = exec_target / ((1 - rate) * mo['w_flow'])
    scale = tot / sum(mo['flow'])
    N = [mo['flow'][m] * scale for m in range(M)]
    net = [[xr(N[m] * mo['mix'][m][i]) for i in range(4)] for m in range(M)]
    ai = [[xr(net[m][i] * (1 - rate)) for i in range(4)] for m in range(M)]
    W = sum(N[m] * mo['durm'][m] for m in range(M)) / sum(N)
    return dict(N=N, net=net, ai=ai, tot=sum(N), W=W, scale=scale)


def _fill(vals, c):
    """문턱 c+1..13 (선정산일 위치 39-c .. 27) 에 미달값(v <= k-1) 배정.

    정산예정일이 축 끝(기준일)까지 고르게 차도록 아직 안 쓴 가장 늦은 정산예정일을 고른다.
    """
    pool = sorted(vals)
    fill, taken = {}, set()
    for k in range(c + 1, 14):
        best = None
        for t in range(len(pool)):
            v = pool[t]
            if v > k - 1:
                continue
            due = WINDOW - k + v
            key = (due in taken, -due)
            if best is None or key < best[0]:
                best = (key, t, due)
        if best is None:
            return None
        fill[k] = pool.pop(best[1])
        taken.add(best[2])
    return fill, pool


def _tail_sums(pool, c):
    """문턱 1..c 에 배정 가능한 c개 값의 합 → 선택 내역.

    오름차순 선택 s_1<=..<=s_c 가 s_j >= j 여야 배정된다.
    같은 조건이 '값 v 이하로 고른 개수 <= v' 다.
    """
    cnt = [0] * 14
    for v in pool:
        cnt[v] += 1
    cur = {(0, 0): []}
    for v in range(1, 14):
        nxt = {}
        for (took, tot), pick in cur.items():
            for k in range(0, min(cnt[v], c - took) + 1):
                t2 = took + k
                if t2 > v:
                    continue
                key = (t2, tot + v * k)
                if key not in nxt:
                    nxt[key] = pick + [(v, k)]
        cur = nxt
    return {tot: pick for (took, tot), pick in cur.items() if took == c}


def _pick(sums, c, u):
    """합이 u 인 선택 → 문턱 k -> 값."""
    if u not in sums:
        return None
    vals = []
    for v, k in sums[u]:
        vals.extend([v] * k)
    vals.sort()
    return {j + 1: vals[j] for j in range(c)}


def _spread(head, tail, rnd, dsum_target, lam=2.0):
    """머리 27건 배치.

    ① 정산예정일 13..39 를 한 건씩 채워 일별 표의 하루 건수를 고르게 한다
    ② 그 구간에 든 채권의 금융일수 합을 27 x 플랫폼 평균 금융일수에 맞춰 PSD 쏠림을 없앤다
    """
    tcnt = [0] * 80
    tsum = 0
    for i in range(13):
        t = 27 + i + tail[i]
        tcnt[t] += 1
        if 13 <= t <= 39:
            tsum += tail[i]

    def score(seq):
        cnt = list(tcnt)
        ds = tsum
        for i, v in enumerate(seq):
            cnt[i + v] += 1
            if 13 <= i + v <= 39:
                ds += v
        c = sum(abs(cnt[t] - 1) for t in range(13, 40))
        c += sum(max(0, cnt[t] - 1) for t in range(1, 13))
        return c + lam * abs(ds - dsum_target), c

    starts = [sorted(head), sorted(head, reverse=True)] + \
             [sorted(head, key=lambda x: rnd.random()) for _ in range(6)]
    best = None
    for seq in starts:
        cur, _c = score(seq)
        for _ in range(1500):
            i, j = rnd.randrange(27), rnd.randrange(27)
            if seq[i] == seq[j]:
                continue
            seq[i], seq[j] = seq[j], seq[i]
            nc, _c2 = score(seq)
            if nc <= cur:
                cur = nc
            else:
                seq[i], seq[j] = seq[j], seq[i]
        if best is None or cur < best[0]:
            best = (cur, list(seq))
        if cur == 0:
            break
    return best


def _realize(vals, c, u, rnd, dsum_target):
    f = _fill(vals, c)
    if f is None:
        return None
    fill, pool = f
    chosen = _pick(_tail_sums(pool, c), c, u)
    if chosen is None:
        return None
    best = None
    for _ in range(12):
        rest = list(pool)
        for v in chosen.values():
            rest.remove(v)
        tail = [(chosen[k] if k <= c else fill[k]) for k in range(13, 0, -1)]
        sc, head = _spread(rest, tail, rnd, dsum_target)
        seq = head + tail
        got = [seq[i] for i in range(WINDOW) if seq[i] >= WINDOW - i]
        if len(seq) != WINDOW or len(got) != c or sum(got) != u:
            continue
        if best is None or sc < best[0]:
            best = (sc, seq)
        if sc == 0:
            break
    return best[1] if best else None


def _hit(w, opts, target, rnd, cap=3000000):
    n = len(w)
    idx = sorted(range(n), key=lambda t: -w[t])
    mn, mx = [0] * (n + 1), [0] * (n + 1)
    for t in range(n - 1, -1, -1):
        j = idx[t]
        mn[t] = mn[t + 1] + w[j] * min(opts[j])
        mx[t] = mx[t + 1] + w[j] * max(opts[j])
    cnt = [0]

    def dfs(t, rem):
        if cnt[0] > cap:
            return None
        cnt[0] += 1
        if t == n:
            return {} if rem == 0 else None
        if rem < mn[t] or rem > mx[t]:
            return None
        j = idx[t]
        mid = (rem - (mn[t + 1] + mx[t + 1]) / 2) / w[j]
        for x in sorted(opts[j], key=lambda v: (abs(v - mid), rnd.random())):
            r = dfs(t + 1, rem - w[j] * x)
            if r is not None:
                r[j] = x
                return r
        return None

    return dfs(0, target)


def assign(cyc, mo, mn, exec_target, seed=20260901):
    import random
    rnd = random.Random(seed)
    raw = cyc['raw']
    pairs = [(m, i) for m in range(mo['M']) for i in range(4)]
    AI = [mn['ai'][m][i] for (m, i) in pairs]
    SAI = sum(AI)
    W = mn['W']

    stot_opts, off_of = [], []
    for (m, i) in pairs:
        seq, dd = raw[ORDER[i]], {}
        for o in range(365):
            dd.setdefault(sum(seq[(o + j) % 365] for j in range(WINDOW)), o)
        stot_opts.append(sorted(dd))
        off_of.append(dd)
    sol1 = _hit(AI, stot_opts, round(WINDOW * W * SAI), rnd)
    assert sol1 is not None, '슬라이스 해 없음'
    offs = {pairs[j]: off_of[j][sol1[j]] for j in range(len(pairs))}
    vals = {k: [raw[ORDER[k[1]]][(offs[k] + j) % 365] for j in range(WINDOW)] for k in pairs}

    SUMS, copts = {}, []
    for (m, i) in pairs:
        a, ok = round(mo['dur'][ORDER[i]]), []
        for c in range(max(1, a - 1), a + 3):
            f = _fill(vals[(m, i)], c)
            if f is None:
                continue
            ss = _tail_sums(f[1], c)
            if ss:
                SUMS[(m, i, c)] = ss
                ok.append(c)
        copts.append(ok)
    sol2 = _hit(AI, copts, exec_target, rnd)
    assert sol2 is not None, '미회수 건수 해 없음'
    cs = [sol2[j] for j in range(len(pairs))]

    A = sum(AI[t] * cs[t] for t in range(len(pairs)))
    wu = (sum(AI[t] * mo['ed2'][ORDER[pairs[t][1]]] for t in range(len(pairs)))
          / sum(AI[t] * mo['dur'][ORDER[pairs[t][1]]] for t in range(len(pairs))))
    uopts = []
    for t, k in enumerate(pairs):
        ss = SUMS[(k[0], k[1], cs[t])]
        nat = cs[t] * mo['ed2'][ORDER[k[1]]] / mo['dur'][ORDER[k[1]]]
        near = sorted(x for x in ss if int(nat) - 3 <= x <= int(nat) + 4)
        uopts.append(near or sorted(ss))
    sol3 = _hit(AI, uopts, round(wu * A), rnd)
    assert sol3 is not None, '미회수 d합 해 없음'

    seqs = {}
    for t, k in enumerate(pairs):
        s = _realize(vals[k], cs[t], sol3[t], rnd,
                     27 * mo['dur'][ORDER[k[1]]])
        assert s is not None, ('배치 실패', k)
        seqs[k] = s
    return seqs


# ══════════════════════════════════════════════════════════════════
# 3. 서식
# ══════════════════════════════════════════════════════════════════
HEAD = PatternFill('solid', fgColor='1F3864')
SUB = PatternFill('solid', fgColor='D9E2F3')
CHK = PatternFill('solid', fgColor='FFF2CC')
BAD = PatternFill('solid', fgColor='FFC7CE')
GOOD = PatternFill('solid', fgColor='C6EFCE')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FW = Font(color='FFFFFF', bold=True, size=10)
FB = Font(bold=True, size=10)
F_ = Font(size=10)
M0 = '#,##0'
M2 = '#,##0.00'
D6 = '0.000000'
D8 = '0.00000000'
P2 = '0.00'
DT = 'yyyy-mm-dd'

# ── 아래첨자 표기 ──────────────────────────────────────────────────
#   이 openpyxl(3.1.5)은 CellRichText + InlineFont(vertAlign='subscript') 로
#   셀 안 일부 글자만 아래첨자로 내릴 수 있다. 그래서 괄호 규약이 아니라 서식으로 간다.
#   글자열 자체는 바꾸지 않는다 — 평문으로 읽으면 지금과 같은 글자다.
#   (본체, 아래첨자) 로만 가른다. 평문으로 읽으면 `SBd-1` 처럼 붙어 있다.
#   수식 문자열(`=`로 시작)·시트 이름·이름정의에는 쓸 수 없어 그 자리는 평문으로 둔다.
#   대문자 `D` 는 금융일수, 소문자 `d` 는 오늘 날짜다 (dm_0831/symbol_rule_0831.md).
#   `d-1` 은 「어제 날짜」가 아니라 「정산예정일이 어제인 대상정산금채권 집합」이다.
#   `산식` 시트 B열의 대표 정의서 원문만 옛 표기(`D-1`)를 그대로 둔다 — 인용이라 손대지 않는다.
SUBSCRIPT = [
    ('SMRd-1', 'SMR', 'd-1'), ('SBd-1', 'SB', 'd-1'), ('SAd-1', 'SA', 'd-1'),
    ('SMd-1', 'SM', 'd-1'), ('SDd-1', 'SD', 'd-1'),
    ('Bd-1i', 'B', 'd-1i'), ('Ad-1i', 'A', 'd-1i'), ('Md-1i', 'M', 'd-1i'),
    ('Dd-1i', 'D', 'd-1i'),
    ('SLi', 'SL', 'i'), ('SAi', 'SA', 'i'),
    ('Api', 'A', 'pi'), ('Dpi', 'D', 'pi'),
    ('Ai', 'A', 'i'), ('Di', 'D', 'i'),
]
SUBMAP = {k: (b, s) for k, b, s in SUBSCRIPT}
SUBRE = re.compile(r'(?<![A-Za-z0-9])(%s)(?![A-Za-z0-9])'
                   % '|'.join(re.escape(k) for k, _, _ in SUBSCRIPT))


def rich(s, bold=False, white=False):
    """라벨 문자열의 기호를 아래첨자 런으로 가른다. 기호가 없으면 원래 문자열 그대로."""
    if not isinstance(s, str) or not s or s.startswith('='):
        return s
    col = 'FFFFFF' if white else None
    base = InlineFont(sz=10, b=bold, color=col)
    low = InlineFont(sz=10, b=bold, color=col, vertAlign='subscript')
    parts, pos = [], 0
    for m in SUBRE.finditer(s):
        head_, tail = SUBMAP[m.group(1)]
        if m.start() > pos:
            parts.append(TextBlock(base, s[pos:m.start()]))
        parts.append(TextBlock(base, head_))
        parts.append(TextBlock(low, tail))
        pos = m.end()
    if not parts:
        return s
    if pos < len(s):
        parts.append(TextBlock(base, s[pos:]))
    return CellRichText(parts)


def head(ws, row, labels, start=1, fill=HEAD):
    for i, t in enumerate(labels):
        c = ws.cell(row=row, column=start + i,
                    value=rich(t, bold=True, white=fill is HEAD))
        c.fill = fill
        c.font = FW if fill is HEAD else FB
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30


def put(ws, row, col, value, fmt=None, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=rich(value, bold=bold))
    c.font = FB if bold else F_
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    c.border = BOX
    return c


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════
# 4. 통합문서 조립
# ══════════════════════════════════════════════════════════════════
def build(idle=None, tag=''):
    """idle 을 주면 그 유휴 비율로 한 벌 더 만든다(파일명에 tag 를 붙인다).

    총 투자자산은 ledger_facts.json 의 `total` 하나다. 두 칸(총 투자자산 · 유휴 비율)에서
      순현금        = 총 투자자산 x 유휴 비율
      투자실행액 목표 = 총 투자자산 - 순현금        (= 총 투자자산 x (1 - 유휴 비율))
    로 갈라진다. 파이썬도 엑셀도 같은 순서로 나눠 반올림 잔차가 두 벌에 같게 남는다.
    """
    cyc = read_cycle()
    plat = read_plat()
    roster, asof, book, cashv = read_book()
    rates = {k: (plat['unpaid'][k], plat['overpaid'][k]) for k in ORDER}
    rate_pct = read_rate_pct()
    rate = rate_pct / 100
    defs = read_definitions()
    word = read_word_lines()
    fx = read_facts()
    sign = read_sign_pending()
    presets = read_presets()

    # 화면 값 — 전부 ledger_facts.json 에서 온다(SEED 대응표).
    scr_exec = S(fx, '투자실행액')            # 화면대조 블록이 쓰는 배포 화면 값
    scr_cash = S(fx, '순현금')
    total_asset = S(fx, '투자자산')
    assert (scr_exec, scr_cash) == (book, cashv), '원장 상수와 화면 값이 어긋난다'
    assert scr_exec + scr_cash == total_asset, '투자실행액 + 순현금 != 투자자산'
    if idle is None:
        idle_ratio = float(D(scr_cash) / D(total_asset))
        exec_target, cash = scr_exec, scr_cash
    else:
        idle_ratio = float(idle)
        cash = int(D(repr(total_asset * idle_ratio)).quantize(D(1), rounding=ROUND_HALF_UP))
        exec_target = total_asset - cash
    scr_m = S(fx, '가맹점별 값')
    scr_share = ratios([m[MCOL['amount']] for m in scr_m], scr_exec)
    SCR = {m[MCOL['name']]: dict(amount=m[MCOL['amount']], w=m[MCOL['w']], s=m[MCOL['s']],
                                 ty=m[MCOL['ty']], tier=m[MCOL['tier']],
                                 flow=m[MCOL['flow']], share=float(scr_share[i]))
           for i, m in enumerate(scr_m)}
    assert set(SCR) == set(r['name'] for r in roster), '로스터 상호가 화면과 다르다'
    _ord = {m[MCOL['name']]: i for i, m in enumerate(scr_m)}      # 화면 표 순서(금액 내림차순)
    roster.sort(key=lambda r: _ord[r['name']])
    week = facts_rollup(fx, '2026-08-21', '2026-08-27')
    full = facts_rollup(fx, *S(fx, '원장 구간'))
    _n = len(roster)
    avg_daily = int(round(sum(r['flow'] for r in roster) / _n))

    mo = base_model(cyc, roster, plat, rate)
    mn = money(mo, exec_target, rate)
    seqs = assign(cyc, mo, mn, exec_target)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    NAME = {}

    def name(key, sheet, ref):
        NAME[key] = "'%s'!%s" % (sheet, ref) if ' ' in sheet else '%s!%s' % (sheet, ref)
        wb.defined_names.add(DefinedName(key, attr_text=NAME[key]))

    # ── 입력 ────────────────────────────────────────────────────
    ws = wb.create_sheet('입력')
    widths(ws, {'A': 16, 'B': 26, 'C': 20, 'D': 8, 'E': 46, 'F': 12, 'G': 14,
                'H': 14, 'I': 18})
    head(ws, 1, ['구분', '항목', '값', '단위', '출처'])
    #   값 표의 행 번호를 먼저 못 박는다 — 아래 수식이 서로를 가리킨다.
    IR = dict(rate=2, ratep=3, year=4, total=5, idle=6, exec=7, cash=8, avg=9,
              asof=10, axis=11, dfrom=12, dto=13, qfrom=14, qto=15, sfrom=16, sto=17,
              n3=18, f5=19)
    D_ = 'C%d' % IR['asof']
    rows = [
        ('요율', '유동화투자자의 할인율', rate, '비율', 'ceo_definitions.md [1번 이미지]'),
        ('요율', '할인율(%)', '=C%d*100' % IR['rate'], '%', 'C%d x 100' % IR['rate']),
        ('요율', '연일수', 365, '일', 'ceo_definitions.md ty수익율'),
        ('규모', '총 투자자산', total_asset, '원', fsrc('투자자산')),
        ('규모', '유휴 비율', idle_ratio, '비율',
         '순현금 / 투자자산 (%s / %s)' % (SEED['순현금'], SEED['투자자산'])),
        ('규모', '투자실행액 목표', '=C%d-C%d' % (IR['total'], IR['cash']), '원',
         'C%d - C%d = 총 투자자산 x (1 - 유휴 비율)' % (IR['total'], IR['cash'])),
        ('규모', '순현금', '=ROUND(C%d*C%d,0)' % (IR['total'], IR['idle']), '원',
         'C%d x C%d = 총 투자자산 x 유휴 비율' % (IR['total'], IR['idle'])),
        ('규모', '가맹점 평균 하루 선정산액', avg_daily, '원',
         'daily_ledger.py BOOKROWS — Σ flow / 곳수'),
        ('기간', '기준일 d', asof, '날짜', 'daily_ledger.py ASOF'),
        ('기간', '선정산일 축 시작', '=%s-%d' % (D_, WINDOW - 1), '날짜', '축 %d일' % WINDOW),
        ('기간', '일별 표 시작', '=%s-26' % D_, '날짜', '전량 커버 구간 시작'),
        ('기간', '일별 표 종료', '=%s' % D_, '날짜', '기준일'),
        ('기간', '조회기간 시작', '=%s-6' % D_, '날짜', 'ceo_definitions.md default 일주일'),
        ('기간', '조회기간 종료', '=%s' % D_, '날짜', '기준일'),
        ('기간', 'S표본 시작', '=%s-20' % D_, '날짜', 'ceo_definitions.md 표본집합 d-20'),
        ('기간', 'S표본 종료', '=%s-11' % D_, '날짜', 'ceo_definitions.md 표본집합 d-11'),
        #   ⑥ 이 쓰는 ③ 은 대표가 「상단 현황의 기간 전체 숫자」까지만 좁혔다. 칸이 비어 있고
        #   비어 있는 동안 ⑥ 은 「미확정」을 낸다. 값을 넣으면 ⑥ 이 그 자리에서 살아난다.
        ('산식', '③ 지시 대상 (미확정)', None, '원',
         'ceo_definitions.md [2번] 이미지의 ③ — 「상단 현황의 기간 전체 숫자」까지만 좁혀짐. '
         '칸 미확정 (U-03 · F-23)'),
        #   ⑤ 는 대표가 수식을 새로 써서 다시 주기로 했다(A-01). 새 수식이 오면 이 한 칸만 고친다.
        #   지금 들어 있는 것은 대표 정의서 원문 산식이다. 실제 수식은 기간집계 행이 정해진 뒤
        #   되돌아 적는다(아래 「⑤ 산식 되돌아 적기」).
        ('산식', '⑤ 산식 (미확정 · 대표 재작성 대기)', None, '%',
         'ceo_definitions.md [2번] 이미지의 ⑤ — 이 칸 하나만 갈아 끼우면 기간집계 ⑤ 와 '
         '⑥ 이 따라온다. 대표 재작성 대기 (A-01)'),
    ]
    fmts = [D8, P2, M0, M0, '0.0%', M0, M0, M0, DT, DT, DT, DT, DT, DT, DT, DT, M0, P2]
    keys = ['할인율', '할인율퍼센트', '연일수', '총투자자산', '유휴비율',
            '투자실행액목표', '순현금', '가맹점평균',
            '기준일', '축시작', '일별시작', '일별종료', '조회시작', '조회종료',
            '표본시작', '표본종료', '지시대상3', '산식5']
    assert len(rows) == len(fmts) == len(keys) == len(IR)
    for i, (g, lab, val, unit, src) in enumerate(rows):
        r = 2 + i
        put(ws, r, 1, g)
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 3, val, fmts[i], fill=CHK)
        put(ws, r, 4, unit)
        put(ws, r, 5, src)
        name(keys[i], '입력', '$C$%d' % r)
    assert '=%s' % D_ == '=C%d' % (2 + keys.index('기준일'))

    SWH = IR['f5'] + 2                                    # 스위치 표 머리 행
    head(ws, SWH, ['스위치', '선택', '선택지', '산출 위치'], fill=SUB)
    sw = [('① W금융일수 모집단', '대상정산금채권 전체',
           '대상정산금채권 전체 / 만기 도래분만 / 미회수 잔량만', '기간집계 B21'),
          ('② W 표기 자릿수', 2, '1 / 2', '기간집계 B22'),
          ('③ 가맹점 수', len(roster), '1 ~ %d' % len(roster), '가맹점 G열'),
          ('④ 산출 방향', 'A', 'A 자산에서 출발 / B 가맹점에서 출발', '가맹점 P열')]
    swkeys = ['SW_모집단', 'SW_자릿수', 'SW_가맹점수', 'SW_방향']
    for i, (lab, val, opt, where) in enumerate(sw):
        r = SWH + 1 + i
        put(ws, r, 1, lab, bold=True)
        put(ws, r, 2, val, fill=CHK)
        put(ws, r, 3, opt)
        put(ws, r, 4, where)
        name(swkeys[i], '입력', '$B$%d' % r)
    for rng, opts in [('B%d' % (SWH + 1), '"대상정산금채권 전체,만기 도래분만,미회수 잔량만"'),
                      ('B%d' % (SWH + 2), '"1,2"'),
                      ('B%d' % (SWH + 3),
                       '"%s"' % ','.join(str(x) for x in range(1, len(roster) + 1))),
                      ('B%d' % (SWH + 4), '"A,B"')]:
        dv = DataValidation(type='list', formula1=opts, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[rng])

    # ── 플랫폼 구성비 — 대표님이 그 자리에서 바꿀 입력 4칸 ────────
    #   워드 4~6번 문단이 w금융일수의 가중치를 금액(Ai)으로 못 박았다(2026-08-31 결정).
    #   그래서 시드값이 금액 실측(Figma 2782:5879)이다. MAU 기준은 가중치 대조 시트에서 대조만.
    PLH = SWH + 6                             # 플랫폼 구성비 표 머리 행
    PL0, PLS = PLH + 1, PLH + 5               # 첫 행 · 합계 행
    head(ws, PLH, ['코드', '플랫폼', '구성비 (입력)', '미지급률', '과지급률', '출처'], fill=SUB)
    for i, p in enumerate(ORDER):
        r = PL0 + i
        put(ws, r, 1, p)
        put(ws, r, 2, LABEL[p], bold=True)
        put(ws, r, 3, plat['book_mix'][p], D8, fill=CHK)
        put(ws, r, 4, rates[p][0], D6, fill=CHK)
        put(ws, r, 5, rates[p][1], D6, fill=CHK)
        put(ws, r, 6, 'platform_duration.py BOOK_MIX · UNPAID · OVERPAID')
    put(ws, PLS, 2, '합계', bold=True)
    put(ws, PLS, 3, '=SUM(C%d:C%d)' % (PL0, PLS - 1), D8, bold=True, fill=CHK)
    name('플랫폼코드범위', '입력', '$A$%d:$A$%d' % (PL0, PLS - 1))
    name('플랫폼명범위', '입력', '$B$%d:$B$%d' % (PL0, PLS - 1))
    name('구성비범위', '입력', '$C$%d:$C$%d' % (PL0, PLS - 1))
    name('미지급률범위', '입력', '$D$%d:$D$%d' % (PL0, PLS - 1))
    name('과지급률범위', '입력', '$E$%d:$E$%d' % (PL0, PLS - 1))

    # ── 금액 실측 — 로스터 구성비의 원천(하루 1가맹점치 순지급액) ──
    MEH = PLS + 2                             # 금액 실측 표 머리 행
    ME0, MES, MED = MEH + 1, MEH + 5, MEH + 6
    head(ws, MEH, ['플랫폼', '하루 순지급액(원)', '구성비', '배달 3사 내부 배분', '출처'],
         fill=SUB)
    for i, p in enumerate(ORDER):
        r = ME0 + i
        put(ws, r, 1, LABEL[p], bold=True)
        put(ws, r, 2, plat['measured'][p], M0, fill=CHK)
        put(ws, r, 3, '=B%d/$B$%d' % (r, MES), D8)
        put(ws, r, 4, '' if p == 'card' else '=B%d/$B$%d' % (r, MED), D8)
        put(ws, r, 5, plat['measured_src'] if p == 'card' else '')
    put(ws, MES, 1, '합계', bold=True)
    put(ws, MES, 2, '=SUM(B%d:B%d)' % (ME0, MES - 1), M0, bold=True)
    put(ws, MES, 3, '=SUM(C%d:C%d)' % (ME0, MES - 1), D8, bold=True)
    put(ws, MED, 1, '배달 3사 합', bold=True)
    put(ws, MED, 2, '=SUM(B%d:B%d)' % (ME0 + 1, MES - 1), M0, bold=True)
    put(ws, MED, 3, '=B%d/B%d' % (MED, MES), D8, bold=True)
    put(ws, MED, 4, '=SUM(D%d:D%d)' % (ME0 + 1, MES - 1), D8, bold=True)
    put(ws, MED, 5, '배달앱/전체 (금액 실측)')
    name('실측금액범위', '입력', '$B$%d:$B$%d' % (ME0, MES - 1))
    name('배달배분범위', '입력', '$D$%d:$D$%d' % (ME0 + 1, MES - 1))
    name('배달앱금액비중', '입력', '$C$%d' % MED)

    # ── 대표 엑셀 MAU — 데이터로 쓰지 않는 참고값(가중치 대조용) ──
    MAH = MED + 2                             # MAU 표 머리 행
    MA0, MAS, MAD = MAH + 1, MAH + 4, MAH + 5
    head(ws, MAH, ['배달앱', 'MAU', '배달 3사 내부 배분', '출처'], fill=SUB)
    for i, (nm, v) in enumerate(cyc['mau']):
        r = MA0 + i
        put(ws, r, 1, nm, bold=True)
        put(ws, r, 2, v, M0, fill=CHK)
        put(ws, r, 3, '=B%d/$B$%d' % (r, MAS), D8)
        put(ws, r, 4, '%s 비중 D4:F4' % cyc['src_xlsx'])
    put(ws, MAS, 1, '계', bold=True)
    put(ws, MAS, 2, '=SUM(B%d:B%d)' % (MA0, MAS - 1), M0, bold=True)
    put(ws, MAS, 3, '=SUM(C%d:C%d)' % (MA0, MAS - 1), D8, bold=True)
    name('MAU배분범위', '입력', '$C$%d:$C$%d' % (MA0, MAS - 1))
    put(ws, MAD, 1, '배달앱/전체', bold=True)
    put(ws, MAD, 2, 1 - plat['card_share'], D8, fill=CHK)
    put(ws, MAD, 3, '=1-B%d' % MAD, D8)
    put(ws, MAD, 4, '%s 비중 H5 — 카드 = 1 - 이 값' % cyc['src_xlsx'])
    name('배달앱비중', '입력', '$B$%d' % MAD)

    RT = MAD + 2                              # 로스터 표 머리 행
    head(ws, RT, ['순번', '사업자ID', '상호', '업종', '품목', '계약일', 'flow 상수',
                  '배달 의존도 b', '하루 선정산액(방향 B)', '계약 상태', '계약 생성일'],
         fill=SUB)
    for i, r0 in enumerate(roster):
        r = RT + 1 + i
        put(ws, r, 1, i + 1)
        put(ws, r, 2, r0['mid'])
        put(ws, r, 3, r0['name'], bold=True)
        put(ws, r, 4, r0['biz'])
        put(ws, r, 5, r0['item'])
        put(ws, r, 6, date(*[int(x) for x in r0['contract'].split('-')]), DT)
        put(ws, r, 7, r0['flow'], M0, fill=CHK)
        put(ws, r, 8, float(r0['b']), D8, fill=CHK)
        put(ws, r, 9, r0['flow'], M0, fill=CHK)
        put(ws, r, 10, '서명 대기' if r0['mid'] in sign else '서명 완료',
            fill=CHK if r0['mid'] in sign else None)
        put(ws, r, 11, date(*[int(x) for x in sign[r0['mid']].split('-')])
            if r0['mid'] in sign else '', DT)
    R9 = RT + len(roster)
    put(ws, R9 + 1, 3, '합계', bold=True)
    put(ws, R9 + 1, 7, '=SUM(G%d:G%d)' % (RT + 1, R9), M0, bold=True)
    put(ws, R9 + 1, 9, '=SUM(I%d:I%d)' % (RT + 1, R9), M0, bold=True)
    put(ws, R9 + 1, 10, '=COUNTIF(J%d:J%d,"서명 대기")&"건 대기"' % (RT + 1, R9), bold=True)
    name('순번범위', '입력', '$A$%d:$A$%d' % (RT + 1, R9))
    name('상호범위', '입력', '$C$%d:$C$%d' % (RT + 1, R9))
    name('flow범위', '입력', '$G$%d:$G$%d' % (RT + 1, R9))
    name('b범위', '입력', '$H$%d:$H$%d' % (RT + 1, R9))
    name('하루입력범위', '입력', '$I$%d:$I$%d' % (RT + 1, R9))
    name('계약상태범위', '입력', '$J$%d:$J$%d' % (RT + 1, R9))

    # ── 엑셀 내려받기 프리셋 6종 — 종료일은 전부 기준일에서 끊긴다 ──
    PR = R9 + 3
    head(ws, PR, ['프리셋', '집계 단위', '검색대상기간', '시작(수식)', '종료(수식)', '일수',
                  '화면 시작', '화면 종료', '차', '출처'], fill=SUB)
    pfml = {'week': ('=기준일-6', '=기준일'),
            'month': ('=DATE(YEAR(기준일),MONTH(기준일),1)', '=기준일'),
            'w4': ('=기준일-WEEKDAY(기준일,3)-21', '=기준일'),
            'w12': ('=기준일-WEEKDAY(기준일,3)-77', '=기준일'),
            'm3': ('=DATE(YEAR(기준일),MONTH(기준일)-2,1)', '=기준일'),
            'm6': ('=DATE(YEAR(기준일),MONTH(기준일)-5,1)', '=기준일')}
    for i, (k, gran, lab, frm, to) in enumerate(presets):
        r = PR + 1 + i
        put(ws, r, 1, k)
        put(ws, r, 2, gran)
        put(ws, r, 3, lab, bold=True)
        put(ws, r, 4, pfml[k][0], DT)
        put(ws, r, 5, pfml[k][1], DT)
        put(ws, r, 6, '=E%d-D%d+1' % (r, r), M0)
        put(ws, r, 7, date(*[int(x) for x in frm.split('-')]), DT)
        put(ws, r, 8, date(*[int(x) for x in to.split('-')]), DT)
        put(ws, r, 9, '=(D%d-G%d)+(E%d-H%d)' % (r, r, r, r), M0, fill=CHK)
        put(ws, r, 10, 'build_xlsx.py PRESETS')
    PRL = PR + len(presets)
    put(ws, PRL + 1, 3, '차 합계', bold=True)
    put(ws, PRL + 1, 9, '=SUM(I%d:I%d)' % (PR + 1, PRL), M0, bold=True)
    widths(ws, {'J': 14, 'K': 14})

    # ── 플랫폼 ──────────────────────────────────────────────────
    ws = wb.create_sheet('플랫폼')
    widths(ws, dict({'A': 8, 'B': 14}, **{get_column_letter(3 + i): 6 for i in range(13)}))
    widths(ws, {'P': 12, 'Q': 14, 'R': 15})
    head(ws, 1, ['코드', '플랫폼'] + list(range(1, 14)) + ['도수 합', 'Σ(d x 도수)',
                                                          'Σ(d² x 도수)'])
    for i, p in enumerate(ORDER):
        r = 2 + i
        put(ws, r, 1, p)
        put(ws, r, 2, LABEL[p], bold=True)
        for d in range(1, 14):
            put(ws, r, 2 + d, cyc['freq'][p].get(d, 0), M0, fill=CHK)
        put(ws, r, 16, '=SUM(C%d:O%d)' % (r, r), M0, bold=True)
        put(ws, r, 17, '=SUMPRODUCT($C$1:$O$1,C%d:O%d)' % (r, r), M0)
        put(ws, r, 18, '=SUMPRODUCT($C$1:$O$1,$C$1:$O$1,C%d:O%d)' % (r, r), M0)
    put(ws, 6, 2, '합계', bold=True)
    put(ws, 6, 16, '=SUM(P2:P5)', M0, bold=True)

    head(ws, 8, ['코드', '플랫폼', '구성비', '평균 금융일수 E[d]', 'E[d²]', 'E[d²]/E[d]',
                 '미지급률', '과지급률', '미지급-과지급', '대표 엑셀 평균', '차'], fill=SUB)
    for i, p in enumerate(ORDER):
        r, s = 9 + i, 2 + i
        put(ws, r, 1, p)
        put(ws, r, 2, LABEL[p], bold=True)
        put(ws, r, 3, '=INDEX(구성비범위,%d)' % (i + 1), D8)
        put(ws, r, 4, '=Q%d/P%d' % (s, s), D8)
        put(ws, r, 5, '=R%d/P%d' % (s, s), D8)
        put(ws, r, 6, '=E%d/D%d' % (r, r), D6)
        put(ws, r, 7, '=INDEX(미지급률범위,%d)' % (i + 1), D6)
        put(ws, r, 8, '=INDEX(과지급률범위,%d)' % (i + 1), D6)
        put(ws, r, 9, '=G%d-H%d' % (r, r), D6)
        put(ws, r, 10, cyc['avg'][p], D8)
        put(ws, r, 11, '=D%d-J%d' % (r, r), '0.00E+00', fill=CHK)
    put(ws, 13, 2, '가중', bold=True)
    put(ws, 13, 3, '=SUM(C9:C12)', D8, bold=True)
    put(ws, 13, 4, '=SUMPRODUCT(C9:C12,D9:D12)', D8, bold=True)
    put(ws, 13, 5, '=SUMPRODUCT(C9:C12,E9:E12)', D8, bold=True)
    put(ws, 13, 6, '=E13/D13', D6, bold=True)
    name('평균금융일수범위', '플랫폼', '$D$9:$D$12')
    name('Ed2범위', '플랫폼', '$E$9:$E$12')

    head(ws, 15, ['항목', '값', '산식'], fill=SUB)
    pf = [('구성비 합', '=SUM(구성비범위)', 'SUM(구성비)', D8),
          ('W금융일수 (구성비 x 평균 금융일수)', '=SUMPRODUCT(구성비범위,평균금융일수범위)',
           'Σ 구성비 x 평균 금융일수', D8),
          ('대표 엑셀 MAU 시장 평균 참고값 H41', cyc['wavg'], '%s 정산주기 H41' % cyc['src_xlsx'], D8),
          ('차', '=B17-B18', 'B17 - B18', '0.00E+00'),
          ('가중 E[d²]', '=SUMPRODUCT(구성비범위,Ed2범위)', 'Σ 구성비 x E[d²]', D8),
          ('미회수 이론 W', '=B20/B17', '가중 E[d²] / W', D6),
          ('Ty수익율(%) — 구성비 W 기준', '=할인율퍼센트*연일수/ROUND(B17,SW_자릿수)',
           '할인율 x 365 / 표기 W', P2)]
    for i, (lab, val, f, fmt) in enumerate(pf):
        r = 16 + i
        put(ws, r, 1, lab, bold=True)
        put(ws, r, 2, val, fmt, fill=CHK)
        put(ws, r, 3, f)
    widths(ws, {'A': 30, 'B': 22, 'C': 34})

    # ── 가맹점 ──────────────────────────────────────────────────
    ws = wb.create_sheet('가맹점')
    cols = ['순번', '사업자ID', '상호', '업종', '품목', '계약일', '사용', 'flow 상수',
            '배달 의존도 b', '카드', '배민', '쿠팡이츠', '요기요', '구성비 합',
            '가맹점 w금융일수', '하루 선정산액', '하루 Ai', '투자금액', '비중 raw', '표기 W',
            '가중 미지급률', '가중 과지급률', 'S입금부족율(%)', 'Ty수익율(%)', '구간',
            '비중 내림(pp)', '잔차(pp)', '잔차 순위', '비중 표기(%)', '계약 상태']
    head(ws, 1, cols)
    n = len(roster)
    last = 2 + n                      # 가맹점 합계 행
    RATE_ROW = last + 5              # 산출 블록의 '배율 (방향 A)' 행
    KROW = last + 1                  # 최대잉여법 — 나눠 줄 눈금 수 k
    for i, r0 in enumerate(roster):
        r, src = 2 + i, RT + 1 + i
        put(ws, r, 1, '=입력!A%d' % src)
        put(ws, r, 2, '=입력!B%d' % src)
        put(ws, r, 3, '=입력!C%d' % src, bold=True)
        put(ws, r, 4, '=입력!D%d' % src)
        put(ws, r, 5, '=입력!E%d' % src)
        put(ws, r, 6, '=입력!F%d' % src, DT)
        put(ws, r, 7, '=IF(A%d<=SW_가맹점수,1,0)' % r)
        put(ws, r, 8, '=입력!G%d' % src, M0)
        put(ws, r, 9, '=입력!H%d' % src, D8)
        put(ws, r, 10, '=G{0}*(1-I{0})'.format(r), D6)
        for j in range(3):
            put(ws, r, 11 + j, '=G{0}*I{0}*INDEX(배달배분범위,{1})'.format(r, j + 1), D6)
        put(ws, r, 14, '=SUM(J%d:M%d)' % (r, r), D6)
        put(ws, r, 15, '=' + '+'.join('%s%d*INDEX(평균금융일수범위,%d)' % (get_column_letter(10 + j), r,
                                                                 j + 1) for j in range(4)), D6)
        put(ws, r, 16,
            '=IF(SW_방향="A",G{0}*H{0}*$B${2},G{0}*입력!I{1})'.format(r, src, RATE_ROW), M0)
        put(ws, r, 17, '=P%d*(1-할인율)' % r, M0)
        put(ws, r, 18, '=P{0}*(1-할인율)*O{0}'.format(r), M0)
        put(ws, r, 19, '=IF($R$%d=0,0,R%d/$R$%d)' % (last, r, last), '0.0%')
        put(ws, r, 20, '=IF(G%d=0,"",ROUND(O%d,SW_자릿수))' % (r, r))
        put(ws, r, 21, '=' + '+'.join('%s%d*INDEX(미지급률범위,%d)' % (get_column_letter(10 + j),
                                                                     r, j + 1)
                                      for j in range(4)), D8)
        put(ws, r, 22, '=' + '+'.join('%s%d*INDEX(과지급률범위,%d)' % (get_column_letter(10 + j),
                                                                     r, j + 1)
                                      for j in range(4)), D8)
        put(ws, r, 23, '=IF(G%d=0,"",(U%d-V%d)/(1-할인율)*100)' % (r, r, r), P2)
        put(ws, r, 24, '=IF(G%d=0,"",할인율퍼센트*연일수/T%d)' % (r, r), P2)
        put(ws, r, 25, '=IF(G%d=0,"",IF(P%d>=5000000,"고액",IF(P%d>=2000000,"평범","소액")))'
            % (r, r, r))
        # 비중 표기 — 최대잉여법(roster16_model.ratios 와 같은 규칙).
        #   0.1pp 로 내린 뒤 남는 눈금 k 개를 잔차가 큰 행부터 하나씩 준다. 합이 정확히 100.0 이고
        #   어느 행도 0.1pp 를 넘게 밀리지 않는다.
        put(ws, r, 26, '=IF(G%d=0,"",ROUNDDOWN(S%d*1000,0)/10)' % (r, r), '0.0')
        put(ws, r, 27, '=IF(G%d=0,"",S%d*100-Z%d)' % (r, r, r), D6)
        put(ws, r, 28, '=IF(G%d=0,"",RANK(AA%d,$AA$2:$AA$%d,0)+COUNTIF($AA$2:AA%d,AA%d)-1)'
            % (r, r, last - 1, r, r), M0)
        put(ws, r, 29, '=IF(G%d=0,"",Z%d+IF(AB%d<=$B$%d,0.1,0))' % (r, r, r, KROW), '0.0',
            fill=CHK)
        put(ws, r, 30, '=입력!J%d' % src)
    for col, f in [(7, 'SUM'), (8, 'SUM'), (16, 'SUM'), (17, 'SUM'), (18, 'SUM'), (19, 'SUM')]:
        L = get_column_letter(col)
        put(ws, last, col, '=%s(%s2:%s%d)' % (f, L, L, last - 1),
            '0.0%' if col == 19 else M0, bold=True)
    put(ws, last, 3, '합계', bold=True)
    put(ws, last, 15, '=IF(P%d=0,0,SUMPRODUCT(P2:P%d,O2:O%d)/P%d)' % (last, last - 1, last - 1,
                                                                      last), D8, bold=True)
    put(ws, last, 20, '=ROUND(O%d,SW_자릿수)' % last, bold=True)
    put(ws, last, 26, '=SUM(Z2:Z%d)' % (last - 1), '0.0', bold=True)
    put(ws, last, 29, '=SUM(AC2:AC%d)' % (last - 1), '0.0', bold=True, fill=CHK)
    put(ws, last, 30, '=COUNTIF(AD2:AD%d,"서명 대기")&"건 대기"' % (last - 1), bold=True)
    put(ws, KROW, 1, '나눠 줄 눈금 수 k (0.1pp)', bold=True)
    put(ws, KROW, 2, '=ROUND((100-Z%d)/0.1,0)' % last, M0, fill=CHK)
    put(ws, KROW, 3, '(100 - Σ 비중 내림) / 0.1')

    b0 = last + 2
    head(ws, b0, ['항목', '값', '산식'], fill=SUB)
    gf = 'SUMPRODUCT(G2:G%d,H2:H%d)' % (last - 1, last - 1)
    gfd = 'SUMPRODUCT(G2:G%d,H2:H%d,O2:O%d)' % (last - 1, last - 1, last - 1)
    blk = [('flow 가중 w금융일수 W_flow', '=IF(%s=0,0,%s/%s)' % (gf, gfd, gf),
            'Σ(사용 x flow x w금융일수) / Σ(사용 x flow)', D8),
           ('하루 선정산액 합계 (방향 A)', '=IF(B%d=0,0,투자실행액목표/((1-할인율)*B%d))'
            % (b0 + 1, b0 + 1), '투자실행액 / (1-r) / W_flow', M2),
           ('배율 (방향 A)', '=IF(%s=0,0,B%d/%s)' % (gf, b0 + 2, gf),
            '하루 선정산액 합계 / Σ(사용 x flow)', D8),
           ('하루 선정산액 합계 (적용)', '=P%d' % last, 'Σ 하루 선정산액', M2),
           ('금액가중 W금융일수', '=O%d' % last, 'Σ(하루선정산액 x w금융일수) / Σ 하루선정산액', D8),
           ('투자실행액 (모집단 산식)', '=B%d*(1-할인율)*B%d' % (b0 + 4, b0 + 5),
            '합계 x (1-r) x W', M2),
           ('가맹점 수 (적용)', '=SUM(G2:G%d)' % (last - 1), 'Σ 사용', M0),
           ('가맹점 수 (방향 A 산출)', '=IF(가맹점평균=0,0,B%d/가맹점평균)' % (b0 + 4),
            '하루 선정산액 합계 / 가맹점 평균', '0.0000'),
           ('곳당 평균 하루 선정산액', '=IF(B%d=0,0,B%d/B%d)' % (b0 + 7, b0 + 4, b0 + 7),
            '합계 / 가맹점 수', M2),
           ('순현금', '=순현금', '입력', M0),
           ('투자자산', '=B%d+B%d' % (b0 + 6, b0 + 10), '투자실행액 + 순현금', M2),
           ('투자실행액 비중', '=B%d/B%d' % (b0 + 6, b0 + 11), '투자실행액 / 투자자산', '0.0%'),
           ('순현금 비중', '=B%d/B%d' % (b0 + 10, b0 + 11), '순현금 / 투자자산', '0.0%')]
    for i, (lab, val, f, fmt) in enumerate(blk):
        r = b0 + 1 + i
        put(ws, r, 1, lab, bold=True)
        put(ws, r, 2, val, fmt, fill=CHK)
        put(ws, r, 3, f)
    MB = {lab: b0 + 1 + i for i, (lab, *_ ) in enumerate(blk)}

    c0 = b0 + len(blk) + 2
    head(ws, c0, ['플랫폼', '로스터 금액가중 구성비', '대표 구성비', '차'], fill=SUB)
    for i in range(4):
        r, L = c0 + 1 + i, get_column_letter(10 + i)
        put(ws, r, 1, LABEL[ORDER[i]], bold=True)
        put(ws, r, 2, '=IF($P$%d=0,0,SUMPRODUCT($P$2:$P$%d,%s2:%s%d)/$P$%d)'
            % (last, last - 1, L, L, last - 1, last), D8)
        put(ws, r, 3, '=INDEX(구성비범위,%d)' % (i + 1), D8)
        put(ws, r, 4, '=B%d-C%d' % (r, r), '0.00E+00', fill=CHK)
    MIXROW = c0 + 1                 # 로스터 금액가중 구성비 첫 행 — 가중치 대조 시트가 읽는다
    widths(ws, {'A': 20, 'B': 22, 'C': 20, 'D': 12, 'E': 10, 'F': 12, 'G': 7, 'H': 12,
                'I': 12, 'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 12, 'P': 15,
                'Q': 14, 'R': 14, 'S': 10, 'T': 9, 'U': 12, 'V': 12, 'W': 12, 'X': 12,
                'Y': 8, 'Z': 14, 'AA': 12, 'AB': 10, 'AC': 13, 'AD': 11})

    # ── 채권 ────────────────────────────────────────────────────
    ws = wb.create_sheet('채권')
    ch = ['번호', '가맹점', '플랫폼', '가맹점#', '플랫폼#', '선정산일', '정산예정일',
          '금융일수 Di', '순지급액', 'Ai', '채권매입수수료', '미지급액', '과지급액',
          '차감액', '투자수익 Md-1i', '상환액 Bd-1i', '미회수', '만기 도래', 'Ai x Di',
          'S표본', '조회기간']
    head(ws, 1, ch)
    recs = []
    for j in range(WINDOW):
        for m in range(n):
            for i in range(4):
                recs.append((j, m, i, seqs[(m, i)][j]))
    r = 1
    NP = n * 4                       # 하루치 채권 수 = 가맹점 x 플랫폼
    #   값 셀은 금융일수 Di(H열) 하나뿐이다. 번호·가맹점#·플랫폼#·선정산일은 번호에서 나온다.
    for (j, m, i, d) in recs:
        r += 1
        put(ws, r, 1, '=ROW()-1', M0)
        put(ws, r, 2, '=INDEX(상호범위,D%d)' % r)
        put(ws, r, 3, '=INDEX(플랫폼명범위,E%d)' % r)
        put(ws, r, 4, '=MOD(INT((A%d-1)/4),%d)+1' % (r, n), M0)
        put(ws, r, 5, '=MOD(A%d-1,4)+1' % r, M0)
        put(ws, r, 6, '=축시작+INT((A%d-1)/%d)' % (r, NP), DT)
        put(ws, r, 7, '=F%d+H%d' % (r, r), DT)
        put(ws, r, 8, d)
        put(ws, r, 9, '=ROUND(INDEX(가맹점!$P$2:$P$%d,D%d)*INDEX(가맹점!$J$2:$M$%d,D%d,E%d),0)'
            % (last - 1, r, last - 1, r, r), M0)
        put(ws, r, 10, '=ROUND(I%d*(1-할인율),0)' % r, M0)
        put(ws, r, 11, '=ROUNDDOWN(I%d*할인율,0)' % r, M0)
        put(ws, r, 12, '=ROUND(I%d*INDEX(미지급률범위,E%d),0)' % (r, r), M0)
        put(ws, r, 13, '=ROUND(I%d*INDEX(과지급률범위,E%d),0)' % (r, r), M0)
        put(ws, r, 14, '=MAX(0,L%d-M%d)' % (r, r), M0)
        put(ws, r, 15, '=K%d-N%d' % (r, r), M0)
        put(ws, r, 16, '=I%d-N%d' % (r, r), M0)
        put(ws, r, 17, '=IF(G%d>기준일,1,0)' % r)
        put(ws, r, 18, '=IF(G%d<=기준일,1,0)' % r)
        put(ws, r, 19, '=J%d*H%d' % (r, r), M0)
        put(ws, r, 20, '=IF(AND(F%d>=표본시작,F%d<=표본종료),1,0)' % (r, r))
        put(ws, r, 21, '=IF(AND(G%d>=조회시작,G%d<=조회종료),1,0)' % (r, r))
    NROW = r
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:U%d' % NROW
    widths(ws, {'A': 7, 'B': 18, 'C': 12, 'D': 8, 'E': 8, 'F': 12, 'G': 12, 'H': 10,
                'I': 12, 'J': 12, 'K': 14, 'L': 11, 'M': 11, 'N': 11, 'O': 12, 'P': 12,
                'Q': 8, 'R': 9, 'S': 14, 'T': 8, 'U': 9})
    RNG = {c: '채권!$%s$2:$%s$%d' % (c, c, NROW) for c in
           'FGHIJKLMNOPQRSTU'}

    # ── 일별 ────────────────────────────────────────────────────
    ws = wb.create_sheet('일별')
    #   E열이 화면 값이다 — Md-1i = 채권매입수수료 - max(0, 미지급금 - 과지급금).
    #   F열(차감 제외)은 대조용으로만 둔다.
    head(ws, 1, ['정산예정일', '건수', 'SBd-1 상환액', 'SAd-1 투자실행금',
                 'SMd-1 투자수익 (화면 값)', 'SMd-1 투자수익(차감 제외 · 참고)',
                 'SMRd-1 투자수익율', 'SDd-1 W금융일수', 'ty수익율(%)', 'EC 순현금',
                 'Σ(Ai x Di)', '부족액 차감', '차감/수수료(%)'])
    NDAY = 27
    for k in range(NDAY):
        r = 2 + k
        put(ws, r, 1, '=일별시작+%d' % k, DT)
        put(ws, r, 2, '=COUNTIFS(%s,$A%d)' % (RNG['G'], r), M0)
        put(ws, r, 3, '=SUMIFS(%s,%s,$A%d)' % (RNG['P'], RNG['G'], r), M0)
        put(ws, r, 4, '=SUMIFS(%s,%s,$A%d)' % (RNG['J'], RNG['G'], r), M0)
        put(ws, r, 5, '=SUMIFS(%s,%s,$A%d)' % (RNG['O'], RNG['G'], r), M0)
        put(ws, r, 6, '=SUMIFS(%s,%s,$A%d)' % (RNG['K'], RNG['G'], r), M0)
        put(ws, r, 7, '=IF(D%d=0,0,E%d/D%d)' % (r, r, r), D8)
        put(ws, r, 8, '=IF(D%d=0,0,K%d/D%d)' % (r, r, r), D6)
        put(ws, r, 9, '=IF(H%d=0,0,G%d*연일수/H%d*100)' % (r, r, r), P2)
        put(ws, r, 10, '=순현금', M0)
        put(ws, r, 11, '=SUMIFS(%s,%s,$A%d)' % (RNG['S'], RNG['G'], r), M0)
        put(ws, r, 12, '=SUMIFS(%s,%s,$A%d)' % (RNG['N'], RNG['G'], r), M0)
        put(ws, r, 13, '=IF(F%d=0,0,L%d/F%d*100)' % (r, r, r), P2)
    DL = 1 + NDAY
    put(ws, DL + 1, 1, '합계', bold=True)
    for col in (2, 3, 4, 5, 6, 10, 11, 12):
        L = get_column_letter(col)
        put(ws, DL + 1, col, '=SUM(%s2:%s%d)' % (L, L, DL), M0, bold=True)
    put(ws, DL + 1, 13, '=IF(F%d=0,0,L%d/F%d*100)' % (DL + 1, DL + 1, DL + 1), P2, bold=True)
    ws.freeze_panes = 'A2'
    widths(ws, {'A': 13, 'B': 8, 'C': 16, 'D': 17, 'E': 20, 'F': 24, 'G': 16, 'H': 16,
                'I': 12, 'J': 14, 'K': 16, 'L': 13, 'M': 15})
    DRG = {c: "일별!$%s$2:$%s$%d" % (c, c, DL) for c in 'ABCDEFGHIJKLM'}

    # ── 기간집계 ────────────────────────────────────────────────
    ws = wb.create_sheet('기간집계')
    head(ws, 1, ['항목', '값', '산식', '출처', '화면 반영'])
    crit = '%s,">="&조회시작,%s,"<="&조회종료' % (DRG['A'], DRG['A'])
    #   차감 반영이 정본이다 — 화면(일별 E열 · 수익 카드)이 이 값을 쓴다.
    #   차감 제외 벌은 대조용으로만 남긴다.
    #   ⑤ 는 대표가 회의에서 수식을 새로 써서 다시 주기로 했다(2026-08-31 · A-01).
    #   산식이 실체로 들어 있는 자리는 `입력` 시트 한 칸(C%d)뿐이고 여기는 그 칸을 참조만 한다.
    #   ⑥ 은 그 ⑤ 셀과 ③ 칸을 참조한다. ③ 칸이 비어 있는 동안 `미확정` 을 낸다.
    MARK = {'PSM 투자수익 (정의 · 차감 반영)': '정본',
            'PSM 투자수익 (차감 제외)': '참고',
            'PSMR (정의)': '정본', 'PSMR (차감 제외)': '참고',
            '④ 투자실행금액 대비 ty수익율 (정의)': '정본', '④ (차감 제외)': '참고',
            '⑤ 투자자산 대비 ty수익율 (정의)': '미확정',
            '⑥ 투자실행금액 대비 ty수익율 (정의)': '미확정',
            'PSA 투자실행금': '정본', 'PSD': '정본', 'PSC 순현금 합': '정본',
            '② 상환액': '정본', '부족액 차감 합': '정본'}
    per_rows = [
        ('조회기간 시작', '=조회시작', '입력', DT, ''),
        ('조회기간 종료', '=조회종료', '입력', DT, ''),
        ('조회 일수', '=조회종료-조회시작+1', '종료 - 시작 + 1', M0, ''),
        ('PSA 투자실행금', '=SUMIFS(%s,%s)' % (DRG['D'], crit), 'Σ SAd-1', M0,
         'ceo_definitions.md [2번] 투자 실행금(PSA)'),
        ('PSM 투자수익 (정의 · 차감 반영)', '=SUMIFS(%s,%s)' % (DRG['E'], crit),
         'Σ SMd-1 = Σ(수수료 - max(0,미지급-과지급))', M0,
         'ceo_definitions.md [2번] Md-1i'),
        ('PSM 투자수익 (차감 제외)', '=SUMIFS(%s,%s)' % (DRG['F'], crit), 'Σ 채권매입수수료',
         M0, '현행 daily_ledger.py:298'),
        ('② 상환액', '=SUMIFS(%s,%s)' % (DRG['C'], crit), 'Σ SBd-1', M0,
         'ceo_definitions.md [2번] 이미지의 ②'),
        ('Σ(Ai x Di)', '=SUMIFS(%s,%s)' % (DRG['K'], crit), 'Σ 일별 Σ(Ai x Di)', M0, ''),
        ('PSMR (정의)', '=IF(B5=0,0,B6/B5)', 'PSM / PSA', D8,
         'ceo_definitions.md [2번] PSMR'),
        ('PSMR (차감 제외)', '=IF(B5=0,0,B7/B5)', 'PSM(차감 제외) / PSA', D8, ''),
        ('PSD', '=IF(B5=0,0,B9/B5)', 'Σ(Api x Dpi) / PSA', D6,
         'ceo_definitions.md [2번] PSD'),
        ('④ 투자실행금액 대비 ty수익율 (정의)', '=IF(B12=0,0,B10*연일수/B12*100)',
         'PSMR x 365 / PSD', P2, 'ceo_definitions.md [2번] 이미지의 ④'),
        ('④ (차감 제외)', '=IF(B12=0,0,B11*연일수/B12*100)', 'PSMR(차감 제외) x 365 / PSD',
         P2, ''),
        ('PSC 순현금 합', '=SUMIFS(%s,%s)' % (DRG['J'], crit), 'Σ EC', M0,
         'ceo_definitions.md [2번] PSC'),
        #   ⑤ · ⑥ 의 수식 문자열은 per_rows 를 다 세운 뒤 행 번호를 알고 나서 되돌아 적는다.
        ('⑤ 투자자산 대비 ty수익율 (정의)', '', '', P2,
         'ceo_definitions.md [2번] 이미지의 ⑤ — 대표 재작성 대기 (A-01)'),
        ('⑥ 투자실행금액 대비 ty수익율 (정의)', '', '', P2,
         'ceo_definitions.md [2번] 이미지의 ⑥ — 대표 주석 「계산식 다시 확인해볼것」 '
         '(TP-66 · F-23)'),
        ('부족액 차감 합', '=SUMIFS(%s,%s)' % (DRG['L'], crit),
         'Σ max(0, 미지급금 - 과지급금)', M0, 'ceo_definitions.md [2번] Md-1i'),
        ('차감 / 채권매입수수료(%)', '=IF(B7=0,0,B18/B7*100)', '부족액 차감 합 / Σ 수수료',
         P2, ''),
        ('검산 — 수수료 - 차감 - PSM(정의)', '=B7-B18-B6', '0 이면 정의대로다', M0, ''),
        #   대표 DM 2026-08-31 16:45 의 식을 이 통합문서 값으로 재 본다.
        #   (가) 대표 근사식과 (나) 원식은 분모가 달라 1/(1-할인율) 배만큼 갈린다.
        #   어느 쪽이 정본인지는 대표 확인 대기라 판정하지 않는다. 두 값을 나란히 둔다.
        ('DM 16:45 (가) 대표 근사식(%)', '=(할인율-IF(B5=0,0,B18/B5))*100',
         '(할인율 - Σ max(0, 미지급금-과지급금) / PSA) x 100', D6,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 (나) 원식 PSMR(%)', '=B10*100',
         'PSM / PSA x 100 — 분모가 투자실행액이라 순지급액 기준보다 1/(1-할인율) 배', D6,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 (나) - (가) (%p)', '', '두 식의 차', D8,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 (나) / (가) (배)', '', '실측 갈림 배수', D8,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 1 / (1 - 할인율) (배)', '=1/(1-할인율)', '이론 갈림 배수', D8,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 차감합', '=B18',
         'Σ max(0, 미지급금-과지급금). 0 이면 (가) 가 할인율 그대로다', M0,
         '대표 DM 2026-08-31 16:45'),
        #   차감합이 0 인 자리에서 두 식이 얼마나 갈리는지. 대표 원문이 「같다」고 한 자리다.
        ('DM 16:45 차감합 0 일 때 (가) (%)', '=할인율*100', '할인율', D6,
         '대표 DM 2026-08-31 16:45'),
        ('DM 16:45 차감합 0 일 때 (나) (%)', '=할인율/(1-할인율)*100',
         '할인율 / (1 - 할인율) — 원식은 분모가 순지급액이 아니라 투자실행액이다', D6,
         '대표 DM 2026-08-31 16:45'),
    ]
    #   행 번호가 정해진 뒤에야 쓸 수 있는 수식을 되돌아 적는다.
    PER = {lab: 2 + i for i, (lab, *_) in enumerate(per_rows)}
    #   아래 수식들이 위 리터럴의 B5·B10·B12·B13·B15·B18 을 그대로 쓴다. 자리가 밀리면 죽는다.
    assert (PER['PSA 투자실행금'], PER['PSM 투자수익 (정의 · 차감 반영)'],
            PER['PSM 투자수익 (차감 제외)'], PER['PSMR (정의)'], PER['PSD'],
            PER['④ 투자실행금액 대비 ty수익율 (정의)'], PER['PSC 순현금 합'],
            PER['부족액 차감 합']) == (5, 6, 7, 10, 12, 13, 15, 18), PER
    R5, R6 = (PER['⑤ 투자자산 대비 ty수익율 (정의)'],
              PER['⑥ 투자실행금액 대비 ty수익율 (정의)'])
    RGA, RWN = PER['DM 16:45 (가) 대표 근사식(%)'], PER['DM 16:45 (나) 원식 PSMR(%)']
    RDF, RRT = PER['DM 16:45 (나) - (가) (%p)'], PER['DM 16:45 (나) / (가) (배)']
    LATE = {
        #   ⑤ 는 계산하지 않는다. `입력` 시트의 ⑤ 산식 칸 하나를 가져다 쓴다.
        '⑤ 투자자산 대비 ty수익율 (정의)': ('=입력!C%d' % IR['f5'], '입력 C%d' % IR['f5']),
        #   ⑥ 은 그 ⑤ 셀과 ③ 칸을 실제로 참조한다. ③ 칸이 비면 「미확정」.
        '⑥ 투자실행금액 대비 ty수익율 (정의)': (
            '=IF(COUNT(입력!C%d)=0,"미확정",B%d/입력!C%d*연일수/B%d)'
            % (IR['n3'], PER['④ 투자실행금액 대비 ty수익율 (정의)'], IR['n3'], R5),
            '(④ / ③) x 365 / ⑤ — ③ 은 입력 C%d, ⑤ 는 B%d' % (IR['n3'], R5)),
        'DM 16:45 (나) - (가) (%p)': ('=B%d-B%d' % (RWN, RGA), '(나) - (가)'),
        'DM 16:45 (나) / (가) (배)': ('=IF(B%d=0,0,B%d/B%d)' % (RGA, RWN, RGA),
                                      '(나) / (가)'),
    }
    for lab, (f_, note) in LATE.items():
        i = PER[lab] - 2
        row = list(per_rows[i])
        row[1], row[2] = f_, note
        per_rows[i] = tuple(row)
    for i, (lab, val, f, fmt, src) in enumerate(per_rows):
        r = 2 + i
        put(ws, r, 1, lab, bold=True)
        put(ws, r, 2, val, fmt, fill=CHK)
        put(ws, r, 3, f)
        put(ws, r, 4, src)
        put(ws, r, 5, MARK.get(lab, ''), bold=MARK.get(lab) in ('정본', '미확정'),
            fill=GOOD if MARK.get(lab) == '정본'
            else (CHK if MARK.get(lab) == '미확정' else None))
    b1 = 2 + len(per_rows) + 1
    head(ws, b1, ['항목', '값', '산식', '출처'], fill=SUB)
    bal = [
        ('대상정산금채권 Σ Ai', '=SUM(%s)' % RNG['J'], 'Σ Ai (선정산일 축 전량)', M0, ''),
        ('대상정산금채권 Σ(Ai x Di)', '=SUM(%s)' % RNG['S'], 'Σ Ai x Di', M0, ''),
        ('W금융일수 — 대상정산금채권 전체', '=B%d/B%d' % (b1 + 2, b1 + 1),
         'Σ(Ai x Di) / Σ Ai', D8, 'ceo_definitions.md [1번] w금융일수'),
        # 「만기 도래분」은 우리 분석 조어다. 원문은 정산예정일이 전일자(D-1)인 하루치만 쓰고
        #   기준일까지 누적해 도래한 모집단을 부르는 낱말이 없다. 출처 칸에 조어임을 남긴다.
        ('만기 도래분 Σ Ai', '=SUMPRODUCT(%s,%s)' % (RNG['J'], RNG['R']), 'Σ Ai x 도래', M0,
         '조어 — 원문 대응 낱말 없음'),
        ('만기 도래분 Σ(Ai x Di)', '=SUMPRODUCT(%s,%s)' % (RNG['S'], RNG['R']), '', M0,
         '조어 — 원문 대응 낱말 없음'),
        ('W금융일수 — 만기 도래분만', '=IF(B%d=0,0,B%d/B%d)' % (b1 + 4, b1 + 5, b1 + 4),
         '', D8, '조어 — 원문 대응 낱말 없음'),
        ('미회수 Σ Ai', '=SUMPRODUCT(%s,%s)' % (RNG['J'], RNG['Q']), 'Σ Ai x 미회수', M0,
         'ceo_definitions.md [1번] 투자 실행액'),
        ('미회수 Σ(Ai x Di)', '=SUMPRODUCT(%s,%s)' % (RNG['S'], RNG['Q']), '', M0, ''),
        ('W금융일수 — 미회수 잔량만', '=IF(B%d=0,0,B%d/B%d)' % (b1 + 7, b1 + 8, b1 + 7),
         '', D8, ''),
        ('W금융일수 raw (스위치 ① 적용)',
         '=IF(SW_모집단="미회수 잔량만",B%d,IF(SW_모집단="만기 도래분만",B%d,B%d))'
         % (b1 + 9, b1 + 6, b1 + 3), '스위치 ①', D8, ''),
        ('W금융일수 표기 (스위치 ② 적용)', '=ROUND(B%d,SW_자릿수)' % (b1 + 10), '스위치 ②',
         'General', ''),
        ('Ty수익율(%)', '=할인율퍼센트*연일수/B%d' % (b1 + 11), '할인율 x 365 / 표기 W', P2,
         'ceo_definitions.md [1번] ty수익율'),
        ('투자실행액', '=B%d' % (b1 + 7), '미회수 Σ Ai', M0, ''),
        ('순현금', '=순현금', '입력', M0, 'ceo_definitions.md [1번] 순현금'),
        ('투자자산', '=B%d+B%d' % (b1 + 13, b1 + 14), '투자실행액 + 순현금', M0, ''),
        ('투자실행액 비중', '=B%d/B%d' % (b1 + 13, b1 + 15), '', '0.0%', ''),
        ('순현금 비중', '=B%d/B%d' % (b1 + 14, b1 + 15), '', '0.0%', ''),
        ('S표본 Σ(미지급-과지급)', '=SUMPRODUCT(%s-%s,%s)' % (RNG['L'], RNG['M'], RNG['T']),
         'Σ SLi', M0, 'ceo_definitions.md [1번] SLi'),
        ('S표본 Σ Ai', '=SUMPRODUCT(%s,%s)' % (RNG['J'], RNG['T']), 'Σ SAi', M0, ''),
        ('S입금부족율(%)', '=IF(B%d=0,0,B%d/B%d*100)' % (b1 + 19, b1 + 18, b1 + 19),
         'Σ SLi / Σ SAi', P2, 'ceo_definitions.md [1번] S입금부족율'),
        ('미회수 건수', '=SUM(%s)' % RNG['Q'], '', M0, ''),
        ('채권 건수', '=COUNT(%s)' % RNG['H'], '', M0, ''),
        # 스위치 ① 를 돌리지 않고도 세 모집단의 Ty 를 나란히 본다.
        ('Ty — 대상정산금채권 전체(%)', '=할인율퍼센트*연일수/ROUND(B%d,SW_자릿수)' % (b1 + 3),
         '할인율 x 365 / ROUND(W전체)', P2, ''),
        ('Ty — 만기 도래분만(%)',
         '=IF(B%d=0,0,할인율퍼센트*연일수/ROUND(B%d,SW_자릿수))' % (b1 + 6, b1 + 6),
         '할인율 x 365 / ROUND(W도래)', P2, '조어 — 원문 대응 낱말 없음'),
        ('Ty — 미회수 잔량만(%)',
         '=IF(B%d=0,0,할인율퍼센트*연일수/ROUND(B%d,SW_자릿수))' % (b1 + 9, b1 + 9),
         '할인율 x 365 / ROUND(W미회수)', P2, ''),
        #   대표 DM 2026-08-31 16:27 — ty수익률에 연 환산·일 환산 두 갈래가 있다.
        #   위 `Ty수익율(%)` 이 연 환산이다. 회전수와 일 환산을 그 옆에 세운다.
        ('1년 회전수 (365 / 표기 W)', '=연일수/B%d' % (b1 + 11), '365 / 표기 W', D6,
         '대표 DM 2026-08-31 16:27'),
        ('Ty수익율 — 연 환산(%)', '=B%d' % (b1 + 12), '할인율 x 365 / 표기 W', P2,
         '대표 DM 2026-08-31 16:27'),
        ('Ty수익율 — 일 환산(%)', '=할인율퍼센트/B%d' % (b1 + 11), '할인율 / 표기 W', D6,
         '대표 DM 2026-08-31 16:27'),
        ('검산 — 일 환산 x 365 = 연 환산', '=할인율퍼센트/B%d*연일수-B%d' % (b1 + 11, b1 + 12),
         '0 이면 두 갈래가 같은 값에서 갈린다', D8, '대표 DM 2026-08-31 16:27'),
    ]
    for i, (lab, val, f, fmt, src) in enumerate(bal):
        r = b1 + 1 + i
        put(ws, r, 1, lab, bold=True)
        put(ws, r, 2, val, fmt, fill=CHK)
        put(ws, r, 3, f)
        put(ws, r, 4, src)
    AGG = {lab: b1 + 1 + i for i, (lab, *_) in enumerate(bal)}
    widths(ws, {'A': 38, 'B': 22, 'C': 42, 'D': 40, 'E': 11})

    # ── 가중치 대조 ─────────────────────────────────────────────
    #   이번 미팅의 논점. w금융일수의 가중치가 금액(Ai)이냐 이용자 수(MAU)냐.
    ws = wb.create_sheet('가중치 대조')
    widths(ws, {'A': 20, 'B': 34, 'C': 17, 'D': 17, 'E': 17, 'F': 17, 'G': 15, 'H': 46,
                'I': 28})

    head(ws, 1, ['기준', '항목', '값', '', '', '', '', '출처'])
    wr = [
        ('대표 워드', 'w금융일수 산식', word[0][1], '용어 정의.docx [1번 이미지] 4번 문단'),
        ('대표 워드', 'Ai 정의', word[1][1], '용어 정의.docx [1번 이미지] 5번 문단'),
        ('대표 워드', 'Di 정의', word[2][1], '용어 정의.docx [1번 이미지] 6번 문단'),
        ('대표 워드', '가중치', '금액(Ai)', 'Σ Ai x Di / Σ Ai 의 분모가 금액이다'),
        ('대표 엑셀', '플랫폼 비중 근거', 'MAU 기반 시장 점유율',
         '%s 비중 시트' % cyc['src_xlsx']),
        ('대표 엑셀', '가중치', '이용자 수', '%s 비중 D4:F4' % cyc['src_xlsx']),
        ('대표 엑셀', 'MAU 배분', '배민 %s · 쿠팡이츠 %s · 요기요 %s → x 배달앱/전체, 카드 = 나머지'
         % tuple(format(v, ',') for _, v in cyc['mau']),
         '%s 비중 D4:H5' % cyc['src_xlsx']),
        ('판정', '어느 쪽을 따르는가', '워드를 따른다 (2026-08-31 사용자 결정)',
         '“그럼 워드가 맞아 워드 방식으로 다시 계산해”'),
    ]
    for i, (g, lab, val, s0) in enumerate(wr):
        r = 2 + i
        put(ws, r, 1, g, bold=True)
        put(ws, r, 2, lab, bold=True)
        c = put(ws, r, 3, val, fill=GOOD if g == '판정' else None)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        put(ws, r, 8, s0)
    ws.row_dimensions[2].height = 28

    # 두 기준의 구성비 · W · Ty 를 나란히
    T0 = 2 + len(wr) + 1
    head(ws, T0, ['플랫폼', '평균 금융일수 E[d]', '(가) 금액 실측 (입력)', '(나) 엑셀 MAU',
                  '(다) 로스터 금액가중', '(가) - (나)', '(다) - (가)', '출처'], fill=SUB)
    for i, p in enumerate(ORDER):
        r = T0 + 1 + i
        put(ws, r, 1, LABEL[p], bold=True)
        put(ws, r, 2, '=INDEX(평균금융일수범위,%d)' % (i + 1), D8)
        put(ws, r, 3, '=INDEX(구성비범위,%d)' % (i + 1), D8, fill=CHK)
        put(ws, r, 4, '=1-배달앱비중' if p == 'card'
            else '=배달앱비중*INDEX(MAU배분범위,%d)' % i, D8)
        put(ws, r, 5, '=가맹점!B%d' % (MIXROW + i), D8)
        put(ws, r, 6, '=C%d-D%d' % (r, r), D8, fill=CHK)
        put(ws, r, 7, '=E%d-C%d' % (r, r), '0.00E+00', fill=CHK)
        put(ws, r, 8, '입력 C%d (입력 셀) · %s · 가맹점 B%d'
            % (PL0 + i, '입력 B%d 배달앱/전체' % MAD if p == 'card'
               else '입력 C%d' % (MA0 + i - 1), MIXROW + i))
    TS = T0 + 5
    put(ws, TS, 1, '합', bold=True)
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS, col, '=SUM(%s%d:%s%d)' % (L, T0 + 1, L, T0 + 4), D8, bold=True)
    put(ws, TS + 1, 1, 'W금융일수', bold=True)
    put(ws, TS + 1, 2, 'Σ 구성비 x 평균 금융일수')
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS + 1, col, '=SUMPRODUCT(평균금융일수범위,%s%d:%s%d)' % (L, T0 + 1, L, T0 + 4),
            D6, bold=True, fill=CHK)
    put(ws, TS + 1, 6, '=C%d-D%d' % (TS + 1, TS + 1), D6, fill=CHK)
    put(ws, TS + 1, 7, '=E%d-C%d' % (TS + 1, TS + 1), '0.00E+00', fill=CHK)
    put(ws, TS + 2, 1, 'W금융일수 표기', bold=True)
    put(ws, TS + 2, 2, 'ROUND(W, 스위치 ②)')
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS + 2, col, '=ROUND(%s%d,SW_자릿수)' % (L, TS + 1), '0.00', bold=True)
    put(ws, TS + 3, 1, 'Ty수익율(%)', bold=True)
    put(ws, TS + 3, 2, '할인율 x 365 / 표기 W')
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS + 3, col, '=할인율퍼센트*연일수/%s%d' % (L, TS + 2), P2, bold=True, fill=CHK)
    put(ws, TS + 3, 6, '=C%d-D%d' % (TS + 3, TS + 3), P2, fill=CHK)
    put(ws, TS + 3, 7, '=E%d-C%d' % (TS + 3, TS + 3), P2, fill=CHK)
    put(ws, TS + 4, 1, '미회수 잔량 W', bold=True)
    put(ws, TS + 4, 2, 'Σ 구성비 x E[d²] / W')
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS + 4, col, '=SUMPRODUCT(Ed2범위,%s%d:%s%d)/%s%d'
            % (L, T0 + 1, L, T0 + 4, L, TS + 1), D6, bold=True)
    put(ws, TS + 5, 1, '미회수 잔량 Ty(%)', bold=True)
    put(ws, TS + 5, 2, '할인율 x 365 / ROUND(미회수 W)')
    for col in (3, 4, 5):
        L = get_column_letter(col)
        put(ws, TS + 5, col, '=할인율퍼센트*연일수/ROUND(%s%d,SW_자릿수)' % (L, TS + 4),
            P2, bold=True)
    WROW = dict(mix=T0 + 1, sum=TS, w=TS + 1, wr=TS + 2, ty=TS + 3,
                openw=TS + 4, openty=TS + 5)

    N0 = TS + 7
    head(ws, N0, ['확인', '내용', '', '', '', '', '', '출처'], fill=SUB)
    notes = [
        ('설계 사실', '로스터 8곳의 금액 구성비를 금액 실측(%s)과 같게 맞췄다. '
                      '그래서 (다)와 (가)의 차가 0 이다.' % plat['measured_src'],
         'platform_duration.py BOOK_MIX · daily_ledger.py BOOKROWS'),
        ('확인 문항 1', '엑셀의 플랫폼 비중은 시장 평균 참고값인가, 화면 계산에 쓰는 값인가. '
                        '워드는 금액으로 가중하라 하고 엑셀은 이용자 수로 가중돼 있어 둘이 다르다.',
         '%s 비중 시트 · 용어 정의.docx [1번 이미지] 4~5번 문단' % cyc['src_xlsx']),
        ('확인 문항 2', '배달앱/전체 = 0.35 와 카드 65% 는 매출액 기준인가 이용자 수 기준인가.',
         '%s 비중 H5 · D4:F4' % cyc['src_xlsx']),
    ]
    for i, (k, t, s0) in enumerate(notes):
        r = N0 + 1 + i
        put(ws, r, 1, k, bold=True)
        c = put(ws, r, 2, t)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        put(ws, r, 8, s0)
        ws.row_dimensions[r].height = 30

    # 배달앱/전체 민감도 — (나) 엑셀 MAU 쪽 계산에 붙인다
    V0 = N0 + len(notes) + 2
    head(ws, V0, ['배달앱/전체', '카드', '배민', '쿠팡이츠', '요기요', 'W금융일수',
                  'Ty수익율(%)', '기준선(0.35) 대비', '비고'], fill=SUB)
    vrows = [(0.30, None), (0.35, None), (0.40, None), (None, '=배달앱비중'),
             (None, '=배달앱금액비중')]
    vlab = ['', '', '', '현재 입력 (입력 B%d)' % MAD,
            '금액 실측 배달앱/전체 (입력 C%d)' % MED]
    for i, (v, f) in enumerate(vrows):
        r = V0 + 1 + i
        put(ws, r, 1, v if f is None else f, D8, fill=CHK if f is None else None)
        put(ws, r, 2, '=1-$A%d' % r, D8)
        for j in range(3):
            put(ws, r, 3 + j, '=$A%d*INDEX(MAU배분범위,%d)' % (r, j + 1), D8)
        put(ws, r, 6, '=' + '+'.join('%s%d*INDEX(평균금융일수범위,%d)'
                                     % (get_column_letter(2 + j), r, j + 1) for j in range(4)),
            D6, bold=True)
        put(ws, r, 7, '=할인율퍼센트*연일수/ROUND(F%d,SW_자릿수)' % r, P2, bold=True)
        put(ws, r, 8, '=TEXT(F%d-$F$%d,"0.0000")&"일 · "&TEXT(G%d-$G$%d,"0.00")&"%%p"'
            % (r, V0 + 2, r, V0 + 2))
        put(ws, r, 9, vlab[i])
    VEND = V0 + len(vrows)
    put(ws, VEND + 1, 1, '배달 3사 내부 배분', bold=True)
    put(ws, VEND + 1, 2, '=TEXT(INDEX(MAU배분범위,1),"0.0000")&" : "&'
                         'TEXT(INDEX(MAU배분범위,2),"0.0000")&" : "&'
                         'TEXT(INDEX(MAU배분범위,3),"0.0000")')
    put(ws, VEND + 1, 8, '%s 비중 D5:F5 (MAU 비)' % cyc['src_xlsx'])
    SENS = V0 + 1

    # ── 화면대조 ────────────────────────────────────────────────
    ws = wb.create_sheet('화면대조')
    head(ws, 1, ['구분', '항목', '엑셀 값', '산출 경로', '화면 값', '차이', '판정', '출처'])

    def A(k):
        return '기간집계!B%d' % AGG[k]

    def P(k):
        return '기간집계!B%d' % PER[k]

    def G(k):
        return '가맹점!B%d' % MB[k]

    # ── 블록 0 — 화면 값 원본 (ledger_facts.json) ────────────────
    #   이 블록은 배포된 화면이 실제로 띄우는 값이다. 유휴 비율을 돌린 벌에서도 여기는
    #   화면 값 그대로 둔다 — 모델이 화면과 어디서 갈라지는지가 판정 칸에 드러나야 한다.
    r1_ = float(D(repr(scr_exec / (scr_exec + scr_cash) * 100)).quantize(
        D('0.1'), rounding=ROUND_HALF_UP))
    r2_ = float(D(repr(scr_cash / (scr_exec + scr_cash) * 100)).quantize(
        D('0.1'), rounding=ROUND_HALF_UP))
    ded_pct = float(D(repr(full['ded'] / full['fee'] * 100)).quantize(
        D('0.01'), rounding=ROUND_HALF_UP))
    sv_rows = [
        ('투자실행액', scr_exec, M0, fsrc('투자실행액')),
        ('순현금', scr_cash, M0, fsrc('순현금')),
        ('투자자산', S(fx, '투자자산'), M0, fsrc('투자자산')),
        ('투자실행액 비중(%)', r1_, '0.0', '%s ÷ %s (roster16_model EXEC_SHARE)'
         % (SEED['투자실행액'], SEED['투자자산'])),
        ('순현금 비중(%)', r2_, '0.0', '%s ÷ %s (roster16_model CASH_SHARE)'
         % (SEED['순현금'], SEED['투자자산'])),
        ('W금융일수 raw', float(S(fx, 'W금융일수 raw')), D6, fsrc('W금융일수 raw')),
        ('W금융일수 표기', float(S(fx, 'W금융일수 표기')), '0.00', fsrc('W금융일수 표기')),
        ('Ty수익율(%)', float(S(fx, 'Ty수익율')), P2, fsrc('Ty수익율')),
        ('S입금부족율 raw(%)', float(S(fx, 'S입금부족율 raw')), D6, fsrc('S입금부족율 raw')),
        ('S입금부족율 표기(%)', float(S(fx, 'S입금부족율 표기')), P2, fsrc('S입금부족율 표기')),
        ('할인율(%)', float(S(fx, '할인율(%)')), P2, fsrc('할인율(%)')),
        ('하루 평균 투자실행금', S(fx, '하루 평균 투자실행금'), M0, fsrc('하루 평균 투자실행금')),
        ('원장 일수', S(fx, '원장 일수'), M0, fsrc('원장 일수')),
        ('채권 건수', S(fx, '채권 건수'), M0, fsrc('채권 건수')),
        ('미회수 채권 건수', S(fx, '미회수 채권 건수'), M0, fsrc('미회수 채권 건수')),
        ('채권 Di 최소', S(fx, '채권 Di 범위')[0], M0, fsrc('채권 Di 범위') + '[0]'),
        ('채권 Di 최대', S(fx, '채권 Di 범위')[1], M0, fsrc('채권 Di 범위') + '[1]'),
        ('로스터 건수', len(scr_m), M0, fsrc('가맹점별 값') + ' 길이'),
        ('서명 대기 건수', len(sign), M0, 'roster16_model.py SIGN_PENDING'),
        ('주간 일수', S(fx, '주간 일수'), M0, fsrc('주간 일수')),
        ('주간 PSA', S(fx, '주간 PSA'), M0, fsrc('주간 PSA')),
        ('주간 PSM', S(fx, '주간 PSM'), M0, fsrc('주간 PSM')),
        ('주간 상환액', S(fx, '주간 상환액'), M0, fsrc('주간 상환액')),
        ('주간 채권매입수수료', week['fee'], M0, fsrc('일자별 값') + ' [5] 합'),
        ('주간 부족액 차감', week['ded'], M0, fsrc('일자별 값') + ' [6] 합'),
        ('주간 PSD raw', float(S(fx, '주간 PSD raw')), D6, fsrc('주간 PSD raw')),
        ('주간 PSD 표기', float(S(fx, '주간 PSD 표기')), '0.00', fsrc('주간 PSD 표기')),
        ('주간 PSC', S(fx, '주간 PSC'), M0, fsrc('주간 PSC')),
        ('주간 ④(%)', float(S(fx, '주간 ④')), P2, fsrc('주간 ④')),
        ('주간 ⑤(%)', float(S(fx, '주간 ⑤')), P2, fsrc('주간 ⑤')),
        ('전 구간 PSA', S(fx, '전 구간 PSA'), M0, fsrc('전 구간 PSA')),
        ('전 구간 PSM', S(fx, '전 구간 PSM'), M0, fsrc('전 구간 PSM')),
        ('전 구간 채권매입수수료', full['fee'], M0, fsrc('일자별 값') + ' [5] 합'),
        ('전 구간 부족액 차감', full['ded'], M0, fsrc('일자별 값') + ' [6] 합'),
        ('전 구간 차감/수수료(%)', ded_pct, P2, fsrc('일자별 값') + ' [6] 합 ÷ [5] 합'),
        ('전 구간 PSD raw', round(full['wraw'], 6), D6,
         fsrc('일자별 값') + ' Σ([0] x [2]) ÷ Σ[2]'),
        ('전 구간 PSD 표기', float(S(fx, '전 구간 PSD 표기')), '0.00', fsrc('전 구간 PSD 표기')),
        ('전 구간 PSC', S(fx, '전 구간 PSC'), M0, fsrc('전 구간 PSC')),
        ('전 구간 ④(%)', float(S(fx, '전 구간 ④')), P2, fsrc('전 구간 ④')),
        ('전 구간 ⑤(%)', float(S(fx, '전 구간 ⑤')), P2, fsrc('전 구간 ⑤')),
    ]
    r = 1
    SV, SVF = {}, {}
    for (lab, val, fmt, s0) in sv_rows:
        r += 1
        put(ws, r, 1, '화면 값 원본')
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 4, s0)
        put(ws, r, 5, val, fmt, fill=CHK)
        put(ws, r, 8, 'ledger_facts.json')
        SV[lab] = r
        SVF[lab] = '$E$%d' % r
    for i, m in enumerate(scr_m):
        for key, val, fmt in [('투자금액', m[MCOL['amount']], M0),
                              ('W', float(m[MCOL['w']]), '0.00'),
                              ('Ty수익율(%)', float(m[MCOL['ty']]), P2),
                              ('S입금부족율(%)', float(m[MCOL['s']]), P2),
                              ('비중(%)', float(scr_share[i]), '0.0')]:
            r += 1
            lab = '%s — %s' % (m[MCOL['name']], key)
            put(ws, r, 1, '화면 값 원본')
            put(ws, r, 2, lab, bold=True)
            put(ws, r, 4, fsrc('가맹점별 값') + '[%d]' % i)
            put(ws, r, 5, val, fmt, fill=CHK)
            put(ws, r, 8, 'ledger_facts.json')
            SV[lab] = r
            SVF[lab] = '$E$%d' % r

    def V(k):
        return SVF[k]

    # ── 블록 1 — 불변식 · 엑셀 모델이 구성으로 맞추는 자리 ────────
    r += 1
    head(ws, r, ['구분', '항목', '엑셀 값', '산출 경로', '화면 값', '차이', '판정', '출처'],
         fill=SUB)
    cmp_rows = [
        ('항등식', '좌변 — 투자실행액 (채권 미회수 Σ Ai)', '=' + A('미회수 Σ Ai'),
         '기간집계 B%d' % AGG['미회수 Σ Ai'], '=' + G('투자실행액 (모집단 산식)'),
         '우변 — 하루 선정산액 합계 x (1-r) x W. 채권 시트의 미회수 건수는 기본 조합에서 '
         '푼 정수라 스위치 ③④를 바꾸면 좌변이 그만큼 어긋난다.', M0, 0.5),
        ('불변식', '투자실행액', '=' + A('투자실행액'), '기간집계 B%d' % AGG['투자실행액'],
         '=' + V('투자실행액'), '화면대조 E%d' % SV['투자실행액'], M0, 0.5),
        ('불변식', '순현금', '=' + A('순현금'), '기간집계 B%d' % AGG['순현금'],
         '=' + V('순현금'), '화면대조 E%d' % SV['순현금'], M0, 0.5),
        ('불변식', '투자자산', '=' + A('투자자산'), '기간집계 B%d' % AGG['투자자산'],
         '=' + V('투자자산'), '화면대조 E%d' % SV['투자자산'], M0, 0.5),
        ('불변식', '투자실행액 비중(%)', '=ROUND(' + A('투자실행액 비중') + '*100,1)',
         '기간집계 B%d' % AGG['투자실행액 비중'], '=' + V('투자실행액 비중(%)'),
         '화면대조 E%d' % SV['투자실행액 비중(%)'], '0.0', 0.005),
        ('불변식', '순현금 비중(%)', '=ROUND(' + A('순현금 비중') + '*100,1)',
         '기간집계 B%d' % AGG['순현금 비중'], '=' + V('순현금 비중(%)'),
         '화면대조 E%d' % SV['순현금 비중(%)'], '0.0', 0.005),
        ('불변식', 'W금융일수 (표기)', '=' + A('W금융일수 표기 (스위치 ② 적용)'),
         '기간집계 B%d' % AGG['W금융일수 표기 (스위치 ② 적용)'], '=' + V('W금융일수 표기'),
         '화면대조 E%d' % SV['W금융일수 표기'], '0.00', 0.0005),
        ('불변식', 'W금융일수 (raw · 소수 6자리)',
         '=ROUND(' + A('W금융일수 raw (스위치 ① 적용)') + ',6)',
         '기간집계 B%d' % AGG['W금융일수 raw (스위치 ① 적용)'], '=' + V('W금융일수 raw'),
         '화면대조 E%d' % SV['W금융일수 raw'], D6, 0.0000005),
        ('불변식', 'Ty수익율(%)', '=' + A('Ty수익율(%)'), '기간집계 B%d' % AGG['Ty수익율(%)'],
         '=' + V('Ty수익율(%)'), '화면대조 E%d' % SV['Ty수익율(%)'], P2, 0.005),
        ('불변식', '로스터 건수', '=' + G('가맹점 수 (적용)'),
         '가맹점 B%d' % MB['가맹점 수 (적용)'], '=' + V('로스터 건수'),
         '화면대조 E%d' % SV['로스터 건수'], M0, 0.5),
        ('불변식', '서명 대기 건수', '=COUNTIF(계약상태범위,"서명 대기")',
         '입력 J열', '=' + V('서명 대기 건수'), '화면대조 E%d' % SV['서명 대기 건수'], M0, 0.5),
        ('불변식', '비중 합(%)', '=SUM(가맹점!AC2:AC%d)' % (last - 1),
         '가맹점 AC열 (최대잉여법)', 100.0, '비중 합 100.0 불변', '0.0', 0.005),
        ('불변식', '채권 Di 최소', '=MIN(%s)' % RNG['H'], '채권 H열', '=' + V('채권 Di 최소'),
         '화면대조 E%d' % SV['채권 Di 최소'], M0, 0.5),
        ('불변식', '채권 Di 최대', '=MAX(%s)' % RNG['H'], '채권 H열', '=' + V('채권 Di 최대'),
         '화면대조 E%d' % SV['채권 Di 최대'], M0, 0.5),
        ('불변식', '플랫폼 구성비 합', '=SUM(구성비범위)',
         '입력 C%d:C%d' % (PL0, PLS - 1), 1.0,
         '합 1 (구성상)', D8, 0.0000001),
        ('교차', '미회수 잔량 W — 채권 vs 도수표',
         '=' + A('W금융일수 — 미회수 잔량만'),
         '기간집계 B%d' % AGG['W금융일수 — 미회수 잔량만'], '=플랫폼!B21',
         '플랫폼 B21 = 가중 E[d²] / W', D6, 0.0005),
        ('교차', '미회수 잔량 Ty(%)', '=' + A('Ty — 미회수 잔량만(%)'),
         '기간집계 B%d' % AGG['Ty — 미회수 잔량만(%)'],
         "='가중치 대조'!C%d" % WROW['openty'], '가중치 대조 C%d' % WROW['openty'], P2, 0.02),
        ('교차', 'W금융일수 — 구성비 x 평균 금융일수', '=플랫폼!B17', '플랫폼 B17',
         '=' + V('W금융일수 raw'), '화면대조 E%d' % SV['W금융일수 raw'], D6, 0.0000005),
        ('교차', '로스터 금액가중 구성비 — 카드', '=가맹점!B%d' % MIXROW,
         '가맹점 B%d' % MIXROW, '=INDEX(구성비범위,1)', '입력 C%d' % PL0, D8, 0.0000001),
        ('불변식', 'S입금부족율(%)', '=ROUND(' + A('S입금부족율(%)') + ',2)',
         '기간집계 B%d' % AGG['S입금부족율(%)'], '=' + V('S입금부족율 표기(%)'),
         '화면대조 E%d' % SV['S입금부족율 표기(%)'], P2, 0.005),
        #   입력 두 칸(총 투자자산 · 유휴 비율)에서 갈라진 목표를 채권 원장이 그대로 맞추는가.
        #   Di 를 다시 풀지 않고 유휴 비율만 손으로 바꾸면 여기서 원 단위 잔차가 드러난다.
        ('불변식', '투자실행액 = 입력 목표', '=' + A('투자실행액'),
         '기간집계 B%d' % AGG['투자실행액'], '=투자실행액목표',
         '입력 C%d = C%d - C%d' % (IR['exec'], IR['total'], IR['cash']), M0, 0.5),
        ('불변식', '투자자산 = 총 투자자산', '=' + A('투자자산'),
         '기간집계 B%d' % AGG['투자자산'], '=총투자자산',
         '입력 C%d' % IR['total'], M0, 0.5),
    ]
    #   가맹점 행의 W·S 는 구성비에서 바로 나온다 — 원장 틸트를 타지 않아 화면과 같은 값이다.
    for i, m in enumerate(scr_m):
        nm = m[MCOL['name']]
        cmp_rows.append(('가맹점', '%s — W 표기' % nm, '=가맹점!T%d' % (2 + i),
                         '가맹점 T%d' % (2 + i), '=' + V('%s — W' % nm),
                         '화면대조 E%d' % SV['%s — W' % nm], '0.00', 0.0005))
    for i, m in enumerate(scr_m):
        nm = m[MCOL['name']]
        cmp_rows.append(('가맹점', '%s — S입금부족율(%%)' % nm,
                         '=ROUND(가맹점!W%d,2)' % (2 + i), '가맹점 W%d' % (2 + i),
                         '=' + V('%s — S입금부족율(%%)' % nm),
                         '화면대조 E%d' % SV['%s — S입금부족율(%%)' % nm], P2, 0.005))
    for (g, lab, val, path, scr, s0, fmt, tol) in cmp_rows:
        r += 1
        put(ws, r, 1, g)
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 3, val, fmt, fill=CHK)
        put(ws, r, 4, path)
        put(ws, r, 5, scr, fmt)
        put(ws, r, 6, '=IFERROR(C%d-E%d,"")' % (r, r), fmt)
        put(ws, r, 7, '=IFERROR(IF(ABS(F%d)<=%s,"일치","차이"),"")' % (r, repr(tol)))
        put(ws, r, 8, s0)

    # ── 블록 2 — 화면 값 자체 정합 (화면 값끼리 산식으로 되짚는다) ─
    r += 1
    head(ws, r, ['구분', '항목', '산식으로 되짚은 값', '산식', '화면 값', '차이', '판정',
                 '출처'], fill=SUB)
    self_rows = [
        ('투자자산', '=%s+%s' % (V('투자실행액'), V('순현금')), '투자실행액 + 순현금',
         V('투자자산'), M0, 0.5, 'ceo_definitions.md [1번] 비중'),
        ('투자실행액 비중(%)', '=ROUND(%s/(%s+%s)*100,1)'
         % (V('투자실행액'), V('투자실행액'), V('순현금')),
         '투자실행액 / (투자실행액 + 순현금)', V('투자실행액 비중(%)'), '0.0', 0.005,
         'ceo_definitions.md [1번] 투자 실행액의 비중'),
        ('순현금 비중(%)', '=ROUND(%s/(%s+%s)*100,1)'
         % (V('순현금'), V('투자실행액'), V('순현금')), '순현금 / (투자실행액 + 순현금)',
         V('순현금 비중(%)'), '0.0', 0.005, 'ceo_definitions.md [1번] 순현금의 비중'),
        ('W금융일수 표기', '=ROUND(%s,2)' % V('W금융일수 raw'), 'ROUND(W raw, 2)',
         V('W금융일수 표기'), '0.00', 0.0005, '9-C 표기 규칙'),
        ('Ty수익율(%)', '=ROUND(%s*연일수/%s,2)' % (V('할인율(%)'), V('W금융일수 표기')),
         '할인율 x 365 / 표기 W', V('Ty수익율(%)'), P2, 0.005,
         'ceo_definitions.md [1번] ty수익율'),
        ('S입금부족율 표기(%)', '=ROUND(%s,2)' % V('S입금부족율 raw(%)'), 'ROUND(S raw, 2)',
         V('S입금부족율 표기(%)'), P2, 0.005, '9-C 표기 규칙'),
        ('주간 PSM = 수수료 - 차감', '=%s-%s' % (V('주간 채권매입수수료'), V('주간 부족액 차감')),
         'Σ 채권매입수수료 - Σ max(0, 미지급-과지급)', V('주간 PSM'), M0, 0.5,
         'ceo_definitions.md [2번] Md-1i'),
        ('주간 상환액 - PSM = PSA', '=%s-%s' % (V('주간 상환액'), V('주간 PSM')),
         'Σ Bd-1i - Σ Md-1i = Σ Ad-1i', V('주간 PSA'), M0, 0.5,
         'ceo_definitions.md [2번] Bd-1i · Md-1i · Ad-1i'),
        ('주간 PSC', '=%s*%s' % (V('순현금'), V('주간 일수')), '순현금 x 조회 일수',
         V('주간 PSC'), M0, 0.5, 'ceo_definitions.md [2번] PSC'),
        ('주간 PSD 표기', '=ROUND(%s,2)' % V('주간 PSD raw'), 'ROUND(PSD raw, 2)',
         V('주간 PSD 표기'), '0.00', 0.0005, '9-C 표기 규칙'),
        ('주간 ④(%)', '=ROUND(%s/%s*100*연일수/%s,2)'
         % (V('주간 PSM'), V('주간 PSA'), V('주간 PSD raw')), 'PSMR x 365 / PSD',
         V('주간 ④(%)'), P2, 0.005, 'ceo_definitions.md [2번] 이미지의 ④'),
        ('전 구간 PSM = 수수료 - 차감',
         '=%s-%s' % (V('전 구간 채권매입수수료'), V('전 구간 부족액 차감')),
         'Σ 채권매입수수료 - Σ 차감', V('전 구간 PSM'), M0, 0.5,
         'ceo_definitions.md [2번] Md-1i'),
        ('전 구간 차감/수수료(%)', '=ROUND(%s/%s*100,2)'
         % (V('전 구간 부족액 차감'), V('전 구간 채권매입수수료')), 'Σ 차감 / Σ 수수료',
         V('전 구간 차감/수수료(%)'), P2, 0.005, 'ceo_definitions.md [2번] Md-1i'),
        ('전 구간 PSC', '=%s*%s' % (V('순현금'), V('원장 일수')), '순현금 x 원장 일수',
         V('전 구간 PSC'), M0, 0.5, 'ceo_definitions.md [2번] PSC'),
        ('전 구간 PSD 표기', '=ROUND(%s,2)' % V('전 구간 PSD raw'), 'ROUND(PSD raw, 2)',
         V('전 구간 PSD 표기'), '0.00', 0.0005, '9-C 표기 규칙'),
        ('전 구간 ④(%)', '=ROUND(%s/%s*100*연일수/%s,2)'
         % (V('전 구간 PSM'), V('전 구간 PSA'), V('전 구간 PSD raw')), 'PSMR x 365 / PSD',
         V('전 구간 ④(%)'), P2, 0.005, 'ceo_definitions.md [2번] 이미지의 ④'),
        ('하루 평균 투자실행금', '=ROUND(%s/%s,0)' % (V('전 구간 PSA'), V('원장 일수')),
         '전 구간 PSA / 원장 일수', V('하루 평균 투자실행금'), M0, 0.5,
         'roster16_model.py DAY_AVG'),
        ('가맹점 투자금액 합', '=' + '+'.join(V('%s — 투자금액' % m[MCOL['name']])
                                              for m in scr_m),
         'Σ 가맹점별 투자금액', V('투자실행액'), M0, 0.5, 'roster16_model.py EXEC'),
        ('가맹점 비중 합(%)', '=' + '+'.join(V('%s — 비중(%%)' % m[MCOL['name']])
                                             for m in scr_m),
         'Σ 가맹점별 비중 (최대잉여법)', 100.0, '0.0', 0.005, 'roster16_model.py ratios'),
    ]
    for i, m in enumerate(scr_m):
        self_rows.append(
            ('%s — Ty = 할인율 x 365 / W' % m[MCOL['name']],
             '=ROUND(%s*연일수/%s,2)' % (V('할인율(%)'), V('%s — W' % m[MCOL['name']])),
             '할인율 x 365 / 표기 W', V('%s — Ty수익율(%%)' % m[MCOL['name']]), P2, 0.005,
             'ceo_definitions.md [1번] ty수익율'))
    for (lab, val, f, ref, fmt, tol, s0) in self_rows:
        r += 1
        put(ws, r, 1, '화면 정합')
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 3, val, fmt, fill=CHK)
        put(ws, r, 4, f)
        put(ws, r, 5, ('=' + ref) if isinstance(ref, str) else ref, fmt)
        put(ws, r, 6, '=IFERROR(C%d-E%d,"")' % (r, r), fmt)
        put(ws, r, 7, '=IFERROR(IF(ABS(F%d)<=%s,"일치","차이"),"")' % (r, repr(tol)))
        put(ws, r, 8, s0)
    LASTC = r

    # ── 블록 3 — 모델 재현 잔차 (0 이 아닌 자리와 그 이유) ────────
    r += 2
    head(ws, r, ['구분', '항목', '엑셀 값', '산출 경로', '화면 값', '차이', '잔차율(%)',
                 '사유'], fill=SUB)
    TILT = ('화면 원장은 요일·주차 규모 계수(SIZE)와 배달 의존도 틸트(TILT)로 하루치가 흔들리고, '
            '미회수 앵커가 기준일에 걸린다. 이 통합문서는 하루치가 평평한 40일 창 모델이다.')
    ROUNDUP = ('화면 원장은 채권을 정산예정일 버킷까지 쪼개 건당 금액이 작다. 미지급·과지급을 원 단위로 '
               '반올림하면 작은 건에서 차가 커진다. 이 통합문서는 (가맹점 x 플랫폼 x 선정산일) '
               '한 건이라 반올림 영향이 거의 없다.')
    SCOPE = ('모집단이 다르다. 이 통합문서는 선정산일 40일 x %d곳 x 4플랫폼 = %d건, '
             '화면 원장은 193일 x %d곳 x 4플랫폼 x 정산예정일 버킷 = %s건.'
             % (n, n * 4 * WINDOW, n, format(S(fx, '채권 건수'), ',')))
    res_rows = [
        ('하루 평균 투자실행금', '=' + G('하루 선정산액 합계 (적용)') + '*(1-할인율)',
         '가맹점 B%d x (1-r)' % MB['하루 선정산액 합계 (적용)'], V('하루 평균 투자실행금'),
         M0, TILT),
        ('조회기간 PSA', '=' + P('PSA 투자실행금'), '기간집계 B%d' % PER['PSA 투자실행금'],
         V('주간 PSA'), M0, TILT),
        ('조회기간 PSD', '=' + P('PSD'), '기간집계 B%d' % PER['PSD'], V('주간 PSD raw'),
         D6, TILT),
        ('조회기간 PSM (차감 반영)', '=' + P('PSM 투자수익 (정의 · 차감 반영)'),
         '기간집계 B%d' % PER['PSM 투자수익 (정의 · 차감 반영)'], V('주간 PSM'), M0,
         TILT + ' ' + ROUNDUP),
        ('조회기간 ④(%)', '=' + P('④ 투자실행금액 대비 ty수익율 (정의)'),
         '기간집계 B%d' % PER['④ 투자실행금액 대비 ty수익율 (정의)'], V('주간 ④(%)'), P2,
         TILT + ' ' + ROUNDUP),
        ('조회기간 ⑤(%)', '=' + P('⑤ 투자자산 대비 ty수익율 (정의)'),
         '기간집계 B%d' % PER['⑤ 투자자산 대비 ty수익율 (정의)'], V('주간 ⑤(%)'), P2,
         TILT + ' ' + ROUNDUP),
        ('조회기간 PSC', '=' + P('PSC 순현금 합'), '기간집계 B%d' % PER['PSC 순현금 합'],
         V('주간 PSC'), M0, '조회 일수가 같으면 차가 0 이다.'),
        ('차감 / 수수료(%)', '=' + P('차감 / 채권매입수수료(%)'),
         '기간집계 B%d' % PER['차감 / 채권매입수수료(%)'], V('전 구간 차감/수수료(%)'), P2,
         ROUNDUP),
        #   표기(ROUND 2자리)로 보면 0.07 = 0.07 이라 차가 가려진다. raw 로 세워 둔다.
        ('S입금부족율 raw(%)', '=' + A('S입금부족율(%)'),
         '기간집계 B%d' % AGG['S입금부족율(%)'], V('S입금부족율 raw(%)'), D6,
         TILT + ' ' + ROUNDUP),
        ('채권 건수', '=' + A('채권 건수'), '기간집계 B%d' % AGG['채권 건수'],
         V('채권 건수'), M0, SCOPE),
        ('미회수 건수', '=' + A('미회수 건수'), '기간집계 B%d' % AGG['미회수 건수'],
         V('미회수 채권 건수'), M0, SCOPE),
    ]
    for i, m in enumerate(scr_m):
        nm = m[MCOL['name']]
        res_rows.append(('%s — 투자금액' % nm, '=가맹점!R%d' % (2 + i),
                         '가맹점 R%d' % (2 + i), V('%s — 투자금액' % nm), M0, TILT))
    for i, m in enumerate(scr_m):
        nm = m[MCOL['name']]
        res_rows.append(('%s — 비중 표기(%%)' % nm, '=가맹점!AC%d' % (2 + i),
                         '가맹점 AC%d' % (2 + i), V('%s — 비중(%%)' % nm), '0.0', TILT))
    RES0 = r + 1
    for (lab, val, path, ref, fmt, why) in res_rows:
        r += 1
        put(ws, r, 1, '모델 잔차')
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 3, val, fmt, fill=CHK)
        put(ws, r, 4, path)
        put(ws, r, 5, '=' + ref, fmt)
        put(ws, r, 6, '=IFERROR(C%d-E%d,"")' % (r, r), fmt)
        put(ws, r, 7, '=IFERROR(IF(E%d=0,"",F%d/E%d*100),"")' % (r, r, r), P2)
        c = put(ws, r, 8, why)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    RESL = r

    ws.conditional_formatting.add('G2:G%d' % LASTC,
                                  CellIsRule(operator='equal', formula=['"차이"'], fill=BAD))
    ws.conditional_formatting.add('G2:G%d' % LASTC,
                                  CellIsRule(operator='equal', formula=['"일치"'], fill=GOOD))
    ws.freeze_panes = 'A2'
    widths(ws, {'A': 12, 'B': 36, 'C': 20, 'D': 34, 'E': 20, 'F': 16, 'G': 10, 'H': 60})

    # ── 산식 ────────────────────────────────────────────────────
    ws = wb.create_sheet('산식')
    head(ws, 1, ['출처 절', '대표 정의서 원문', '이 통합문서의 셀'])
    MAPS = [
        ('투자 실행액', '기간집계 B%d · 화면대조 %d행' % (AGG['투자실행액'], SV['투자실행액'])),
        ('유동화투자자의 할인율', '입력 C%d · C%d' % (IR['rate'], IR['ratep'])),
        ('순현금', '입력 C%d · 기간집계 B%d' % (IR['cash'], AGG['순현금'])),
        ('w금융일수 =', '기간집계 B%d · 가중치 대조 C%d — 가중치는 금액(Ai)'
         % (AGG['W금융일수 — 대상정산금채권 전체'], WROW['w'])),
        ('Ai =', '채권 J열 — w금융일수의 가중치'),
        ('Di =', '채권 H열'),
        ('대상정산금채권:', '채권 시트 전 행 (선정산일 축 %d일)' % WINDOW),
        ('각 정산금채권의 ID', '채권 D·E·F열'),
        ('플랫폼ID', '입력 A%d:A%d' % (PL0, PLS - 1)),
        #   원문은 `D` 를 현재일자로 쓴다. 이 통합문서는 `d` 로 쓴다 —
        #   `D` 는 금융일수 (dm_0831/symbol_rule_0831.md · 이서준 지시 2026-08-31).
        ('D = 현재일자', '입력 C%d — 이 통합문서 표기는 `d`. `D` 는 금융일수 (채권 H열)'
         % IR['asof']),
        ('비중', '기간집계 B%d · B%d' % (AGG['투자실행액 비중'], AGG['순현금 비중'])),
        ('매일 자정일 지나면', '일별 시트 전 행'),
        ('상기 배치작업이 완료된 후', '기간집계 2~%d행' % (1 + len(per_rows))),
        ('금융일수 =', '채권 H열 (정산주기.xlsx N6:Q370 실측)'),
        ('S입금부족율', '기간집계 B%d' % AGG['S입금부족율(%)']),
        ('SLi =', '채권 L열 - M열'),
        ('표본집합:', '채권 T열 · 입력 C%d:C%d' % (IR['sfrom'], IR['sto'])),
        ('SAi =', '채권 J열'),
        ('ty수익율 =', '기간집계 B%d' % AGG['Ty수익율(%)']),
        ('투자 실행액의 비중', '기간집계 B%d' % AGG['투자실행액 비중']),
        ('순현금의 비중', '기간집계 B%d' % AGG['순현금 비중']),
        ('전일자(D-1) 대상정산금채권의 상환액', '일별 C열'),
        ('Σ BD-1i', '채권 P열 = 순지급액 - 차감액'),
        ('전일자(D-1) 대상정산금채권의 투자실행금', '일별 D열'),
        ('Σ AD-1i', '채권 J열'),
        ('전일자(D-1) 대상정산금채권의 투자수익(SMD-1)', '일별 E열'),
        ('MD-1i =', '채권 O열 = 채권매입수수료 - 차감액'),
        ('전일자(D-1) 대상정산금채권의 투자수익율', '일별 G열'),
        ('전일자(D-1) 대상정산금채권의 w금융일수', '일별 H열'),
        ('AD-1i =', '채권 J열'),
        ('DD-1i =', '채권 H열'),
        ('전일자(D-1) 대상정산금채권의 ty수익율', '일별 I열'),
        ('전일자(D-1) 순현금 (EC)', '일별 J열'),
        ('투자 실행금(PSA)', '기간집계 B%d' % PER['PSA 투자실행금']),
        ('투자수익(PSM)', '기간집계 B%d · B%d'
         % (PER['PSM 투자수익 (정의 · 차감 반영)'], PER['PSM 투자수익 (차감 제외)'])),
        ('투자실행금액 대비 ty수익율(이미지의 ④)', '기간집계 B%d · B%d'
         % (PER['④ 투자실행금액 대비 ty수익율 (정의)'], PER['④ (차감 제외)'])),
        ('PSMR =', '기간집계 B%d · B%d' % (PER['PSMR (정의)'], PER['PSMR (차감 제외)'])),
        ('PSD =', '기간집계 B%d' % PER['PSD']),
        ('Api :', '채권 J열 x U열'),
        ('Dpi:', '채권 H열'),
        ('투자자산 대비 ty수익율 (이미지의 ⑤)',
         '기간집계 B%d ← 입력 C%d (⑤ 산식 칸 · 미확정 · 대표 재작성 대기 A-01). '
         '통합문서 안에서 ⑤ 를 계산하는 자리는 그 한 칸뿐이다'
         % (PER['⑤ 투자자산 대비 ty수익율 (정의)'], IR['f5'])),
        ('PSC =', '기간집계 B%d' % PER['PSC 순현금 합']),
        ('상환액 (이미지의 ②)', '기간집계 B%d · 일별 C열' % PER['② 상환액']),
        #   ⑥ 은 ④·⑤ 를 실제로 참조한다. ③ 만 지시 대상이 정해지지 않아 입력 칸을 비워 두었고,
        #   그동안 ⑥ 셀은 「미확정」을 낸다(레지스터 TP-66 · 확인 문항 F-23 · U-03).
        ('투자실행금액 대비 ty수익율 (이미지의 ⑥)',
         '기간집계 B%d = (B%d / 입력 C%d) x 365 / B%d — ③ 지시 대상 미확정이라 입력 C%d 가 '
         '비어 있고 그동안 「미확정」. 대표 주석 「계산식 다시 확인해볼것」'
         % (PER['⑥ 투자실행금액 대비 ty수익율 (정의)'],
            PER['④ 투자실행금액 대비 ty수익율 (정의)'], IR['n3'],
            PER['⑤ 투자자산 대비 ty수익율 (정의)'], IR['n3'])),
    ]
    r = 1
    for sec, line in defs:
        r += 1
        put(ws, r, 1, sec)
        put(ws, r, 2, line)
        cell = ''
        for key, ref in MAPS:
            if line.startswith(key):
                cell = ref
                break
        put(ws, r, 3, cell, fill=CHK if cell else None)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical='top')
    widths(ws, {'A': 16, 'B': 86, 'C': 40})
    ws.freeze_panes = 'A2'

    # ── 스위치 산출 위치 — 기간집계 행이 정해진 뒤 되돌아 적는다 ──
    _in = wb['입력']
    put(_in, SWH + 1, 4, '기간집계 B%d' % AGG['W금융일수 raw (스위치 ① 적용)'])
    put(_in, SWH + 2, 4, '기간집계 B%d' % AGG['W금융일수 표기 (스위치 ② 적용)'])

    # ── ⑤ 산식 되돌아 적기 — 통합문서에서 ⑤ 를 계산하는 유일한 칸 ──
    #   지금 들어 있는 것은 대표 정의서 원문 산식 `(④ x PSA) / (PSA + PSC)` 다.
    #   대표가 새 수식을 주면 이 한 칸만 갈아 끼운다. 기간집계 ⑤ 와 ⑥ 이 따라 움직인다.
    F5 = ('=IF(기간집계!B%d+기간집계!B%d=0,0,'
          '기간집계!B%d*기간집계!B%d/(기간집계!B%d+기간집계!B%d))'
          % (PER['PSA 투자실행금'], PER['PSC 순현금 합'],
             PER['④ 투자실행금액 대비 ty수익율 (정의)'], PER['PSA 투자실행금'],
             PER['PSA 투자실행금'], PER['PSC 순현금 합']))
    put(_in, IR['f5'], 3, F5, P2, fill=CHK)

    # ── 읽는 법 ─────────────────────────────────────────────────
    ws = wb.create_sheet('읽는 법')
    widths(ws, {'A': 6, 'B': 18, 'C': 66, 'D': 56, 'E': 18})
    RG = 'B%d' % AGG['W금융일수 raw (스위치 ① 적용)']
    RW = 'B%d' % AGG['W금융일수 표기 (스위치 ② 적용)']
    guide = [
        ('읽는 법', '시트 차례 · 화면 값 되짚는 경로 3줄 · 시뮬레이션 입력 칸 · '
                     '기호 규칙 · 대표 DM 2026-08-31',
         '이 시트'),
        ('입력', '가정값 %d칸 · 스위치 4칸 · 플랫폼 구성비 · 금액 실측 · 대표 엑셀 MAU 참고값 · '
                 '로스터 %d곳 · 엑셀 프리셋 %d종' % (len(rows), len(roster), len(presets)),
         'C%d:C%d · B%d:B%d · C%d:C%d · A%d:K%d'
         % (IR['rate'], IR['sto'], SWH + 1, SWH + 4, PL0, PLS - 1, RT + 1, R9)),
        ('플랫폼', '금융일수 도수분포(1~13일) · 평균 금융일수 · 구성비 가중 W금융일수',
         'C2:O5 · D9:D12 · B17'),
        ('가맹점', '로스터 %d곳의 플랫폼 배분 · 하루 선정산액 · 투자금액 · W · S · Ty · '
                   '비중(최대잉여법)' % len(roster),
         'J2:M%d · O2:O%d · P2:P%d · AC2:AC%d' % (last - 1, last - 1, last - 1, last - 1)),
        ('채권', '(가맹점 x 플랫폼 x 선정산일 %d일) %s건 원장 — 값 셀은 금융일수 Di 하나뿐'
                 % (WINDOW, format(NROW - 1, ',')),
         'H열 Di · J열 Ai · Q열 미회수 · S열 Ai x Di'),
        ('일별', '정산예정일 %d일치 집계 — 상환액 · 투자실행금 · 투자수익 · W · ty수익율'
                 % NDAY, 'C2:E%d · H2:H%d · I2:I%d' % (DL, DL, DL)),
        ('기간집계', '조회기간 집계 2~%d행 · 모집단별 W·Ty·S %d~%d행'
                     % (1 + len(per_rows), b1 + 1, b1 + len(bal)),
         'B%d · B%d · B%d · %s · %s · B%d'
         % (PER['PSA 투자실행금'], PER['④ 투자실행금액 대비 ty수익율 (정의)'],
            PER['⑤ 투자자산 대비 ty수익율 (정의)'], RG, RW, AGG['S입금부족율(%)'])),
        ('가중치 대조', '금액 실측 · 대표 엑셀 MAU · 로스터 금액가중 세 기준을 나란히',
         'C%d:E%d · C%d:E%d' % (WROW['w'], WROW['w'], WROW['ty'], WROW['ty'])),
        ('화면대조', '화면 값 원본 %d행 · 불변식 · 화면 정합 %d행 · 모델 잔차 %d행'
                     % (len(sv_rows), LASTC - 1, RESL - RES0 + 1),
         'E열 화면 값 · G열 판정 · %d~%d행 모델 잔차' % (RES0, RESL)),
        ('산식', '대표 정의서 %d줄 ↔ 이 통합문서의 셀' % len(defs), 'C열'),
    ]
    head(ws, 1, ['순서', '시트', '무엇', '주요 칸'])
    for i, (sh, what, where) in enumerate(guide):
        r = 2 + i
        put(ws, r, 1, i + 1, M0)
        put(ws, r, 2, sh, bold=True)
        c = put(ws, r, 3, what)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        put(ws, r, 4, where)
    G1 = 1 + len(guide)

    T1 = G1 + 2
    head(ws, T1, ['#', '화면 값', '값', '되짚는 경로'], fill=SUB)
    trace = [
        ('투자실행액', '=화면대조!E%d' % SV['투자실행액'], M0,
         '입력 C%d → 가맹점 B%d → 가맹점 P2:P%d → 채권 J2:J%d (Q열=1) → 기간집계 B%d → B%d'
         % (IR['exec'], MB['하루 선정산액 합계 (방향 A)'], last - 1, NROW,
            AGG['미회수 Σ Ai'], AGG['투자실행액'])),
        ('W금융일수 표기', '=화면대조!E%d' % SV['W금융일수 표기'], '0.00',
         '플랫폼 C2:O5 → 플랫폼 D9:D12 → 가맹점 O2:O%d → 채권 S2:S%d → 기간집계 %s → %s'
         % (last - 1, NROW, RG, RW)),
        ('Ty수익율(%)', '=화면대조!E%d' % SV['Ty수익율(%)'], P2,
         '입력 C%d · 입력 C%d · 기간집계 %s → 기간집계 B%d'
         % (IR['ratep'], IR['year'], RW, AGG['Ty수익율(%)'])),
    ]
    for i, (lab, val, fmt, path) in enumerate(trace):
        r = T1 + 1 + i
        put(ws, r, 1, i + 1, M0)
        put(ws, r, 2, lab, bold=True)
        put(ws, r, 3, val, fmt, fill=CHK)
        put(ws, r, 4, path)
    T1L = T1 + len(trace)

    T2 = T1L + 2
    head(ws, T2, ['#', '바꾸는 칸', '항목', '지금 값', '움직이는 값'], fill=SUB)
    sim = [
        ('입력 C%d' % IR['total'], '총 투자자산', '=총투자자산', M0,
         '입력 C%d · C%d → 채권 I·J열 → 기간집계 B%d · B%d · B%d'
         % (IR['exec'], IR['cash'], AGG['투자실행액'], AGG['순현금'], AGG['투자자산'])),
        ('입력 C%d' % IR['idle'], '유휴 비율', '=유휴비율', '0.0%',
         '입력 C%d · C%d → 기간집계 B%d · B%d · B%d(⑤)'
         % (IR['exec'], IR['cash'], AGG['투자실행액 비중'], AGG['순현금 비중'],
            PER['⑤ 투자자산 대비 ty수익율 (정의)'])),
        ('입력 B%d' % (SWH + 1), '스위치 ① W금융일수 모집단', '=SW_모집단', 'General',
         '기간집계 %s · %s · B%d' % (RG, RW, AGG['Ty수익율(%)'])),
        ('입력 B%d' % (SWH + 2), '스위치 ② W 표기 자릿수', '=SW_자릿수', M0,
         '기간집계 %s · B%d' % (RW, AGG['Ty수익율(%)'])),
        ('입력 B%d' % (SWH + 3), '스위치 ③ 가맹점 수', '=SW_가맹점수', M0,
         '가맹점 G열 → 가맹점 B%d · B%d'
         % (MB['가맹점 수 (적용)'], MB['하루 선정산액 합계 (적용)'])),
        ('입력 B%d' % (SWH + 4), '스위치 ④ 산출 방향', '=SW_방향', 'General',
         '가맹점 P열 → 가맹점 B%d' % MB['하루 선정산액 합계 (적용)']),
        ('입력 C%d:C%d' % (PL0, PLS - 1), '플랫폼 구성비', '=SUM(구성비범위)', D8,
         "플랫폼 B17 · 가맹점 O열 · '가중치 대조' C%d" % WROW['w']),
        ('입력 C%d:C%d' % (IR['qfrom'], IR['qto']), '조회기간', '=조회종료-조회시작+1', M0,
         '기간집계 2~%d행' % (1 + len(per_rows))),
        ('입력 C%d' % IR['asof'], '기준일 d', '=기준일', DT,
         '채권 F·G열 · 일별 A열 · 채권 Q·R·T·U열'),
        ('입력 C%d' % IR['n3'], '③ 지시 대상 (미확정)', '=지시대상3', M0,
         '기간집계 B%d(⑥) — 비어 있는 동안 ⑥ 은 「미확정」'
         % PER['⑥ 투자실행금액 대비 ty수익율 (정의)']),
        ('입력 C%d' % IR['f5'], '⑤ 산식 (미확정 · 대표 재작성 대기)', '=산식5', P2,
         '기간집계 B%d(⑤) → 기간집계 B%d(⑥) → 화면대조 모델 잔차 「조회기간 ⑤(%%)」'
         % (PER['⑤ 투자자산 대비 ty수익율 (정의)'],
            PER['⑥ 투자실행금액 대비 ty수익율 (정의)'])),
        ('입력 B%d' % MAD, '배달앱/전체 (대표 엑셀 MAU 쪽 계산)', '=배달앱비중', D8,
         "'가중치 대조' D%d · D%d — 참고값 계열" % (WROW['w'], WROW['ty'])),
    ]
    for i, (where, lab, val, fmt, moves) in enumerate(sim):
        r = T2 + 1 + i
        put(ws, r, 1, i + 1, M0)
        put(ws, r, 2, where, bold=True)
        put(ws, r, 3, lab)
        put(ws, r, 4, val, fmt, fill=CHK)
        put(ws, r, 5, moves)
    T2L = T2 + len(sim)

    # ── 기호 규칙 — 2026-08-31 확정 (dm_0831/symbol_rule_0831.md) ──
    T3 = T2L + 2
    head(ws, T3, ['#', '기호', '뜻', '이 통합문서의 자리', '출처'], fill=SUB)
    sym = [
        ('D', '금융일수 (대문자)', '채권 H열 Di · 기간집계 %s' % RG,
         '이서준 지시 2026-08-31'),
        ('d', '오늘 날짜 (소문자)', '입력 C%d 기준일 d' % IR['asof'],
         '이서준 지시 2026-08-31 · 대표 DM 15:15'),
        ('d-1', '정산예정일이 어제인 대상정산금채권 집합. 날짜가 아니라 조건이다',
         '일별 시트 한 줄 = 정산예정일 하루치 집합', '대표 DM 2026-08-31 15:15'),
        ('S', 'Sum 또는 Sample. 표본으로 쓸 때는 d-20 ~ d-11',
         '입력 C%d:C%d · 채권 T열' % (IR['sfrom'], IR['sto']),
         '대표 DM 2026-08-31 15:15'),
        ('P', '기간(period). 정산예정일이 선택한 기간에 드는 것들의 합계',
         '기간집계 2~%d행' % (1 + len(per_rows)), '대표 DM 2026-08-31 15:15'),
        ('R', '비율', '기간집계 B%d PSMR' % PER['PSMR (정의)'],
         '대표 DM 2026-08-31 15:15'),
        ('M', '투자 수익', '채권 O열 Md-1i · 일별 E열',
         '대표 DM 2026-08-31 15:15'),
        ('L', '입금부족액 (미지급금 - 환급금)', '채권 L열 - M열 · 기간집계 B%d'
         % AGG['S표본 Σ(미지급-과지급)'], '대표 DM 2026-08-31 15:15'),
        ('i', '채권번호. 특정 가맹점 · 특정 날짜 · 특정 플랫폼의 정산금채권번호',
         '채권 A열', '대표 DM 2026-08-31 14:45'),
        ('w', '가중(weight). w금융일수 = 가중평균금융일수',
         '기간집계 %s · %s' % (RG, RW), '대표 DM 2026-08-31 16:41'),
        ('원문 표기 (대문자)', '대표 정의서 원문은 현재일자를 대문자로 쓴다. '
                                '인용이라 산식 시트 B열에만 그대로 남는다',
         '산식 B열', 'ceo_definitions.md'),
    ]
    for i, (k, mean, where, s0) in enumerate(sym):
        r = T3 + 1 + i
        put(ws, r, 1, i + 1, M0)
        put(ws, r, 2, k, bold=True)
        c = put(ws, r, 3, mean)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        put(ws, r, 4, where)
        put(ws, r, 5, s0)
    T3L = T3 + len(sym)

    # ── 대표 DM 2026-08-31 — 인용은 원문 그대로 ──────────────────
    T4 = T3L + 2
    head(ws, T4, ['#', 'DM 시각', '대표 원문', '이 통합문서의 자리',
                  '이 통합문서의 값'], fill=SUB)
    dm = [
        ('2026-08-31 16:45',
         'ty수익률이 위는 할인율, 아래는 투자수익률 smr인 이유는 위는 예상치 아래는 실제 '
         '결과치라서 실제로 얻은 수익률(할인율)을 나타내는 것.',
         '위 = 기간집계 B%d Ty수익율(할인율 기준 예상치) · '
         '아래 = 기간집계 B%d PSMR(실적치)' % (AGG['Ty수익율(%)'], PER['PSMR (정의)']), ''),
        ('2026-08-31 16:45',
         '투자수익률은 할인율 -max(0, 미지급금-과지급금)/투자실행액 이므로 미지급-과지급이 '
         '0이면 할인율이 되므로 같다.',
         '기간집계 B%d (가) 대표 근사식(%%) · B%d (나) 원식 PSMR(%%) · B%d 차 · '
         'B%d 실측 갈림 배수 · B%d 이론 1/(1-할인율) · B%d 차감합 · '
         'B%d·B%d 차감합 0 일 때 두 값. '
         '두 식은 분모가 달라 갈린다 — 어느 쪽이 정본인지는 대표 확인 대기'
         % (RGA, RWN, RDF, RRT, PER['DM 16:45 1 / (1 - 할인율) (배)'],
            PER['DM 16:45 차감합'], PER['DM 16:45 차감합 0 일 때 (가) (%)'],
            PER['DM 16:45 차감합 0 일 때 (나) (%)']),
         '=기간집계!B%d' % RDF),
        ('2026-08-31 16:27',
         '365를 w 금융일수로 나누면 1년동안 몇 번 굴릴 수 있는가 => 여기에 수익률(할인율) '
         '곱하면 이게 ty 수익률이다.',
         '기간집계 B%d 1년 회전수 · B%d Ty 연 환산'
         % (AGG['1년 회전수 (365 / 표기 W)'], AGG['Ty수익율 — 연 환산(%)']), ''),
        ('2026-08-31 16:27', '2. ty 수익률 → 일 환산 수익률',
         '기간집계 B%d Ty 일 환산 · B%d 검산(일 환산 x 365 - 연 환산)'
         % (AGG['Ty수익율 — 일 환산(%)'], AGG['검산 — 일 환산 x 365 = 연 환산']),
         '=기간집계!B%d' % AGG['검산 — 일 환산 x 365 = 연 환산']),
        ('2026-08-31 16:19', '(가맹점에서 채권을 사올 때는 1%를 떼고 사오므로)',
         '요율 미확정 (C1). 입력 C%d 할인율은 ledger_facts.json rate 그대로다. '
         '1%% 를 이 통합문서 어디에도 넣지 않았다' % IR['rate'], ''),
        ('2026-08-31 16:19', '실제로는 이 금액보다 선정산 실행액이 작다.',
         '같은 자리 — 선정산 제외액 · 이체/플랫폼 수수료는 이 통합문서 모집단 밖이다', ''),
    ]
    for i, (t, quote, where, judge) in enumerate(dm):
        r = T4 + 1 + i
        put(ws, r, 1, i + 1, M0)
        put(ws, r, 2, t, bold=True)
        c = put(ws, r, 3, quote)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c2 = put(ws, r, 4, where)
        c2.alignment = Alignment(wrap_text=True, vertical='top')
        put(ws, r, 5, judge, fill=CHK if judge else None)
    ws.freeze_panes = 'A2'
    wb.move_sheet('읽는 법', offset=-(len(wb.sheetnames) - 1))

    # ── 저장 — 정본과 사용자 폴더에 바이트 동일하게 ───────────────
    out = OUT if not tag else OUT.replace('.xlsx', '_%s.xlsx' % tag)
    out2 = OUT2 if not tag else OUT2.replace('.xlsx', '_%s.xlsx' % tag)
    wb.save(out)
    os.makedirs(os.path.dirname(out2), exist_ok=True)
    shutil.copyfile(out, out2)
    assert open(out, 'rb').read() == open(out2, 'rb').read(), '두 파일이 다르다'
    return dict(out=out, out2=out2, bytes=os.path.getsize(out),
                sheets=wb.sheetnames, nrec=NROW - 1, nday=NDAY, last=last,
                merchants=n, ncmp=LASTC - 1, nres=RESL - RES0 + 1,
                total_asset=total_asset, idle=idle_ratio,
                exec_target=exec_target, cash=cash, avg_daily=avg_daily,
                w_screen=S(fx, 'W금융일수 표기'), ty_screen=S(fx, 'Ty수익율'),
                mix={LABEL[p]: plat['book_mix'][p] for p in ORDER},
                measured_w=plat['measured_w'], mau_w=plat['mau_w'])


if __name__ == '__main__':
    #   인자 없으면 정본 한 벌. `유휴 20 25 30` 처럼 주면 그 비율로 한 벌씩 더 만든다.
    args = sys.argv[1:]
    if args and args[0] == '유휴':
        for pct in args[1:]:
            info = build(idle=float(pct) / 100, tag='유휴%s' % pct)
            print(json.dumps(info, ensure_ascii=False, indent=1))
    else:
        info = build()
        print(json.dumps(info, ensure_ascii=False, indent=1))
