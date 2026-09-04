# -*- coding: utf-8 -*-
"""대표 정의서 [2번 이미지] 일별 배치 기호 8개 기계 검사.

원문: ceo_definitions.md [2번 이미지]
    BD-1i  = 순지급액 - max(0, 미지급금 - 과지급금)
    AD-1i  = 순지급액 x (1 - 할인율)
    MD-1i  = 채권매입수수료 - max(0, 미지급금 - 과지급금)
    DD-1i  = 금융일수
    SBD-1  = SUM BD-1i        SAD-1 = SUM AD-1i        SMD-1 = SUM MD-1i
    SDD-1  = (SUM AD-1i x DD-1i) / SUM AD-1i
    SMRD-1 = SMD-1 / SAD-1    ty수익율 = SMRD-1 x 365 / SDD-1

검사 5항목
    1  BD-1i - AD-1i = MD-1i 가 원장에서 성립하는가 (채권층 · 하루층 · 전 구간층)
    2  SMRD-1 · ty수익율이 화면 · 엑셀 · 용어정의서에서 같은 값인가
    3  SDD-1 의 가중치가 AD-1i(금액)인가 — 건수 가중이 섞인 자리가 없는가
    4  기호 8개의 한글 이름이 산출물 다섯 곳에서 일관된가
    5  BD-1i 의 max(0, 미지급금-과지급금) 이 실제로 0 에서 잘리는가
    6  자기시험 — 위 검사기들이 일부러 깨뜨린 값을 잡아내는가

판정 규칙
    · FAIL 1건 이상이면 종료코드 1.
    · try/except 로 오류를 SKIP 으로 삼키지 않는다. 예외는 FAIL 로 센다.
    · 파일 부재는 FAIL 이다 (rd() 가 예외를 던지고 그 예외가 FAIL 로 기록된다).
    · 「위반 0건」류 검사는 검사 대상이 0건이 아님을 함께 판정한다.

verify_weighting.js 와의 경계
    verify_weighting.js 는 [1번] 축 — 책 전체 W금융일수(3.04)의 가중 기준과 MAU 참고값 갈라적기를 본다.
    이 검증기는 [2번] 축 — 하루치 SDD-1 · 기간 PSD 의 가중치가 AD-1i 인가만 본다. 겹치지 않는다.

허용 오차 (임의값 아님 · 반올림 규칙에서 유도)
    채권 1건  |BD-1i - AD-1i - MD-1i| = |net x (1-r) - ai| <= 0.5
              ai = ROUND_HALF_UP(...) · net = ROUND_HALF_UP(ai/(1-r)) 이므로
              net = ai/(1-r) + e, |e| <= 0.5  ->  net(1-r) - ai = e(1-r), |.| <= 0.5(1-r) < 0.5
    하루      위 잔차가 그날 채권 수 n 만큼 쌓이므로 0.5 x n. 원장이 하루 수수료를 한 번 내림하므로 +1.
    전 구간   같은 규칙에 n = 구간 전체 채권 수.

산출: verify_batch_symbols_result.json
"""
import io, json, os, re, subprocess, sys, tempfile, traceback
from decimal import Decimal as D, ROUND_HALF_UP

from openpyxl.utils import get_column_letter as CL

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

APP_REPO = '/Users/semi/cursor/payhug-investor-admin'
PROTO_REPO = '/Users/semi/cursor/payhug-investor-prototype'
XLSX = os.path.join(BASE, '검산_투자자어드민_20260901.xlsx')
CHROME = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'

# 기호 8개 — 이 목록이 검사 범위다. 표기는 정본 `dm_0901/symbol_rule_0901.md` 가 정한다.
# 여기에 표를 또 두면 규칙이 바뀔 때 두 곳이 어긋난다 — 2026-08-31 에 그렇게 깨졌고,
# 2026-09-01 정본 전환 때 `SBd-1` 계열을 이고 있던 이 파일이 또 낡았다. 표는 정본에만 둔다.
RULE = os.path.join(BASE, 'dm_0901', 'symbol_rule_0901.md')
ROUND_RULE = os.path.join(BASE, 'dm_0901', 'rounding_rule_0901.md')
# 아래첨자 → 평문. 정본 「표기 형태」 표 — 평문(엑셀 열머리·코드·DB)은 `Aᵢ` 를 `Ai` 로 적는다.
_SUBS = {'ᵢ': 'i'}


def _plain(s):
    for a, b in _SUBS.items():
        s = s.replace(a, b)
    return s


def _canon():
    """정본에서 낱개 기호·하루 집계 기호를 읽는다.

    낱개는 「기호 전건」 표의 `낱개` 행, 하루 집계는 「갈아 끼우는 표」의 정본 칸에서 온다.
    정본이 없거나 표 모양이 바뀌면 조용히 통과시키지 않고 여기서 멎는다.
    """
    if not os.path.exists(RULE):
        raise SystemExit('!! 기호 정본이 없다 — ' + RULE)
    txt = io.open(RULE, encoding='utf-8').read()
    one = {}
    for name, sym in re.findall(r'\|\s*낱개\s*\|\s*([^|]+?)\s*\|\s*`([^`]+)`\s*\|', txt):
        one[_plain(sym)] = name.strip()
    m = re.search(r'\|\s*`SB_\{D-1\}`[^|]*\|\s*([^|]+?)\s*\|', txt)
    if not m:
        raise SystemExit('!! 정본에서 하루 집계 갈아끼움 줄을 못 읽었다')
    day = re.findall(r'`([^`]+)`', m.group(1))          # B(d-1) A(d-1) M(d-1) MR(d-1)
    m = re.search(r'\|\s*`SD_\{D-1\}`[^|]*\|\s*`([^`]+)`\s*\|', txt)
    if not m:
        raise SystemExit('!! 정본에서 wD 갈아끼움 줄을 못 읽었다')
    wd = m.group(1)                                     # wD(d-1)
    if len(day) != 4:
        raise SystemExit('!! 하루 집계 정본 표기가 4개가 아니다 — %s' % day)
    miss = [x for x in ('Bi', 'Ai', 'Mi', 'Di') if x not in one]
    if miss:
        raise SystemExit('!! 정본 낱개 표에 없는 기호 — %s' % miss)
    return one, day, wd


def _round_digits():
    """비율·일수를 몇 자리까지 남기는지 정본에서 읽는다 — `dm_0901/rounding_rule_0901.md`."""
    if not os.path.exists(ROUND_RULE):
        raise SystemExit('!! 반올림 정본이 없다 — ' + ROUND_RULE)
    txt = io.open(ROUND_RULE, encoding='utf-8').read()
    m = re.search(r'비율·일수\s*—\s*소수\s*(\d+)\s*자리', txt)
    if not m:
        raise SystemExit('!! 반올림 정본에서 자릿수를 못 읽었다')
    return int(m.group(1))


def _unround(f):
    """`=IF(X=0,0,ROUND(몸통,N))` 에서 (`=IF(X=0,0,몸통)`, N) 을 돌려준다.

    ROUND 가 없으면 (원문, None). 몸통을 그대로 남기므로 산식이 바뀌면 문언 대조가 잡는다.
    """
    s = str(f).replace(' ', '')
    m = re.match(r'^(=IF\([A-Z]+\d+=0,0,)ROUND\((.+),(\d+)\)\)$', s)
    if m:
        return m.group(1) + m.group(2) + ')', int(m.group(3))
    return s, None


def _rnd(v, n):
    """엑셀 ROUND 와 같은 반올림 — 파이썬 기본은 은행가 반올림이라 .5 에서 갈린다."""
    if n is None:
        return v
    from decimal import Decimal, ROUND_HALF_UP
    return float(Decimal(repr(v)).quantize(Decimal(1).scaleb(-n), rounding=ROUND_HALF_UP))


ONE_NAME, DAY4, WD = _canon()
RD = _round_digits()
SYMBOLS = ['Bi', 'Ai', 'Mi', 'Di', DAY4[0], DAY4[1], DAY4[2], WD]
SYM = dict(zip(['B', 'A', 'M', 'D', 'SB', 'SA', 'SM', 'SD'], SYMBOLS))
SMR = DAY4[3]
# 대표 원문은 대문자 평문 표기(SBD-1)다. 원문은 sha256 잠금이고 인용을 손대지 않으므로
# 이 짝은 원문을 훑을 때만 쓴다 — 우리 표기가 바뀌어도 여기는 그대로다.
CEO_SYMBOLS = ['BD-1i', 'AD-1i', 'MD-1i', 'DD-1i', 'SBD-1', 'SAD-1', 'SMD-1', 'SDD-1']
CEO_OF = dict(zip(SYMBOLS, CEO_SYMBOLS))

R = []          # {sec, name, pass, detail}
_SEEN = set()


def chk(sec, name, ok, detail=''):
    key = (sec, name)
    assert key not in _SEEN, '검사 이름 중복: %s' % (key,)
    _SEEN.add(key)
    R.append({'sec': sec, 'name': name, 'pass': bool(ok), 'detail': str(detail)})
    return bool(ok)


def rd(path, binary=False):
    """파일 판독 — 없으면 예외를 던진다. 빈 문자열로 넘기지 않는다."""
    if not os.path.exists(path):
        raise IOError('파일 없음: %s' % path)
    if binary:
        with open(path, 'rb') as fp:
            return fp.read()
    with io.open(path, 'r', encoding='utf-8') as fp:
        t = fp.read()
    if not t.strip():
        raise IOError('파일이 비었다: %s' % path)
    return t


def r2(x):
    return D(str(x)).quantize(D('0.01'), rounding=ROUND_HALF_UP)


def num(s):
    m = re.search(r'-?[\d,]*\d(?:\.\d+)?', str(s).replace(' ', ''))
    if not m:
        return None
    return float(m.group(0).replace(',', ''))


# ══════════════════════════════════════════════════════════════════
# 검사기 — 주입 시험이 가능하도록 데이터를 인자로 받는다.
# ══════════════════════════════════════════════════════════════════
def bond_residuals(rows, rate):
    """채권 1건마다 BD-1i - AD-1i - MD-1i.
       BD = net - ded, AD = ai, MD = net*rate - ded  ->  잔차 = net*(1-rate) - ai."""
    out = []
    for r in rows:
        bd = D(r['net']) - D(r['ded'])
        ad = D(r['ai'])
        md = D(r['net']) * rate - D(r['ded'])
        out.append(bd - ad - md)
    return out


def day_identity(day_rows, rate, nfloor=1):
    """하루 단위 세 갈래 잔차. day_rows = [{n, net, ai, ded, repay, exec, profit, fee}]

    nfloor = 그 묶음 안에 든 「하루 수수료 내림」 횟수. 원장은 하루마다 한 번 내림하므로
    하루 행은 1, 여러 날을 합친 전 구간 행은 그 일수다. 내림 한 번이 잔차를 (-1, 0] 만큼
    남기므로 허용 폭이 일수에 비례한다 — 손으로 고른 값이 아니라 내림 횟수에서 나온다.

    상환액은 순지급액에서 부족액을 한 줄로 뺀 값이다(dm_0901 규칙 2 · 확정).
      repay  정의식과 원 단위 완전일치를 본다.
      emit   화면 두 칸(투자실행금 + 투자 수익)을 더한 값과의 어긋남. 쪼개면 Ai 반올림
             0.5 x n 과 채권매입수수료 내림 nfloor 가 남으므로 0 이 아닌 것이 규칙이고,
             그 두 자리에서 나올 수 있는 폭을 넘는지만 본다."""
    bad = {'def': [], 'emit': [], 'repay': [], 'profit': []}
    for g in day_rows:
        n = g['n']
        res = D(g['net']) * (D(1) - rate) - D(g['ai'])          # SUM BD - SUM AD - SUM MD
        if abs(res) > D('0.5') * n:
            bad['def'].append((g['d'], str(res), str(D('0.5') * n)))
        gap = D(g['repay'] - g['exec'] - g['profit'])            # 쪼개 더한 값과의 어긋남
        if abs(gap) > D('0.5') * n + nfloor:
            bad['emit'].append((g['d'], int(gap), str(D('0.5') * n + nfloor)))
        dr = D(g['repay']) - (D(g['net']) - D(g['ded']))         # 원장 상환액 <-> 정의식
        if dr != 0:
            bad['repay'].append((g['d'], str(dr)))
        dp = D(g['profit']) - (D(g['net']) * rate - D(g['ded']))  # 원장 투자수익 <-> 정의식
        if not (D(-nfloor) < dp <= D('0.000001')):
            bad['profit'].append((g['d'], str(dp), '한계 (-%d, 0]' % nfloor))
    return bad


def w_by_weight(rows, weight):
    """weight = 'ai'(AD-1i 금액) | 'cnt'(건수) | 'net'(순지급액)."""
    if weight == 'cnt':
        return D(sum(r['di'] for r in rows)) / D(len(rows))
    k = 'ai' if weight == 'ai' else 'net'
    tot = sum(r[k] for r in rows)
    return D(sum(r[k] * r['di'] for r in rows)) / D(tot)


def clamp_eval(unpaid, over):
    """대표 정의서의 max(0, 미지급금 - 과지급금)."""
    return max(0, unpaid - over)


def clamp_broken(unpaid, over):
    """max 를 뺀 잘못된 구현 — 자기시험용."""
    return unpaid - over


def clamp_check(fn, cases):
    """미지급 < 과지급 인 사례에서 0 으로 잘리는가. 안 잘리면 위반."""
    bad = []
    for up, ov in cases:
        v = fn(up, ov)
        if up < ov and v != 0:
            bad.append((up, ov, v))
        if v < 0:
            bad.append((up, ov, v))
    return bad


def ty_rows_check(rows, days=365):
    """행마다 ty == round(profit/exec*100*365/w, 2) 인가. rows=[{d,exec,profit,w,ty}]

    ty 는 소수 6자리 W 로 내고 화면에는 2자리만 보인다(dm_0901 규칙 1 · 확정).
    표에 적힌 2자리 W 로 되짚으면 ty 가 |ty| x 0.005 / (W - 0.005) 만큼 어긋난다 —
    거기에 ty 자신의 표기 반올림 0.005 를 더한 폭까지만 봐 준다.
    """
    bad = []
    for r in rows:
        if not r['exec'] or not r['w']:
            bad.append((r['d'], 'exec/w 0'))
            continue
        want = float(r2(D(str(r['profit'])) / D(str(r['exec'])) * 100 * D(days) / D(str(r['w']))))
        slack = abs(r['ty']) * 0.005 / (r['w'] - 0.005) + 0.005
        if abs(want - r['ty']) > slack:
            bad.append((r['d'], r['ty'], want))
    return bad


def name_check(table):
    """table = {symbol: {source: name}}. 같은 기호에 이름이 둘 이상이면 위반."""
    bad = []
    for sym in sorted(table):
        got = table[sym]
        names = sorted(set(got.values()))
        if len(names) > 1:
            bad.append((sym, {k: v for k, v in got.items()}))
    return bad


# ══════════════════════════════════════════════════════════════════
# 화면 조작 — 헤드리스 크롬. CDP 드라이버를 임시 파일로 떨어뜨려 돌린다.
# ══════════════════════════════════════════════════════════════════
DRIVER_JS = r"""
const http = require('http'), fs = require('fs'), path = require('path'), os = require('os');
const { spawn } = require('child_process');
const CFG = JSON.parse(process.argv[2]);
const PORT = 8790 + (process.pid % 80), DPORT = 9490 + (process.pid % 80);
const MIME = {'.html':'text/html; charset=utf-8', '.css':'text/css; charset=utf-8',
  '.js':'text/javascript', '.png':'image/png', '.webp':'image/webp', '.pdf':'application/pdf',
  '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'};

const server = http.createServer((req, res) => {
  const url = decodeURIComponent(req.url.split('?')[0]);
  const seg = url.split('/').filter(Boolean);
  const t = CFG.targets.find(x => x.key === seg[0]);
  if(!t){ res.writeHead(404); res.end('no target'); return; }
  const p = path.join(t.root, seg.slice(1).join('/'));
  fs.readFile(p, (e, b) => {
    if(e){ res.writeHead(404); res.end('nope'); return; }
    res.writeHead(200, {'Content-Type': MIME[path.extname(p)] || 'application/octet-stream'});
    res.end(b);
  });
});

let msgId = 0, ws, pending = new Map();
const consoleErrors = [];
function send(method, params){
  const id = ++msgId;
  ws.send(JSON.stringify({id, method, params: params || {}}));
  return new Promise((res, rej) => pending.set(id, {res, rej}));
}
async function evalJS(expr){
  const r = await send('Runtime.evaluate', {expression:'(function(){' + expr + '})()',
    returnByValue:true, awaitPromise:true});
  if(r.exceptionDetails) throw new Error('page eval: ' + JSON.stringify(
    (r.exceptionDetails.exception && r.exceptionDetails.exception.description) || r.exceptionDetails.text));
  return r.result.value;
}
const sleep = ms => new Promise(r => setTimeout(r, ms));
/* 고정 대기로 값을 읽으면 느린 회차에 null 이 잡힌다. 조건이 참이 될 때까지 기다린다. */
async function waitFor(expr, label, ms){
  const end = Date.now() + (ms || 15000);
  let last = null;
  while(Date.now() < end){
    try { last = await evalJS('return !!(' + expr + ');'); } catch(e){ last = false; }
    if(last) return true;
    await sleep(120);
  }
  throw new Error('대기 실패: ' + label + ' (' + expr + ')');
}
function must(v, label){
  if(v === null || v === undefined) throw new Error('읽지 못한 자리: ' + label);
  return v;
}

/* 페이지 안에 심는 판독기 — 표를 텍스트 그대로 걷는다. */
const HELP = `
  window.__num = function(s){
    var m = String(s).replace(/,/g,'').match(/-?\\d+(?:\\.\\d+)?/);
    return m ? parseFloat(m[0]) : null;
  };
  window.__tbl = function(box){
    if(!box) return null;
    var t = box.querySelector('table'); if(!t) return null;
    var cells = tr => [].map.call(tr.children, td => td.textContent.trim());
    return {head:[].map.call(t.querySelectorAll('thead th'), e => e.textContent.trim()),
            body:[].map.call(t.querySelectorAll('tbody tr'), cells),
            foot:[].map.call(t.querySelectorAll('tfoot tr'), cells)};
  };
  window.__wrapTbl = function(scope, title){
    var ws = scope.querySelectorAll('.tbl-wrap'), i, h;
    for(i = 0; i < ws.length; i++){
      h = ws[i].querySelector('.tbl-head');
      if(h && h.textContent.indexOf(title) >= 0) return window.__tbl(ws[i]);
    }
    return null;
  };
  window.__set = function(id, v){
    var el = document.getElementById(id);
    if(!el) throw new Error('없는 입력칸 ' + id);
    el.value = v;
    el.dispatchEvent(new Event('input', {bubbles:true}));
    el.dispatchEvent(new Event('change', {bubbles:true}));
    return el.value;
  };
  window.__click = function(sel){
    var el = document.querySelector(sel);
    if(!el) throw new Error('없는 컨트롤 ' + sel);
    el.click(); return true;
  };
`;

async function readProfit(){
  return await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-profit"]');
    var box = sec.querySelector('[data-mount="pf-tbl"]');
    var stat = sec.querySelector('[data-mount="pf-stat"]');
    var cards = [].map.call(stat.querySelectorAll('.stat'), function(s){
      var lab = s.querySelector('.summary-label');
      return {label: lab ? lab.textContent.trim() : '',
              vals: [].map.call(s.querySelectorAll('.summary-value'), function(v){ return v.textContent.trim(); }),
              sub: (s.querySelector('.summary-sub') || {textContent:''}).textContent.trim(),
              period: (s.querySelector('.stat-period') || {textContent:''}).textContent.trim()};
    });
    var from = sec.querySelector('[data-mount="pf-from"]').value;
    var to = sec.querySelector('[data-mount="pf-to"]').value;
    var gran = (sec.querySelector('[data-act="pf-gran"].is-on') || {dataset:{}}).dataset.gran || null;
    var title = (sec.querySelector('[data-mount="pf-tbl-title"]') || {textContent:''}).textContent.trim();
    return {from:from, to:to, gran:gran, title:title, cards:cards, tbl:window.__tbl(box)};
  `);
}

async function readSim(){
  return await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-sim"]');
    var out = sec.querySelector('[data-mount="sim-out"]');
    return {bonds: window.__wrapTbl(out, '채권별 산출'),
            daily: window.__wrapTbl(out, '일별 투자수익'),
            status: window.__wrapTbl(out, '현황'),
            cards: [].map.call(out.querySelectorAll('.stat'), function(s){
              var lab = s.querySelector('.summary-label');
              return {label: lab ? lab.textContent.trim() : '',
                      vals: [].map.call(s.querySelectorAll('.summary-value'), function(v){ return v.textContent.trim(); }),
                      period: (s.querySelector('.stat-period') || {textContent:''}).textContent.trim()};
            }),
            inputs: {unpaid:document.getElementById('sim-unpaid').value,
                     over:document.getElementById('sim-over').value,
                     r:document.getElementById('sim-r').value}};
  `);
}

async function runSim(unpaid, over){
  await evalJS(`window.__set('sim-unpaid', ${JSON.stringify(String(unpaid))});`);
  await sleep(80);
  await evalJS(`window.__set('sim-over', ${JSON.stringify(String(over))});`);
  await sleep(80);
  /* 실행 버튼을 누르면 화면이 먼저 다시 그려지고(진행 표시) 계산은 300ms 뒤에 온다.
     표가 떴다는 것만 보고 읽으면 직전 회차의 결과를 읽는다 — 결과 객체가 바뀌는 것을 기다린다. */
  await evalJS(`window.__simStamp = (window.SIM && window.SIM.result) || null; return 1;`);
  await evalJS(`window.__click('[data-act="sim-run"]');`);
  await waitFor(`window.SIM && window.SIM.running === false && window.SIM.result
                 && window.SIM.result !== window.__simStamp
                 && document.querySelector('section.screen[data-screen="invest-sim"] [data-mount="sim-out"] table')`,
                '시뮬 새 결과 (미지급 ' + unpaid + ' / 과지급 ' + over + ')');
  await sleep(150);
  const r = await readSim();
  must(r.bonds, '시뮬 채권별 산출 표');
  must(r.daily, '시뮬 일별 투자수익 표');
  return r;
}

async function drive(t, base){
  const out = {key:t.key};
  await send('Page.navigate', {url: base + '/' + t.key + '/' + t.page});
  await waitFor(`document.querySelector('section.screen[data-screen="invest-profit"]')`,
                t.key + ' 화면 적재', 30000);
  await sleep(400);
  await evalJS(HELP + 'return 1;');

  out.viewport = await evalJS('return {w:window.innerWidth, h:window.innerHeight};');

  /* 투자 수익 — 씨앗 기간(일별 일주일) */
  await evalJS(`location.hash = '#invest-profit'; return 1;`);
  await waitFor(`document.querySelector('section.screen[data-screen="invest-profit"] [data-mount="pf-tbl"] tbody tr')`,
                t.key + ' 투자 수익 일별 표');
  out.profitWeek = must((await readProfit()), t.key + ' profitWeek');
  must(out.profitWeek.tbl, t.key + ' profitWeek 표');

  /* 직접입력으로 원장 전 구간 — 날짜 칸에 값을 넣고 change 를 흘린다 */
  await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-profit"]');
    var f = sec.querySelector('[data-mount="pf-from"]'), t = sec.querySelector('[data-mount="pf-to"]');
    f.value = '2026-03-01'; f.dispatchEvent(new Event('change', {bubbles:true}));
    return 1;
  `);
  await sleep(200);
  await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-profit"]');
    var t = sec.querySelector('[data-mount="pf-to"]');
    t.value = '2026-08-27'; t.dispatchEvent(new Event('change', {bubbles:true}));
    return 1;
  `);
  await sleep(400);
  await evalJS(`window.__click('section.screen[data-screen="invest-profit"] [data-act="pf-search"]');`);
  await waitFor(`document.querySelectorAll('section.screen[data-screen="invest-profit"] [data-mount="pf-tbl"] tbody tr').length > 100`,
                t.key + ' 전 구간 180행');
  out.profitFull = must((await readProfit()), t.key + ' profitFull');
  must(out.profitFull.tbl, t.key + ' profitFull 표');

  /* 월별 단위 */
  await evalJS(`window.__click('section.screen[data-screen="invest-profit"] [data-act="pf-gran"][data-gran="monthly"]');`);
  await waitFor(`document.querySelector('section.screen[data-screen="invest-profit"] [data-mount="pf-tbl-title"]').textContent.indexOf('월별') >= 0`,
                t.key + ' 월별 전환');
  await sleep(150);
  out.profitMonthly = must((await readProfit()), t.key + ' profitMonthly');
  must(out.profitMonthly.tbl, t.key + ' profitMonthly 표');

  /* 투자 시뮬레이션 — 씨앗 / 클램프 음수 / 클램프 양수.
     원본(app)만 — 시연본은 투자 시뮬레이션을 뺐다(step7 시뮬 제거 2026-09-04 · PROTO_DROPPED).
     시연본에서는 그 화면이 없음을 확인만 한다(되살아나면 FAIL 로 잡히게 값을 남긴다). */
  if(t.key === 'app'){
    await evalJS(`location.hash = '#invest-sim'; return 1;`);
    await waitFor(`document.getElementById('sim-unpaid') && document.querySelector('section.screen[data-screen="invest-sim"] [data-mount="sim-out"]')`,
                  t.key + ' 시뮬레이션 화면');
    out.simSeed = await runSim('0.08', '0.01');
    out.simClampNeg = await runSim('0.01', '0.08');
    out.simClampPos = await runSim('0.20', '0.01');
  } else {
    out.simAbsent = await evalJS(`return {section: !!document.querySelector('section.screen[data-screen="invest-sim"]'),
      nav: !!document.querySelector('.nav-item[data-menu="invest-sim"]'),
      fn: typeof window.simRun === 'function' || typeof window.SIM !== 'undefined'};`);
  }
  return out;
}

async function main(){
  await new Promise(r => server.listen(PORT, r));
  const base = 'http://127.0.0.1:' + PORT;
  const profile = fs.mkdtempSync(path.join(os.tmpdir(), 'phbs-'));
  /* macOS 함정 — --window-size 의 세로는 크롬 창 전체다. 뷰포트는 87px 작다.
     --screenshot 플래그를 붙이면 이 편차가 사라져 보정이 없어진 것처럼 오판한다. 붙이지 않는다. */
  const chrome = spawn(CFG.chrome, ['--headless=new', '--remote-debugging-port=' + DPORT,
    '--user-data-dir=' + profile, '--no-first-run', '--no-default-browser-check',
    '--disable-gpu', '--window-size=1440,1287', 'about:blank'], {stdio:'ignore'});

  let targets = null;
  for(let i = 0; i < 60 && !targets; i++){
    await sleep(300);
    try {
      targets = await new Promise((res, rej) => {
        http.get({host:'127.0.0.1', port:DPORT, path:'/json'}, r => {
          let d = ''; r.on('data', c => d += c); r.on('end', () => res(JSON.parse(d)));
        }).on('error', rej);
      });
    } catch(e){ targets = null; }
  }
  if(!targets) throw new Error('크롬 CDP 접속 실패');
  const page = targets.find(t => t.type === 'page');
  ws = new WebSocket(page.webSocketDebuggerUrl);
  await new Promise(r => ws.addEventListener('open', r));
  ws.addEventListener('message', ev => {
    const m = JSON.parse(ev.data);
    if(m.id && pending.has(m.id)){ pending.get(m.id).res(m.result); pending.delete(m.id); return; }
    if(m.method === 'Runtime.consoleAPICalled' && (m.params.type === 'error' || m.params.type === 'warning'))
      consoleErrors.push(m.params.type + ': ' + m.params.args.map(a => a.value || a.description || a.type).join(' '));
    if(m.method === 'Runtime.exceptionThrown')
      consoleErrors.push('exception: ' + ((m.params.exceptionDetails.exception &&
        m.params.exceptionDetails.exception.description) || m.params.exceptionDetails.text));
  });
  await send('Runtime.enable'); await send('Page.enable');

  const res = {targets:{}, console:consoleErrors};
  for(const t of CFG.targets) res.targets[t.key] = await drive(t, base);
  process.stdout.write(JSON.stringify(res));
  try { chrome.kill(); } catch(e){}
  server.close();
  process.exit(0);
}
main().catch(e => { process.stdout.write(JSON.stringify({error:String(e && e.stack || e)})); process.exit(3); });
"""


def run_driver():
    if not os.path.exists(CHROME):
        raise IOError('크롬 없음: %s' % CHROME)
    for root, page in ((APP_REPO, 'app.html'), (PROTO_REPO, 'index.html')):
        if not os.path.exists(os.path.join(root, page)):
            raise IOError('화면 파일 없음: %s' % os.path.join(root, page))
    cfg = {'chrome': CHROME,
           'targets': [{'key': 'app', 'root': APP_REPO, 'page': 'app.html'},
                       {'key': 'proto', 'root': PROTO_REPO, 'page': 'index.html'}]}
    fd, p = tempfile.mkstemp(suffix='.js', prefix='phbs-driver-')
    with os.fdopen(fd, 'w') as fp:
        fp.write(DRIVER_JS)
    try:
        cp = subprocess.run(['node', p, json.dumps(cfg)], stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE, timeout=420)
        raw = cp.stdout.decode('utf-8', 'replace').strip()
        if not raw:
            raise IOError('드라이버 무응답 rc=%s err=%s' % (cp.returncode, cp.stderr.decode('utf-8', 'replace')[:400]))
        out = json.loads(raw)
        if out.get('error'):
            raise IOError('드라이버 예외: %s' % out['error'][:600])
        return out
    finally:
        os.unlink(p)


# ══════════════════════════════════════════════════════════════════
# 1. 항등식 BD-1i - AD-1i = MD-1i
# ══════════════════════════════════════════════════════════════════
def sec1(L):
    dl, facts = L['ledger'], L['facts']
    rate = dl.RATE
    recv = dl.RECEIVABLES
    chk('1', '원장 채권 모집단 0건 아님', len(recv) > 0, '채권 %d건' % len(recv))

    res = bond_residuals(recv, rate)
    worst = max(res, key=lambda x: abs(x))
    chk('1', '채권층 |BD-1i - AD-1i - MD-1i| <= 0.5원',
        all(abs(x) <= D('0.5') for x in res),
        '채권 %d건 · 최대 잔차 %s원 (한계 0.5 = 원 단위 반올림 e x (1-r))' % (len(res), worst))

    # 하루 단위 — 정산예정일이 원장 구간(2026-03-01 ~ 기준일)에 드는 채권만
    lo, hi = facts['ledgerSpan']
    agg = {}
    for r in recv:
        d = r['due'].strftime('%Y-%m-%d')
        if not (lo <= d <= hi):
            continue
        g = agg.setdefault(d, {'d': d, 'n': 0, 'net': 0, 'ai': 0, 'ded': 0})
        g['n'] += 1
        g['net'] += r['net']
        g['ai'] += r['ai']
        g['ded'] += r['ded']
    emit = dict((r['d'], r) for r in dl.LEDGER)
    chk('1', '하루층 대상 일수 = 원장 일수',
        sorted(agg) == sorted(emit) and len(agg) == facts['ledgerDays'],
        '집계 %d일 · 원장 %d일 · facts %d일' % (len(agg), len(emit), facts['ledgerDays']))

    days = []
    for d in sorted(agg):
        g, e = agg[d], emit[d]
        days.append(dict(g, repay=e['repay'], exec=e['exec'], profit=e['profit'], fee=e['fee']))
    bad = day_identity(days, rate)
    chk('1', '하루층 정의식 |SUM BD - SUM AD - SUM MD| <= 0.5 x n',
        not bad['def'], '어긋난 날 %d건 %s' % (len(bad['def']), bad['def'][:3]))
    chk('1', '하루층 |상환액 - 투자실행금 - 투자수익| <= 0.5 x n + 내림횟수',
        not bad['emit'], '어긋난 날 %d건 %s' % (len(bad['emit']), bad['emit'][:3]))
    chk('1', '하루층 원장 상환액 <-> 정의식 SUM(순지급액 - 차감) (원 단위 완전일치)',
        not bad['repay'], '어긋난 날 %d건 %s' % (len(bad['repay']), bad['repay'][:3]))
    chk('1', '하루층 원장 투자수익 <-> 정의식 SUM(수수료 - 차감) (내림 1원 이내)',
        not bad['profit'], '어긋난 날 %d건 %s' % (len(bad['profit']), bad['profit'][:3]))

    # 전 구간
    tot = {'d': '전 구간', 'n': sum(g['n'] for g in days),
           'net': sum(g['net'] for g in days), 'ai': sum(g['ai'] for g in days),
           'ded': sum(g['ded'] for g in days), 'repay': sum(g['repay'] for g in days),
           'exec': sum(g['exec'] for g in days), 'profit': sum(g['profit'] for g in days),
           'fee': sum(g['fee'] for g in days)}
    tb = day_identity([tot], rate, nfloor=len(days))   # 하루 수수료 내림이 일수만큼 쌓인다
    tres = D(tot['net']) * (D(1) - rate) - D(tot['ai'])
    chk('1', '전 구간 정의식 |SUM BD - SUM AD - SUM MD| <= 0.5 x n',
        not tb['def'], '잔차 %s원 · 한계 %s원 · 채권 %d건' % (tres, D('0.5') * tot['n'], tot['n']))
    chk('1', '전 구간 |상환액 - 투자실행금 - 투자수익| <= 0.5 x n + 내림횟수',
        not tb['emit'], '차 %d원 · 한계 %s원 (상환 %d · 실행 %d · 수익 %d)'
        % (tot['repay'] - tot['exec'] - tot['profit'], D('0.5') * tot['n'] + len(days),
           tot['repay'], tot['exec'], tot['profit']))
    chk('1', '전 구간 원장 상환액 <-> 정의식 (원 단위 완전일치)', not tb['repay'],
        '차 %s원' % (D(tot['repay']) - (D(tot['net']) - D(tot['ded']))))
    chk('1', '전 구간 원장 투자수익 <-> 정의식', not tb['profit'],
        '차 %s원 · 한계 (-%d, 0] = 하루 수수료 내림 %d회'
        % (D(tot['profit']) - (D(tot['net']) * rate - D(tot['ded'])), len(days), len(days)))
    chk('1', '전 구간 합계가 facts 와 같음',
        tot['exec'] == facts['fullExec'] and tot['profit'] == facts['fullProfit'],
        'exec %d/%d · profit %d/%d' % (tot['exec'], facts['fullExec'], tot['profit'], facts['fullProfit']))

    # ledger_facts.json 이 원장 값을 그대로 실었는가 (바이트 대조)
    tbd = facts['tyByDate']
    miss = [d for d in emit if d not in tbd]
    off = []
    for d, e in emit.items():
        if d not in tbd:
            continue
        want = [str(e['w']), str(e['ty']), e['exec'], e['profit'], e['repay'], e['fee'], e['ded']]
        if tbd[d] != want:
            off.append((d, tbd[d], want))
    chk('1', 'ledger_facts.tyByDate 가 원장 행과 바이트 일치',
        not miss and not off and len(tbd) == len(emit),
        '누락 %d · 불일치 %d · %s' % (len(miss), len(off), off[:2]))
    return {'days': days, 'total': tot, 'agg': agg}


# ══════════════════════════════════════════════════════════════════
# 2. SMRD-1 · ty수익율 — 화면 · 엑셀 · 용어정의서
# ══════════════════════════════════════════════════════════════════
DAY = '2026-08-21'      # 용어정의서가 실값 예시로 든 날. 씨앗 기간(일별 일주일) 첫 행이다.


def parse_profit_rows(tbl):
    """화면 일별 표 -> [{d, repay, exec, profit, w, ty}]"""
    out = []
    for row in tbl['body']:
        out.append({'d': row[0], 'repay': num(row[1]), 'exec': num(row[2]),
                    'profit': num(row[3]), 'w': num(row[4]), 'ty': num(row[5])})
    return out


def sec2(L, drv):
    facts = L['facts']
    got = {}          # 산출물 -> {SM, SA, SMR, SD, ty}

    # ── 화면 ──────────────────────────────────────────────
    for key in ('app', 'proto'):
        t = drv['targets'][key]
        chk('2', '[%s] 뷰포트 1440x1200 (macOS 87px 보정)' % key,
            t['viewport'] == {'w': 1440, 'h': 1200}, str(t['viewport']))
        for tag, blk in (('씨앗 일주일', t['profitWeek']), ('직접입력 전 구간', t['profitFull'])):
            rows = parse_profit_rows(blk['tbl'])
            chk('2', '[%s] %s 행수 0건 아님' % (key, tag), len(rows) > 0, '%d행' % len(rows))
            bad = ty_rows_check(rows)
            chk('2', '[%s] %s 행 ty = SMRD-1 x 365 / SDD-1' % (key, tag),
                not bad, '어긋난 행 %d건 %s' % (len(bad), bad[:3]))
            # 행값이 원장과 같은가 — 화면은 3.10 을 3.1 로 읽어 오므로 수치로 맞춘다
            off = []
            for r in rows:
                if r['d'] not in facts['tyByDate']:
                    continue
                f = facts['tyByDate'][r['d']]
                if (abs(r['w'] - float(f[0])) > 1e-9 or abs(r['ty'] - float(f[1])) > 1e-9
                        or int(r['exec']) != f[2] or int(r['profit']) != f[3]
                        or int(r['repay']) != f[4]):
                    off.append((r['d'], [r['w'], r['ty'], r['exec'], r['profit'], r['repay']], f[:5]))
            unknown = [r['d'] for r in rows if r['d'] not in facts['tyByDate']]
            chk('2', '[%s] %s 행값 <-> 원장 tyByDate' % (key, tag),
                not off and not unknown, '불일치 %d %s · 원장에 없는 날짜 %d %s'
                % (len(off), off[:2], len(unknown), unknown[:3]))
        # 합계 행
        blk = t['profitWeek']
        foot = blk['tbl']['foot'][0]
        f_repay, f_exec, f_profit, f_w, f_ty = (num(foot[1]), num(foot[2]), num(foot[3]),
                                                num(foot[4]), num(foot[5]))
        rows = parse_profit_rows(blk['tbl'])
        want_ty = float(r2(D(str(f_profit)) / D(str(f_exec)) * 100 * D(365) / D(str(f_w))))
        chk('2', '[%s] 합계 행 ty = PSMR x 365 / PSD' % key,
            abs(want_ty - f_ty) <= 0.005, '화면 %.2f · 되짚기 %.2f' % (f_ty, want_ty))
        chk('2', '[%s] 합계 행 <-> facts 주간값' % key,
            int(f_exec) == facts['weekExec'] and int(f_profit) == facts['weekProfit']
            and int(f_repay) == facts['weekRepay'] and '%.2f' % f_ty == facts['weekTy'],
            'exec %d/%d · profit %d/%d · ty %.2f/%s'
            % (f_exec, facts['weekExec'], f_profit, facts['weekProfit'], f_ty, facts['weekTy']))
        # 그날 행
        one = [r for r in rows if r['d'] == DAY]
        chk('2', '[%s] %s 행이 화면에 있음' % (key, DAY), len(one) == 1, '%d건' % len(one))
        if one:
            got['화면(%s)' % key] = {'SM': one[0]['profit'], 'SA': one[0]['exec'],
                                     'SMR': one[0]['profit'] / one[0]['exec'] * 100,
                                     'SD': one[0]['w'], 'ty': one[0]['ty']}

    # ── 엑셀 ──────────────────────────────────────────────
    import audit_xlsx_check as AX
    if not os.path.exists(XLSX):
        raise IOError('검산 엑셀 없음: %s' % XLSX)
    bk = AX.Book(XLSX)
    ws = bk.wb['일별']
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {}                       # M(d-1) 은 정본(E)·참고(F) 두 칸에 걸리므로 처음 자리를 쓴다
    # 열머리는 「기호 이름」 꼴이라 첫 공백 앞이 기호다. 괄호에서 더 자르면
    # `B(d-1)` 이 `B` 로 뭉개져 하루 집계와 낱개가 같은 열쇠가 된다.
    for i, h in enumerate(hdr):
        if h:
            col.setdefault(str(h).split(' ')[0], i + 1)
    chk('2', '엑셀 일별 열머리에 %s · %s · %s · %s · %s 이 있고 자리가 맞음'
        % (SYM['SB'], SYM['SA'], SYM['SM'], SMR, SYM['SD']),
        col.get(SYM['SB']) == 3 and col.get(SYM['SA']) == 4 and col.get(SYM['SM']) == 5
        and col.get(SMR) == 7 and col.get(SYM['SD']) == 8, str(hdr))
    nrow = 0
    fbad, vbad = [], []
    for r in range(2, ws.max_row):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        nrow += 1
        fg, fh, fi = (str(ws.cell(r, c).value) for c in (7, 8, 9))
        # 비율·일수는 정본 반올림 규칙(소수 RD자리)을 걸어도 되고 안 걸어도 된다.
        # 몸통이 정본 산식인지만 보고, 값 대조는 그 셀이 실제로 선언한 자릿수로 맞춘다 —
        # 반올림 유무를 값 쪽에서 흘려 주면 산식이 바뀌어도 통과한다.
        gx, gn = _unround(fg)
        hx, hn = _unround(fh)
        ix, ino = _unround(fi)
        # G 는 백분율 열이다(열머리 「MR(d-1) 투자수익율(%%)」). 정본 산식은 M÷A 라는
        # 비율이나, 6자리 반올림을 백분율 기준으로 걸어야 원장과 맞는다 —
        # 원장 daily_ledger 는 r6(pf / ex * 100) 로 끊는다(dm_0901 규칙 1).
        if gx != '=IF(D%d=0,0,E%d/D%d*100)' % (r, r, r):
            fbad.append(('G%d' % r, fg))
        if hx != '=IF(D%d=0,0,K%d/D%d)' % (r, r, r):
            fbad.append(('H%d' % r, fh))
        # I 는 G 가 이미 백분율이라 *100 을 다시 걸지 않는다.
        if ix != '=IF(H%d=0,0,G%d*연일수/H%d)' % (r, r, r):
            fbad.append(('I%d' % r, fi))
        d_, e_, g_, h_, i_ = (bk.cell('일별', 'D%d' % r), bk.cell('일별', 'E%d' % r),
                              bk.cell('일별', 'G%d' % r), bk.cell('일별', 'H%d' % r),
                              bk.cell('일별', 'I%d' % r))
        if abs(g_ - _rnd(e_ / d_ * 100, gn)) > 1e-12:
            vbad.append(('G%d' % r, g_, _rnd(e_ / d_ * 100, gn)))
        if abs(i_ - _rnd(g_ * 365 / h_, ino)) > 1e-9:
            vbad.append(('I%d' % r, i_, _rnd(g_ * 365 / h_, ino)))
    chk('2', '엑셀 일별 행수 0건 아님', nrow > 0, '%d행' % nrow)
    chk('2', '엑셀 일별 수식 문언 = %s = %s/%s · ty = %s x 365 / %s' % (SMR, SYM['SM'], SYM['SA'], SMR, SYM['SD']),
        not fbad, '어긋난 셀 %d건 %s' % (len(fbad), fbad[:3]))
    chk('2', '엑셀 일별 평가값이 그 수식과 같음', not vbad, '어긋난 셀 %d건 %s' % (len(vbad), vbad[:3]))
    # 정본 반올림 — `wD(d-1)` 은 소수 RD자리까지 남긴 값을 다음 계산에 넣는다.
    # 근거 dm_0901/rounding_rule_0901.md 규칙 1 · symbol_rule_0901.md 「갈림 두 자리」.
    # 여기서 반올림이 빠지면 기간집계 PwD 가 화면·원장과 갈린다(3.108481 ↔ 3.107588).
    nord = [('H%d' % r, str(ws.cell(r, 8).value)) for r in range(2, ws.max_row)
            if ws.cell(r, 1).value is not None and _unround(ws.cell(r, 8).value)[1] != RD]
    chk('2', '엑셀 일별 %s 이 정본 자릿수(소수 %d자리) 반올림을 걸었음' % (SYM['SD'], RD),
        nrow > 0 and not nord, '안 건 셀 %d건 %s' % (len(nord), nord[:3]))

    # 화면대조 시트 — 출처가 [2번] 인 행 전건 판정
    cw = bk.wb['화면대조']
    tgt = []
    for r in range(1, cw.max_row + 1):
        src = cw.cell(r, 8).value
        if isinstance(src, str) and '[2번]' in src:
            tgt.append(r)
    chk('2', '엑셀 화면대조에 [2번] 출처 행 0건 아님', len(tgt) > 0, '%d행' % len(tgt))
    verdict = [(r, bk.cell('화면대조', 'G%d' % r)) for r in tgt]
    chk('2', '엑셀 화면대조 [2번] 행 전건 「일치」',
        all(v == '일치' for _, v in verdict),
        '%s' % [(r, v) for r, v in verdict if v != '일치'][:5])

    # 엑셀이 되짚은 주간 ④ <-> 화면 카드 ④ (원장을 거치지 않는 직결 대조)
    row4 = [r for r in tgt if cw.cell(r, 2).value == '주간 ④(%)']
    chk('2', '엑셀 화면대조에 「주간 ④(%)」 행 있음', len(row4) == 1, str(row4))
    x4 = bk.cell('화면대조', 'C%d' % row4[0]) if row4 else None
    tycards = [c for c in drv['targets']['app']['profitWeek']['cards'] if c['label'] == '연환산수익률']
    chk('2', '화면 주간 카드에 연환산수익률 두 칸 있음',
        len(tycards) == 1 and len(tycards[0]['vals']) == 2, str(tycards))
    scr4 = num(tycards[0]['vals'][0]) if tycards and len(tycards[0]['vals']) == 2 else None
    chk('2', '엑셀 되짚은 주간 ④ = 화면 카드 ④', x4 is not None and scr4 is not None
        and abs(x4 - scr4) <= 0.005, '엑셀 %s · 화면 %s' % (x4, scr4))

    # ── 용어정의서 ────────────────────────────────────────
    for label, path in (('용어정의서(원고)', os.path.join(BASE, 'glossary_manuscript.md')),
                        ('용어정의서(배포 glossary.html)', os.path.join(APP_REPO, 'glossary.html'))):
        # 아래첨자 조판(A_{D-1} · A<sub>D\u22121</sub>)을 걷어 괄호 표기로 맞춘 뒤 대조한다.
        txt = deflate(rd(path))
        pat = {
            'SM': r'SM\(D-1\)\s*=\s*([\d,]+)',
            'SA': r'SA\s*=\s*([\d,]+)',
            'SMR': r'SMR\s*=\s*[\d,]+\s*(?:÷|/)\s*[\d,]+\s*[\r\n=\s]*([\d.]+)',
            'SD': r'SD\s*=\s*([\d.]+)일',
            'ty': r'0\.038051%\s*×\s*119\.416658\s*=\s*([\d.]+)%',
        }
        v = {}
        for k, p in pat.items():
            m = re.search(p, txt)
            chk('2', '[%s] %s 실값 판독' % (label, k), m is not None, p)
            if m:
                v[k] = float(m.group(1).replace(',', ''))
        if len(v) == 5:
            v['SMR'] = v['SMR'] * 100          # 0.00038051 -> 0.038051 (%)
            got[label] = {'SM': v['SM'], 'SA': v['SA'], 'SMR': v['SMR'],
                          'SD': v['SD'], 'ty': round(v['ty'], 2)}
            chk('2', '[%s] SMRD-1 = SMD-1 / SAD-1' % label,
                abs(v['SM'] / v['SA'] * 100 - v['SMR']) <= 5e-6,
                '%s / %s = %.8f%% · 문서 %.6f%%' % (v['SM'], v['SA'], v['SM'] / v['SA'] * 100, v['SMR']))
            chk('2', '[%s] ty = SMRD-1 x 365 / SDD-1' % label,
                abs(v['SMR'] * 365 / v['SD'] - v['ty']) <= 0.005,
                '%.6f x 365 / %s = %.4f · 문서 %.4f' % (v['SMR'], v['SD'], v['SMR'] * 365 / v['SD'], v['ty']))

    # ── 삼각 대조 ─────────────────────────────────────────
    lg = facts['tyByDate'][DAY]
    got['원장(ledger_facts)'] = {'SM': float(lg[3]), 'SA': float(lg[2]),
                                 'SMR': float(lg[3]) / float(lg[2]) * 100,
                                 'SD': float(lg[0]), 'ty': float(lg[1])}
    chk('2', '대조 대상 산출물 4곳 모두 값을 냈음', len(got) == 5,
        '%d곳: %s' % (len(got), sorted(got)))
    # SD 는 화면·원장이 소수 2자리 표기값을, 문서는 6자리 계산값을 싣는다(dm_0901 규칙 1).
    # 두 자리로 끊었을 때 같은지를 본다 — 글자까지 같기를 요구하면 규칙과 부딪힌다.
    for k in ('SM', 'SA', 'SD', 'ty'):
        vals = dict((s, got[s][k]) for s in got)
        nd = 2 if k == 'SD' else 4
        chk('2', '%s 삼각 대조 — %s' % (DAY, k),
            len(set(round(x, nd) for x in vals.values())) == 1, str(vals))
    smrs = dict((s, round(got[s]['SMR'], 6)) for s in got)
    chk('2', '%s 삼각 대조 — SMRD-1(%%)' % DAY, len(set(smrs.values())) == 1, str(smrs))
    return got


# ══════════════════════════════════════════════════════════════════
# 3. SDD-1 의 가중치는 AD-1i(금액)인가
# ══════════════════════════════════════════════════════════════════
def sec3(L, drv, s1):
    dl, facts = L['ledger'], L['facts']
    byday = {}
    for r in dl.RECEIVABLES:
        d = r['due'].strftime('%Y-%m-%d')
        if d in s1['agg']:
            byday.setdefault(d, []).append(r)
    emit = dict((r['d'], r) for r in dl.LEDGER)

    off_ai, same_cnt, off_net = [], [], []
    for d in sorted(byday):
        rows = byday[d]
        w_ai = r2(w_by_weight(rows, 'ai'))
        w_cnt = r2(w_by_weight(rows, 'cnt'))
        w_net = r2(w_by_weight(rows, 'net'))
        if w_ai != emit[d]['w']:
            off_ai.append((d, str(w_ai), str(emit[d]['w'])))
        if w_cnt == emit[d]['w']:
            same_cnt.append((d, str(w_cnt)))
        if w_net != emit[d]['w']:
            off_net.append((d, str(w_net), str(emit[d]['w'])))
    chk('3', '원장 대상 일수 0건 아님', len(byday) == facts['ledgerDays'] and byday,
        '%d일' % len(byday))
    chk('3', '원장 SDD-1 = SUM(AD-1i x DD-1i) / SUM AD-1i 전건',
        not off_ai, '어긋난 날 %d건 %s' % (len(off_ai), off_ai[:3]))
    chk('3', '건수 가중으로는 어긋난다 (검사에 판별력이 있음)',
        not same_cnt, '건수 가중과 같아진 날 %d건 %s' % (len(same_cnt), same_cnt[:3]))
    chk('3', '순지급액 가중도 AD-1i 가중과 같은 값 (Ai = net x (1-r) 비례)',
        not off_net, '어긋난 날 %d건 %s' % (len(off_net), off_net[:3]))

    # 엑셀 — 가중치 열이 Ai 인가
    import audit_xlsx_check as AX
    bk = AX.Book(XLSX)
    bs = bk.wb['채권']
    hdr = dict((str(bs.cell(1, c).value), c) for c in range(1, bs.max_column + 1))
    chk('3', '엑셀 채권 시트에 「Ai」·「금융일수 Di」·「Ai x Di」 열이 있음',
        all(k in hdr for k in ('Ai', '금융일수 Di', 'Ai x Di')), str(sorted(hdr)))
    cS = hdr.get('Ai x Di')
    cJ = hdr.get('Ai')
    cH = hdr.get('금융일수 Di')
    sbad, n = [], 0
    for r in range(2, bs.max_row + 1):
        f = str(bs.cell(r, cS).value).replace(' ', '')
        if bs.cell(r, 1).value is None:
            continue
        n += 1
        want = '=%s%d*%s%d' % (CL(cJ), r, CL(cH), r)
        if f != want:
            sbad.append(('%s%d' % (CL(cS), r), f, want))
    chk('3', '엑셀 채권 곱 열이 0건 아님', n > 0, '%d행' % n)
    chk('3', '엑셀 채권 곱 열 = Ai x Di 전행 (다른 열 곱 0건)',
        not sbad, '어긋난 행 %d건 %s' % (len(sbad), sbad[:3]))

    ds = bk.wb['일별']
    hbad, nb = [], 0
    for r in range(2, ds.max_row):
        if ds.cell(r, 1).value is None:
            continue
        nb += 1
        f0 = str(ds.cell(r, 8).value)
        f, _n = _unround(f0)                            # 정본 반올림은 몸통을 감싸기만 한다
        if f != '=IF(D%d=0,0,K%d/D%d)' % (r, r, r):     # D = SUM Ai, B = 건수
            hbad.append(('H%d' % r, f0))
        if 'B%d' % r in f:
            hbad.append(('H%d 건수 참조' % r, f0))
    chk('3', '엑셀 일별 %s 분모가 %s(D열)이며 건수(B열)가 아님' % (SYM['SD'], SYM['SA']),
        nb > 0 and not hbad, '%d행 · 어긋남 %d건 %s' % (nb, len(hbad), hbad[:3]))

    hv = []
    for r in range(2, ds.max_row):
        if ds.cell(r, 1).value is None:
            continue
        h_ = bk.cell('일별', 'H%d' % r)
        k_ = bk.cell('일별', 'K%d' % r)
        d_ = bk.cell('일별', 'D%d' % r)
        b_ = bk.cell('일별', 'B%d' % r)
        _n = _unround(ds.cell(r, 8).value)[1]            # 그 셀이 선언한 자릿수로 맞춘다
        if abs(h_ - _rnd(k_ / d_, _n)) > 1e-12:
            hv.append(('H%d' % r, h_, _rnd(k_ / d_, _n)))
        if abs(h_ - _rnd(k_ / b_, _n)) < 1e-9:           # 건수 가중과 우연히 같으면 판별 불가
            hv.append(('H%d 건수가중과 동일' % r, h_))
    chk('3', '엑셀 일별 %s 평가값 = SUM(Ai x Di) / SUM Ai · 건수 가중과 다름' % SYM['SD'],
        not hv, '어긋남 %d건 %s' % (len(hv), hv[:3]))

    # 화면 — 시뮬레이션 일별/합계 W 가 금액 가중인가
    #   원본(app)만. 시연본은 투자 시뮬레이션을 뺐다(step7 2026-09-04 · PROTO_DROPPED) — 없음을 따로 판정한다.
    pa_ = drv['targets']['proto'].get('simAbsent') or {}
    chk('3', '[proto] 투자 시뮬레이션 화면·메뉴·전역 함수 없음 (PROTO_DROPPED)',
        bool(pa_) and not pa_['section'] and not pa_['nav'] and not pa_['fn'], str(pa_))
    for key in ('app',):
        sim = drv['targets'][key]['simSeed']
        bonds = [b for b in sim['bonds']['body'] if b[1] == '기간 내']
        chk('3', '[%s] 시뮬레이션 기간 내 채권 0건 아님' % key, len(bonds) > 0, '%d건' % len(bonds))
        amts = [num(b[5]) for b in bonds]                # 투자실행금 = AD-1i
        dis = [num(b[4]) for b in bonds]                 # 금융일수 = DD-1i
        w_ai = sum(a * d for a, d in zip(amts, dis)) / sum(amts)
        w_cnt = sum(dis) / len(dis)
        foot = sim['daily']['foot'][0]
        shown = num(foot[4])
        chk('3', '[%s] 시뮬 합계 PSD = SUM(AD-1i x DD-1i) / SUM AD-1i' % key,
            abs(shown - w_ai) <= 0.005, '화면 %s · 금액가중 %.4f' % (shown, w_ai))
        chk('3', '[%s] 시뮬 합계 PSD 가 건수 가중과 다름 (판별력)' % key,
            abs(w_ai - w_cnt) > 0.005 and abs(shown - w_cnt) > 0.005,
            '금액 %.4f · 건수 %.4f · 화면 %s' % (w_ai, w_cnt, shown))

        # 투자 수익 화면 합계 W — 투자실행금 가중이어야 한다
        blk = drv['targets'][key]['profitWeek']
        rows = parse_profit_rows(blk['tbl'])
        fw = num(blk['tbl']['foot'][0][4])
        want = sum(r['w'] * r['exec'] for r in rows) / sum(r['exec'] for r in rows)
        plain = sum(r['w'] for r in rows) / len(rows)
        chk('3', '[%s] 투자 수익 합계 W = 투자실행금 가중평균' % key,
            abs(fw - want) <= 0.005, '화면 %s · 가중 %.5f' % (fw, want))
        chk('3', '[%s] 투자 수익 합계 W 가 단순평균과 다름 (판별력)' % key,
            abs(want - plain) > 1e-9, '가중 %.6f · 단순 %.6f' % (want, plain))

        # 월별 집계도 같은 축이다 — 달 행의 W 가 투자실행금 가중인가
        mb = drv['targets'][key]['profitMonthly']
        chk('3', '[%s] 월별 단위로 전환됨' % key, mb['title'] == '월별 투자수익', mb['title'])
        mrows = [{'d': r[0], 'repay': num(r[1]), 'exec': num(r[2]), 'profit': num(r[3]),
                  'w': num(r[4]), 'ty': num(r[5])} for r in mb['tbl']['body']]
        chk('3', '[%s] 월별 행수 0건 아님' % key, len(mrows) > 0, '%d행' % len(mrows))
        fw2 = dict((d, (w, ty)) for d, w, ty in L['facts']['monthTy'])
        fex = dict(L['facts']['monthExec'])
        offm = [r['d'] for r in mrows
                if r['d'] not in fw2 or abs(r['w'] - float(fw2[r['d']][0])) > 1e-9
                or abs(r['ty'] - float(fw2[r['d']][1])) > 1e-9 or int(r['exec']) != fex.get(r['d'])]
        chk('3', '[%s] 월별 행 W·Ty·투자실행금 <-> 원장 month_rollup' % key,
            not offm and len(mrows) == len(fw2), '불일치 %s · 화면 %d행 · 원장 %d행'
            % (offm[:3], len(mrows), len(fw2)))
        # 달 안의 일별 행에서 되짚기 — 금액 가중이어야 하고 단순평균과는 달라야 한다
        drows = parse_profit_rows(drv['targets'][key]['profitFull']['tbl'])
        bad_w, no_disc = [], []
        for r in mrows:
            inside = [x for x in drows if x['d'][:7] == r['d']]
            if not inside:
                bad_w.append((r['d'], '일별 행 없음'))
                continue
            wt = sum(x['w'] * x['exec'] for x in inside) / sum(x['exec'] for x in inside)
            pl = sum(x['w'] for x in inside) / len(inside)
            if abs(r['w'] - float(r2(wt))) > 1e-9:
                bad_w.append((r['d'], r['w'], float(r2(wt))))
            if abs(wt - pl) <= 1e-9:
                no_disc.append((r['d'], wt, pl))
        chk('3', '[%s] 월별 W = 그 달 일별 행의 투자실행금 가중평균' % key,
            not bad_w, '어긋난 달 %d건 %s' % (len(bad_w), bad_w[:3]))
        chk('3', '[%s] 월별 W 가 단순평균과 다름 (판별력)' % key,
            not no_disc, '구분 안 되는 달 %d건 %s' % (len(no_disc), no_disc[:3]))


# ══════════════════════════════════════════════════════════════════
# 4. 기호 8개 이름 일관
# ══════════════════════════════════════════════════════════════════
def norm_name(s):
    """symbol_glossary.json 이 스스로 등록한 별칭 규칙만 적용한다 —
       w(원문 표기) / W(화면 표기) 는 같은 용어로 등록돼 있다. 그 밖의 차이는 손대지 않는다."""
    s = s.strip()
    s = re.sub(r'^W금융일수$', 'w금융일수', s)
    return s


def deflate(t):
    """아래첨자 조판을 걷어 낸 대조용 표기로 되돌린다.

    문서·데이터가 A_{d-1,i}(마크다운) · A<sub>d\u22121,&thinsp;i</sub>(HTML) 로 쓰는 것을
    이 검사가 쓰는 괄호 표기 A(D-1)i 로 맞춘다. 원문 평문 표기(AD-1i)는 건드리지 않는다.

    2026-08-31 기호 규칙으로 아래첨자의 전일자가 소문자 d 로 바뀌었다(A_{d-1,i} · SA_{d-1}).
    옛 대문자 표기도 아직 남아 있는 산출물이 있어 두 표기를 다 접는다. 접는 자리는 첨자뿐이고,
    앞자리 글자(A · B · M · D)는 그대로 둔다 — 앞의 D 는 금융일수라 대문자다.
    """
    t = (t.replace('&thinsp;', '').replace('&#8201;', '').replace('\u2009', '')
          .replace('\u2212', '-'))
    t = re.sub(r'<sub>[Dd]-1,\s*i</sub>', '(D-1)i', t)
    t = re.sub(r'<sub>[Dd]-1</sub>', '(D-1)', t)
    t = re.sub(r'<sub>p,\s*i</sub>', 'pi', t)
    t = re.sub(r'<sub>i</sub>', 'i', t)
    t = (t.replace('_{d-1,i}', '(D-1)i').replace('_{d-1}', '(D-1)')
          .replace('_{D-1,i}', '(D-1)i').replace('_{D-1}', '(D-1)')
          .replace('_{p,i}', 'pi').replace('_i', 'i'))
    return t


def sec4(L):
    table = dict((s, {}) for s in SYMBOLS)
    plain = dict(zip(['B(D-1)i', 'A(D-1)i', 'M(D-1)i', 'D(D-1)i',
                      'SB(D-1)', 'SA(D-1)', 'SM(D-1)', 'SD(D-1)'], SYMBOLS))

    # (1) symbol_glossary.json
    sg = json.loads(rd(os.path.join(BASE, 'symbol_glossary.json')))
    alias = {}
    for s in sg['symbols']:
        canon = deflate(s['symbol'])
        if canon in plain:
            table[plain[canon]]['symbol_glossary.json'] = norm_name(s['ko_name'])
        alias[canon] = ([canon, s['symbol']]
                        + [a.split(' ')[0] for a in s.get('aliases', [])])
    # norm_name() 이 W금융일수 -> w금융일수 로 접는 근거. 사전이 두 표기를 스스로 등록해 두지 않았다면
    # 그 접기는 검사를 느슨하게 만드는 짓이 되므로, 근거가 사라지면 여기서 FAIL 이 난다.
    walias = [s for s in sg['symbols'] if s['symbol'] == 'w']
    chk('4', 'symbol_glossary.json 이 w금융일수 · W금융일수 두 표기를 같은 용어로 등록했음 '
             '(이름 대조에서 대소문자를 접는 근거)',
        len(walias) == 1
        and any(a.startswith('w금융일수') for a in walias[0].get('aliases', []))
        and any(a.startswith('W금융일수') for a in walias[0].get('aliases', [])),
        str(walias[0].get('aliases') if walias else None))
    chk('4', 'symbol_glossary.json 에서 8기호 전건 판독',
        len([s for s in SYMBOLS if 'symbol_glossary.json' in table[s]]) == 8,
        str([s for s in SYMBOLS if 'symbol_glossary.json' not in table[s]]))

    # (2) symbol_glossary.md — 기호 표의 「기호 | 한국어 이름」 칸
    md = deflate(rd(os.path.join(BASE, 'symbol_glossary.md')))
    for canon, flat in plain.items():
        m = re.search(r'^\|\s*`' + re.escape(canon) + r'`\s*\|\s*\*\*(.+?)\*\*\s*\|', md, re.M)
        if m:
            table[flat]['symbol_glossary.md'] = norm_name(m.group(1))
    chk('4', 'symbol_glossary.md 에서 8기호 전건 판독',
        len([s for s in SYMBOLS if 'symbol_glossary.md' in table[s]]) == 8,
        str([s for s in SYMBOLS if 'symbol_glossary.md' not in table[s]]))

    # (3) 배포 glossary.html — 기호 표의 「한국어 이름」 칸
    gh = deflate(rd(os.path.join(APP_REPO, 'glossary.html')))
    for canon, flat in plain.items():
        m = re.search(r'<code>' + re.escape(canon) + r'</code></td><td><b>(.+?)</b>', gh)
        if m:
            table[flat]['glossary.html(배포)'] = norm_name(m.group(1))
    chk('4', '배포 glossary.html 에서 8기호 전건 판독',
        len([s for s in SYMBOLS if 'glossary.html(배포)' in table[s]]) == 8,
        str([s for s in SYMBOLS if 'glossary.html(배포)' not in table[s]]))

    # (4) 용어정의서 원문 ceo_definitions.md — 「이름(기호)」 꼴
    cd = rd(os.path.join(BASE, 'ceo_definitions.md'))
    #     「대상정산금채권의 상환액(SBD-1)」 처럼 앞에 한정어가 붙는다. 공백 뒤 마지막 낱말이 이름이다.
    for flat in SYMBOLS:
        for m in re.finditer(r'([가-힣wW][가-힣A-Za-z0-9]*)\('
                             + re.escape(CEO_OF[flat]) + r'\)', cd):
            table[flat]['ceo_definitions.md'] = norm_name(m.group(1))
    named = [s for s in SYMBOLS if 'ceo_definitions.md' in table[s]]
    chk('4', '원문 ceo_definitions.md 에서 이름 붙은 기호 판독 0건 아님',
        len(named) > 0, '이름 붙은 기호 %s · 원문이 이름을 안 붙인 기호 %s'
        % (named, [s for s in SYMBOLS if s not in named]))

    # (5) 검산 엑셀 — 「기호 이름」 꼴 열머리
    import audit_xlsx_check as AX
    bk = AX.Book(XLSX)
    hit = 0
    # 열머리 행만 본다. 아무 칸이나 훑으면 `가중치 대조!B3` 「Ai 정의」 같은 산식 설명 줄이
    # 기호의 이름으로 새어 들어와 출처 대조가 '정의' 로 뭉개진다.
    for wsx in bk.wb.worksheets:
        for c in wsx[1]:
            v = c.value
            if not isinstance(v, str) or v.startswith('='):
                continue
            for flat in SYMBOLS:
                m = re.match(r'^' + re.escape(flat) + r'\s+([가-힣A-Za-z][가-힣A-Za-z0-9 ]*?)\s*(?:\(|$)', v.strip())
                # 곱 열 `Ai x Di` 는 뒤가 이름이 아니라 다른 기호다 — 이름으로 세지 않는다.
                if m and not any(s in m.group(1) for s in SYMBOLS):
                    table[flat]['검산 엑셀'] = norm_name(m.group(1))
                    hit += 1
    chk('4', '검산 엑셀에서 「기호 이름」 꼴 열머리 0건 아님', hit > 0, '%d곳' % hit)

    bad = name_check(table)
    chk('4', '기호 8개에 붙은 한글 이름이 출처별로 갈리지 않음', not bad, str(bad))
    chk('4', '기호 8개가 두 곳 이상에서 이름을 얻었음',
        all(len(table[s]) >= 2 for s in SYMBOLS),
        str({s: sorted(table[s]) for s in SYMBOLS if len(table[s]) < 2}))

    # 추가 — 산출물이 쓰는 기호 표기가 symbol_glossary.json 에 등록돼 있는가
    reg = set()
    for canon, forms in alias.items():
        for f in forms:
            reg.add(f)
    coined = {}
    for label, path in (('app.html', os.path.join(APP_REPO, 'app.html')),
                        ('시연본 index.html', os.path.join(PROTO_REPO, 'index.html'))):
        txt = rd(path)
        for m in re.finditer(r'\[2번 이미지\][^\n]*', txt):
            for tok in re.findall(r'\b(?:S?[A-Z]{1,3}(?:\(D-1\))?i?)\b', m.group(0)):
                if tok in ('D',):
                    continue
                if tok not in reg and tok in ('Bi', 'Mi', 'Ai', 'Di'):
                    coined.setdefault(tok, []).append(label)
    xl = set()
    bsh = bk.wb['채권']
    for c in range(1, bsh.max_column + 1):
        h = str(bsh.cell(1, c).value or '')
        for tok in re.findall(r'\b([A-Z][a-z]?i)\b', h):
            if tok not in reg:
                xl.add(tok)
                coined.setdefault(tok, []).append('검산 엑셀 채권 시트')
    chk('4', '[2번] 축 기호 표기가 전부 symbol_glossary.json 에 등록돼 있음',
        not coined, '미등록 표기 %s' % coined)

    # 추가 — 엑셀에서 기호를 머리에 단 열이 그 기호의 값을 담는가.
    #        「참고」·「차감 제외」 표식이 붙은 열은 통합문서 스스로 정본/참고를 가른 자리라 예외로 두되,
    #        표식 자체가 있는지를 판정한다(표식 없이 다른 값을 담으면 FAIL).
    ds = bk.wb['일별']
    misuse = []
    for c in range(1, ds.max_column + 1):
        h = str(ds.cell(1, c).value or '')
        if not h.startswith(SYM['SM']):
            continue
        f2 = str(ds.cell(2, c).value)
        holds_def = '채권!$O$' in f2                      # O = 투자수익 Mi = 수수료 - 차감
        marked = ('참고' in h) or ('차감 제외' in h)
        if not holds_def and not marked:
            misuse.append((ds.cell(1, c).coordinate, h, f2[:50]))
    # 머리에 그 기호를 단 열이 하나도 없으면 판정이 빈 채로 통과한다 — 그것도 FAIL 로 세운다.
    marked_cols = [c for c in range(1, ds.max_column + 1)
                   if str(ds.cell(1, c).value or '').startswith(SYM['SM'])]
    chk('4', '엑셀에서 %s 을 머리에 달고 다른 값을 담은 열에 「참고」 표식이 있음' % SYM['SM'],
        bool(marked_cols) and not misuse, '%s 열 %d개 · 표식 없는 오용 %s'
        % (SYM['SM'], len(marked_cols), misuse))
    return table


# ══════════════════════════════════════════════════════════════════
# 5. max(0, 미지급금 - 과지급금) 클램프
# ══════════════════════════════════════════════════════════════════
def sec5(L, drv):
    dl = L['ledger']

    # 5-a 현행 데이터에서 음수 차감 0건 + 클램프가 발동한 곳수
    neg = [r for r in dl.RECEIVABLES if r['ded'] < 0]
    bind = [r for r in dl.RECEIVABLES if r['unpaid'] < r['over']]
    chk('5', '원장에 음수 차감 0건', not neg, '음수 %d건' % len(neg))
    # 현행 데이터만으로는 클램프가 발동하지 않는다 — 그 사실 자체를 판정해 둔다.
    # 이 판정이 참인 동안은 「데이터에 음수가 없다」가 클램프의 증거가 될 수 없고,
    # 아래 행태 시험이 유일한 증거다. 요율표가 뒤집히면 이 판정이 FAIL 로 바뀌어 알려 준다.
    chk('5', '현행 요율표는 플랫폼 전건 미지급률 > 과지급률 (데이터만으로는 클램프가 발동하지 않음)',
        all(dl.OVERPAID[k] < dl.UNPAID[k] for k in dl.ORDER) and len(bind) == 0,
        '미지급<과지급 채권 %d건 / %d건 · 요율 %s'
        % (len(bind), len(dl.RECEIVABLES),
           [(k, str(dl.UNPAID[k]), str(dl.OVERPAID[k])) for k in dl.ORDER]))

    # 5-b 행태 시험 — 요율을 뒤집어 _build() 를 다시 돌린다
    up0, ov0 = dl.UNPAID, dl.OVERPAID
    try:
        dl.UNPAID, dl.OVERPAID = ov0, up0
        rebuilt = dl._build()
    finally:
        dl.UNPAID, dl.OVERPAID = up0, ov0
    flipped_neg = [r for r in rebuilt if r['ded'] < 0]
    flipped_nz = [r for r in rebuilt if r['ded'] != 0]
    flipped_bind = [r for r in rebuilt if r['unpaid'] < r['over']]
    chk('5', '원장 행태 시험 — 요율을 뒤집으면 미지급 < 과지급 채권이 실제로 생긴다',
        len(flipped_bind) > 0, '%d건 / %d건' % (len(flipped_bind), len(rebuilt)))
    chk('5', '원장 행태 시험 — 그 채권들의 차감이 전건 0 (음수가 흐르지 않음)',
        not flipped_neg and not flipped_nz,
        '음수 %d건 · 0 아닌 값 %d건 %s' % (len(flipped_neg), len(flipped_nz),
                                          [(r['unpaid'], r['over'], r['ded']) for r in flipped_nz[:3]]))
    rp = [D(r['net']) - D(r['ded']) for r in rebuilt]
    chk('5', '원장 행태 시험 — BD-1i 가 순지급액 그대로 (초과 과지급이 상환액에 얹히지 않음)',
        all(b == D(r['net']) for b, r in zip(rp, rebuilt)),
        '어긋난 채권 %d건' % sum(1 for b, r in zip(rp, rebuilt) if b != D(r['net'])))

    # 5-c 엑셀 수식 문언
    import audit_xlsx_check as AX
    bk = AX.Book(XLSX)
    bs = bk.wb['채권']
    hdr = dict((str(bs.cell(1, c).value), c) for c in range(1, bs.max_column + 1))
    cN, cL, cM = hdr.get('차감액'), hdr.get('미지급액'), hdr.get('과지급액')
    chk('5', '엑셀 채권 시트에 차감액·미지급액·과지급액 열이 있음',
        None not in (cN, cL, cM), str(sorted(hdr)))
    nbad, n = [], 0
    for r in range(2, bs.max_row + 1):
        if bs.cell(r, 1).value is None:
            continue
        n += 1
        f = str(bs.cell(r, cN).value).replace(' ', '')
        want = '=MAX(0,%s%d-%s%d)' % (CL(cL), r, CL(cM), r)
        if f != want:
            nbad.append(('%s%d' % (CL(cN), r), f, want))
    chk('5', '엑셀 차감액 열 = MAX(0, 미지급액 - 과지급액) 전행',
        n > 0 and not nbad, '%d행 · 어긋남 %d건 %s' % (n, len(nbad), nbad[:3]))

    # 5-d 엑셀 행태 시험 — 사본에서 미지급률 < 과지급률 로 바꿔 재평가
    import shutil
    tmp = os.path.join(tempfile.gettempdir(), 'phbs_clamp_probe.xlsx')
    shutil.copyfile(XLSX, tmp)          # 원본은 수정 금지 대상이라 사본에서만 만진다
    import openpyxl
    wb2 = openpyxl.load_workbook(tmp)
    inp = wb2['입력']
    # 요율 표의 자리를 세지 않는다. 입력 시트는 칸이 늘면 행이 밀린다 —
    # 2026-08-31 에 두 칸이 늘면서 24~27 이 28~31 로 밀려 이 시험이 통째로 헛돌았다.
    hrow = ucol = ocol = None
    for r in range(1, inp.max_row + 1):
        row = [str(inp.cell(r, c).value or '') for c in range(1, inp.max_column + 1)]
        if '미지급률' in row and '과지급률' in row:
            hrow, ucol, ocol = r, row.index('미지급률') + 1, row.index('과지급률') + 1
            break
    if hrow is None:
        raise SystemExit('!! 입력 시트에서 미지급률·과지급률 열을 못 찾았다')
    swapped = 0
    for r in range(hrow + 1, inp.max_row + 1):
        u, o = inp.cell(r, ucol).value, inp.cell(r, ocol).value
        if not isinstance(u, (int, float)) or not isinstance(o, (int, float)):
            break
        inp.cell(r, ucol).value, inp.cell(r, ocol).value = o, u
        swapped += 1
    if swapped == 0:
        raise SystemExit('!! 요율 행을 한 줄도 못 뒤집었다 — 표 모양이 바뀌었다')
    wb2.save(tmp)
    bk2 = AX.Book(tmp)
    nz, nn, chg = [], 0, 0
    for r in range(2, bs.max_row + 1):
        if bs.cell(r, 1).value is None:
            continue
        nn += 1
        l_ = bk2.cell('채권', '%s%d' % (CL(cL), r))
        m_ = bk2.cell('채권', '%s%d' % (CL(cM), r))
        v = bk2.cell('채권', '%s%d' % (CL(cN), r))
        if l_ < m_:
            chg += 1
        if v != 0:
            nz.append((r, l_, m_, v))
    os.unlink(tmp)
    chk('5', '엑셀 행태 시험 — 요율을 뒤집으면 미지급 < 과지급 행이 생긴다',
        chg > 0, '%d행 / %d행' % (chg, nn))
    chk('5', '엑셀 행태 시험 — 그 행들의 차감액이 전건 0',
        not nz, '0 아닌 행 %d건 %s' % (len(nz), nz[:3]))

    # 5-e 화면 행태 시험 — 시뮬레이션 미지급률 < 과지급률
    #   원본(app)만 — 시연본은 투자 시뮬레이션을 뺐다(step7 2026-09-04 · PROTO_DROPPED). 3절이 없음을 판정한다.
    for key in ('app',):
        t = drv['targets'][key]
        neg = t['simClampNeg']['bonds']['body']
        pos = t['simClampPos']['bonds']['body']
        chk('5', '[%s] 시뮬 클램프 시험 행 0건 아님' % key,
            len(neg) > 0 and len(pos) > 0, '음수쪽 %d행 · 양수쪽 %d행' % (len(neg), len(pos)))
        # 열: #, 구분, 플랫폼, 순지급액, 금융일수, 투자실행금, 채권매입수수료, 미지급 차감, 투자수익, 상환액
        bad = [r for r in neg if num(r[7]) != 0]
        chk('5', '[%s] 미지급률 < 과지급률 -> 미지급 차감 전건 0' % key,
            not bad, '0 아닌 행 %d건 %s' % (len(bad), bad[:2]))
        badb = [r for r in neg if num(r[9]) != num(r[3])]
        chk('5', '[%s] 그때 BD-1i = 순지급액 (초과 과지급이 얹히지 않음)' % key,
            not badb, '어긋난 행 %d건 %s' % (len(badb), badb[:2]))
        badm = [r for r in neg if num(r[8]) != num(r[6])]
        chk('5', '[%s] 그때 MD-1i = 채권매입수수료' % key,
            not badm, '어긋난 행 %d건 %s' % (len(badm), badm[:2]))
        # 판별력 — 컨트롤이 죽지 않았는가
        pz = [r for r in pos if num(r[7]) <= 0]
        chk('5', '[%s] 미지급률 > 과지급률 -> 미지급 차감 전건 > 0 (컨트롤 살아 있음)' % key,
            not pz, '0 이하 행 %d건 %s' % (len(pz), pz[:2]))
        # 입력칸이 0.20 을 0.2 로 정규화하므로 수치로 맞춘다
        ni, pi = t['simClampNeg']['inputs'], t['simClampPos']['inputs']
        chk('5', '[%s] 두 시험의 입력값이 실제로 화면에 들어갔음' % key,
            float(ni['unpaid']) == 0.01 and float(ni['over']) == 0.08
            and float(pi['unpaid']) == 0.20 and float(pi['over']) == 0.01
            and float(ni['unpaid']) < float(ni['over']) and float(pi['unpaid']) > float(pi['over']),
            '%s / %s' % (ni, pi))
        # 채권 1건 항등식이 화면에서도 성립 (내림 2회 -> [0, 2) 원)
        allb = t['simSeed']['bonds']['body']
        res = [(num(r[9]) - num(r[5]) - num(r[8])) for r in allb]
        chk('5', '[%s] 화면 채권 1건 BD-1i - AD-1i - MD-1i 가 [0,2)원 (내림 2회)' % key,
            allb and all(0 <= x < 2 for x in res), '잔차 %s' % sorted(set(res)))


# ══════════════════════════════════════════════════════════════════
# 6. 자기시험 — 일부러 깨뜨린 값을 검사기가 잡는가
# ══════════════════════════════════════════════════════════════════
def sec6(L, s1):
    dl = L['ledger']
    rate = dl.RATE

    # 6-a 채권층
    sample = [dict(r) for r in dl.RECEIVABLES[:2000]]
    clean = bond_residuals(sample, rate)
    sample[7]['net'] += 1
    dirty = bond_residuals(sample, rate)
    chk('6', '자기시험 채권층 — 깨끗한 벌은 통과',
        all(abs(x) <= D('0.5') for x in clean), '최대 %s' % max(clean, key=lambda x: abs(x)))
    chk('6', '자기시험 채권층 — net+1 을 심으면 잡힌다',
        any(abs(x) > D('0.5') for x in dirty),
        '최대 %s' % max(dirty, key=lambda x: abs(x)))

    # 6-b/c 하루층 · 전 구간층
    days = [dict(g) for g in s1['days']]
    chk('6', '자기시험 하루층 — 깨끗한 벌은 통과',
        not any(day_identity(days, rate).values()), '')
    d2 = [dict(g) for g in days]
    d2[3]['profit'] += 1
    b2 = day_identity(d2, rate)
    chk('6', '자기시험 하루층 — 투자수익 +1 을 심으면 잡힌다',
        bool(b2['profit']), str({k: len(v) for k, v in b2.items()}))
    d3 = [dict(g) for g in days]
    d3[10]['repay'] -= 1
    b3 = day_identity(d3, rate)
    chk('6', '자기시험 하루층 — 상환액 -1 을 심으면 잡힌다', bool(b3['repay']),
        str({k: len(v) for k, v in b3.items()}))
    lim = int(D('0.5') * s1['total']['n'])
    t2 = dict(s1['total'])
    t2['ai'] += lim + 1000                 # 전 구간 잔차 한계(0.5 x n)를 확실히 넘기는 크기
    b4 = day_identity([t2], rate, nfloor=len(days))
    chk('6', '자기시험 전 구간층 — 투자실행금을 한계+1,000원 밀면 잡힌다', bool(b4['def']),
        '심은 값 %d원 · 한계 %d원' % (lim + 1000, lim))
    t3 = dict(s1['total'])
    t3['profit'] -= len(days) + 10         # 내림 허용폭(일수)을 넘기는 크기
    b5 = day_identity([t3], rate, nfloor=len(days))
    chk('6', '자기시험 전 구간층 — 투자수익을 내림 허용폭 밖으로 밀면 잡힌다',
        bool(b5['profit']), '허용폭 (-%d, 0]' % len(days))

    # 6-d 가중치 검사기
    rows = [{'ai': 100, 'net': 100, 'di': 1}, {'ai': 1, 'net': 1, 'di': 10}]
    wai = w_by_weight(rows, 'ai')
    wcn = w_by_weight(rows, 'cnt')
    chk('6', '자기시험 가중 — 금액 가중과 건수 가중이 다른 값을 낸다',
        r2(wai) != r2(wcn), '금액 %s · 건수 %s' % (r2(wai), r2(wcn)))
    chk('6', '자기시험 가중 — 건수 가중을 정답으로 두면 어긋난다',
        r2(wcn) != r2(wai), '')

    # 6-e 클램프 검사기
    cases = [(100, 30), (30, 100), (50, 50)]
    chk('6', '자기시험 클램프 — 올바른 구현은 통과',
        not clamp_check(clamp_eval, cases), '')
    chk('6', '자기시험 클램프 — max 를 뺀 구현은 잡힌다',
        bool(clamp_check(clamp_broken, cases)), str(clamp_check(clamp_broken, cases)))

    # 6-f 이름 검사기
    chk('6', '자기시험 이름 — 같은 이름뿐이면 통과',
        not name_check({'SBD-1': {'a': '상환액', 'b': '상환액'}}), '')
    chk('6', '자기시험 이름 — 다른 이름을 심으면 잡힌다',
        bool(name_check({'SBD-1': {'a': '상환액', 'b': '회수액'}})), '')

    # 6-g ty 행 검사기
    good = [{'d': 'x', 'exec': 25941773, 'profit': 9871, 'w': 3.06, 'ty': 4.54}]
    badr = [{'d': 'x', 'exec': 25941773, 'profit': 9871, 'w': 3.06, 'ty': 4.99}]
    chk('6', '자기시험 ty — 맞는 행은 통과', not ty_rows_check(good), '')
    chk('6', '자기시험 ty — 틀린 행을 심으면 잡힌다', bool(ty_rows_check(badr)), str(ty_rows_check(badr)))

    # 6-h 파일 판독기 — 없는 파일은 예외(=FAIL)이지 빈 문자열이 아니다
    raised = False
    try:
        rd(os.path.join(BASE, '__없는파일__.md'))
    except IOError:
        raised = True
    chk('6', '자기시험 판독기 — 없는 파일은 예외로 끊긴다 (빈 문자열로 넘어가지 않음)', raised, '')
    raised2 = False
    try:
        empty = os.path.join(tempfile.gettempdir(), 'phbs_empty.md')
        open(empty, 'w').close()
        rd(empty)
    except IOError:
        raised2 = True
    finally:
        if os.path.exists(empty):
            os.unlink(empty)
    chk('6', '자기시험 판독기 — 빈 파일도 예외로 끊긴다', raised2, '')


# ══════════════════════════════════════════════════════════════════
def main():
    load = {}
    try:
        import daily_ledger
        load['ledger'] = daily_ledger
        load['facts'] = json.loads(rd(os.path.join(BASE, 'ledger_facts.json')))
        chk('0', '원장·사실값 적재', True, '채권 %d건 · 원장 %d일'
            % (len(daily_ledger.RECEIVABLES), len(daily_ledger.LEDGER)))
    except Exception as e:
        chk('0', '원장·사실값 적재', False, '%s: %s' % (type(e).__name__, e))
        dump()
        return 1

    drv = None
    try:
        drv = run_driver()
        chk('0', '헤드리스 크롬 조작 (app · proto)', True,
            '화면 %d곳 · 콘솔 %d건' % (len(drv['targets']), len(drv.get('console') or [])))
        chk('0', '화면 콘솔 오류 0건', not (drv.get('console') or []), str((drv.get('console') or [])[:3]))
    except Exception as e:
        chk('0', '헤드리스 크롬 조작 (app · proto)', False,
            '%s: %s' % (type(e).__name__, str(e)[:400]))

    s1 = None
    for name, fn in (('1', lambda: sec1(load)),
                     ('2', lambda: sec2(load, drv)),
                     ('3', lambda: sec3(load, drv, s1)),
                     ('4', lambda: sec4(load)),
                     ('5', lambda: sec5(load, drv)),
                     ('6', lambda: sec6(load, s1))):
        try:
            out = fn()
            if name == '1':
                s1 = out
        except Exception as e:
            chk(name, '항목 %s 실행' % name, False,
                '%s: %s | %s' % (type(e).__name__, str(e)[:300],
                                 traceback.format_exc().strip().splitlines()[-3:]))
    return dump()


def dump():
    fails = [r for r in R if not r['pass']]
    by = {}
    for r in R:
        b = by.setdefault(r['sec'], [0, 0])
        b[0] += 1
        b[1] += 0 if r['pass'] else 1
    out = {'total': len(R), 'fail': len(fails),
           'bySection': dict((k, {'검사': v[0], 'FAIL': v[1]}) for k, v in sorted(by.items())),
           'cases': R}
    with io.open(os.path.join(BASE, 'verify_batch_symbols_result.json'), 'w', encoding='utf-8') as fp:
        fp.write(json.dumps(out, ensure_ascii=False, indent=1))
    print('검사 %d건 · FAIL %d건' % (len(R), len(fails)))
    for k in sorted(by):
        print('  항목 %s  검사 %3d  FAIL %d' % (k, by[k][0], by[k][1]))
    if fails:
        print()
        for f in fails:
            print('FAIL [%s] %s' % (f['sec'], f['name']))
            print('      %s' % f['detail'][:400])
    return 1 if fails else 0


if __name__ == '__main__':
    sys.exit(main())
