#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""금지 낱말·어투 기계 검사.

목록은 dm_0901/banned_words.md 한 곳이다. 이 파일에는 낱말을 적지 않는다.
목록이 바뀌면 검사가 따라 바뀐다.

검사 대상 — 사람이 읽는 산출물과 그 원고
  ~/Downloads/payhug_용어정의서/*.docx *.html
  ~/Downloads/payhug_검산엑셀/*.xlsx
  payhug-investor-admin/*.html
  final_terms.json · termsdoc_seed.json · meeting_0901/steps_all.json
  glossary_manuscript.md · capability_manuscript.md · ceo_inquiry.md
  feasibility.md · ceoq_seed.json

목록 읽는 법
  「## 쓰지 않는 말과 대신 쓸 말」 표에서 1열의 백틱 낱말을 금지어로, 2열을 대체어로 읽는다.
  2열에 「그 자리에서」 또는 「제목」이 들어 있으면 그 행은 자리 한정 검사다 —
  제목·열 이름·라벨·단독 칸에 선 것만 잡고 문장 안은 두다.
  1열에 물결표(~)나 괄호 조건이 붙은 행은 기계로 가르지 못해 수동 항목으로 센다.
  낱글자 한 자짜리 금지어는 앞뒤가 글자·숫자가 아닐 때만 잡는다.

가려내는 자리
  인용    대표 정의서 원문 그대로인 대목 · quote/orig/ceo_source 필드 · 출처를 단 따옴표 인용
  기존표기 대조표의 「기존 표기」 「옛 표기」 「현행 표기」 칸
  대체어  이 목록이 대신 쓰라고 한 말 (S입금부족율 · 투자실행액 · 화면 표시값 …)
  미확정  「미확정」 표시는 미확정 3대(C1·C2·C4) 때문에 붙인 것이라 잡지 않는다

어투 검사
  값 뒤에 대시로 설명을 붙인 칸 · 마크다운 굵게·백틱이 글자로 찍힌 자리 ·
  값 없는 칸을 —, -, N/A 로 채운 자리 · 가운뎃점으로 나열하고 종결 없이 끝낸 문장

돌리는 법
  python3 verify_banned.py            최신 판
  python3 verify_banned.py --all      날짜 박은 지난 판까지
  python3 verify_banned.py --self     판별력 시험 (심어 놓고 잡는지)
"""

import glob
import html as H
import io
import json
import os
import re
import sys
import zipfile

PIPE = os.path.dirname(os.path.abspath(__file__))
REPO = '/Users/semi/cursor/payhug-investor-admin'
DL = os.path.expanduser('~/Downloads')
LIST = os.path.join(PIPE, 'dm_0901', 'banned_words.md')
CEO = os.path.join(PIPE, 'ceo_definitions.md')

# 다른 조가 손대는 자리 — 세기만 하고 소관을 적는다
OWNER = {
    'final_terms.json': '설명 문장 조',
    'final-terms.html': '설명 문장 조',
    '용어기호정리': '설명 문장 조',
    'app.html': '화면·계산식 조',
    'calc.html': '화면·계산식 조',
    'invest-sim': '화면·계산식 조',
    'invest-profit': '화면·계산식 조',
}


def owner(path):
    base = os.path.basename(path)
    for k, v in OWNER.items():
        if base.startswith(k) or base == k:
            return v
    return ''


# ══════════════════════════════════════════════════════════════════
#  1. 목록 읽기 — dm_0901/banned_words.md
# ══════════════════════════════════════════════════════════════════

TICK = re.compile(r'`([^`]+)`')


def _rows(md, head):
    """「## head」 절의 첫 표를 [(칸,칸,…)] 로 돌려준다."""
    m = re.search(r'^##\s+' + re.escape(head) + r'\s*$', md, re.M)
    if not m:
        raise SystemExit('목록에 「%s」 절이 없다: %s' % (head, LIST))
    body = md[m.end():]
    nx = re.search(r'^##\s', body, re.M)
    if nx:
        body = body[:nx.start()]
    out = []
    for ln in body.split('\n'):
        ln = ln.strip()
        if not ln.startswith('|'):
            continue
        cells = [c.strip() for c in ln.strip('|').split('|')]
        if all(re.fullmatch(r':?-{2,}:?', c) for c in cells):
            continue
        out.append(cells)
    return out[1:] if out else []          # 첫 줄은 열 이름


def load_list():
    md = io.open(LIST, encoding='utf-8').read()
    words, subs, manual = [], set(), []
    for cells in _rows(md, '쓰지 않는 말과 대신 쓸 말'):
        raw, alt = cells[0], (cells[1] if len(cells) > 1 else '')
        note = ' '.join(cells[1:])
        # 2열이 「그 자리에서」·「제목」이라고 하면 제목·열 이름·라벨 자리만 본다
        spot = ('그 자리에서' in alt) or ('제목' in alt)
        # 1열이 「단독」이라고 하면 그 칸이 통째로 그 글자일 때만 본다
        alone = '단독' in raw
        for t in TICK.findall(raw):
            if t.startswith('~') or t.startswith('#'):
                manual.append(t)
                continue
            words.append(dict(w=t, alt=alt, spot=spot, alone=alone, note=note))
        for t in TICK.findall(alt):
            subs.add(t)
    tone = []
    for cells in _rows(md, '어투'):
        tone.append((cells[0], cells[1] if len(cells) > 1 else ''))
    # 긴 낱말을 먼저 본다 — 「갈려 나온다」가 「갈려」에 먹히지 않게
    words.sort(key=lambda d: -len(d['w']))
    return words, subs, manual, tone


# ══════════════════════════════════════════════════════════════════
#  2. 대상 파일 → (자리, 글) 조각
#     자리 kind: 'head'(제목·열이름·라벨·단독칸) · 'text'(문장) · 'quote'(인용)
# ══════════════════════════════════════════════════════════════════

CEOT = re.sub(r'\s+', '', io.open(CEO, encoding='utf-8').read())

# 출처를 단 따옴표 인용 — "…"(`SB s3`) · 「…」 + 각주·원문·스토리보드 표식
CITED = re.compile(r'[“"「]([^”"」\n]{8,})[”"」]')
# 백틱 인용 뒤에 괄호로 출처를 단 자리 — `…` (2026-08-28 미팅 결론)
CITEDT = re.compile(r'`([^`\n]{8,})`\s*[(（]')
SRCMARK = re.compile(r'각주|원문|스토리보드|약관|제\d+조|정의서|SB s\d|노드 \d')
QKEY = re.compile(r'quote|orig|ceo_source|원문|인용', re.I)
OLDCOL = re.compile(r'기존 ?표기|옛 ?표기|현행 ?표기|바꾸기 전|갈아 ?끼우기 전|before')


def _is_ceo(s):
    f = re.sub(r'\s+', '', s)
    return len(f) >= 8 and f in CEOT


def frag(kind, text, where, js=False):
    return dict(kind=kind, t=text, at=where, js=js)


def read_md(p):
    out = []
    lines = io.open(p, encoding='utf-8').read().split('\n')
    fence = None
    tbl_head = None
    for i, ln in enumerate(lines, 1):
        at = '%s:%d' % (os.path.basename(p), i)
        f = re.match(r'^\s*(```|~~~)(.*)$', ln)
        if f:
            fence = None if fence else (f.group(2).strip() or ' ')
            continue
        if fence is not None:
            out.append(frag('quote' if fence.startswith('원문') else 'text', ln, at))
            continue
        if ln.startswith('#'):
            out.append(frag('head', ln.lstrip('#').strip(), at))
            continue
        if ln.strip().startswith('|'):
            cells = [c.strip() for c in ln.strip().strip('|').split('|')]
            if all(re.fullmatch(r':?-{2,}:?', c) for c in cells if c):
                tbl_head = tbl_head or []
                continue
            if tbl_head is None:
                tbl_head = cells                      # 열 이름 줄
                for c in cells:
                    out.append(frag('head', c, at))
                continue
            for j, c in enumerate(cells):
                col = tbl_head[j] if j < len(tbl_head) else ''
                k = 'quote' if OLDCOL.search(col) or OLDCOL.search(cells[0]) else 'text'
                if k == 'text' and len(c) <= 8 and c:
                    k = 'head'                        # 짧은 칸은 라벨 자리로 본다
                out.append(frag(k, c, at))
            continue
        tbl_head = None
        lab = re.match(r'^\s*[-*]?\s*\*\*([^*]{1,20})\*\*', ln)
        #   굵게 쓴 문장 첫머리(「투자자 몫이다.」)는 라벨이 아니다 — 종결 어미로 가른다
        if lab and not re.search(r'(다|까|요)\.?$', lab.group(1).strip()):
            out.append(frag('head', lab.group(1), at))
        out.append(frag('text', ln, at))
    return out


def strip_html(t):
    t = re.sub(r'<style\b.*?</style>', ' ', t, flags=re.S | re.I)
    t = re.sub(r'<!--.*?-->', ' ', t, flags=re.S)
    return t


def read_html(p):
    src = strip_html(io.open(p, encoding='utf-8').read())
    out = []
    # <q>·<blockquote> 는 인용 자리다 — 글자 그대로 두는 곳이라 잡지 않는다
    for m in re.finditer(r'<(q|blockquote)\b[^>]*>(.*?)</\1>', src, re.S | re.I):
        out.append(frag('quote', H.unescape(re.sub(r'<[^>]+>', ' ', m.group(2))),
                        os.path.basename(p)))
    src = re.sub(r'<(q|blockquote)\b[^>]*>.*?</\1>', ' ', src, flags=re.S | re.I)
    for m in re.finditer(r'<(th|h1|h2|h3|h4|h5|h6|caption|legend)\b[^>]*>(.*?)</\1>',
                         src, re.S | re.I):
        out.append(frag('head', H.unescape(re.sub(r'<[^>]+>', ' ', m.group(2))),
                        os.path.basename(p)))
    rest = re.sub(r'<(th|h1|h2|h3|h4|h5|h6|caption|legend)\b[^>]*>.*?</\1>', ' ',
                  src, flags=re.S | re.I)
    # 글줄 안 태그는 지우고 블록 태그만 줄바꿈으로 — 인용 부호가 두 줄로 갈리지 않게
    rest = re.sub(r'</?(b|i|u|em|strong|code|span|a|sub|sup|small|mark|abbr|kbd)\b[^>]*>',
                  '', rest, flags=re.I)
    #   <script> 안은 화면에 글자로 뜨지 않는 코드가 섞인다 — 금지어는 보되 어투는 안 본다
    # 편집판이 자기 원고를 품는 블록은 화면에 안 나온다 — 낱말 검사에서 뺀다.
    rest = re.sub(r'(?is)<script[^>]*id="(?:seed|src)"[^>]*>.*?</script>', ' ', rest)
    scripts = re.findall(r'<script\b[^>]*>(.*?)</script>', rest, re.S | re.I)
    rest = re.sub(r'<script\b[^>]*>.*?</script>', ' ', rest, flags=re.S | re.I)
    body = H.unescape(re.sub(r'<[^>]+>', '\n', rest))
    for ln in body.split('\n'):
        if ln.strip():
            out.append(frag('text', ln, os.path.basename(p)))
    for sc in scripts:
        for ln in H.unescape(sc).split('\n'):
            if ln.strip():
                out.append(frag('text', ln, os.path.basename(p), js=True))
    return out


def read_json(p):
    data = json.load(io.open(p, encoding='utf-8'))
    out = []
    base = os.path.basename(p)

    # 화면에 안 나오는 내부 분류 키 — 생성기가 갈래를 가르는 데만 쓴다.
    # 값이 문서에 찍히지 않으므로 낱말 검사에서 뺀다.
    INNER = {'kind'}

    def walk(node, path, quoted):
        if isinstance(node, dict):
            for k, v in node.items():
                if str(k) in INNER:
                    continue
                out.append(frag('head', str(k), '%s %s' % (base, path)))
                walk(v, path + '.' + str(k), quoted or bool(QKEY.search(str(k))))
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, '%s[%d]' % (path, i), quoted)
        elif isinstance(node, str):
            k = 'quote' if (quoted or _is_ceo(node)) else \
                ('head' if len(node) <= 8 else 'text')
            out.append(frag(k, node, '%s %s' % (base, path)))
    walk(data, '', False)
    return out


def read_docx(p):
    z = zipfile.ZipFile(p)
    out = []
    for n in z.namelist():
        if not (n.startswith('word/') and n.endswith('.xml')):
            continue
        xml = z.read(n).decode('utf-8')
        for m in re.finditer(r'<w:(p|tc)\b.*?</w:\1>', xml, re.S):
            t = re.sub(r'<[^>]+>', '', m.group(0))
            t = H.unescape(t).strip()
            if not t:
                continue
            k = 'quote' if _is_ceo(t) else ('head' if len(t) <= 8 else 'text')
            out.append(frag(k, t, os.path.basename(p)))
    return out


def read_xlsx(p):
    from openpyxl import load_workbook
    wb = load_workbook(p, read_only=True, data_only=False)
    out = []
    base = os.path.basename(p)
    for ws in wb:
        for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
            for ci, v in enumerate(row, 1):
                if not isinstance(v, str) or not v.strip():
                    continue
                at = '%s %s!R%dC%d' % (base, ws.title, ri, ci)
                if _is_ceo(v):
                    out.append(frag('quote', v, at))
                elif len(v) <= 10:
                    out.append(frag('head', v, at))
                else:
                    out.append(frag('text', v, at))
    wb.close()
    return out


READER = {'.md': read_md, '.html': read_html, '.json': read_json,
          '.docx': read_docx, '.xlsx': read_xlsx}


STAMP = re.compile(r'^(?P<n>.+?)_(?P<d>\d{8})(?P<t>_[0-9A-Za-z]+)?(?P<e>\.[a-z]+)$')


def newest(paths):
    """날짜를 박은 산출물은 최신 판만 본다 — 지난 판은 이미 나간 종이다."""
    best, plain = {}, []
    for p in paths:
        m = STAMP.match(os.path.basename(p))
        if not m:
            plain.append(p)
            continue
        key = (m.group('n'), m.group('t') or '', m.group('e'))
        if key not in best or m.group('d') > best[key][0]:
            best[key] = (m.group('d'), p)
    return plain + [v[1] for v in best.values()]


def targets(allver=False):
    ps = []
    ps += sorted(glob.glob(os.path.join(DL, 'payhug_용어정의서', '*.docx')))
    ps += sorted(glob.glob(os.path.join(DL, 'payhug_용어정의서', '*.html')))
    ps += sorted(glob.glob(os.path.join(DL, 'payhug_검산엑셀', '*.xlsx')))
    ps += sorted(glob.glob(os.path.join(REPO, '*.html')))
    for n in ['final_terms.json', 'termsdoc_seed.json', 'meeting_0901/steps_all.json',
              'glossary_manuscript.md', 'capability_manuscript.md', 'ceo_inquiry.md',
              'feasibility.md', 'ceoq_seed.json']:
        ps.append(os.path.join(PIPE, n))
    ps = [p for p in ps if os.path.exists(p)]
    return sorted(ps if allver else newest(ps))


# ══════════════════════════════════════════════════════════════════
#  3. 금지어 검사
# ══════════════════════════════════════════════════════════════════

WORD = re.compile(r'[0-9A-Za-z가-힣]')


def hits(fr, words, subs):
    """조각 하나에서 금지어를 찾는다. (낱말, 자리, 문맥) 목록."""
    t = fr['t']
    if fr['kind'] == 'quote':
        return []
    # 대체어가 덮은 자리는 가린다 — S입금부족율 안의 S 처럼
    mask = list(t)
    for s in sorted(subs, key=len, reverse=True):
        for m in re.finditer(re.escape(s), t):
            for i in range(m.start(), m.end()):
                mask[i] = '\x00'
    # 출처를 밝힌 따옴표 인용도 가린다 — 인용은 글자 그대로 두는 자리다
    if SRCMARK.search(t):
        for m in CITED.finditer(t):
            for i in range(m.start(1), m.end(1)):
                mask[i] = '\x00'
    for m in CITEDT.finditer(t):
        for i in range(m.start(1), m.end(1)):
            mask[i] = '\x00'
    # 대표 원문 그대로인 대목도 가린다
    if _is_ceo(t):
        return []
    masked = ''.join(mask)
    stripped = t.strip()
    out = []
    taken = [False] * len(t)
    for d in words:
        w = d['w']
        if d.get('alone'):
            # 낱글자 단독 — 그 칸이 통째로 그 글자일 때만
            if stripped != w or fr['kind'] != 'head':
                continue
            out.append((w, fr['kind'], stripped))
            continue
        for m in re.finditer(re.escape(w), masked):
            a, b = m.start(), m.end()
            if any(taken[a:b]):
                continue
            if len(w) == 1 and w.isascii():
                # 낱글자 기호는 앞뒤가 글자·숫자면 다른 낱말 속이다. 한글은 조사가 붙으니 빼지 않는다
                if (a and WORD.match(t[a - 1])) or (b < len(t) and WORD.match(t[b])):
                    continue
            if d['spot'] and fr['kind'] != 'head':
                continue
            for i in range(a, b):
                taken[i] = True
            out.append((w, fr['kind'], t[max(0, a - 34):b + 34].strip()))
    return out


# ══════════════════════════════════════════════════════════════════
#  4. 어투 검사
# ══════════════════════════════════════════════════════════════════

# 값 뒤에 대시로 설명을 붙인 자리 — 대시 뒤가 완결된 설명 문장일 때만 잡는다.
#   「80,000,000원 — 현황 표 같은 칸」처럼 자리·범위를 가리키는 대시는 설명이 아니다.
TONE_DASH = re.compile(r'[0-9][0-9,.]*\s*(원|%|%p|일|건|배|곳|명)\s+—\s+'
                       r'(?P<tail>[^—\n]{12,}?(다|까|요))\s*[.]?\s*$')
TONE_MD = re.compile(r'\*\*[^*\n]{1,40}\*\*|`[^`\n]{1,40}`')
TONE_EMPTY = re.compile(r'^\s*(—|-|–|N/?A|없음\s*—)\s*$')
TONE_LIST = re.compile(r'^[^\n]{40,}·[^\n]*[가-힣A-Za-z0-9)\]」』]$')
# 조사가 붙어 문장으로 읽히는 자리만 — 목차·경로 나열은 뺀다
JOSA = re.compile(r'[은는이가을를로써]\s|[에서와과의]\s\S')
ENDING = re.compile(r'(다|음|임|함|것|중|기준|여부|까지|이상|이하)$')


def tone(fr, ext):
    t = fr['t'].strip()
    out = []
    if TONE_DASH.search(t):
        out.append(('값 뒤에 대시로 설명', t[:80]))
    if ext in ('.html', '.docx', '.xlsx') and not fr.get('js') and TONE_MD.search(t):
        out.append(('마크다운 표시가 글자로 찍힘', t[:80]))
    if TONE_EMPTY.match(t):
        out.append(('값 없는 칸을 채움', t[:40]))
    if fr['kind'] == 'text' and not fr.get('js') and TONE_LIST.match(t) \
            and JOSA.search(t) and not ENDING.search(t) \
            and not t.endswith(('.', ':', '?', '!', ')', '」', '』')):
        out.append(('가운뎃점 나열에 종결이 없음', t[:80]))
    return out


# ══════════════════════════════════════════════════════════════════
#  5. 실행
# ══════════════════════════════════════════════════════════════════

#  기계로 잘라 말하지 못하는 갈래 — 세되 통과 여부는 여기에 걸지 않는다.
#    값 없는 칸: 배포 화면이 실제로 「-」를 찍는 자리가 있어 문서와 갈라 볼 수 없다.
#    가운뎃점 나열: 목차·경로 나열과 설명 문장을 글자만으로 가르지 못한다.
SOFT = {'값 없는 칸을 채움', '가운뎃점 나열에 종결이 없음'}


def scan(paths=None, allver=False):
    words, subs, manual, tone_rows = load_list()
    paths = paths or targets(allver)
    wcnt, wfile, samples = {}, {}, {}
    tcnt, tsample = {}, {}
    for p in paths:
        ext = os.path.splitext(p)[1].lower()
        rd = READER.get(ext)
        if not rd:
            continue
        try:
            frs = rd(p)
        except Exception as e:                       # noqa: BLE001
            print('  읽기 실패 %s — %s' % (p, e))
            continue
        for fr in frs:
            for w, kind, ctx in hits(fr, words, subs):
                wcnt[w] = wcnt.get(w, 0) + 1
                wfile.setdefault(w, {})
                key = os.path.basename(p)
                wfile[w][key] = wfile[w].get(key, 0) + 1
                samples.setdefault(w, []).append('%s | %s' % (fr['at'], ctx))
            for lab, ctx in tone(fr, ext):
                tcnt[lab] = tcnt.get(lab, 0) + 1
                tsample.setdefault(lab, []).append('%s | %s' % (fr['at'], ctx))
    return words, manual, wcnt, wfile, samples, tcnt, tsample


def report(paths=None, quiet=False, allver=False):
    words, manual, wcnt, wfile, samples, tcnt, tsample = scan(paths, allver)
    total = sum(wcnt.values())
    if not quiet:
        print('금지 낱말 — 목록 %s · 검사 %d개 낱말'
              % (os.path.relpath(LIST, PIPE), len(words)))
        print('%-22s %6s  %s' % ('낱말', '건수', '자리'))
        print('-' * 78)
        for d in words:
            w = d['w']
            n = wcnt.get(w, 0)
            if not n:
                continue
            own = {}
            for f, c in wfile[w].items():
                o = owner(f) or '이 조'
                own[o] = own.get(o, 0) + c
            spot = '제목·열이름만' if d['spot'] else '전건'
            print('%-22s %6d  %s  [%s]'
                  % (w, n, ' · '.join('%s %d' % kv for kv in sorted(own.items())), spot))
            for s in samples[w][:3]:
                print('        %s' % s[:150])
        if manual:
            print('\n수동 항목 (기계로 가르지 못함) — %s' % ' · '.join(manual))
        print('\n어투')
        print('-' * 78)
        if not tcnt:
            print('  0건')
        for lab, n in sorted(tcnt.items(), key=lambda x: -x[1]):
            mark = '  (참고)' if lab in SOFT else ''
            print('%-28s %6d%s' % (lab, n, mark))
            for s in tsample[lab][:3]:
                print('        %s' % s[:150])
        hard = sum(n for lab, n in tcnt.items() if lab not in SOFT)
        soft = sum(n for lab, n in tcnt.items() if lab in SOFT)
        print('\n금지 낱말 %d건 · 어투 %d건 (참고 %d건은 셈에서 뺀다)'
              % (total, hard, soft))
    hard = sum(n for lab, n in tcnt.items() if lab not in SOFT)
    return total, hard, wcnt, tcnt


# ══════════════════════════════════════════════════════════════════
#  판별력 시험 — 심어 놓고 잡는지, 가려낼 자리에선 조용한지
# ══════════════════════════════════════════════════════════════════

def selftest():
    import shutil
    import tempfile
    words, subs, manual, _ = load_list()
    tmp = tempfile.mkdtemp(prefix='banned_')
    ok, bad = 0, []

    def probe(name, text, expect, kind_hint=None):
        nonlocal ok
        p = os.path.join(tmp, name)
        io.open(p, 'w', encoding='utf-8').write(text)
        _, _, wc, _, _, _, _ = scan([p])
        got = sum(wc.values())
        if (got > 0) == expect:
            ok += 1
        else:
            bad.append('%s — 잡길 %s 했는데 %d건' % (name, '기대' if expect else '안 기대', got))

    W = [d['w'] for d in words]
    pick = lambda s: next(w for w in W if w == s)

    # (가) 심으면 잡는가
    probe('p1.md', '이 값은 %s 로 만든다.\n' % pick('재료'), True)
    probe('p2.md', '| 갈래 | 뜻 |\n|---|---|\n| %s | 현재일자 |\n' % pick('상수'), True)
    probe('p3.md', '## %s 풀이\n' % pick('낱개'), True)
    probe('p4.md', '가맹점별 표의 %s 열을 본다고 적는다.\n' % pick('투자금액'), True)
    probe('p5.md', '| 항목 | 값 |\n|---|---|\n| %s | 3.04 |\n' % pick('몫'), True)
    probe('p6.json', json.dumps({'note': '두 갈래로 %s 나온다' % pick('갈려')},
                                ensure_ascii=False), True)
    probe('p7.html', '<table><tr><th>%s</th></tr></table>' % pick('화면 값'), True)

    # (나) 가려낼 자리에서 조용한가
    probe('q1.md', '```원문:\n- w금융일수 = Σ Ai x Di / Σ Ai\n```\n', False)
    probe('q2.md', '| 기존 표기 | 정본 |\n|---|---|\n| %s | 화면 표시값 |\n'
          % pick('화면 값'), False)
    probe('q3.md', '화면 열머리는 `S입금부족율` 이고 `w금융일수` 는 그 옆이다.\n', False)
    probe('q4.md', '이 값은 미확정이라 「미확정」으로 둔다.\n', False)
    probe('q5.md', '그 하루 몫만 떼어 본다. 투자자 몫은 그 안의 일부다.\n', False)
    probe('q6.md', '스토리보드 각주는 "보유한 정산채권들의 %s 을 가중한 평균"(`SB s3`) '
                   '이라 적었다.\n' % pick('투자금액'), False)
    probe('q7.md', '**투자자 몫이다.** 정산상품에 배정된 요율이 할인율이다.\n', False)
    probe('p8.md', '이 값을 %s로 삼는다.\n' % pick('재료'), True)   # 조사가 붙어도 잡는다

    # (다) 어투
    def tprobe(name, text, expect):
        nonlocal ok
        p = os.path.join(tmp, name)
        io.open(p, 'w', encoding='utf-8').write(text)
        _, _, _, _, _, tc, _ = scan([p])
        got = sum(tc.values())
        if (got > 0) == expect:
            ok += 1
        else:
            bad.append('%s — 어투 잡길 %s 했는데 %d건'
                       % (name, '기대' if expect else '안 기대', got))

    tprobe('t1.html', '<p>16.34% — 카드사 70건만 놓고 낸 값이라 다르다</p>', True)
    tprobe('t2.html', '<p>**자기가 가중한 값 바로 앞**</p>', True)
    tprobe('t3.md', '| 항목 | 값 |\n|---|---|\n| 상환액 | — |\n', True)
    tprobe('t4.md', '값은 3.04일이고 그 옆 칸은 13.21% 다.\n', False)

    shutil.rmtree(tmp, ignore_errors=True)
    print('판별력 시험 — %d/%d 통과' % (ok, ok + len(bad)))
    for b in bad:
        print('  FAIL %s' % b)
    return not bad


if __name__ == '__main__':
    if '--self' in sys.argv:
        sys.exit(0 if selftest() else 1)
    n, t, _, _ = report(allver='--all' in sys.argv)
    sys.exit(0 if (n == 0 and t == 0) else 1)
