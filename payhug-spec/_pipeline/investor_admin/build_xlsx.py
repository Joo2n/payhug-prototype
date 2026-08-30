# -*- coding: utf-8 -*-
"""투자자 어드민 엑셀 14종 생성기 — roster16_model 값 + 원본 어드민(lib/excel.ts) 헤더 서식.

값 원천 : roster16_model.py (이 스크립트는 숫자를 만들지 않고 받아 쓰기만 한다)
서식 원천: /Users/semi/cursor/payhug-admin-web/lib/excel.ts `fillSheet`
파일명   : {내용}_{시작일}_{종료일}.xlsx · 날짜 YYYY-MM-DD
           (원본 예: payhug-admin-web/app/settlement/overview/TransferRecordsTab.tsx:318)

워크북 구성: 단일 시트 14파일.
  원본 excel.ts 는 downloadExcel(단일 시트)·downloadExcelSheets(다중 시트)를 모두 제공하되,
  실제 사용처인 `이체내역`·`차액정산`은 downloadExcel(단일 시트)로 낸다.
  투자자 어드민은 다운로드 버튼 1개가 화면의 표 1개에 대응하므로 단일 시트가 원본 정합.
  다중 시트로 묶으면 버튼 1개가 화면에 없는 표까지 싣게 되어 downloadExcelSheets 의 전제
  (한 화면의 여러 섹션을 한 번에 내림)와 어긋난다.

투자 수익은 화면에서 도달 가능한 프리셋 조합마다 자기 파일을 낸다 — 6조합 x (표 · 수익 현황) = 12벌.
  집계 단위만 보고 파일을 고르면 같은 단위 안에서 기간이 갈리는 두 조합(일별 일주일·금월)이
  한 파일을 가리켜, 화면은 27행 금월인데 파일은 일주일치가 되는 자리가 생긴다.
  파일을 고르는 열쇠는 집계 단위 + 시작일 + 종료일 셋이다 — 통합본 xlsKey() 와 같은 열쇠다.
  프리셋 밖 기간(직접입력)은 실물을 찍지 않는다. 화면이 그 자리에서 다운로드를 잠근다.

실행: python3 build_xlsx.py
      → assets/xlsx 14파일 재생성 + 구 파일명 삭제 + 미리보기 4종 파일바(파일명·크기·생성일시) 동기화
"""
import io, os, re, sys
import datetime as _dt
from decimal import Decimal as D, ROUND_HALF_UP

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from roster16_model import (ROSTER, SHARES, EXEC, CASH, TOTAL, W_W, S_W, TY_W,
                            EXEC_SHARE, CASH_SHARE, DAILY, DSUM, MSUM, ty_asset,
                            r1, r2, ty, f)
import daily_ledger

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = '/Users/semi/cursor/payhug-investor-admin'
XDIR = os.path.join(ROOT, 'assets', 'xlsx')

# ── 헤더 서식 — 원본 lib/excel.ts `fillSheet` 규격 ─────────────────
#   cell.fill      = { type:"pattern", pattern:"solid", fgColor:{ argb:"FF6366F1" } }
#   cell.font      = { bold:true, color:{ argb:"FFFFFFFF" }, size:11 }
#   cell.alignment = { horizontal:"center", vertical:"middle" }
#   headerRow.height = 24
FACE        = '맑은 고딕'   # 기존 4파일 실측 서체 — 서식 교체 대상 아님
HEAD_FILL   = PatternFill(fill_type='solid', fgColor='FF6366F1')
HEAD_FONT   = Font(name=FACE, bold=True, color='FFFFFFFF', size=11)
# ExcelJS 의 vertical:"middle" 은 OOXML 상 vertical="center" — openpyxl 표기로 옮긴 값.
HEAD_ALIGN  = Alignment(horizontal='center', vertical='center')
HEAD_HEIGHT = 24
HEAD_ROW    = 3          # 1행 병합 제목 · 2행 공백 · 3행 헤더 (기존 구조 유지)

# ── 본문·제목·합계·각주 서식 (기존 4파일 실측값 유지) ──────────────
TITLE_FONT = Font(name=FACE, bold=True, color='111827', size=13)
TITLE_ROW_HEIGHT = 22
BODY_FONT  = Font(name=FACE, color='374151', size=10)
TOTAL_FONT = Font(name=FACE, bold=True, color='111827', size=10)
TOTAL_FILL = PatternFill(fill_type='solid', fgColor='F9FAFB')
NOTE_FONT  = Font(name=FACE, color='9CA3AF', size=9)
THIN       = Side(style='thin', color='D0D7DE')
BOX        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RIGHT      = Alignment(horizontal='right')

FMT_AMT, FMT_DAY, FMT_PCT2, FMT_PCT1 = '#,##0', '0.00', '0.00%', '0.0%'

# 2행은 제목·머리글 사이 여백으로 비워 둔다. 3행 이하 좌표가 움직이지 않는다.
NOTICE_ROW = 2


def put_notice(ws, ncols):
    return

def pct(v):
    """퍼센트 표기값(3.72 등) → 셀 저장값(0.0372). Decimal 로 나눠 부동소수 잔차를 막는다."""
    return float(D(str(v)) / D(100))

# ── 시트 조립 헬퍼 ────────────────────────────────────────────────
def new_sheet(wb, title, headline, ncols, widths, freeze):
    ws = wb.active
    ws.title = title
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord('A') + i)].width = w
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=headline)
    c.font = TITLE_FONT
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT
    if freeze:
        ws.freeze_panes = freeze
    return ws

def put_header(ws, labels):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=HEAD_ROW, column=i, value=label)
        c.fill, c.font, c.alignment, c.border = HEAD_FILL, HEAD_FONT, HEAD_ALIGN, BOX
    ws.row_dimensions[HEAD_ROW].height = HEAD_HEIGHT

def put_row(ws, r, cells, total=False):
    """cells = [(값, 숫자서식|None, 정렬|None), ...] · None 이면 빈 셀(테두리만)."""
    for i, item in enumerate(cells, start=1):
        c = ws.cell(row=r, column=i)
        c.font, c.border = (TOTAL_FONT if total else BODY_FONT), BOX
        if total:
            c.fill = TOTAL_FILL
        if item is None:
            continue
        v, fmt, align = item
        c.value = v
        if fmt:
            c.number_format = fmt
        if align:
            c.alignment = align

def put_note(ws, r, text):
    return          # 시트에 설명을 적지 않는다

def save(wb, filename):
    path = os.path.join(XDIR, filename)
    wb.save(path)
    return path

# ── 주·달 버킷 ────────────────────────────────────────────────────
#   화면의 rollupBy 와 같은 산식이다. 원장은 daily_ledger 하나뿐이고 여기서 묶기만 한다.
#   W금융일수 = 투자실행금 가중평균.
#   Ty수익율  = SMR x 365 / SD  (대표 정의서).  일자별 ty 를 다시 가중평균하지 않는다 —
#              그러면 같은 행에 적힌 W 로 되짚을 수 없다(ty_bucket_fix.md).
def _mon_start(d):
    x = _dt.date(*map(int, d.split('-')))
    return (x - _dt.timedelta(days=x.weekday())).isoformat()

def _sun_end(d):
    return (_dt.date(*map(int, _mon_start(d).split('-'))) + _dt.timedelta(days=6)).isoformat()

def bucket(rows, label):
    ex = sum(r['exec'] for r in rows)
    pf = sum(r['profit'] for r in rows)
    rp = sum(r['repay'] for r in rows)
    w  = sum(D(str(r['w'])) * D(r['exec']) for r in rows) / D(ex)
    t  = (D(pf) / D(ex) * D(100)) * D(365) / w
    return dict(d=label, repay=rp, exec=ex, profit=pf, w=w, ty=t)


def rollup(frm, to, keyf, labelf):
    days = [r for r in daily_ledger.LEDGER if frm <= r['d'] <= to]
    order, group = [], {}
    for r in days:
        k = keyf(r['d'])
        if k not in group:
            group[k] = []
            order.append(k)
        group[k].append(r)
    order.sort()
    return [bucket(group[k], labelf(group[k][0]['d'])) for k in order]


# ── 화면에서 도달 가능한 조회 조합 ────────────────────────────────
#   통합본 build_app.py 의 PRESET_RANGE · PRESET_GRAN · PRESET_LABEL 과 같은 값이다.
#   종료일은 전부 기준일에서 끊는다 — 마지막 버킷이 기준일 뒤 빈 날짜를 이고 있으면
#   같은 크기의 앞 버킷들과 나란히 놓였을 때 급락으로 읽힌다.
BASE_DATE = '2026-08-27'
INVESTOR  = '㈜테스트인베스트'


def _add(d, n):
    return (_dt.date(*map(int, d.split('-'))) + _dt.timedelta(days=n)).isoformat()


def _m_first(d):
    return d[:7] + '-01'


def _m_shift(d, n):
    x = _dt.date(*map(int, _m_first(d).split('-')))
    y, m = x.year, x.month + n
    while m < 1:
        y, m = y - 1, m + 12
    while m > 12:
        y, m = y + 1, m - 12
    return '%04d-%02d-01' % (y, m)


def _week_label(to):
    # 주 라벨 = 월요일 ~ 그 주 일요일. 조회 종료일에서 끊긴 주는 종료일까지만 적는다.
    return lambda d: _mon_start(d) + ' ~ ' + min(_sun_end(d), to)[5:]


# (프리셋 키, 집계 단위, 검색대상기간 라벨, 시작일, 종료일)
PRESETS = [
    ('week',  'daily',   '일주일', _add(BASE_DATE, -6),              BASE_DATE),
    ('month', 'daily',   '금월',       _m_first(BASE_DATE),              BASE_DATE),
    ('w4',    'weekly',  '4주',            _add(_mon_start(BASE_DATE), -21), BASE_DATE),
    ('w12',   'weekly',  '12주',           _add(_mon_start(BASE_DATE), -77), BASE_DATE),
    ('m3',    'monthly', '3개월',      _m_shift(BASE_DATE, -2),          BASE_DATE),
    ('m6',    'monthly', '6개월',      _m_shift(BASE_DATE, -5),          BASE_DATE),
]

GRAN_NAME  = {'daily': '일별', 'weekly': '주별', 'monthly': '월별'}
GRAN_COL   = {'daily': '정산예정일', 'weekly': '정산예정주',
              'monthly': '정산예정월'}
GRAN_WIDTH = {'daily': 13.5, 'weekly': 21.5, 'monthly': 21.5}


def preset_rows(gran, frm, to):
    # 그 조합의 표 본문 — 화면 pfRows() 와 같은 묶음.
    if gran == 'weekly':
        return rollup(frm, to, _mon_start, _week_label(to))
    if gran == 'monthly':
        return rollup(frm, to, lambda d: d[:7], lambda d: d[:7])
    return [dict(d=r['d'], repay=r['repay'], exec=r['exec'], profit=r['profit'],
                 w=D(str(r['w'])), ty=D(str(r['ty'])))
            for r in daily_ledger.LEDGER if frm <= r['d'] <= to]


def profit_file(gran, frm, to):
    return '%s투자수익_%s_%s.xlsx' % (GRAN_NAME[gran], frm, to)


def status_file(frm, to):
    return '투자수익현황_%s_%s.xlsx' % (frm, to)


def put_bucket_sheet(title, headline, colhead, rows, filename, w0=21.5):
    wb = openpyxl.Workbook()
    ws = new_sheet(wb, title, headline, 6, [w0, 16.5, 16.5, 12.5, 12.5, 11.5], 'A4')
    put_notice(ws, 6)
    put_header(ws, [colhead, '상환액', '투자실행금', '투자 수익', 'W금융일수', 'Ty수익율'])
    r = 4
    for x in rows:
        put_row(ws, r, [(x['d'], None, None), (x['repay'], FMT_AMT, None), (x['exec'], FMT_AMT, None),
                        (x['profit'], FMT_AMT, None), (float(r2(x['w'])), FMT_DAY, None),
                        (pct(r2(x['ty'])), FMT_PCT2, None)])
        r += 1
    tot = bucket(rows, '합계')
    put_row(ws, r, [('합계', None, None), (tot['repay'], FMT_AMT, None), (tot['exec'], FMT_AMT, None),
                    (tot['profit'], FMT_AMT, None), (float(r2(tot['w'])), FMT_DAY, None),
                    (pct(r2(tot['ty'])), FMT_PCT2, None)], total=True)
    put_note(ws, r + 2, '')
    return save(wb, filename)


# ── 1) 투자자산현황 ───────────────────────────────────────────────
def build_assets_status():
    wb = openpyxl.Workbook()
    ws = new_sheet(wb, '투자자산 현황', '투자자산 현황 — 기준일 2026-08-27 / ㈜테스트인베스트',
                   7, [18.5, 16.5, 12.5, 14.5, 11.5, 11.5, 13.5], 'A4')
    put_notice(ws, 7)
    put_header(ws, ['자산 구분', '금액 (원)', 'W금융일수', 'S입금부족율', 'Ty수익율', '비중', '보관'])
    put_row(ws, 4, [('투자실행액', None, None), (EXEC, FMT_AMT, None),
                    (float(r2(W_W)), FMT_DAY, None), (pct(r2(S_W)), FMT_PCT2, None),
                    (pct(r2(TY_W)), FMT_PCT2, None), (pct(EXEC_SHARE), FMT_PCT1, None),
                    ('㈜페이허그', None, None)])
    put_row(ws, 5, [('순현금', None, None), (CASH, FMT_AMT, None), None, None, None,
                    (pct(CASH_SHARE), FMT_PCT1, None), ('㈜쿠콘', None, None)])
    put_row(ws, 6, [('합계 (투자자산)', None, None), (TOTAL, FMT_AMT, None), None, None, None,
                    (1, FMT_PCT1, None), None], total=True)
    put_note(ws, 8, '※ 합계(투자자산) = 투자실행액 + 순현금. '
                    'W금융일수·S입금부족율·Ty수익율은 투자실행액에만 산정.')
    return save(wb, '투자자산현황_2026-08-27_2026-08-27.xlsx')

# ── 2) 가맹점별투자자산 ───────────────────────────────────────────
def build_assets_merchant():
    wb = openpyxl.Workbook()
    ws = new_sheet(wb, '가맹점별 투자자산', '가맹점별 투자자산 — 기준일 2026-08-27 / ㈜테스트인베스트',
                   6, [20.5, 16.5, 12.5, 14.5, 11.5, 11.5], 'A4')
    put_notice(ws, 6)
    put_header(ws, ['가맹점', '투자금액 (원)', 'W금융일수', 'S입금부족율', 'Ty수익율', '비중'])
    r = 4
    for (name, amount, w, s, *_), share in zip(ROSTER, SHARES):
        put_row(ws, r, [(name, None, None), (amount, FMT_AMT, None), (float(w), FMT_DAY, None),
                        (pct(s), FMT_PCT2, None), (pct(ty(w)), FMT_PCT2, None),
                        (pct(share), FMT_PCT1, None)])
        r += 1
    put_row(ws, r, [('합계', None, None), (EXEC, FMT_AMT, None), None, None, None,
                    (1, FMT_PCT1, None)], total=True)
    put_note(ws, r + 2, '※ 비중은 투자실행액 합계(%s원) 대비 각 가맹점 투자금액의 구성비.' % f(EXEC))
    return save(wb, '가맹점별투자자산_2026-08-27_2026-08-27.xlsx')

# ── 3) 투자수익현황 — 집계 단위 3벌(일별·주별·월별) ────────────────
#   화면의 `수익 현황` 카드 한 장이 곧 이 시트다. 카드가 4주를 말하는데 파일이 일주일이면
#   화면과 파일이 다른 기간을 말한다 — 그래서 집계 단위마다 자기 기간 파일을 낸다.
#   머리글의 `기준일`은 문서를 낸 날(BASE_DATE)이고, 조회 구간은 `검색대상기간` 행이 진다.
def put_status_sheet(label, frm, to, ex, pf, ty4, ty5, filename):
    wb = openpyxl.Workbook()
    ws = new_sheet(wb, '투자수익 현황', '투자수익 현황 — 기준일 2026-08-27 / ㈜테스트인베스트',
                   2, [31.5, 35.5], None)
    put_notice(ws, 2)
    put_header(ws, ['항목', '값'])
    put_row(ws, 4, [('검색대상기간', None, None), ('%s (%s ~ %s)' % (label, frm, to), None, RIGHT)])
    put_row(ws, 5, [('투자실행금', None, None), (ex, FMT_AMT, None)])
    put_row(ws, 6, [('투자수익', None, None), (pf, FMT_AMT, None)])
    put_row(ws, 7, [('Ty수익율 (투자실행금액 대비)', None, None), (pct(ty4), FMT_PCT2, None)])
    # ⑤ = (④ × PSA) / (PSA + PSC) — 분모는 기간 유량끼리. roster16_model.ty_asset 산출.
    put_row(ws, 8, [('Ty수익율 (투자자산 대비)', None, None), (pct(ty5), FMT_PCT2, None)])
    put_note(ws, 10, '')
    return save(wb, filename)


def _ec_days(frm, to):
    # EC 합에 들어가는 날수 = 그 구간에 원장이 갖고 있는 일자 수(화면 ecDays 와 같다).
    return len([r for r in daily_ledger.LEDGER if frm <= r['d'] <= to])


# ── 3~6) 투자 수익 — 프리셋 조합마다 표 1벌 · 수익 현황 1벌 ────────
#   집계 단위가 같아도 기간이 다르면 다른 파일이다. 화면이 27행 금월을 보이는데
#   파일이 일주일치면 두 산출물이 다른 기간을 말한다 — 그 자리를 없앤다.
def build_profit_preset(gran, label, frm, to):
    rows = preset_rows(gran, frm, to)
    tot = bucket(rows, '합계')
    table = put_bucket_sheet(
        '%s 투자수익' % GRAN_NAME[gran],
        '%s 투자수익 — %s ~ %s / %s' % (GRAN_NAME[gran], frm, to, INVESTOR),
        GRAN_COL[gran], rows, profit_file(gran, frm, to), GRAN_WIDTH[gran])
    status = put_status_sheet(label, frm, to, tot['exec'], tot['profit'],
                              r2(tot['ty']), ty_asset(tot['ty'], tot['exec'], _ec_days(frm, to)),
                              status_file(frm, to))
    return [table, status]


def build_profit_all():
    out = []
    for _key, gran, label, frm, to in PRESETS:
        out += build_profit_preset(gran, label, frm, to)
    return out

# ── 미리보기 파일바 동기화 ────────────────────────────────────────
# 화면 : (미리보기 HTML, 엑셀 파일명)
PREVIEW = [('xls-assets-status.html',   '투자자산현황_2026-08-27_2026-08-27.xlsx'),
           ('xls-assets-merchant.html', '가맹점별투자자산_2026-08-27_2026-08-27.xlsx'),
           ('xls-profit-status.html',   '투자수익현황_2026-08-21_2026-08-27.xlsx'),
           ('xls-profit-daily.html',    '일별투자수익_2026-08-21_2026-08-27.xlsx')]

LEGACY = ['투자자산_현황_20260827.xlsx', '가맹점별_투자자산_20260827.xlsx',
          '투자수익_현황_20260827.xlsx', '일별_투자수익_20260827.xlsx']


def wanted():
    """이번 판이 내는 파일 이름 전량 — 여기 없는 .xlsx 는 옛 기간의 잔존물이라 지운다."""
    names = ['투자자산현황_2026-08-27_2026-08-27.xlsx', '가맹점별투자자산_2026-08-27_2026-08-27.xlsx']
    for _key, gran, _label, frm, to in PRESETS:
        names += [profit_file(gran, frm, to), status_file(frm, to)]
    return names

def sync_preview(html, filename):
    import time
    path = os.path.join(XDIR, filename)
    size = '%.1f KB' % (os.path.getsize(path) / 1024.0)
    made = time.strftime('%Y-%m-%d %H:%M', time.localtime(os.path.getmtime(path)))
    p = os.path.join(ROOT, html)
    s = io.open(p, encoding='utf-8').read()
    s = re.sub(r'(<div class="fb-name">)[^<]+(</div>)', r'\g<1>%s\g<2>' % filename, s)
    s = re.sub(r'(href="assets/xlsx/)[^"]+(" download)', r'\g<1>%s\g<2>' % filename, s)
    s = re.sub(r'(<div class="fb-meta">\s*<span>)[^<]+(</span>)', r'\g<1>%s\g<2>' % size, s)
    s = re.sub(r'(생성일시 <span class="mono">)[^<]+(</span>)', r'\g<1>%s\g<2>' % made, s)
    io.open(p, 'w', encoding='utf-8').write(s)
    return size, made

if __name__ == '__main__':
    made = [build_assets_status(), build_assets_merchant()] + build_profit_all()
    keep = set(wanted())
    assert len(keep) == len(made), '파일 이름이 겹친다 — %d ≠ %d' % (len(keep), len(made))
    for name in sorted(set(LEGACY) | set(os.listdir(XDIR))):
        if not name.endswith('.xlsx') or name in keep:
            continue
        p = os.path.join(XDIR, name)
        if os.path.exists(p):
            os.remove(p)
            print('  삭제 %s' % name)
    for p in made:
        print('  생성 %s  %d bytes' % (os.path.basename(p), os.path.getsize(p)))
    for html, fn in PREVIEW:
        size, ts = sync_preview(html, fn)
        print('  동기화 %-28s %s · %s · %s' % (html, fn, size, ts))
