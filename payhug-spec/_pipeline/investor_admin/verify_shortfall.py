# -*- coding: utf-8 -*-
"""S입금부족율의 집계 방식·표본 구간 기계 검사.

대표 정의서 원문 (ceo_definitions.md [1번 이미지])
    S입금부족율 = Σ SLi / Σ SAi
    SLi = 표본집합에 속하는 ID가 i인 대상정산금채권의 미지급액 - 과지급액
    SAi = 표본집합에 속하는 ID가 i인 대상정산금채권의 순지급액 x (1 - 할인율)
    표본집합: 선정산일이 D-20부터 D-11인 기간에 속하는 대상정산금채권들의 집합

2026-08-31 대표 회의가 확정한 것 — 근거는 meeting_0831/meeting_20260831_raw.txt 원문
    (00:22:30) 조현준 「그 S 입금 부정률이라는게 샘플 입금 부정률인 거거든」
               → S 는 Sample 의 S, 변수 이름이다. Σ(Summation) 가 아니다.
    (00:22:30) 조현준 「이게 어제만의 상황일 수도 있으니까 한 열흘치를 묶어서 계산하는 거야 샘플로」
               → 표본은 열흘치. 구간 폭은 그대로다.
    (00:26:18) 조현준 「전체 투자 금액은 300만 원이 되는 거고 … 입금 부족액은
               오늘 입금 부족액, 내일 입금 부족액 합쳐서 3만 원이더라. 그러면
               300만 원 분의 3만 원 하면 아 샘플 입금 부정률은 1%다」
               → 기간 집계는 분자·분모를 각각 더한 뒤 나눈다. 하루별 비율의 평균이 아니다.
    (00:26:18) 조현준 「썸으로 쓴 기억은 없어」
               → 산식의 계산 규칙은 바뀌지 않았다. 바뀐 것은 S 를 읽는 법뿐이다.

    판정: (가) 표기 정리. 산식 변경 없음 · 표본 구간 변경 없음.
    이 검증기는 그 판정을 원장·화면·엑셀에 못으로 박는다.

검사 6항목
    1  원장의 S 가 방식1(Σ SLi / Σ SAi, 기간 전체 합산 비율)로 계산되는가
    2  표본 구간이 선정산일 D-20 ~ D-11 열흘인가 — 원장·화면·엑셀·용어정의서
    3  방식1 과 방식2(하루별 비율의 평균)의 값이 갈리는가 · 화면값이 어느 쪽인가
    4  회의 근거 인용이 회의록 원문에 실재하는가 (근거가 사라지면 기준이 뜬다)
    5  헤드리스 크롬으로 실제 화면을 세워 S 칸·모집단 툴팁을 읽는가
    6  자기시험 — 다른 방식으로 계산한 값을 넣으면 위 검사들이 잡는가

판정 규칙
    · FAIL 1건 이상이면 종료코드 1.
    · try/except 로 오류를 SKIP 으로 삼키지 않는다. 예외는 FAIL 로 센다.
    · 파일 부재는 FAIL 이다 (rd() 가 예외를 던진다).
    · 「위반 0건」류 검사는 검사 대상이 0건이 아님을 함께 판정한다.

verify_batch_symbols.py 와의 경계
    verify_batch_symbols.py 는 [2번] 축 — 일별 배치 기호 8개(SB·SA·SM·SMR·SD)를 본다.
    이 검증기는 [1번] 축의 S입금부족율 하나만 본다. 겹치는 검사가 없다.

산출: verify_shortfall_result.json
"""
import io, json, os, re, subprocess, sys, tempfile, traceback
from datetime import timedelta
from decimal import Decimal as D, ROUND_HALF_UP

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)

APP_REPO = '/Users/semi/cursor/payhug-investor-admin'
PROTO_REPO = '/Users/semi/cursor/payhug-investor-prototype'
XLSX = os.path.join(BASE, '검산_투자자어드민_20260901.xlsx')
MEETING = os.path.join(BASE, 'meeting_0831', 'meeting_20260831_raw.txt')
CEO = os.path.join(BASE, 'ceo_definitions.md')
CHROME = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'

# 회의록에서 이 검증기의 기준을 떠받치는 인용. 원문에서 사라지면 기준이 근거를 잃는다.
QUOTES = [
    ('00:13:23', 'S는 샘플을 샘플에 S를'),
    ('00:22:30', '그 S 입금 부정률이라는게 샘플 입금 부정률인 거거든'),
    ('00:22:30', '한 열흘치를 묶어서 계산하는 거야'),
    ('00:26:18', '전체 투자 금액은 300만 원이 되는 거고'),
    ('00:26:18', '300만 원 분의 3만 원'),
    ('00:26:18', '썸으로 쓴 기억은 없어'),
    ('00:27:40', 'i부터 s까지'),
]

# 대표 정의서 원문 문장 — 표본 구간의 원천이다.
SAMPLE_SENTENCE = '표본집합: 선정산일이 D-20부터 D-11인 기간에 속하는 대상정산금채권들의 집합'
RATIO_SENTENCE = 'S입금부족율 = Σ SLi / Σ SAi'

# 화면 툴팁이 말해야 하는 모집단 문언. 건수는 원장에서 읽는다 — 여기 적지 않는다.
# 2026-08-31 기호 규칙 — 날짜 쪽은 소문자 d 다 (dm_0831/symbol_rule_0831.md).
# 바로 위 SAMPLE_SENTENCE 는 대표 원문 인용이라 대문자 그대로 둔다.
POP_S_TEXT = '선정산일이 기준일 20일 전 ~ 11일 전인 표본'

# macOS 함정 — --window-size=1440,H 의 실제 뷰포트는 1440x(H-87). 87 을 더해 창을 띄운다.
# --screenshot 플래그를 붙이면 편차가 0 이 되어 보정이 사라진 것처럼 오판한다. 붙이지 않는다.
VIEWPORT = (1440, 1200)
MAC_CHROME_H = 87

SAMPLE_DAYS = 10          # 회의 「열흘치」 (00:22:30)
SAMPLE_FROM = 20          # D-20
SAMPLE_TO = 11            # D-11

R = []
_SEEN = set()


def chk(sec, name, ok, detail=''):
    key = (sec, name)
    assert key not in _SEEN, '검사 이름 중복: %s' % (key,)
    _SEEN.add(key)
    R.append({'sec': sec, 'name': name, 'pass': bool(ok), 'detail': str(detail)})
    return bool(ok)


def rd(path):
    if not os.path.exists(path):
        raise IOError('파일 없음: %s' % path)
    with io.open(path, 'r', encoding='utf-8') as fp:
        t = fp.read()
    if not t.strip():
        raise IOError('파일이 비었다: %s' % path)
    return t


def q6(x):
    return D(x).quantize(D('0.000001'), rounding=ROUND_HALF_UP)


def q2(x):
    return D(x).quantize(D('0.01'), rounding=ROUND_HALF_UP)


# ══════════════════════════════════════════════════════════════════
# 집계 방식 두 가지 — 순수 함수. 항목 6 이 이 함수들로 자기시험한다.
# ══════════════════════════════════════════════════════════════════
def agg_ratio(rows):
    """방식1 — 기간 전체 합산 비율.  Σ SLi / Σ SAi  (단위 %)

    대표 (00:26:18) 「전체 투자 금액은 300만 원 … 부족액 합쳐서 3만 원 … 300만 원 분의 3만 원」
    """
    sa = sum(r['sa'] for r in rows)
    sl = sum(r['sl'] for r in rows)
    return (D(sl) / D(sa) * 100) if sa else D(0)


def mean_of_ratios(rows, key):
    """방식2 — 묶음(하루·가맹점)마다 비율을 내어 단순평균. 채택하지 않은 방식이다."""
    g = {}
    for r in rows:
        g.setdefault(r[key], []).append(r)
    parts = [agg_ratio(v) for v in g.values() if sum(x['sa'] for x in v)]
    return (sum(parts) / D(len(parts))) if parts else D(0)


def compose(parts):
    """방식1 의 정의적 성질 — 조각을 분모로 가중해 합치면 전체와 같다.
    parts = [(SA, 비율%), ...]"""
    tot = sum(p[0] for p in parts)
    if not tot:
        return D(0)
    return sum(D(p[0]) * p[1] for p in parts) / D(tot)


def facts_match(facts, m1):
    """항목 1·3 이 쓰는 대조 그 자체. 항목 6 이 같은 함수를 손댄 값으로 부른다."""
    return {'raw': facts.get('sRaw') == str(q6(m1)),
            'disp': facts.get('s') == str(q2(m1))}


def screen_merch_match(tbl, merchants):
    """화면 가맹점표의 S 열이 원장 명단과 전건 같은가. 항목 6 이 손댄 표로 부른다."""
    if not tbl or '입금부족률' not in tbl['head']:
        return False, None, None
    mi = tbl['head'].index('입금부족률')
    got = [(r[0], r[mi]) for r in tbl['body']]
    want = [(m[0], '%s%%' % m[3]) for m in merchants]
    return got == want, got, want


def norm(recv):
    """원장 채권 -> 검사용 행. sa = SAi, sl = SLi."""
    return [{'sa': r['ai'], 'sl': r['unpaid'] - r['over'],
             'adv': r['adv'].strftime('%Y-%m-%d'),
             'due': r['due'].strftime('%Y-%m-%d'),
             'mid': r['mid']} for r in recv]


# ══════════════════════════════════════════════════════════════════
# 화면 조작 드라이버
# ══════════════════════════════════════════════════════════════════
DRIVER_JS = r"""
const http = require('http'), fs = require('fs'), path = require('path'), os = require('os');
const { spawn } = require('child_process');
const DL = require(CFGDLPATH);
const CFG = JSON.parse(process.argv[2]);
const PORT = 8870 + (process.pid % 70), DPORT = 9570 + (process.pid % 70);
const MIME = {'.html':'text/html; charset=utf-8', '.css':'text/css; charset=utf-8',
  '.js':'text/javascript', '.png':'image/png', '.webp':'image/webp', '.pdf':'application/pdf',
  '.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'};

const server = http.createServer((req, res) => {
  const url = decodeURIComponent(req.url.split('?')[0]);
  const seg = url.split('/').filter(Boolean);
  const t = CFG.targets.find(x => x.key === seg[0]);
  if(!t){ res.writeHead(404); res.end('no target'); return; }
  const p = path.join(t.root, seg.slice(1).join('/'));
  fs.readFile(p, (e, b) => {
    if(e){ res.writeHead(404); res.end('nope'); return; }
    res.writeHead(200, {'Content-Type': MIME[path.extname(p)] || 'application/octet-stream'});
    res.end(b);
  });
});

let msgId = 0, ws, pending = new Map();
const consoleErrors = [];
function send(method, params){
  const id = ++msgId;
  ws.send(JSON.stringify({id, method, params: params || {}}));
  return new Promise((res, rej) => pending.set(id, {res, rej}));
}
async function evalJS(expr){
  const r = await send('Runtime.evaluate', {expression:'(function(){' + expr + '})()',
    returnByValue:true, awaitPromise:true});
  if(r.exceptionDetails) throw new Error('page eval: ' + JSON.stringify(
    (r.exceptionDetails.exception && r.exceptionDetails.exception.description) || r.exceptionDetails.text));
  return r.result.value;
}
const sleep = ms => new Promise(r => setTimeout(r, ms));
async function waitFor(expr, label, ms){
  const end = Date.now() + (ms || 15000);
  while(Date.now() < end){
    let ok = false;
    try { ok = await evalJS('return !!(' + expr + ');'); } catch(e){ ok = false; }
    if(ok) return true;
    await sleep(120);
  }
  throw new Error('대기 실패: ' + label + ' (' + expr + ')');
}
function must(v, label){
  if(v === null || v === undefined) throw new Error('읽지 못한 자리: ' + label);
  return v;
}

async function readAssets(){
  return await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-assets"]');
    if(!sec) throw new Error('투자 자산 화면 없음');
    function tbl(box){
      if(!box) return null;
      /* 자리(mount)가 표 자신인 경우와 표를 감싼 상자인 경우가 둘 다 있다. */
      var t = (box.tagName === 'TABLE') ? box : box.querySelector('table');
      if(!t) return null;
      var cells = tr => [].map.call(tr.children, td => td.textContent.trim());
      return {head:[].map.call(t.querySelectorAll('thead th'), function(e){
                var a = e.querySelector('.tip-anchor');
                return a ? a.textContent.trim() : e.textContent.trim(); }),
              pop:[].map.call(t.querySelectorAll('thead th'), function(e){
                var p = e.querySelector('.tip-panel');
                return p ? p.textContent.trim() : null; }),
              body:[].map.call(t.querySelectorAll('tbody tr'), cells)};
    }
    var st = sec.querySelector('[data-mount="ia-status"]');
    var me = sec.querySelector('[data-mount="ia-merch"]');
    return {status: tbl(st), merch: tbl(me)};
  `);
}

async function drive(t, base){
  const out = {key:t.key};
  await send('Page.navigate', {url: base + '/' + t.key + '/' + t.page});
  await waitFor(`document.querySelector('section.screen[data-screen="invest-assets"]')`,
                t.key + ' 화면 적재', 30000);
  await sleep(400);
  await waitFor('window.innerWidth === ' + CFG.view[0] + ' && window.innerHeight === ' + CFG.view[1],
                t.key + ' 뷰포트 ' + CFG.view.join('x') + ' 안정화', 20000);
  out.viewport = await evalJS('return {w:window.innerWidth, h:window.innerHeight};');
  await evalJS(`location.hash = '#invest-assets'; return 1;`);
  await waitFor(`document.querySelector('section.screen[data-screen="invest-assets"] [data-mount="ia-status"] tbody tr')
                 && document.querySelector('section.screen[data-screen="invest-assets"] [data-mount="ia-merch"] tbody tr')`,
                t.key + ' 투자 자산 두 표');
  await sleep(200);
  out.assets = must(await readAssets(), t.key + ' assets');
  must(out.assets.status, t.key + ' 현황표');
  must(out.assets.merch, t.key + ' 가맹점표');
  /* 열머리 툴팁은 CSS :hover 로 뜬다. 합성 MouseEvent 로는 :hover 가 걸리지 않는다 —
     CDP 로 실제 포인터를 그 좌표에 옮겨야 뜬다. 안 옮기고 읽으면 display:none 을 읽는다. */
  const rect = await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-assets"]');
    var ths = sec.querySelectorAll('[data-mount="ia-status"] thead th');
    for(var i = 0; i < ths.length; i++){
      var a = ths[i].querySelector('.tip-anchor');
      if(a && a.textContent.trim() === '입금부족률'){
        a.scrollIntoView({block:'center'});
        var r = a.getBoundingClientRect();
        return {x: Math.round(r.left + r.width/2), y: Math.round(r.top + r.height/2)};
      }
    }
    return null;
  `);
  if(!rect) throw new Error(t.key + ' 현황표에 입금부족률 툴팁 앵커가 없다');
  out.tipHoverAt = rect;
  await send('Input.dispatchMouseEvent', {type:'mouseMoved', x:rect.x, y:rect.y,
    button:'none', buttons:0, clickCount:0});
  await sleep(200);
  out.tipVisible = await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-assets"]');
    var ths = sec.querySelectorAll('[data-mount="ia-status"] thead th');
    for(var i = 0; i < ths.length; i++){
      var a = ths[i].querySelector('.tip-anchor');
      if(a && a.textContent.trim() === '입금부족률'){
        var p = ths[i].querySelector('.tip-panel');
        var cs = getComputedStyle(p);
        return {text:p.textContent.trim(), display:cs.display,
                w:p.getBoundingClientRect().width, h:p.getBoundingClientRect().height};
      }
    }
    return null;
  `);
  /* 포인터를 치우면 다시 숨는가 — 늘 떠 있는 패널을 「툴팁이 뜬다」로 오판하지 않게 한다. */
  await send('Input.dispatchMouseEvent', {type:'mouseMoved', x:5, y:5,
    button:'none', buttons:0, clickCount:0});
  await sleep(200);
  out.tipHidden = await evalJS(`
    var sec = document.querySelector('section.screen[data-screen="invest-assets"]');
    var ths = sec.querySelectorAll('[data-mount="ia-status"] thead th');
    for(var i = 0; i < ths.length; i++){
      var a = ths[i].querySelector('.tip-anchor');
      if(a && a.textContent.trim() === '입금부족률')
        return getComputedStyle(ths[i].querySelector('.tip-panel')).display;
    }
    return null;
  `);
  return out;
}

async function main(){
  await new Promise(r => server.listen(PORT, r));
  const base = 'http://127.0.0.1:' + PORT;
  const dl = DL.dir();
  const profile = fs.mkdtempSync(path.join(os.tmpdir(), 'phsf-'));
  /* macOS 함정 — --window-size 의 세로는 크롬 창 전체다. 뷰포트는 87px 작다.
     --screenshot 플래그를 붙이면 편차가 0 이 되어 보정이 사라진 것처럼 오판한다. 붙이지 않는다. */
  const chrome = spawn(CFG.chrome, ['--headless=new', '--remote-debugging-port=' + DPORT]
    .concat(DL.args(dl, profile))
    .concat(['--no-first-run', '--no-default-browser-check', '--disable-gpu',
             '--window-size=' + CFG.win.join(','), 'about:blank']), {stdio:'ignore'});

  let targets = null;
  for(let i = 0; i < 60 && !targets; i++){
    await sleep(300);
    try {
      targets = await new Promise((res, rej) => {
        http.get({host:'127.0.0.1', port:DPORT, path:'/json'}, r => {
          let d = ''; r.on('data', c => d += c); r.on('end', () => res(JSON.parse(d)));
        }).on('error', rej);
      });
    } catch(e){ targets = null; }
  }
  if(!targets) throw new Error('크롬 CDP 접속 실패');
  const page = targets.find(t => t.type === 'page');
  ws = new WebSocket(page.webSocketDebuggerUrl);
  await new Promise(r => ws.addEventListener('open', r));
  ws.addEventListener('message', ev => {
    const m = JSON.parse(ev.data);
    if(m.id && pending.has(m.id)){ pending.get(m.id).res(m.result); pending.delete(m.id); return; }
    if(m.method === 'Runtime.consoleAPICalled' && (m.params.type === 'error' || m.params.type === 'warning'))
      consoleErrors.push(m.params.type + ': ' + m.params.args.map(a => a.value || a.description || a.type).join(' '));
    if(m.method === 'Runtime.exceptionThrown')
      consoleErrors.push('exception: ' + ((m.params.exceptionDetails.exception &&
        m.params.exceptionDetails.exception.description) || m.params.exceptionDetails.text));
  });
  await send('Runtime.enable'); await send('Page.enable');
  await DL.behavior(send, dl);
  /* --window-size 만 믿으면 회차에 따라 창이 다른 크기로 뜬다(1113 · 775 를 관측).
     붙은 뒤 같은 크기를 한 번 더 못 박고, 뷰포트가 그 값이 될 때까지 기다린다.
     보정 87px 은 그대로 둔다 — 없애면 macOS 함정이 사라진 것처럼 보인다. */
  const wid = (await send('Browser.getWindowForTarget', {targetId: page.id})).windowId;
  await send('Browser.setWindowBounds', {windowId: wid,
    bounds: {left: 0, top: 0, width: CFG.win[0], height: CFG.win[1], windowState: 'normal'}});
  await sleep(300);

  const res = {targets:{}, console:consoleErrors};
  for(const t of CFG.targets) res.targets[t.key] = await drive(t, base);
  res.dropped = fs.readdirSync(dl);
  process.stdout.write(JSON.stringify(res));
  try { chrome.kill(); } catch(e){}
  DL.clean(dl);
  server.close();
  process.exit(0);
}
main().catch(e => { process.stdout.write(JSON.stringify({error:String(e && e.stack || e)})); process.exit(3); });
"""


def run_driver():
    if not os.path.exists(CHROME):
        raise IOError('크롬 없음: %s' % CHROME)
    for root, page in ((APP_REPO, 'app.html'), (PROTO_REPO, 'index.html')):
        if not os.path.exists(os.path.join(root, page)):
            raise IOError('화면 파일 없음: %s' % os.path.join(root, page))
    dlmod = os.path.join(BASE, 'chrome_dl.js')
    if not os.path.exists(dlmod):
        raise IOError('chrome_dl.js 없음 — 붙이지 않으면 사용자 다운로드 폴더에 파일이 쌓인다')
    cfg = {'chrome': CHROME, 'view': list(VIEWPORT),
           'win': [VIEWPORT[0], VIEWPORT[1] + MAC_CHROME_H],
           'targets': [{'key': 'app', 'root': APP_REPO, 'page': 'app.html'},
                       {'key': 'proto', 'root': PROTO_REPO, 'page': 'index.html'}]}
    src = DRIVER_JS.replace('CFGDLPATH', json.dumps(dlmod))
    fd, p = tempfile.mkstemp(suffix='.js', prefix='phsf-driver-')
    with os.fdopen(fd, 'w') as fp:
        fp.write(src)
    try:
        cp = subprocess.run(['node', p, json.dumps(cfg)], stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE, timeout=420)
        raw = cp.stdout.decode('utf-8', 'replace').strip()
        if not raw:
            raise IOError('드라이버 무응답 rc=%s err=%s'
                          % (cp.returncode, cp.stderr.decode('utf-8', 'replace')[:400]))
        out = json.loads(raw)
        if out.get('error'):
            raise IOError('드라이버 예외: %s' % out['error'][:600])
        return out
    finally:
        os.unlink(p)


# ══════════════════════════════════════════════════════════════════
# 1. 원장의 S 가 방식1 인가
# ══════════════════════════════════════════════════════════════════
def sec1(L):
    dl, facts = L['ledger'], L['facts']
    rows = norm(dl._SAMPLE)
    chk('1', '표본 모집단 0건 아님', len(rows) > 0, '표본 채권 %d건' % len(rows))

    # (1) 행태 판정 — 원장이 쓰는 함수를 인공 표본으로 직접 부른다.
    #     SA 가 다른 두 묶음을 주면 합산비율(0.75%)과 하루별평균(0.8333…%)이 갈린다.
    synth = [{'ai': 10000, 'unpaid': 100, 'over': 0, 'k': 'a'},
             {'ai': 30000, 'unpaid': 200, 'over': 0, 'k': 'b'}]
    got = dl.shortfall(synth)
    want1 = D(300) / D(40000) * 100          # 0.75
    want2 = (D(100) / D(10000) + D(200) / D(30000)) / 2 * 100
    chk('1', 'daily_ledger.shortfall 이 합산비율(방식1)로 답한다',
        got == want1 and got != want2,
        '반환 %s · 방식1 %s · 방식2 %s' % (got, want1, q6(want2)))

    # (2) 원장 실측 재계산 — 검증기가 독립으로 낸 값과 완전일치
    m1 = agg_ratio(rows)
    chk('1', '원장 S_RAW = 검증기 독립 재계산 (완전일치)',
        dl.S_RAW == m1, '원장 %s · 재계산 %s' % (dl.S_RAW, m1))

    # (3) ledger_facts 바이트 대조
    fm = facts_match(facts, m1)
    chk('1', 'ledger_facts.sRaw = 방식1 6자리 표기',
        fm['raw'], 'facts %s · 방식1 %s' % (facts['sRaw'], q6(m1)))
    chk('1', 'ledger_facts.s = 방식1 2자리 표기',
        fm['disp'], 'facts %s · 방식1 %s' % (facts['s'], q2(m1)))

    # (4) 가맹점 행 8건 전건 — 행마다 그 가맹점 표본의 방식1
    mrows = facts['merchants']
    chk('1', '가맹점 행 0건 아님', len(mrows) > 0, '%d곳' % len(mrows))
    bad = []
    for m in mrows:
        name = m[0]
        mine = [r for r in rows if r['mid'] == _mid_of(dl, name)]
        want = str(q2(agg_ratio(mine)))
        if m[3] != want:
            bad.append((name, m[3], want))
    chk('1', '가맹점 %d곳 S 가 전건 방식1' % len(mrows), not bad, '어긋남 %d건 %s' % (len(bad), bad[:3]))

    # (5) 방식1 의 정의적 성질 — 조각을 분모 가중으로 합치면 전체와 같다.
    #     단순평균으로는 복원되지 않는다. 세 축(가맹점·선정산일·홀짝)에서 모두 본다.
    for axis, label in (('mid', '가맹점'), ('adv', '선정산일')):
        g = {}
        for r in rows:
            g.setdefault(r[axis], []).append(r)
        parts = [(sum(x['sa'] for x in v), agg_ratio(v)) for v in g.values()]
        chk('1', '분모 가중 합성 = 전체 (%s %d조각)' % (label, len(parts)),
            compose(parts) == m1, '합성 %s · 전체 %s' % (q6(compose(parts)), q6(m1)))
    half = [rows[i::2] for i in (0, 1)]
    parts = [(sum(x['sa'] for x in h), agg_ratio(h)) for h in half]
    chk('1', '분모 가중 합성 = 전체 (홀짝 2조각)',
        compose(parts) == m1, '합성 %s · 전체 %s' % (q6(compose(parts)), q6(m1)))

    return {'m1': m1, 'rows': rows}


def _mid_of(dl, name):
    for x in dl.MERCHANTS:
        if x['name'] == name:
            return x['mid']
    raise KeyError('원장에 없는 가맹점: %s' % name)


# ══════════════════════════════════════════════════════════════════
# 2. 표본 구간 — 선정산일 D-20 ~ D-11 열흘
# ══════════════════════════════════════════════════════════════════
def sec2(L, s1, drv):
    dl, facts = L['ledger'], L['facts']
    rows = s1['rows']
    lo, hi = dl.SAMPLE

    chk('2', 'SAMPLE 시작 = D-%d' % SAMPLE_FROM,
        lo == dl.ASOF - timedelta(days=SAMPLE_FROM),
        '%s · 기준일 %s' % (lo, dl.ASOF))
    chk('2', 'SAMPLE 끝 = D-%d' % SAMPLE_TO,
        hi == dl.ASOF - timedelta(days=SAMPLE_TO),
        '%s · 기준일 %s' % (hi, dl.ASOF))
    chk('2', '구간 폭 = 열흘 (회의 00:22:30)',
        (hi - lo).days + 1 == SAMPLE_DAYS, '%d일' % ((hi - lo).days + 1))

    # 자르는 축이 선정산일인가 — 정산예정일로 자른 집합과 다르다(판별력)
    by_adv = [r for r in dl.RECEIVABLES if lo <= r['adv'] <= hi]
    by_due = [r for r in dl.RECEIVABLES if lo <= r['due'] <= hi]
    chk('2', '표본을 선정산일 축으로 자른다',
        len(dl._SAMPLE) == len(by_adv) and dl._SAMPLE == by_adv,
        '선정산일 %d건 · _SAMPLE %d건' % (len(by_adv), len(dl._SAMPLE)))
    # 두 축의 건수는 우연히 같다(하루 320건 x 열흘). 건수로는 판별되지 않으므로
    # 채권 자체가 갈리는지와 값이 갈리는지로 본다.
    ids_adv, ids_due = set(map(id, by_adv)), set(map(id, by_due))
    chk('2', '정산예정일 축과 갈린다 (검사에 판별력 있음)',
        bool(ids_adv ^ ids_due) and agg_ratio(norm(by_due)) != agg_ratio(norm(by_adv)),
        '겹치지 않는 채권 %d건 · 값 %s vs %s (건수는 %d = %d 로 같다)'
        % (len(ids_adv ^ ids_due), q6(agg_ratio(norm(by_due))), q6(agg_ratio(norm(by_adv))),
           len(by_due), len(by_adv)))

    chk('2', '표본 하루수 = 열흘 (실제 선정산일 가짓수)',
        len(set(r['adv'] for r in rows)) == SAMPLE_DAYS,
        '%d일 %s' % (len(set(r['adv'] for r in rows)), sorted(set(r['adv'] for r in rows))[:2]))

    chk('2', 'ledger_facts.sampleSpan 바이트 일치',
        facts['sampleSpan'] == [dl.ymd(lo), dl.ymd(hi)],
        '%s vs %s' % (facts['sampleSpan'], [dl.ymd(lo), dl.ymd(hi)]))
    chk('2', 'ledger_facts.sampleReceivables 바이트 일치',
        facts['sampleReceivables'] == len(dl._SAMPLE),
        '%s vs %d' % (facts['sampleReceivables'], len(dl._SAMPLE)))

    # 용어정의서 원문
    ceo = rd(CEO)
    chk('2', 'ceo_definitions.md 에 표본집합 원문이 그대로',
        SAMPLE_SENTENCE in ceo.replace('  ', ' '), SAMPLE_SENTENCE)
    chk('2', 'ceo_definitions.md 에 비율 산식 원문이 그대로',
        RATIO_SENTENCE in ceo, RATIO_SENTENCE)

    # 화면 원천 텍스트 — 툴팁이 뜨는 파일 전건. 명단을 손으로 적지 않고 훑어 찾는다.
    files = []
    for root in (APP_REPO, PROTO_REPO):
        for fn in sorted(os.listdir(root)):
            if not fn.endswith('.html'):
                continue
            t = rd(os.path.join(root, fn))
            if '입금부족률' not in t:
                continue
            if 'POP_S' in t or 'tip-anchor">입금부족률' in t:
                files.append((os.path.join(root, fn), t))
    chk('2', 'S 모집단 툴팁을 가진 화면 파일 0건 아님', len(files) > 0,
        '%d건 %s' % (len(files), [os.path.basename(f) for f, _ in files]))
    n_str = '%s건' % format(len(dl._SAMPLE), ',')
    bad = [os.path.basename(f) for f, t in files if POP_S_TEXT not in t or n_str not in t]
    chk('2', '화면 원천 %d건이 전부 구간·건수를 말한다' % len(files), not bad,
        '어긋남 %s · 기대 「%s」 %s' % (bad, POP_S_TEXT, n_str))

    # 엑셀 — 산식 문언 + 실제 셀 수식이 「두 합의 비」꼴인가
    wb = openpyxl.load_workbook(XLSX)
    txt = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
           for c in row if c.value is not None]
    chk('2', '엑셀 산식 시트에 표본집합 원문', any(SAMPLE_SENTENCE in s for s in txt), SAMPLE_SENTENCE)
    chk('2', '엑셀 산식 시트에 비율 산식 원문', any(RATIO_SENTENCE in s for s in txt), RATIO_SENTENCE)
    xr = xlsx_shortfall(wb)
    chk('2', '엑셀 S 셀이 두 합의 비 꼴 (평균 아님)', ratio_shaped(xr),
        '분자 %s | 분모 %s | 비 %s' % (xr['num'], xr['den'], xr['ratio']))

    # 화면 조작 결과 — 툴팁이 실제로 그 문언인가
    if drv:
        for key, t in sorted(drv['targets'].items()):
            tip = t.get('tipVisible')
            chk('2', '%s 열머리에 포인터를 올리면 툴팁이 뜬다' % key,
                bool(tip) and tip.get('display') == 'block' and tip.get('w', 0) > 0
                and tip.get('h', 0) > 0,
                '좌표 %s · %s' % (t.get('tipHoverAt'), str(tip)[:200]))
            chk('2', '%s 포인터를 치우면 툴팁이 숨는다' % key,
                t.get('tipHidden') == 'none', str(t.get('tipHidden')))
            chk('2', '%s 툴팁 문언이 구간·건수' % key,
                bool(tip) and POP_S_TEXT in tip['text'] and n_str in tip['text'],
                (tip or {}).get('text', ''))
    return {'files': files, 'xlsx': xr, 'nstr': n_str}


def xlsx_shortfall(wb):
    """엑셀에서 S입금부족율 셀과 그 분자·분모 수식을 뽑는다."""
    ws = wb['기간집계']
    cell = None
    for row in ws.iter_rows():
        for c in row:
            if str(c.value).strip() == 'S입금부족율(%)':
                cell = 'B%d' % c.row
    if not cell:
        raise KeyError('기간집계 시트에 S입금부족율(%) 행이 없다')
    f = str(ws[cell].value)
    m = re.search(r'B(\d+)\s*/\s*B(\d+)', f)
    if not m:
        raise KeyError('S 셀 수식이 B?/B? 꼴이 아니다: %s' % f)
    return {'cell': cell, 'ratio': f,
            'num': str(ws['B%s' % m.group(1)].value),
            'den': str(ws['B%s' % m.group(2)].value)}


def ratio_shaped(xr):
    """두 합의 비 꼴인가 — 분자·분모가 각각 SUM 계열이고 어느 쪽에도 AVERAGE 가 없다."""
    if 'AVERAGE' in xr['ratio'].upper() or 'AVERAGE' in xr['num'].upper() or 'AVERAGE' in xr['den'].upper():
        return False
    okn = xr['num'].upper().startswith('=SUM')
    okd = xr['den'].upper().startswith('=SUM')
    return bool(okn and okd and re.search(r'B\d+\s*/\s*B\d+', xr['ratio']))


# ══════════════════════════════════════════════════════════════════
# 3. 방식1 vs 방식2 — 값 차이와 판별력
# ══════════════════════════════════════════════════════════════════
def sec3(L, s1, drv):
    dl, facts = L['ledger'], L['facts']
    rows, m1 = s1['rows'], s1['m1']
    sa = sum(r['sa'] for r in rows)
    sl = sum(r['sl'] for r in rows)

    m2adv = mean_of_ratios(rows, 'adv')     # 하루별(선정산일) 비율의 평균
    m2due = mean_of_ratios(rows, 'due')     # 하루별(정산예정일) 비율의 평균
    m3 = mean_of_ratios(rows, 'mid')        # 가맹점별 비율의 평균

    delta = {}
    for k, v in (('mean_by_adv', m2adv), ('mean_by_due', m2due), ('mean_by_merchant', m3)):
        delta[k] = {'raw': str(q6(v)), 'disp': str(q2(v)),
                    'pp': str((v - m1).quantize(D('0.00000001'))),
                    'won': str((v / 100 * D(sa) - D(sl)).quantize(D('0.01')))}
    delta['sum_over_sum'] = {'raw': str(q6(m1)), 'disp': str(q2(m1)),
                             'pp': '0', 'won': '0',
                             'SL': str(sl), 'SA': str(sa)}

    chk('3', '방식1 · 방식2 가 서로 다른 값을 낸다 (선정산일 축)',
        m1 != m2adv,
        '방식1 %s / 방식2 %s / 차 %s%%p · %s원'
        % (q6(m1), q6(m2adv), delta['mean_by_adv']['pp'], delta['mean_by_adv']['won']))
    chk('3', '방식1 · 방식2 가 서로 다른 값을 낸다 (정산예정일 축)',
        m1 != m2due,
        '방식1 %s / 방식2 %s / 차 %s%%p · %s원'
        % (q6(m1), q6(m2due), delta['mean_by_due']['pp'], delta['mean_by_due']['won']))
    chk('3', '방식1 · 가맹점별 평균이 서로 다른 값을 낸다',
        m1 != m3,
        '방식1 %s / 가맹평균 %s / 차 %s%%p · %s원'
        % (q6(m1), q6(m3), delta['mean_by_merchant']['pp'], delta['mean_by_merchant']['won']))

    # 화면값이 어느 쪽인가 — raw 6자리로 본다.
    # 2자리 표기는 선정산일 축 방식2 를 구별하지 못한다(둘 다 0.07). raw 를 검사 축으로 쓰는 이유다.
    chk('3', '화면 raw = 방식1', facts_match(facts, m1)['raw'],
        'facts %s · 방식1 %s' % (facts['sRaw'], q6(m1)))
    chk('3', '화면 raw != 방식2 (선정산일 축) — raw 자리에서 판별된다',
        facts['sRaw'] != str(q6(m2adv)), 'facts %s · 방식2 %s' % (facts['sRaw'], q6(m2adv)))
    chk('3', '화면 raw != 방식2 (정산예정일 축)',
        facts['sRaw'] != str(q6(m2due)), 'facts %s · 방식2 %s' % (facts['sRaw'], q6(m2due)))
    chk('3', '화면 raw != 가맹점별 평균',
        facts['sRaw'] != str(q6(m3)), 'facts %s · 가맹평균 %s' % (facts['sRaw'], q6(m3)))
    chk('3', '화면 표기 2자리 != 방식2 (정산예정일 축)',
        facts['s'] != str(q2(m2due)), 'facts %s · 방식2 %s' % (facts['s'], q2(m2due)))

    # 화면에 실제로 뜬 글자 — 헤드리스로 읽은 것
    if drv:
        for key, t in sorted(drv['targets'].items()):
            st, me = t['assets']['status'], t['assets']['merch']
            ci = st['head'].index('입금부족률')
            exec_row = [r for r in st['body'] if r and r[0] == '투자실행액']
            cash_row = [r for r in st['body'] if r and r[0] == '순현금']
            chk('3', '%s 현황표 투자실행액 행 S = 원장 표기' % key,
                len(exec_row) == 1 and exec_row[0][ci] == '%s%%' % facts['s'],
                '%s · 기대 %s%%' % (exec_row[0][ci] if exec_row else None, facts['s']))
            chk('3', '%s 현황표 순현금 행 S = -' % key,
                len(cash_row) == 1 and cash_row[0][ci] == '-',
                str(cash_row[0][ci] if cash_row else None))
            ok, got, want = screen_merch_match(me, facts['merchants'])
            chk('3', '%s 가맹점표 S %d행 전건 = 원장' % (key, len(facts['merchants'])),
                ok, '화면 %s · 원장 %s' % ((got or [])[:3], (want or [])[:3]))
    return {'delta': delta, 'm1': m1, 'm2adv': m2adv, 'm2due': m2due, 'm3': m3}


# ══════════════════════════════════════════════════════════════════
# 4. 회의 근거 정박
# ══════════════════════════════════════════════════════════════════
def sec4():
    t = rd(MEETING)
    for ts, q in QUOTES:
        chk('4', '회의록 인용 실재 (%s) %s' % (ts, q[:22]), q in t and ts in t, q)
    # 구간을 바꾸자는 결정이 없었다는 것은 「없음」의 증명이라 문자열로 못 박지 않는다.
    # 대신 회의가 말한 폭(열흘)이 원장 구간 폭과 같은지를 항목 2 가 판정한다.
    chk('4', '회의록 분량 0 아님', len(t) > 10000, '%d자' % len(t))


# ══════════════════════════════════════════════════════════════════
# 5. 화면 조작 결과 자체
# ══════════════════════════════════════════════════════════════════
def sec5(drv):
    if not drv:
        chk('5', '헤드리스 크롬 조작', False, '드라이버 결과 없음')
        return
    chk('5', '화면 2주소 조작', len(drv['targets']) == 2, str(sorted(drv['targets'])))
    chk('5', '화면 콘솔 오류 0건', not drv.get('console'), str((drv.get('console') or [])[:3]))
    chk('5', '다운로드 폴더 오염 0건 (chrome_dl 임시자리)',
        not drv.get('dropped'), str(drv.get('dropped')))
    for key, t in sorted(drv['targets'].items()):
        vp = t['viewport']
        chk('5', '%s 뷰포트 %dx%d (macOS %dpx 보정)' % (key, VIEWPORT[0], VIEWPORT[1], MAC_CHROME_H),
            vp['w'] == VIEWPORT[0] and vp['h'] == VIEWPORT[1],
            '%s · 창 %dx%d' % (vp, VIEWPORT[0], VIEWPORT[1] + MAC_CHROME_H))


# ══════════════════════════════════════════════════════════════════
# 6. 자기시험 — 다른 방식으로 낸 값을 넣으면 잡히는가
# ══════════════════════════════════════════════════════════════════
def sec6(L, s1, s2, s3, drv):
    dl = L['ledger']
    rows, m1 = s1['rows'], s1['m1']
    m2adv, m2due, m3 = s3['m2adv'], s3['m2due'], s3['m3']

    # (가) 방식 자체가 갈리는가 — 인공 데이터로 판별력 확인
    synth = [{'sa': 10000, 'sl': 100, 'k': 'a'}, {'sa': 30000, 'sl': 200, 'k': 'b'}]
    a, b = agg_ratio(synth), mean_of_ratios(synth, 'k')
    chk('6', '인공 표본에서 방식1 0.75% · 방식2 0.833…%',
        a == D('0.75') and q6(b) == D('0.833333'), '방식1 %s · 방식2 %s' % (a, q6(b)))

    # (나) 화면 raw 자리에 다른 방식 값을 심고 항목 1·3 이 쓰는 대조 함수를 그대로 부른다.
    chk('6', 'facts 대조 함수가 원본을 통과시킨다',
        facts_match(L['facts'], m1)['raw'] and facts_match(L['facts'], m1)['disp'],
        str(facts_match(L['facts'], m1)))
    for label, v in (('방식2 선정산일 축', m2adv), ('방식2 정산예정일 축', m2due),
                     ('가맹점별 평균', m3)):
        bad = dict(L['facts'])
        bad['sRaw'] = str(q6(v))
        bad['s'] = str(q2(v))
        got = facts_match(bad, m1)
        chk('6', '화면 raw 에 %s 값을 심으면 대조 실패' % label,
            not got['raw'], '심은 값 %s · 정답 %s · 대조 %s' % (q6(v), q6(m1), got))

    # (다) 원장 함수를 하루별 평균으로 갈아끼우면 facts 대조가 깨지는가
    fake = mean_of_ratios(rows, 'adv')
    chk('6', '원장 S 를 하루별 평균으로 바꾸면 facts.sRaw 대조 실패',
        str(q6(fake)) != L['facts']['sRaw'],
        '바꾼 값 %s · facts %s' % (q6(fake), L['facts']['sRaw']))

    # (라) 표본 구간을 넓히면 건수·값이 갈리는가
    lo, hi = dl.SAMPLE
    wide = [r for r in dl.RECEIVABLES if lo - timedelta(days=1) <= r['adv'] <= hi]
    chk('6', '표본을 D-21 로 하루 넓히면 건수 대조 실패',
        len(wide) != len(dl._SAMPLE), '%d건 vs %d건' % (len(wide), len(dl._SAMPLE)))
    chk('6', '표본을 D-21 로 하루 넓히면 값도 갈린다',
        q6(agg_ratio(norm(wide))) != q6(m1),
        '%s vs %s' % (q6(agg_ratio(norm(wide))), q6(m1)))

    # (마) 자르는 축을 정산예정일로 바꾸면 갈리는가
    due = [r for r in dl.RECEIVABLES if lo <= r['due'] <= hi]
    chk('6', '자르는 축을 정산예정일로 바꾸면 표본·값이 갈린다',
        bool(set(map(id, due)) ^ set(map(id, dl._SAMPLE)))
        and q6(agg_ratio(norm(due))) != q6(m1),
        '겹치지 않는 채권 %d건 · %s vs %s'
        % (len(set(map(id, due)) ^ set(map(id, dl._SAMPLE))),
           q6(agg_ratio(norm(due))), q6(m1)))

    # (바) 화면 문언 판독기 — 구간을 D-10 으로 바꾼 사본을 잡는가
    nstr = s2['nstr']
    tampered = 0
    for path, t in s2['files']:
        bad = t.replace(POP_S_TEXT, '선정산일이 기준일 20일 전 ~ 10일 전인 표본')
        if POP_S_TEXT not in bad:
            tampered += 1
    chk('6', '툴팁 판독기가 D-10 치환본을 전건 잡는다',
        tampered == len(s2['files']) and tampered > 0,
        '%d/%d건' % (tampered, len(s2['files'])))
    # 건수 판독기 — 건수를 지운 사본
    cnt = sum(1 for _, t in s2['files'] if nstr not in t.replace(nstr, 'X건'))
    chk('6', '툴팁 판독기가 건수 치환본을 전건 잡는다',
        cnt == len(s2['files']) and cnt > 0, '%d/%d건' % (cnt, len(s2['files'])))

    # (사) 엑셀 판독기 — AVERAGE 로 바꾼 수식을 잡는가
    xr = dict(s2['xlsx'])
    chk('6', '엑셀 판독기가 원본을 통과시킨다', ratio_shaped(xr), xr['ratio'])
    for k, v in (('ratio', '=AVERAGE(B40:B41)'), ('num', '=AVERAGE(채권!L2:L1281)'),
                 ('den', '=AVERAGE(채권!J2:J1281)')):
        t = dict(xr)
        t[k] = v
        chk('6', '엑셀 판독기가 %s 를 AVERAGE 로 바꾼 사본을 잡는다' % k,
            not ratio_shaped(t), '%s = %s' % (k, v))

    # (자) 화면 가맹점표 판독기 — 한 칸만 바꾼 사본을 잡는가
    if drv:
        for key, t in sorted(drv['targets'].items()):
            me = t['assets']['merch']
            ok0, got0, _ = screen_merch_match(me, L['facts']['merchants'])
            chk('6', '%s 가맹점표 판독기가 원본을 통과시킨다' % key, ok0, str(got0)[:120])
            mi = me['head'].index('입금부족률')
            tam = {'head': me['head'],
                   'body': [list(r) for r in me['body']]}
            tam['body'][0][mi] = '%s%%' % q2(m2due)      # 방식2 값으로 한 칸만 바꿈
            ok1, _, _ = screen_merch_match(tam, L['facts']['merchants'])
            chk('6', '%s 가맹점표 판독기가 한 칸 치환본을 잡는다' % key, not ok1,
                '심은 값 %s%%' % q2(m2due))
            tam2 = {'head': [h for h in me['head'] if h != '입금부족률'],
                    'body': [list(r) for r in me['body']]}
            ok2, _, _ = screen_merch_match(tam2, L['facts']['merchants'])
            chk('6', '%s 가맹점표 판독기가 S 열 삭제본을 잡는다' % key, not ok2,
                'S 열이 사라지면 0건 통과가 아니라 실패')

    # (아) 회의 근거 판독기 — 인용을 지운 사본을 잡는가
    raw = rd(MEETING)
    miss = 0
    for ts, q in QUOTES:
        if q not in raw.replace(q, ''):
            miss += 1
    chk('6', '회의록 판독기가 인용 삭제본을 전건 잡는다',
        miss == len(QUOTES) and miss > 0, '%d/%d건' % (miss, len(QUOTES)))


# ══════════════════════════════════════════════════════════════════
def main():
    L = {}
    try:
        import daily_ledger
        L['ledger'] = daily_ledger
        L['facts'] = json.loads(rd(os.path.join(BASE, 'ledger_facts.json')))
        chk('0', '원장·사실값 적재', True,
            '채권 %d건 · 표본 %d건' % (len(daily_ledger.RECEIVABLES), len(daily_ledger._SAMPLE)))
    except Exception as e:
        chk('0', '원장·사실값 적재', False, '%s: %s' % (type(e).__name__, e))
        return dump()

    drv = None
    try:
        drv = run_driver()
        chk('0', '헤드리스 크롬 조작 (app · proto)', True, '화면 %d곳' % len(drv['targets']))
    except Exception as e:
        chk('0', '헤드리스 크롬 조작 (app · proto)', False,
            '%s: %s' % (type(e).__name__, str(e)[:400]))

    s1 = s2 = s3 = None
    for name, fn in (('1', lambda: sec1(L)),
                     ('2', lambda: sec2(L, s1, drv)),
                     ('3', lambda: sec3(L, s1, drv)),
                     ('4', lambda: sec4()),
                     ('5', lambda: sec5(drv)),
                     ('6', lambda: sec6(L, s1, s2, s3, drv))):
        try:
            out = fn()
            if name == '1':
                s1 = out
            elif name == '2':
                s2 = out
            elif name == '3':
                s3 = out
        except Exception as e:
            chk(name, '항목 %s 실행' % name, False,
                '%s: %s | %s' % (type(e).__name__, str(e)[:300],
                                 traceback.format_exc().strip().splitlines()[-3:]))
    return dump(s3)


def dump(s3=None):
    fails = [r for r in R if not r['pass']]
    by = {}
    for r in R:
        b = by.setdefault(r['sec'], [0, 0])
        b[0] += 1
        b[1] += 0 if r['pass'] else 1
    out = {'total': len(R), 'fail': len(fails),
           'verdict': '(가) 표기 정리 — 산식·표본 구간 변경 없음 (회의 2026-08-31)',
           'bySection': dict((k, {'검사': v[0], 'FAIL': v[1]}) for k, v in sorted(by.items())),
           'delta': (s3 or {}).get('delta'),
           'cases': R}
    with io.open(os.path.join(BASE, 'verify_shortfall_result.json'), 'w', encoding='utf-8') as fp:
        fp.write(json.dumps(out, ensure_ascii=False, indent=1))
    print('검사 %d건 · FAIL %d건' % (len(R), len(fails)))
    for k in sorted(by):
        print('  항목 %s  검사 %3d  FAIL %d' % (k, by[k][0], by[k][1]))
    if out['delta']:
        print()
        print('  방식 간 값 차이 (판정 대상 아님 · 기록)')
        for k in ('sum_over_sum', 'mean_by_adv', 'mean_by_due', 'mean_by_merchant'):
            d = out['delta'][k]
            print('    %-17s raw %s · 표기 %s · 차 %s%%p · %s원' %
                  (k, d['raw'], d['disp'], d['pp'], d['won']))
    if fails:
        print()
        for f in fails:
            print('FAIL [%s] %s' % (f['sec'], f['name']))
            print('      %s' % f['detail'][:400])
    return 1 if fails else 0


if __name__ == '__main__':
    sys.exit(main())
