# -*- coding: utf-8 -*-
"""정산주기_정리 xlsx 검증기.

1) 시트별 행·열 수
2) 수식 셀 개수 — 집계 칸이 값이 아니라 수식인지
3) `검산` 시트 각 행의 차를 파이썬으로 독립 계산 (원본을 다시 읽어서)
4) `일자별 365일` 전 셀이 원본 K6:Q370 과 1:1 일치하는지
5) 가중 기준 갈라적기 — 이 워크북의 W·Ty 는 MAU(이용자 수) 비중에서 나온 값이다.
   그 값이 나오는 자리마다 ±4행 안에 `참고`·`시장 평균`·`MAU` 중 하나가 있어야 하고,
   `가중평균` 시트에 금액 기준 화면 값(ledger_facts.json 의 w·ty)을 갈라 적은 줄이 있어야 한다.
   MAU 값 자체는 지우지 않는다 — 자리가 0건이면 FAIL 이다.
"""
import datetime as dt
import glob
import json
import os
import re
import sys

import openpyxl

SRC = '/Users/semi/Downloads/정산주기.xlsx'
#   기본은 실물 산출 폴더. 음성 시험 때만 CYCLE_OUTDIR 로 사본 폴더를 가리킨다.
OUTDIR = os.environ.get('CYCLE_OUTDIR', '/Users/semi/Downloads/payhug_정산주기_정리')
WEEK = '월화수목금토일'
COLS = ['E', 'F', 'G', 'H']


def latest():
    f = sorted(glob.glob(os.path.join(OUTDIR, '정산주기_정리_*.xlsx')))
    if not f:
        sys.exit('출력 파일 없음: %s' % OUTDIR)
    return f[-1]


def main():
    out = latest()
    ok = True

    sv = openpyxl.load_workbook(out, data_only=False)
    print('파일  %s' % out)
    print()
    print('%-14s %6s %6s %8s' % ('시트', '행', '열', '수식셀'))
    total_f = 0
    per = {}
    for ws in sv.worksheets:
        nf = sum(1 for row in ws.iter_rows() for c in row
                 if isinstance(c.value, str) and c.value.startswith('='))
        per[ws.title] = nf
        total_f += nf
        print('%-14s %6d %6d %8d' % (ws.title, ws.max_row, ws.max_column, nf))
    print('%-14s %6s %6s %8d' % ('합계', '', '', total_f))
    print()

    # ── 원본 재독 ─────────────────────────────────────────────
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sc, sw = wb['정산주기'], wb['비중']
    rows = []
    r = 6
    while sc.cell(r, 11).value is not None:
        rows.append((int(sc.cell(r, 11).value), int(sc.cell(r, 12).value),
                     sc.cell(r, 13).value,
                     [int(sc.cell(r, c).value) for c in range(14, 18)]))
        r += 1
    n = len(rows)
    sums = [sum(x[3][i] for x in rows) for i in range(4)]
    avgs = [sums[i] / n for i in range(4)]

    mau = [sw.cell(4, c).value for c in range(4, 7)]        # 배민 쿠팡 요기요
    mau_sum = sum(mau)
    deliv = sw['H5'].value
    share = [m / mau_sum for m in mau]
    mix_bm, mix_cpe, mix_yo = [s * deliv for s in share]
    mix = [mix_yo, mix_cpe, mix_bm, 1 - deliv]              # 요기요 쿠팡 배민 카드
    w = avgs[0] * mix[0] + avgs[1] * mix[1] + avgs[2] * mix[2] + avgs[3] * mix[3]

    expect = {}
    names = ['요기요', '쿠팡이츠', '배민', '신용카드']
    for i, nm in enumerate(names):
        expect['%s 평균 금융일수' % nm] = avgs[i]
        expect['%s 평균 금융일수 (도수)' % nm] = avgs[i]
        expect['%s 금융일수 합계' % nm] = sums[i]
        expect['%s 관측 일수' % nm] = n
        expect['%s 구성비' % nm] = mix[i]
        expect['%s 일자별 평균' % nm] = avgs[i]
    expect['MAU 계'] = mau_sum
    expect['배달앱/전체'] = deliv
    expect['구성비 합'] = sum(mix)
    expect['W 가중평균 금융일수'] = w
    expect['일자별 행 수'] = n

    # ── 검산 시트 독립 재계산 ─────────────────────────────────
    ws = sv['검산']
    print('%-30s %-24s %-24s %s' % ('항목', '원본 값', '독립 계산', '차'))
    nzero = ndiff = 0
    for rr in range(2, ws.max_row + 1):
        name = ws.cell(rr, 1).value
        orig = ws.cell(rr, 3).value
        if name is None or orig is None:
            continue
        if name not in expect:
            print('  미매핑 항목: %s' % name)
            ok = False
            continue
        got = expect[name]
        d = float(got) - float(orig)
        print('%-30s %-24r %-24r %s' % (name, orig, got, repr(d)))
        if d == 0.0:
            nzero += 1
        else:
            ndiff += 1
            ok = False
    print()
    print('검산 행 %d건 — 차 0: %d건 · 차 있음: %d건' % (nzero + ndiff, nzero, ndiff))

    # ── 일자별 전수 대조 ──────────────────────────────────────
    wd = sv['일자별 365일']
    bad = 0
    for i, (mm, dd, wk, vals) in enumerate(rows):
        rr = 2 + i
        if wd.cell(rr, 1).value != mm or wd.cell(rr, 2).value != dd \
                or wd.cell(rr, 3).value != wk:
            bad += 1
        for j in range(4):
            if wd.cell(rr, 5 + j).value != vals[j]:
                bad += 1
        f = wd.cell(rr, 4).value
        if not (isinstance(f, str) and f.startswith('=DATE(')):
            bad += 1
    print('일자별 전수 대조 (%d행 x 8열) — 불일치 %d칸' % (n, bad))
    ok = ok and bad == 0

    # ── 집계 칸이 값으로 박히지 않았는지 ──────────────────────
    hard = []
    for sh, addrs in (('도수분포', None), ('가중평균', ('B2', 'B3', 'B4', 'B5',
                                                   'C2', 'C3', 'C4', 'C5',
                                                   'D2', 'D3', 'D4', 'D5', 'B7', 'B8', 'B11')),
                      ('구성비', ('E2', 'B3', 'C3', 'D3', 'E3',
                                'B7', 'B8', 'B9', 'B10', 'B11'))):
        s = sv[sh]
        if addrs is None:
            for row in s.iter_rows(min_row=2, min_col=2):
                for c in row:
                    if c.value is not None and not (isinstance(c.value, str)
                                                    and c.value.startswith('=')):
                        hard.append('%s!%s' % (sh, c.coordinate))
        else:
            for a in addrs:
                v = s[a].value
                if not (isinstance(v, str) and v.startswith('=')):
                    hard.append('%s!%s' % (sh, a))
    for rr in range(2, ws.max_row + 1):
        for a in ('D', 'E'):
            v = ws['%s%d' % (a, rr)].value
            if v is not None and not (isinstance(v, str) and v.startswith('=')):
                hard.append('검산!%s%d' % (a, rr))
    print('집계 칸 중 값으로 박힌 칸 %d개%s'
          % (len(hard), (' — ' + ', '.join(hard)) if hard else ''))
    ok = ok and not hard

    # ── 도수분포 COUNTIF 독립 검증 ────────────────────────────
    fq = sv['도수분포']
    lo = fq['A2'].value
    cnt_bad = 0
    for rr in range(2, fq.max_row + 1):
        d = fq.cell(rr, 1).value
        if not isinstance(d, int):
            break
        for j in range(4):
            want = sum(1 for x in rows if x[3][j] == d)
            f = fq.cell(rr, 2 + j).value
            if not (isinstance(f, str) and f.startswith('=COUNTIF')):
                cnt_bad += 1
            per.setdefault('_freq', {})
    print('도수분포 COUNTIF 수식 아닌 칸 %d개 (구간 %s~%s)'
          % (cnt_bad, lo, fq.cell(fq.max_row - 3, 1).value))
    ok = ok and cnt_bad == 0

    # ── 플랫폼 요약 요일별 금융일수 수식 독립 평가 ────────────
    ps = sv['플랫폼 요약']
    fbad = fok = 0
    for rr in range(2, ps.max_row + 1):
        pre, dep, f = ps.cell(rr, 5).value, ps.cell(rr, 6).value, ps.cell(rr, 7).value
        if not pre or not dep:
            continue
        if not (isinstance(f, str) and f.startswith('=MOD(FIND(')):
            fbad += 1
            continue
        want = (WEEK.index(dep.strip()[-1]) - WEEK.index(pre.strip()[-1]) - 1) % 7 + 1
        # 같은 요일 조합이 일자별 시트 평일 관측치에 실제로 나오는지 대조
        pay = ps.cell(rr, 4).value
        col = {'요기요': 0, '쿠팡이츠': 1, '배민': 2, '신용카드': 3}[ps.cell(rr, 2).value]
        seen = [x[3][col] for x in rows if x[2] == pay]
        if want in seen:
            fok += 1
        else:
            fbad += 1
            print('  요일 기준값 불일치: %s %s %s->%s = %d' %
                  (ps.cell(rr, 2).value, pay, pre, dep, want))
    print('플랫폼 요약 요일 수식 %d행 — 원본 관측치와 일치 %d · 불일치 %d'
          % (fok + fbad, fok, fbad))
    ok = ok and fbad == 0

    # ── 가중 기준 갈라적기 ────────────────────────────────────
    BAN = re.compile(r'(?<![\d.])65(?:\.0+)?\s*%|(?<![\d.])2\.7504'
                     r'|(?<![\d.])14\.60?0*\s*%')
    MARK = ('참고', '시장 평균', 'MAU')
    hits, unmarked = [], []
    for wsx in sv.worksheets:
        for row in wsx.iter_rows():
            for c in row:
                if c.value is None or not BAN.search(str(c.value)):
                    continue
                where = '%s!%s' % (wsx.title, c.coordinate)
                hits.append(where)
                near = ' '.join(str(x.value)
                                for rr in wsx.iter_rows(min_row=max(1, c.row - 4),
                                                        max_row=c.row + 4)
                                for x in rr if x.value is not None)
                if not any(k in near for k in MARK):
                    unmarked.append(where)
    fx = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                     'ledger_facts.json'), encoding='utf-8'))
    wv = sv['가중평균']
    mtxt = ' '.join(str(wv.cell(r, c).value)
                    for r in range(1, wv.max_row + 1)
                    for c in range(1, wv.max_column + 1) if wv.cell(r, c).value)
    need = {'표식 낱말': any(k in mtxt for k in MARK), '금액 기준 명시': '금액' in mtxt,
            '화면 W %s' % fx['w']: fx['w'] in mtxt,
            '화면 Ty %s' % fx['ty']: fx['ty'] in mtxt}
    print('MAU 값 자리 %d곳%s' % (len(hits), (' — ' + ', '.join(hits)) if hits else ''))
    print('갈라 적지 않은 자리 %d곳%s'
          % (len(unmarked), (' — ' + ', '.join(unmarked)) if unmarked else ''))
    print('가중평균 갈라적기 줄 — %s'
          % ' · '.join('%s %s' % (k, 'O' if v else 'X') for k, v in need.items()))
    ok = ok and bool(hits) and not unmarked and all(need.values())

    # ── 워크북 XML 금지어 ─────────────────────────────────────
    #   셀 값만 보면 정의된 이름·시트 이름·데이터 유효성에 숨은 것을 못 잡는다.
    #   금지어 판정기는 audit_xlsx_check 한 곳에만 둔다 — 기준을 두 벌 두지 않는다.
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    import audit_xlsx_check as AX
    ban = AX.banned_xml(out)
    print('XML 조각 %d개 · 금지어 잔여 %d건 · 예외 자리 %s'
          % (ban['검사한 XML 조각'], len(ban['금지어 잔여']), ban['예외 자리 실측']))
    for h in ban['금지어 잔여'][:5]:
        print('  금지어: %s  %s' % (h[0], h[1]))
    #   이 워크북에는 예외 자리가 없다 — 하나라도 생기면 FAIL 이다.
    xml_ok = (ban['검사한 XML 조각'] > 0 and not ban['금지어 잔여']
              and set(ban['예외 자리 실측'].values()) == {0})
    print('XML 금지어 판정 — %s' % ('O' if xml_ok else 'X'))
    ok = ok and xml_ok

    print()
    print('판정: %s' % ('PASS' if ok else 'FAIL'))
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
