# -*- coding: utf-8 -*-
"""요율 정정 맵 적용 + 로스터 16건 통일 — 정적 HTML·xlsx 실파일 수정."""
import io, os, re, json, sys
from decimal import Decimal as D
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from roster16_model import (ROSTER, SHARES, EXEC, CASH, TOTAL, W_W, S_W, TY_W,
                            EXEC_SHARE, CASH_SHARE, DAILY, MONTHLY, DSUM, MSUM,
                            AUG_CARD, r1, r2, f)

ROOT = '/Users/semi/cursor/payhug-investor-admin'
PIPE = os.path.dirname(os.path.abspath(__file__))
MAP  = json.load(io.open(os.path.join(PIPE, 'rate_fix_map.json'), encoding='utf-8'))

# 자산 기준이 8건→16건으로 바뀌면서 맵의 Ty(투자자산 대비) 4건은 대체값을 쓴다
OVERRIDE = {
    ('invest-profit.html',            '>3.29<span class="unit">%</span></div>'): '>%s<span class="unit">%%</span></div>' % DSUM['tyAsset'],
    ('invest-profit--monthly.html',   '>3.23<span class="unit">%</span></div>'): '>%s<span class="unit">%%</span></div>' % AUG_CARD['tyAsset'],
    ('xls-profit-status.html', '<tr><th class="row-head">8</th><td>Ty수익율 (투자자산 대비)</td><td class="c-num">3.29%</td>'):
        '<tr><th class="row-head">8</th><td>Ty수익율 (투자자산 대비)</td><td class="c-num">%s%%</td>' % DSUM['tyAsset'],
}
# 로스터 재계산으로 대체되는 파일 — 맵 적용 대상에서 제외하고 표 전체를 다시 만든다
SUPERSEDED = {'invest-assets.html', 'invest-assets--page2.html', 'invest-assets--download.html',
              'invest-assets--cert-confirm.html', 'xls-assets-status.html'}

log = []
def P(s=''): log.append(s); print(s)

# ════════════ 1) 맵 적용 (텍스트) ════════════
def apply_map():
    counts, errs = {}, []
    for x in MAP['fixes']:
        if x['kind'] != 'text' or x['file'] in SUPERSEDED or x['file'] == 'app.html':
            continue
        p = os.path.join(ROOT, x['file'])
        s = io.open(p, encoding='utf-8').read()
        if s.count(x['locator']) != 1:
            errs.append('%s locator %d회' % (x['file'], s.count(x['locator']))); continue
        new = OVERRIDE.get((x['file'], x['corrected']), x['corrected'])
        io.open(p, 'w', encoding='utf-8').write(s.replace(x['locator'], new, 1))
        counts[x['file']] = counts.get(x['file'], 0) + 1
    return counts, errs

# ════════════ 2) 투자자산 화면 — 로스터 16건 ════════════
def merch_rows(sl):
    out = []
    for (n, a, w, s, *_), sh in sl:
        out.append(
            '            <tr>\n'
            '              <td><span class="name">%s</span></td>\n'
            '              <td class="num"><span class="strong">%s</span></td>\n'
            '              <td class="num">%s일</td>\n'
            '              <td class="num">%s%%</td>\n'
            '              <td class="num">%s%%</td>\n'
            '              <td class="num">%s%%</td>\n'
            '            </tr>' % (n, f(a), w, s, ty_of(w), sh))
    return '\n'.join(out)

def ty_of(w):
    from roster16_model import ty
    return ty(w)

PAIRS = list(zip(ROSTER, SHARES))

def swap_tbody(s, idx, body):
    """idx번째 <tbody>…</tbody> 내용을 교체 (0-base)."""
    pos, start = 0, None
    for _ in range(idx + 1):
        start = s.index('<tbody>', pos); pos = start + 7
    end = s.index('</tbody>', pos)
    return s[:start + 7] + '\n' + body + '\n          ' + s[end:]

STATUS_OLD = (
 '            <tr>\n'
 '              <td><span class="name">투자실행액</span></td>\n'
 '              <td class="num"><span class="strong">1,284,500,000</span></td>\n'
 '              <td class="num">11.2일</td>\n'
 '              <td class="num">0.42%</td>\n'
 '              <td class="num">3.59%</td>\n'
 '              <td class="num">92.4%</td>\n'
 '              <td>㈜페이허그</td>\n'
 '            </tr>')
STATUS_NEW = (
 '            <tr>\n'
 '              <td><span class="name">투자실행액</span></td>\n'
 '              <td class="num"><span class="strong">%s</span></td>\n'
 '              <td class="num">%s일</td>\n'
 '              <td class="num">%s%%</td>\n'
 '              <td class="num">%s%%</td>\n'
 '              <td class="num">%s%%</td>\n'
 '              <td>㈜페이허그</td>\n'
 '            </tr>' % (f(EXEC), r1(W_W), r2(S_W), r2(TY_W), EXEC_SHARE))

def fix_assets(fn, page):
    p = os.path.join(ROOT, fn)
    s = io.open(p, encoding='utf-8').read()
    n0 = s
    # 요약 카드
    s = s.replace('<div class="summary-value">1,389,800,000<span class="unit">원</span></div>',
                  '<div class="summary-value">%s<span class="unit">원</span></div>' % f(TOTAL))
    s = s.replace('<div class="summary-value">1,284,500,000<span class="unit">원</span></div>',
                  '<div class="summary-value">%s<span class="unit">원</span></div>' % f(EXEC))
    s = s.replace('<div class="summary-sub">비중 92.4% · 보관 ㈜페이허그</div>',
                  '<div class="summary-sub">비중 %s%% · 보관 ㈜페이허그</div>' % EXEC_SHARE)
    s = s.replace('<div class="summary-sub">비중 7.6% · 보관 ㈜쿠콘</div>',
                  '<div class="summary-sub">비중 %s%% · 보관 ㈜쿠콘</div>' % CASH_SHARE)
    s = s.replace('<div class="summary-value">3.59<span class="unit">%</span></div>',
                  '<div class="summary-value">%s<span class="unit">%%</span></div>' % r2(TY_W))
    s = s.replace('<div class="summary-sub">W금융일수 11.2일 기준</div>',
                  '<div class="summary-sub">W금융일수 %s일 기준</div>' % r1(W_W))
    # 현황 표
    s = s.replace(STATUS_OLD, STATUS_NEW)
    s = s.replace('              <td class="num">7.6%</td>\n              <td>㈜쿠콘</td>',
                  '              <td class="num">%s%%</td>\n              <td>㈜쿠콘</td>' % CASH_SHARE)
    s = s.replace('              <td>합계 (투자자산)</td>\n              <td class="num">1,389,800,000</td>',
                  '              <td>합계 (투자자산)</td>\n              <td class="num">%s</td>' % f(TOTAL))
    # 가맹점별 표 (2번째 tbody)
    sl = PAIRS[0:8] if page == 1 else PAIRS[8:16]
    s = swap_tbody(s, 1, merch_rows(sl))
    # 페이지네이션
    s = fix_pagination(s, page)
    io.open(p, 'w', encoding='utf-8').write(s)
    return s != n0

BTN_RE  = re.compile(r'(?:[ \t]*<button class="page-btn(?: active)?">\d+</button>\n?)+')
ARW_RE  = re.compile(r'<button class="page-arrow"( disabled)?>')

def fix_pagination(s, page, pages=2):
    """페이지 버튼 묶음과 좌·우 화살표 disabled 를 현재 페이지에 맞춘다."""
    m = BTN_RE.search(s)
    if not m: raise SystemExit('page-btn 미발견')
    indent = re.match(r'[ \t]*', m.group(0)).group(0)
    btns = ''.join('%s<button class="page-btn%s">%d</button>\n'
                   % (indent, ' active' if i == page else '', i) for i in range(1, pages + 1))
    s = s[:m.start()] + btns + s[m.end():]
    arrows = list(ARW_RE.finditer(s))
    if len(arrows) != 2: raise SystemExit('page-arrow %d개' % len(arrows))
    for idx, want in ((1, page < pages), (0, page > 1)):      # 뒤에서부터 치환해 offset 유지
        a = arrows[idx]
        rep = '<button class="page-arrow"%s>' % ('' if want else ' disabled')
        s = s[:a.start()] + rep + s[a.end():]
        arrows = list(ARW_RE.finditer(s))
    return s

# ════════════ 3) 증명서 ════════════
def fix_certificate():
    p = os.path.join(ROOT, 'certificate.html')
    s = io.open(p, encoding='utf-8').read()
    rows = []
    for (n, a, w, sv, *_), sh in PAIRS:
        rows.append('              <tr>\n'
                    '                <td>%s</td>\n'
                    '                <td class="num">%s</td>\n'
                    '                <td class="num">%s일</td>\n'
                    '                <td class="num">%s%%</td>\n'
                    '                <td class="num">%s%%</td>\n'
                    '                <td class="num">%s%%</td>\n'
                    '              </tr>' % (n, f(a), w, sv, ty_of(w), sh))
    start = s.index('<tbody>'); end = s.index('</tbody>', start)
    s = s[:start + 7] + '\n' + '\n'.join(rows) + '\n            ' + s[end:]
    s = s.replace('<td class="num">1,284,500,000</td>', '<td class="num">%s</td>' % f(EXEC))
    s = s.replace('<span class="k">대상 가맹점</span><span class="v">8개</span>',
                  '<span class="k">대상 가맹점</span><span class="v">%d개</span>' % len(ROSTER))
    io.open(p, 'w', encoding='utf-8').write(s)

# ════════════ 4) 엑셀 미리보기 2종 ════════════
def fix_xls_assets_status():
    p = os.path.join(ROOT, 'xls-assets-status.html')
    s = io.open(p, encoding='utf-8').read()
    s = s.replace('<td>투자실행액</td><td class="c-num">1,284,500,000</td><td class="c-num">11.2</td><td class="c-num">0.42%</td><td class="c-num">3.59%</td><td class="c-num">92.4%</td>',
                  '<td>투자실행액</td><td class="c-num">%s</td><td class="c-num">%s</td><td class="c-num">%s%%</td><td class="c-num">%s%%</td><td class="c-num">%s%%</td>'
                  % (f(EXEC), r1(W_W), r2(S_W), r2(TY_W), EXEC_SHARE))
    s = s.replace('<td>순현금</td><td class="c-num">105,300,000</td><td></td><td></td><td></td><td class="c-num">7.6%</td>',
                  '<td>순현금</td><td class="c-num">105,300,000</td><td></td><td></td><td></td><td class="c-num">%s%%</td>' % CASH_SHARE)
    s = s.replace('<td>합계 (투자자산)</td><td class="c-num">1,389,800,000</td>',
                  '<td>합계 (투자자산)</td><td class="c-num">%s</td>' % f(TOTAL))
    io.open(p, 'w', encoding='utf-8').write(s)

def fix_xls_assets_merchant():
    p = os.path.join(ROOT, 'xls-assets-merchant.html')
    s = io.open(p, encoding='utf-8').read()
    rows = []
    for i, ((n, a, w, sv, *_), sh) in enumerate(PAIRS):
        rows.append('            <tr><th class="row-head">%d</th><td>%s</td><td class="c-num">%s</td>'
                    '<td class="c-num">%s</td><td class="c-num">%s%%</td><td class="c-num">%s%%</td>'
                    '<td class="c-num">%s%%</td><td class="c-empty"></td></tr>'
                    % (4 + i, n, f(a), w, sv, ty_of(w), sh))
    tot = 4 + len(PAIRS)
    rows.append('            <tr class="r-total"><th class="row-head">%d</th><td>합계</td><td class="c-num">%s</td>'
                '<td></td><td></td><td></td><td class="c-num">100.0%%</td><td class="c-empty"></td></tr>' % (tot, f(EXEC)))
    blank = ('            <tr><th class="row-head">%d</th>' + '<td class="c-empty"></td>' * 7 + '</tr>')
    rows.append(blank % (tot + 1))
    rows.append('            <tr><th class="row-head">%d</th><td class="c-note" colspan="7">'
                '※ 비중은 투자실행액 합계(%s원) 대비 각 가맹점 투자금액의 구성비.</td></tr>' % (tot + 2, f(EXEC)))
    rows.append(blank % (tot + 3))
    rows.append(blank % (tot + 4))
    a = s.index('<tr><th class="row-head">4</th>')
    a = s.rindex('\n', 0, a) + 1
    b = s.index('</tbody>', a)
    b = s.rindex('\n', 0, b) + 1
    s = s[:a] + '\n'.join(rows) + '\n' + s[b:]
    io.open(p, 'w', encoding='utf-8').write(s)

# ════════════ 5) 가맹점·계약기록 건수 ════════════
def fix_counts():
    n = len(ROSTER)
    for fn, page in (('merchants.html', 1), ('merchants--filter-open.html', 1),
                     ('contracts.html', 1), ('contracts--all.html', 1)):
        p = os.path.join(ROOT, fn)
        s = io.open(p, encoding='utf-8').read()
        s = s.replace('총 <b class="mono">8</b>건', '총 <b class="mono">%d</b>건' % n)
        s = fix_pagination(s, page)
        if fn.startswith('contracts'):
            s = s.replace('>8건 선택<', '>%d건 선택<' % n).replace('다운로드 (8)', '다운로드 (%d)' % n)
        io.open(p, 'w', encoding='utf-8').write(s)
    # 검색 적용 상태 — '곱창' 은 왕십리곱창타운까지 2건
    p = os.path.join(ROOT, 'merchants--filtered.html')
    s = io.open(p, encoding='utf-8').read()
    row = ('            <tr class="clickable">\n'
           '              <td class="mono">M2026-0010</td>\n'
           '              <td><span class="name">왕십리곱창타운</span></td>\n'
           '              <td class="mono">112-34-56789</td>\n'
           '              <td>한상철</td>\n'
           '              <td>음식점업</td>\n'
           '              <td>곱창</td>\n'
           '              <td><span class="badge sm badge-primary">A-001</span> ㈜페이허그</td>\n'
           '            </tr>')
    end = s.index('</tbody>')
    ins = s.rindex('</tr>', 0, end) + len('</tr>')
    s = s[:ins] + '\n' + row + s[ins:]
    s = s.replace('총 <b class="mono">1</b>건', '총 <b class="mono">2</b>건')
    io.open(p, 'w', encoding='utf-8').write(s)

# ════════════ 6) xlsx 실파일 ════════════
def fix_xlsx():
    import openpyxl
    from copy import copy
    base = os.path.join(ROOT, 'assets', 'xlsx')
    # 6-1 맵의 cell 항목
    cells = {}
    for x in MAP['fixes']:
        if x['kind'] != 'cell': continue
        cells.setdefault(os.path.basename(x['file']), []).append((x['locator'], x['corrected']))
    for fn, items in cells.items():
        wb = openpyxl.load_workbook(os.path.join(base, fn))
        for loc, val in items:
            sh, addr = loc.split('!')
            wb[sh][addr] = val
        if fn.startswith('투자수익_현황'):
            wb['투자수익 현황']['B8'] = float(DSUM['tyAsset']) / 100
        wb.save(os.path.join(base, fn))
    # 6-2 투자자산 현황
    p = os.path.join(base, '투자자산_현황_20260827.xlsx')
    wb = openpyxl.load_workbook(p); ws = wb['투자자산 현황']
    ws['B4'] = EXEC; ws['C4'] = float(r1(W_W)); ws['D4'] = float(r2(S_W)) / 100
    ws['E4'] = float(r2(TY_W)) / 100; ws['F4'] = float(EXEC_SHARE) / 100
    ws['F5'] = float(CASH_SHARE) / 100; ws['B6'] = TOTAL
    wb.save(p)
    # 6-3 가맹점별 투자자산 — 8행 → 16행
    p = os.path.join(base, '가맹점별_투자자산_20260827.xlsx')
    wb = openpyxl.load_workbook(p); ws = wb['가맹점별 투자자산']
    style = {c: (copy(ws['%s11' % c]._style)) for c in 'ABCDEF'}
    ws.insert_rows(12, 8)
    for i, ((n, a, w, sv, *_), sh) in enumerate(PAIRS):
        r = 4 + i
        ws['A%d' % r] = n; ws['B%d' % r] = a; ws['C%d' % r] = float(w)
        ws['D%d' % r] = float(sv) / 100; ws['E%d' % r] = float(ty_of(w)) / 100
        ws['F%d' % r] = float(sh) / 100
        for c in 'ABCDEF': ws['%s%d' % (c, r)]._style = copy(style[c])
    tot = 4 + len(PAIRS)
    ws['B%d' % tot] = EXEC; ws['F%d' % tot] = 1
    ws['A%d' % (tot + 2)] = '※ 비중은 투자실행액 합계(%s원) 대비 각 가맹점 투자금액의 구성비.' % f(EXEC)
    wb.save(p)

if __name__ == '__main__':
    c, e = apply_map()
    P('== 맵 적용 (텍스트) ==')
    for k in sorted(c): P('  %-32s %d건' % (k, c[k]))
    P('  locator 오류 %d %s' % (len(e), e))
    P('== 로스터 16건 ==')
    for fn, pg in (('invest-assets.html', 1), ('invest-assets--page2.html', 2),
                   ('invest-assets--download.html', 1), ('invest-assets--cert-confirm.html', 1)):
        P('  %-34s 변경 %s (페이지 %d)' % (fn, fix_assets(fn, pg), pg))
    fix_certificate();          P('  certificate.html                   16행 · 대상 16개')
    fix_xls_assets_status();    P('  xls-assets-status.html             현황 3행')
    fix_xls_assets_merchant();  P('  xls-assets-merchant.html           16행 + 합계')
    fix_counts();               P('  merchants·contracts 건수·페이지네이션')
    fix_xlsx();                 P('  assets/xlsx 4종')
    io.open(os.path.join(PIPE, 'roster16_apply.log'), 'w', encoding='utf-8').write('\n'.join(log))
