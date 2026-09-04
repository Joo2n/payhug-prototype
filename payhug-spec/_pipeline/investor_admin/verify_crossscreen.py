# -*- coding: utf-8 -*-
"""화면 간 정합 — 정적 HTML · app.html 데이터셋 · xlsx 실파일이 같은 숫자를 말하는지 대조."""
import io, os, re, json, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from roster16_model import ROSTER, SHARES, EXEC, CASH, TOTAL, W_W, S_W, TY_W, EXEC_SHARE, CASH_SHARE, DAILY, MONTHLY, DSUM, AUG_CARD, r2, ty, f
import openpyxl

ROOT = '/Users/semi/cursor/payhug-investor-admin'
def rd(p): return io.open(os.path.join(ROOT, p), encoding='utf-8').read()
def n(t): return int(str(t).replace(',', ''))
strip = lambda h: re.sub(r'<[^>]+>', '', h).strip()

rows_out, fails = [], 0
def chk(name, got, want):
    global fails
    ok = got == want
    if not ok: fails += 1
    rows_out.append((name, ok, got, want))

WANT_MERCH = [(x[0], x[1], x[2], x[3], str(ty(x[2])), str(sh)) for x, sh in zip(ROSTER, SHARES)]

# ── 정적 HTML: 가맹점 표 ──────────────────────────────────────────
TR = re.compile(r'<tr>\s*<td><span class="name">([^<]+)</span></td>\s*'
                r'<td class="num"><span class="strong">([\d,]+)</span></td>\s*'
                r'<td class="num">([\d.]+)일</td>\s*<td class="num">([\d.]+)%</td>\s*'
                r'<td class="num">([\d.]+)%</td>\s*<td class="num">([\d.]+)%</td>\s*</tr>')
def html_merch(p):
    return [(a, n(b), c, d, e, g) for a, b, c, d, e, g in TR.findall(rd(p))]

p1 = html_merch('invest-assets.html')
chk('invest-assets 로스터', p1, WANT_MERCH)
chk('invest-assets--download 1p',  html_merch('invest-assets--download.html'), WANT_MERCH[:10])
chk('invest-assets--cert-confirm 1p', html_merch('invest-assets--cert-confirm.html'), WANT_MERCH[:10])

# ── 증명서 ────────────────────────────────────────────────────────
CTR = re.compile(r'<tr>\s*<td>([^<]+)</td>\s*<td class="num">([\d,]+)</td>\s*<td class="num">([\d.]+)일</td>\s*'
                 r'<td class="num">([\d.]+)%</td>\s*<td class="num">([\d.]+)%</td>\s*<td class="num">([\d.]+)%</td>\s*</tr>')
cert = rd('certificate.html')
chk('certificate 로스터', [(a, n(b), c, d, e, g) for a, b, c, d, e, g in CTR.findall(cert)], WANT_MERCH)
chk('certificate 대상 건수', re.search(r'대상 가맹점</span><span class="v">(\d+)개', cert).group(1), str(len(ROSTER)))

# ── 엑셀 미리보기 ─────────────────────────────────────────────────
XTR = re.compile(r'<tr><th class="row-head">\d+</th><td>([^<]+)</td><td class="c-num">([\d,]+)</td>'
                 r'<td class="c-num">([\d.]+)</td><td class="c-num">([\d.]+)%</td>'
                 r'<td class="c-num">([\d.]+)%</td><td class="c-num">([\d.]+)%</td>')
chk('xls-assets-merchant 로스터',
    [(a, n(b), c, d, e, g) for a, b, c, d, e, g in XTR.findall(rd('xls-assets-merchant.html'))], WANT_MERCH)

# ── xlsx 실파일 ───────────────────────────────────────────────────
wb = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/가맹점별투자자산_2026-08-27_2026-08-27.xlsx'))
ws = wb['가맹점별 투자자산']
xls = [(ws['A%d' % r].value, ws['B%d' % r].value,
        ('%.2f' % ws['C%d' % r].value), ('%.2f' % (ws['D%d' % r].value * 100)),
        ('%.2f' % (ws['E%d' % r].value * 100)), ('%.1f' % (ws['F%d' % r].value * 100)))
       for r in range(4, 4 + len(ROSTER))]
chk('가맹점별투자자산.xlsx 로스터', xls, WANT_MERCH)
# 합계 행 번호는 로스터 곳수에서 나온다 — 4행부터 시작해 로스터 행 다음 줄이다.
_TOT = 4 + len(ROSTER)
chk('가맹점별투자자산.xlsx 합계',
    (ws['B%d' % _TOT].value, round(ws['F%d' % _TOT].value, 4)), (EXEC, 1))

wb = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/투자자산현황_2026-08-27_2026-08-27.xlsx'))
ws = wb['투자자산 현황']
chk('투자자산현황.xlsx 투자실행액 행',
    (ws['B4'].value, '%.2f' % ws['C4'].value, '%.2f' % (ws['D4'].value * 100),
     '%.2f' % (ws['E4'].value * 100), '%.1f' % (ws['F4'].value * 100)),
    (EXEC, str(r2(W_W)), str(r2(S_W)), str(r2(TY_W)), str(EXEC_SHARE)))
chk('투자자산현황.xlsx 순현금·합계',
    (ws['B5'].value, '%.1f' % (ws['F5'].value * 100), ws['B6'].value),
    (CASH, str(CASH_SHARE), TOTAL))

import build_xlsx as BX
_P = {p[0]: p for p in BX.PRESETS}          # week · w4 · m6 — 화면 기본 프리셋과 같은 묶음
wb = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/' + BX.profit_file('daily', _P['week'][3], _P['week'][4])))
ws = wb['일별 투자수익']
chk('일별투자수익.xlsx 7행',
    [(ws['A%d' % r].value, ws['B%d' % r].value, ws['C%d' % r].value, ws['D%d' % r].value,
      '%.2f' % ws['E%d' % r].value, '%.2f' % (ws['F%d' % r].value * 100)) for r in range(4, 11)],
    [(x['d'], x['repay'], x['exec'], x['profit'], x['w'], str(x['ty'])) for x in DAILY])
chk('일별투자수익.xlsx 합계',
    (ws['B11'].value, ws['C11'].value, ws['D11'].value, '%.2f' % ws['E11'].value, '%.2f' % (ws['F11'].value * 100)),
    (DSUM['repay'], DSUM['exec'], DSUM['profit'], str(DSUM['w']), str(DSUM['ty'])))

wb = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/' + BX.status_file(_P['week'][3], _P['week'][4])))
ws = wb['투자수익 현황']
chk('투자수익현황.xlsx 카드',
    (ws['B5'].value, ws['B6'].value, '%.2f' % (ws['B7'].value * 100), '%.2f' % (ws['B8'].value * 100)),
    (DSUM['exec'], DSUM['profit'], str(DSUM['ty']), str(DSUM['tyAsset'])))

# ── 정적 투자수익 화면 ────────────────────────────────────────────
PR = re.compile(r'<td class="mono">(\d{4}-\d\d(?:-\d\d)?)</td>\s*<td class="num">([\d,]+)</td>\s*'
                r'<td class="num">([\d,]+)</td>\s*<td class="num"><span class="strong">([\d,]+)</span></td>\s*'
                r'<td class="num">([\d.]+)</td>\s*<td class="num">([\d.]+)%</td>')
def html_profit(p):
    return [(a, n(b), n(c), n(d), e, g) for a, b, c, d, e, g in PR.findall(rd(p))]
want_daily = [(x['d'], x['repay'], x['exec'], x['profit'], x['w'], str(x['ty'])) for x in DAILY]
want_mon   = [(x['d'], x['repay'], x['exec'], x['profit'], x['w'], str(x['ty'])) for x in MONTHLY]
chk('invest-profit 일별 7행', html_profit('invest-profit.html'), want_daily)
chk('invest-profit--monthly 월별 6행', html_profit('invest-profit--monthly.html'), want_mon)

# ── 주별·월별 낱장 ↔ 주별·월별 엑셀 실물 ─────────────────────────
#   버킷 라벨(2026-08-17 ~ 08-23 · 2026-08)을 받는 별도 정규식으로 낱장 표를 읽고,
#   같은 자리 엑셀과 값을 맞대 본다. 화면과 파일이 갈라지면 여기서 걸린다.
BR = re.compile(r'<td class="mono">([\d\-~ ]+)</td>\s*<td class="num">([\d,]+)</td>\s*'
                r'<td class="num">([\d,]+)</td>\s*<td class="num"><span class="strong">([\d,]+)</span></td>\s*'
                r'<td class="num">([\d.]+)</td>\s*<td class="num">([\d.]+)%</td>')
def html_bucket(p):
    return [(a.strip(), n(b), n(c), n(d), e, g) for a, b, c, d, e, g in BR.findall(rd(p))]

def xls_bucket(fn, sheet, nrows):
    w = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/' + fn))[sheet]
    return [(str(w['A%d' % r].value), w['B%d' % r].value, w['C%d' % r].value, w['D%d' % r].value,
             '%.2f' % w['E%d' % r].value, '%.2f' % (w['F%d' % r].value * 100))
            for r in range(4, 4 + nrows)]

# got = 낱장 HTML · want = 같은 자리 엑셀. 어느 쪽이 정본인지는 이 자리가 정하지 못한다 —
# 아래 app.html 절의 「경로 지목」 두 줄이 원장 6자리 가중으로 갈라 준다.
XLS_WEEKLY = xls_bucket(BX.profit_file('weekly', _P['w4'][3], _P['w4'][4]), '주별 투자수익', 4)
XLS_MONTHLY = xls_bucket(BX.profit_file('monthly', _P['m6'][3], _P['m6'][4]), '월별 투자수익', 6)
HTML_MONTHLY = html_bucket('invest-profit--monthly.html')
chk('invest-profit--weekly 낱장(got) = 주별투자수익.xlsx(want) 4행',
    html_bucket('invest-profit--weekly.html'), XLS_WEEKLY)
chk('invest-profit--monthly 낱장(got) = 월별투자수익.xlsx(want) 6행',
    HTML_MONTHLY, XLS_MONTHLY)

# ── 주별·월별 낱장의 `수익 현황` 카드 ↔ 같은 기간 현황 엑셀 ───────
#   카드가 4주를 말하는데 링크·파일이 일주일이면 여기서 걸린다.
CARD_NUM = re.compile(r'<div class="summary-label">(?:투자실행금|투자수익)</div>\s*'
                      r'<div class="summary-value">([\d,]+)<span class="unit">원</span>')
CARD_TY = re.compile(r'(?:투자실행금액|투자자산) 대비.*?<div class="summary-value">([\d.]+)'
                     r'<span class="unit">%</span>', re.S)
CARD_PERIOD = re.compile(r'<div class="summary-sub mono">([^<]+)</div>')

def html_status(p):
    s_ = rd(p)
    ex, pf = [n(x) for x in CARD_NUM.findall(s_)]
    t4, t5 = CARD_TY.findall(s_)
    return (CARD_PERIOD.search(s_).group(1), ex, pf, t4, t5)

def xls_status(fn):
    w = openpyxl.load_workbook(os.path.join(ROOT, 'assets/xlsx/' + fn))['투자수익 현황']
    period = w['B4'].value.split('(')[1].rstrip(')')
    return (period, w['B5'].value, w['B6'].value,
            '%.2f' % (w['B7'].value * 100), '%.2f' % (w['B8'].value * 100))

for _f, _x in (('invest-profit.html',            BX.status_file(_P['week'][3], _P['week'][4])),
               ('invest-profit--weekly.html',    BX.status_file(_P['w4'][3], _P['w4'][4])),
               ('invest-profit--monthly.html',   BX.status_file(_P['m6'][3], _P['m6'][4]))):
    chk('%s 현황 카드 = %s' % (_f.replace('.html', ''), _x), html_status(_f), xls_status(_x))
    _href = 'assets/xlsx/' + _x
    chk('%s 현황 카드 링크' % _f.replace('.html', ''), _href in rd(_f), True)

XPR = re.compile(r'<td class="c-num">([\d,]+)</td><td class="c-num">([\d,]+)</td><td class="c-num">([\d,]+)</td>'
                 r'<td class="c-num">([\d.]+)</td><td class="c-num">([\d.]+)%</td>')
chk('xls-profit-daily 7행',
    [(n(a), n(b), n(c), d, e) for a, b, c, d, e in XPR.findall(rd('xls-profit-daily.html'))][:7],
    [(x['repay'], x['exec'], x['profit'], x['w'], str(x['ty'])) for x in DAILY])

sp = rd('xls-profit-status.html')
chk('xls-profit-status 카드',
    (n(re.search(r'투자실행금</td><td class="c-num">([\d,]+)', sp).group(1)),
     n(re.search(r'>투자수익</td><td class="c-num">([\d,]+)', sp).group(1)),
     re.search(r'투자실행금액 대비\)</td><td class="c-num">([\d.]+)%', sp).group(1),
     re.search(r'투자자산 대비\)</td><td class="c-num">([\d.]+)%', sp).group(1)),
    (DSUM['exec'], DSUM['profit'], str(DSUM['ty']), str(DSUM['tyAsset'])))

# ── app.html 데이터셋 ────────────────────────────────────────────
app = rd('app.html')
def jsarr(name):
    a = app.index('var %s = [' % name); b = app.index('];', a)
    return app[a:b]
mm = re.findall(r"name:'([^']+)'.*?amount:(\d+), w:([\d.]+), s:([\d.]+), ty:([\d.]+)", jsarr('MERCHANTS'))
chk('app.html MERCHANTS',
    [(a, int(b), ('%.2f' % float(c)), ('%.2f' % float(d)), ('%.2f' % float(e))) for a, b, c, d, e in mm],
    [(x[0], x[1], x[2], x[3], str(ty(x[2]))) for x in ROSTER])
ar = re.search(r"\{name:'투자실행액', amount:(\d+), w:([\d.]+),\s+s:([\d.]+), ty:([\d.]+)", app)
chk('app.html ASSET_ROWS 투자실행액',
    (int(ar.group(1)), '%.2f' % float(ar.group(2)), '%.2f' % float(ar.group(3)), '%.2f' % float(ar.group(4))),
    (EXEC, str(r2(W_W)), str(r2(S_W)), str(r2(TY_W))))
#   행 끝의 `ad`(= 그 날 Σ Ai×Di · ⑤ 새 분모 재료)는 step7 ⑤ 산식 교체(2026-09-04 · daily_ledger.js_array)로
#   붙었다. 있어도 없어도 읽는다 — 없던 시절 꼴을 못 읽게 좁히지 않는다. [기준 교체 2026-09-04]
_DAILY_RE = re.compile(r"\{d:'([\d-]+)', repay:(\d+), exec:(\d+), profit:(\d+), w:([\d.]+), ty:([\d.]+)(?:, ad:\d+)?\}")
def jsraw(name):
    """app.html 이 실은 일별 원장 원문 — W·Ty 는 6자리 그대로다(build_app.js_array)."""
    return [(a, int(b), int(c), int(d), e, g) for a, b, c, d, e, g in _DAILY_RE.findall(jsarr(name))]
def jsrows(name):
    """화면 표기(2자리)로 자른 것 — 정적 낱장·엑셀의 표기값과 맞댈 때 쓴다."""
    return [(a, b, c, d, ('%.2f' % float(e)), ('%.2f' % float(g))) for a, b, c, d, e, g in jsraw(name)]
# app.html 은 일별 원장 한 벌만 갖고, 월별 표는 그 원장을 달별로 합쳐 만든다.
# 대조 1 — 원장 가운데 기본 조회 기간(일주일) 구간이 정적 화면·xlsx 와 같은 7행인가.
# 대조 2 — 원장을 달별로 합친 값이 월별 정적 화면의 6행과 같은가(합계·가중평균 모두).
LEDGER = jsrows('DAILY')
#   원장을 한 행도 못 읽으면 아래 「⊂ 범위」 검사들이 빈 채로 통과한다 — 0건은 그 자체로 FAIL 이다.
chk('app.html 일별 원장 DAILY 0건 아님 (파서가 행 꼴을 읽는가)', len(LEDGER) > 0, True)
chk('app.html 일별 원장 · 기본 기간 7행',
    [r for r in LEDGER if '2026-08-21' <= r[0] <= '2026-08-27'], want_daily)

def month_rollup(rows, digits=6):
    """통합본 build_app.py 의 rollupBy 와 같은 규칙으로 달별로 묶는다.

    W금융일수는 투자실행금 가중평균, Ty수익율은 그 가중평균에서 되짚는다(PSMR x 365 / PSD).
    일자별 Ty 를 다시 가중평균하면 1/w 의 볼록성 때문에 달 행에 적힌 W 로 되짚을 수 없다.
    만기가 3일대로 짧아지면 그 격차가 0.05%p 까지 벌어져 화면과 대조가 갈린다.

    digits — 가중에 넣는 일별 W 의 자릿수. 6 이 정본이다(dm_0901/rounding_rule_0901.md 규칙 1:
    한 단계마다 6자리에서 끊고 그 값을 다음 계산에 넣는다. 화면에만 2자리로 자른다).
    2 를 넣으면 그 규칙이 폐기한 「표기 2자리 값을 다시 씀」 경로가 되고, 2026-06 이 4.38 → 4.39 로
    한 눈금 뒤집힌다. 판정에 쓰는 것은 6 이고, 2 는 어긋남의 정체를 지목할 때만 쓴다.
    """
    from decimal import Decimal as _D
    agg, order = {}, []
    for d, repay, ex, pr, w, t in rows:
        k = d[:7]
        if k not in agg:
            agg[k] = dict(repay=0, ex=0, pr=0, wx=_D(0)); order.append(k)
        g = agg[k]
        wv = _D(w) if digits == 6 else _D('%.2f' % float(w))
        g['repay'] += repay; g['ex'] += ex; g['pr'] += pr
        g['wx'] += wv * _D(ex)
    out = []
    for k in order:
        g = agg[k]
        wv = g['wx'] / _D(g['ex'])
        tv = (_D(g['pr']) / _D(g['ex']) * _D(100)) * _D(365) / wv
        out.append((k, g['repay'], g['ex'], g['pr'], str(r2(wv)), str(r2(tv))))
    return out
# [기준 교체 2026-09-02] 이 검증기가 스스로 2자리 W 로 가중하고 있었다 — jsrows() 가 원장 6자리를
# '%.2f' 로 잘라 넘겨, 규칙 1 이 폐기한 경로를 검증기가 재현하며 2026-06 을 4.39 로 냈다.
# 화면·낱장은 4.38 이라 FAIL 이 났고, 그 FAIL 은 화면 결함이 아니라 검증기 노후였다.
LEDGER6 = jsraw('DAILY')
chk('app.html 일별 원장 월 롤업 = 월별 6행', month_rollup(LEDGER6), want_mon)

# ── 월별 Ty 한 칸이 세 곳에서 갈릴 때 어느 경로가 틀렸는지 지목한다 ────
#   원장(app.html DAILY) 6자리 가중이 정본이다 — dm_0901/rounding_rule_0901.md 규칙 1.
#   같은 원장을 2자리로 잘라 다시 가중하면 규칙 1 이 폐기한 경로가 되고, 그 경로가 낸 값과
#   엑셀이 같으면 엑셀이 그 폐기 경로를 쓰고 있다는 뜻이다(build_xlsx.bucket 의 D(str(r['w']))).
#   build_xlsx.py 는 이 조의 수정 대상이 아니다 — 지목만 한다.
MON6 = month_rollup(LEDGER6)                 # 정본
MON2 = month_rollup(LEDGER6, digits=2)       # 폐기 경로 재현 — 판정에 쓰지 않는다
_TY6 = [(g[0], g[5]) for g in MON6]
_TY2 = [(g[0], g[5]) for g in MON2]
_TYH = [(g[0], g[5]) for g in HTML_MONTHLY]
_TYX = [(g[0], g[5]) for g in XLS_MONTHLY]
chk('경로 지목 · 월별 Ty — 낱장 invest-profit--monthly.html = 원장 6자리 가중', _TYH, _TY6)
chk('경로 지목 · 월별 Ty — 월별투자수익.xlsx = 원장 6자리 가중 '
    '(어긋나면 build_xlsx.bucket 의 2자리 가중 · 다른 조 담당)', _TYX, _TY6)
# 어긋난 칸이 정확히 폐기 경로가 낸 값인지까지 못 박는다 — 「그냥 다르다」로 끝내지 않는다.
if _TYX != _TY6:
    rows_out.append(('경로 지목 · 엑셀이 낸 값 = 2자리 가중 재현값 (정체 확인)', _TYX == _TY2, _TYX, _TY2))
    if _TYX != _TY2:
        fails += 1

chk('app.html PAGE_SIZE', re.search(r'var PAGE_SIZE = (\d+);', app).group(1), '10')

# ── ⑥(일별·버킷 행 Ty수익율)이 ⑤ 함수를 거치는 자리 — 미확정 표시 ────
#   판정하지 않는다. 달 버킷 ⑥ 이 ty5() 를 거치는 것이 맞는지는 대표 확인 문항
#   Q1(⑤ 산식 재전달)·Q2(③ 이 현황의 어느 칸인가)에 걸려 있다. 어느 쪽이 옳은지 대는 근거가
#   지금 없으므로 값을 정답으로도 오답으로도 세지 않고, 「미확정에 걸린 자리」로 적어 둔다.
#   지금 TY6_PSC 가 0 이라 ⑥ 은 ④ 와 같은 값이다 — 그 0 이 바뀌면 월별 Ty 가 통째로 움직인다.
import daily_ledger as _DL
_PEND = [('⑤ ty_asset', _DL.TY5_STATUS, _DL.TY5_SOURCE),
         ('③ ty_third', _DL.TY3_STATUS, _DL.TY3_SOURCE),
         ('⑥ TY6_PSC', str(_DL.TY6_PSC), 'ty5() 를 거친다 · PSC 0 인 동안 ④ 와 같은 값')]
PENDING_NOTE = [{'자리': a, '상태': b, '출처': c} for a, b, c in _PEND]

# ── W금융일수 현실 범위 — 조현준 슬랙 2026-08-28 실측 2.0 ~ 6.2일 ──
from platform_duration import FLOOR, CEIL
from decimal import Decimal as _DD
def in_range(vals):
    return sorted({str(v) for v in vals if not (_DD(str(FLOOR)) <= _DD(str(v)) <= _DD(str(CEIL)))})
chk('로스터 %d건 W금융일수 ⊂ [%s, %s]' % (len(ROSTER), FLOOR, CEIL), in_range(x[2] for x in ROSTER), [])
chk('일별 원장 %d일 W금융일수 ⊂ [%s, %s]' % (len(LEDGER), FLOOR, CEIL), in_range(r[4] for r in LEDGER), [])
chk('낱장 가맹점 표 W금융일수 ⊂ [%s, %s]' % (FLOOR, CEIL), in_range(x[2] for x in p1), [])
chk('투자실행액 행 W금융일수 ⊂ [%s, %s]' % (FLOOR, CEIL), in_range([r2(W_W)]), [])

# ── 가중치 등재 — 정본 표기는 w금융일수 하나, duration 은 허용 블록 한 곳 ──────
#   이 검사의 이름은 「대출 어휘 0건」이다. 막으려는 것은 선정산을 대출로 읽히게 하는 낱말이지
#   금융 일반어가 아니다. 그 취지가 규칙을 좁히는 근거다.
#
#   2026-08-31 재판정. 앞선 판은 「Duration·가중평균만기·만기」를 문서 전체에서 0건으로 막았고
#   사유는 「근거 없는 조어」였다. 그 사유는 무효다 — 대표가 세 자리에서 직접 썼다.
#     · 대표 DM 2026-08-31 16:41:24 「w금융일수 = 공식 용어로 duration, 우리는 가중 평균 만기일(weight)」
#     · 3차 미팅 00:43:33 「엄밀히는 공식 영어는 듀레이션」
#     · 3차 미팅 00:44:30 「가중 평균 만기해서 웨이티드의 W 를 붙인 것」
#   가중평균 만기일은 매콜리 듀레이션의 정의 그 자체이고, 금리 민감도는 수정 듀레이션이라 별개다.
#   그래서 「뜻이 어긋난다」는 옛 사유도 무효다.
#   낱말을 통짜로 막으면 대표 원문 인용도 못 싣는다 — DM 2026-08-31 15:37 「전일자 만기 (지급예정일)」.
#   인용 자리를 막는 규칙은 그 자체로 틀렸다.
#
#   그래서 이렇게 좁힌다.
#     · 화면 6종 — 지금대로 0건 (D-23)
#     · glossary.html — 허용 블록 밖에서 0건. 허용 블록은 부록 A 낱글자 `w` 항의 각주 한 곳뿐이고,
#       블록 전문을 글자 그대로 못 박아 곳수를 1로 고정한다. **2개 이상이면 FAIL** 이다.
#       같은 블록이 둘이면 count 가 2가 되어 걸리고, 다른 문장으로 자리를 더 만들면
#       블록을 걷어 낸 나머지에 낱말이 남아 걸린다.
#   정본 표기는 `w금융일수` 하나 그대로다. 화면·표·머리글 어디에도 다른 이름을 쓰지 않는다.
BANNED = ('가중평균만기', '평균만기', '만기')          # 한글은 글자 그대로
BANNED_EN = re.compile(r'(?i)duration')                # 영문은 대소문자 가리지 않는다
DUR_NOTE = (
    '공식 용어는 duration, 대표 표현은 가중평균 만기일 — 대표 DM 2026-08-31 16:41:24 '
    '「w금융일수 = 공식 용어로 duration, 우리는 가중 평균 만기일(weight) 가중평균금융일수」 · '
    '3차 미팅 00:43:33 「엄밀히는 공식 영어는 듀레이션」 · 00:44:30 '
    '「가중 평균 만기해서 웨이티드의 W 를 붙인 것」. 가중평균 만기일은 매콜리 듀레이션의 정의 그대로이고, '
    '금리 민감도를 재는 수정 듀레이션과는 다르다. 정본 표기는 w금융일수 하나이고 '
    '화면·표·머리글에는 이 이름만 쓴다.')


def _text(h):
    """태그·스크립트를 걷고 공백을 접은 화면 글자."""
    import html as _H
    return re.sub(r'\s+', ' ', _H.unescape(
        re.sub(r'<[^>]+>', ' ', re.sub(r'<(script|style)\b.*?</\1>', ' ', h, flags=re.S | re.I)))).strip()


gl = rd('glossary.html')
glt = _text(gl)
chk('용어 해설 · 가중치 블록 3종', gl.count('<b>가중치</b>'), 3)
chk('용어 해설 · 가중치가 투자액(= 선정산대금)임을 3곳에서 말함',
    gl.count('투자액(= 선정산대금)'), 3)
chk('용어 해설 · 정본 표기 w금융일수 등재', 'w금융일수' in gl, True)
# (1) 허용 블록은 정확히 한 곳 — 곳수를 못 박는다
chk('용어 해설 · duration 허용 블록 = 1곳 (2곳 이상이면 FAIL · 부록 A 낱글자 w)',
    glt.count(DUR_NOTE), 1)
chk('용어 해설 · 허용 블록에 출처 세 자리(DM 16:41 · 회의 00:43:33 · 00:44:30) 병기',
    all(s in DUR_NOTE and s in glt for s in ('대표 DM 2026-08-31 16:41:24',
                                             '3차 미팅 00:43:33', '00:44:30')), True)
# (2) 그 블록을 통째로 걷어 낸 나머지에는 0건
_rest = glt.replace(DUR_NOTE, '', 1)
chk('용어 해설 · 허용 블록 밖 대출 어휘 0건 (%s)' % ' · '.join(BANNED),
    sorted({w for w in BANNED if w in _rest}), [])
chk('용어 해설 · 허용 블록 밖 duration 0건', BANNED_EN.findall(_rest), [])
# (3) 전체 곳수까지 못 박는다 — 블록 안에서 낱말이 늘어나도 걸린다
chk('용어 해설 · duration 총 곳수 = 허용 블록 안 곳수',
    len(BANNED_EN.findall(glt)), len(BANNED_EN.findall(DUR_NOTE)))
chk('용어 해설 · 「만기」 총 곳수 = 허용 블록 안 곳수',
    glt.count('만기'), DUR_NOTE.count('만기'))
chk('용어 해설 · 플랫폼별 평균 금융일수 실측 4건',
    all(v in gl for v in ('2.0일', '3.4일', '4.7일', '6.2일')), True)
chk('용어 해설 · 출처 병기', '정산주기.xlsx' in gl, True)
# (4) 화면 6종은 예외 없이 0건 — 허용 블록은 용어 해설에만 있다
screens = [rd(x) for x in ('app.html', 'invest-assets.html',
                           'invest-profit.html', 'invest-profit--weekly.html',
                           'invest-profit--monthly.html', 'certificate.html')]
chk('화면에 대출 어휘 없음 (D-23)',
    sorted({w for w in BANNED for t in screens if w in t}), [])
#   영문 duration 은 화면 글자만 본다. 원문 그대로 옮긴 프론트 주석(Toast.tsx duration 0)·
#   Tailwind 클래스(duration-200)·파이썬 상수(DURATION)가 마크업 안에 있고, 그것은 화면 낱말이 아니다.
chk('화면 글자에 duration 없음 (D-23)',
    sorted({m for t in screens for m in BANNED_EN.findall(_text(t))}), [])

# ── 열머리 모집단 툴팁 — W·S 가 자기 모집단을 스스로 말한다 ────────
#   현황표의 W·S·금액 세 칸은 모집단이 서로 달라, 가맹점별 표를 금액으로 가중평균해도
#   현황표 값과 맞아떨어지지 않는다. 건수는 채권 원장 실측이라 화면에 손으로 적을 수 없다.
import daily_ledger as _LG
POP_W_N = f(len(_LG.RECEIVABLES))
POP_S_N = f(_LG.facts()['sampleReceivables'])
POP_TIP = re.compile(r'<span class="tip-anchor">(가중평균 금융일수|입금부족률)</span>'
                     r'<span class="tip-panel">([^<]+)'
                     r'<span class="tip-row"><span>채권 건수</span>'
                     r'<span class="tip-green">([\d,]+)건</span>')
POP_WANT = [('가중평균 금융일수', '대상정산금채권 전체 (발생 기준)', POP_W_N),
            ('입금부족률', '선정산일이 기준일 20일 전 ~ 11일 전인 표본', POP_S_N)]
# 기준일 d 는 확정이라 열머리 두 곳 어디에도 `미확정` 배지가 없다.
PEND_TH = re.compile(r'<span class="tip-anchor">(가중평균 금융일수|입금부족률)</span>.*?'
                     r'</span></span>( <span class="badge sm badge-amber">미확정</span>)?</th>', re.S)
for _p in ('invest-assets.html', 'invest-assets--download.html',
           'invest-assets--cert-confirm.html', 'invest-assets--empty.html'):
    _got = POP_TIP.findall(rd(_p))
    chk('%s 열머리 모집단 툴팁 (현황표·가맹점별 표)' % _p, _got, POP_WANT * 2)
    chk('%s 열머리 미확정 배지 없음' % _p,
        [(_a, bool(_b)) for _a, _b in PEND_TH.findall(rd(_p))],
        [('가중평균 금융일수', False), ('입금부족률', False)] * 2)
# 통합본은 popTh() 가 그리므로 마크업이 아니라 그 재료(POP_W · POP_S)를 대조한다.
_app = rd('app.html')
chk('app.html 열머리 모집단 재료',
    re.findall(r"var (POP_[WS]) = \{of:'([^']+)',\s*n:'([\d,]+)건'(, pend:1)?\}", _app),
    [('POP_W', '대상정산금채권 전체 (발생 기준)', POP_W_N, ''),
     ('POP_S', '선정산일이 기준일 20일 전 ~ 11일 전인 표본', POP_S_N, '')])
chk('app.html 두 표가 같은 popTh 를 쓴다', _app.count("popTh('가중평균 금융일수', POP_W)"), 2)
chk('app.html 두 표가 같은 popTh 를 쓴다 (S)', _app.count("popTh('입금부족률', POP_S)"), 2)
# 모집단이 다르다는 사실 자체 — 두 건수가 같으면 툴팁을 달 이유가 없다
chk('W·S 모집단 건수가 서로 다르다', POP_W_N != POP_S_N, True)
# 미회수 채권 곳수와도 다르다(금액 칸의 모집단) — 세 칸이 각자 다른 집합이다
chk('금액 칸 모집단(미회수)과도 다르다',
    len({len(_LG.RECEIVABLES), _LG.facts()['sampleReceivables'], len(_LG.OPEN)}), 3)

print('== 미확정에 걸린 자리 (판정하지 않는다 · 대표 확인 문항 Q1 ⑤ 산식 · Q2 ③ 이 어느 칸) ==')
for _r in PENDING_NOTE:
    print('  · %s — %s (%s)' % (_r['자리'], _r['상태'], _r['출처']))
print('== 화면 간 정합 %d건 · 불일치 %d ==' % (len(rows_out), fails))
for name, ok, got, want in rows_out:
    print(('  PASS ' if ok else '  FAIL ') + name)
    if not ok:
        print('        got :', got)
        print('        want:', want)
sys.exit(1 if fails else 0)
