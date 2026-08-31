# -*- coding: utf-8 -*-
"""검산 통합문서 독립 검증 — 셀 수식을 파싱해 직접 계산한다.

LibreOffice 가 없으므로 엑셀 수식 파서·계산기를 따로 구현해 값을 낸다.
openpyxl 은 수식 문자열만 읽고, 계산은 여기서 한다. 생성기와 코드를 공유하지 않는다.
"""
import json, math, os, re, sys
from datetime import date, datetime, timedelta
from decimal import Decimal as D, ROUND_HALF_UP, ROUND_DOWN

import openpyxl
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
#   기본은 정본. 음성 시험 때만 AUDIT_XLSX 로 사본을 가리킨다.
XLSX = os.environ.get('AUDIT_XLSX', os.path.join(BASE, '검산_투자자어드민_20260901.xlsx'))
VDIR = os.environ.get('AUDIT_VARIANT_DIR', BASE)
EPOCH = date(1899, 12, 30)


# ── 토크나이저 ────────────────────────────────────────────────────
TOK = re.compile(r"""
 (?P<str>"(?:[^"]|"")*")
|(?P<num>\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)
|(?P<ref>(?:'[^']+'|[A-Za-z가-힣_][A-Za-z0-9가-힣_.]*)!\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)
|(?P<cell>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)
|(?P<func>[A-Za-z가-힣_][A-Za-z0-9가-힣_.]*(?=\())
|(?P<name>[A-Za-z가-힣_][A-Za-z0-9가-힣_.]*)
|(?P<op><>|<=|>=|[-+*/^&<>=(),:])
|(?P<ws>\s+)
""", re.X)


def lex(s):
    out, i = [], 0
    while i < len(s):
        m = TOK.match(s, i)
        if not m:
            raise ValueError('토큰 실패 %r at %d' % (s, i))
        i = m.end()
        if m.lastgroup != 'ws':
            out.append((m.lastgroup, m.group()))
    return out


# ── 값 도우미 ─────────────────────────────────────────────────────
def flat(v):
    if isinstance(v, list):
        o = []
        for x in v:
            o.extend(flat(x))
        return o
    return [v]


def numify(x):
    if x is None or x == '':
        return 0.0
    if isinstance(x, bool):
        return 1.0 if x else 0.0
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, date):
        return float((x - EPOCH).days)
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def shape(v):
    return v if isinstance(v, list) else None


def binop(op, a, b):
    if isinstance(a, list) or isinstance(b, list):
        fa, fb = flat(a) if isinstance(a, list) else None, flat(b) if isinstance(b, list) else None
        n = len(fa) if fa else len(fb)
        out = []
        for i in range(n):
            out.append(binop(op, fa[i] if fa else a, fb[i] if fb else b))
        return out
    if op == '&':
        return _txt(a) + _txt(b)
    if op in ('=', '<>', '<', '>', '<=', '>='):
        if isinstance(a, str) or isinstance(b, str):
            x, y = _txt(a), _txt(b)
        else:
            x, y = numify(a), numify(b)
        return {'=': x == y, '<>': x != y, '<': x < y, '>': x > y,
                '<=': x <= y, '>=': x >= y}[op]
    x, y = numify(a), numify(b)
    if op == '+':
        return x + y
    if op == '-':
        return x - y
    if op == '*':
        return x * y
    if op == '/':
        return float('nan') if y == 0 else x / y
    if op == '^':
        return x ** y
    raise ValueError(op)


def _txt(v):
    if v is None:
        return ''
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, date):
        return str((v - EPOCH).days)
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def xround(x, n):
    n = int(n)
    return float(D(repr(float(x))).quantize(D(1).scaleb(-int(n)), rounding=ROUND_HALF_UP))


def xrounddown(x, n):
    return float(D(repr(float(x))).quantize(D(1).scaleb(-int(n)), rounding=ROUND_DOWN))


def _asdate(v):
    if isinstance(v, list):
        v = flat(v)[0]
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return EPOCH + timedelta(days=int(numify(v)))


def match_crit(val, crit):
    c = _txt(crit)
    m = re.match(r'^(<=|>=|<>|<|>|=)?(.*)$', c)
    op, rhs = m.group(1) or '=', m.group(2)
    try:
        r = float(rhs)
        v = numify(val)
    except ValueError:
        r, v = rhs, _txt(val)
    return {'=': v == r, '<>': v != r, '<': v < r, '>': v > r, '<=': v <= r, '>=': v >= r}[op]


# ── 계산기 ────────────────────────────────────────────────────────
class Book:
    def __init__(self, path):
        self.wb = openpyxl.load_workbook(path, data_only=False)
        self.cache, self.stack = {}, set()
        self.names = {}
        for k, dn in self.wb.defined_names.items():
            self.names[k] = dn.value if hasattr(dn, 'value') else dn.attr_text
        self.nform = 0

    # 셀 값
    def cell(self, sheet, coord):
        key = (sheet, coord)
        if key in self.cache:
            return self.cache[key]
        if key in self.stack:
            raise ValueError('순환 참조 %s!%s' % key)
        self.stack.add(key)
        prev = getattr(self, 'currow', None)
        try:
            v = self.wb[sheet][coord].value
            if isinstance(v, str) and v.startswith('='):
                self.nform += 1
                self.currow = self.wb[sheet][coord].row
                v = self.eval(v[1:], sheet)
            elif isinstance(v, datetime):
                v = v.date()
        finally:
            self.stack.discard(key)
            self.currow = prev
        self.cache[key] = v
        return v

    def rng(self, sheet, a1):
        ws = self.wb[sheet]
        cells = ws[a1.replace('$', '')]
        if not isinstance(cells, tuple):
            return self.cell(sheet, cells.coordinate)
        out = []
        for row in cells:
            if not isinstance(row, tuple):
                row = (row,)
            out.append([self.cell(sheet, c.coordinate) for c in row])
        if len(out) == 1:
            return out[0]
        if all(len(r) == 1 for r in out):
            return [r[0] for r in out]
        return out

    def resolve(self, ref, cur):
        if '!' in ref:
            sh, a1 = ref.split('!', 1)
            return self.rng(sh.strip("'"), a1)
        return self.rng(cur, ref)

    def eval(self, src, sheet):
        return Parser(self, lex(src), sheet).run()

    def call(self, fn, a):
        if fn == 'SUM':
            return sum(numify(x) for x in flat(a))
        if fn == 'SUMPRODUCT':
            cols = [flat(x) for x in a]
            n = max(len(c) for c in cols)
            s = 0.0
            for i in range(n):
                p = 1.0
                for c in cols:
                    p *= numify(c[i] if len(c) > 1 else c[0])
                s += p
            return s
        if fn in ('SUMIFS', 'COUNTIFS', 'SUMIF'):
            if fn == 'SUMIFS':
                tgt, rest = flat(a[0]), a[1:]
            elif fn == 'SUMIF':
                tgt, rest = flat(a[2]) if len(a) > 2 else flat(a[0]), [a[0], a[1]]
            else:
                tgt, rest = None, a
            pairs = [(flat(rest[i]), rest[i + 1]) for i in range(0, len(rest), 2)]
            n = len(pairs[0][0])
            s = 0.0
            for i in range(n):
                if all(match_crit(rg[i], cr) for rg, cr in pairs):
                    s += numify(tgt[i]) if tgt is not None else 1.0
            return s
        if fn == 'COUNT':
            return float(sum(1 for x in flat(a) if isinstance(x, (int, float, date))
                             and not isinstance(x, bool)))
        if fn == 'COUNTA':
            return float(sum(1 for x in flat(a) if x not in (None, '')))
        if fn == 'MIN':
            f = [numify(x) for x in flat(a) if x not in (None, '')]
            return min(f) if f else 0.0
        if fn == 'MAX':
            f = [numify(x) for x in flat(a) if x not in (None, '')]
            return max(f) if f else 0.0
        if fn == 'ROUND':
            return xround(numify(a[0]), numify(a[1]))
        if fn == 'ROUNDDOWN':
            return xrounddown(numify(a[0]), numify(a[1]))
        if fn == 'ABS':
            return abs(numify(a[0]))
        if fn == 'IF':
            c = a[0]
            c = c[0] if isinstance(c, list) else c
            t = bool(c) if isinstance(c, bool) else numify(c) != 0
            return a[1] if t else (a[2] if len(a) > 2 else False)
        if fn == 'IFERROR':
            v = a[0]
            return a[1] if (isinstance(v, float) and math.isnan(v)) else v
        if fn == 'AND':
            return all((x if isinstance(x, bool) else numify(x) != 0) for x in flat(a))
        if fn == 'OR':
            return any((x if isinstance(x, bool) else numify(x) != 0) for x in flat(a))
        if fn == 'INDEX':
            arr = a[0]
            if len(a) == 2:
                return flat(arr)[int(numify(a[1])) - 1]
            r, c = int(numify(a[1])), int(numify(a[2]))
            if arr and isinstance(arr[0], list):
                return arr[r - 1][c - 1]
            return flat(arr)[(r - 1) if len(a) < 3 or c == 1 else (c - 1)]
        if fn == 'TRANSPOSE':
            return a[0]
        if fn == 'MOD':
            x, y = numify(a[0]), numify(a[1])
            return float('nan') if y == 0 else x - y * math.floor(x / y)
        if fn == 'INT':
            return float(math.floor(numify(a[0])))
        if fn == 'ROW':
            return float(self.currow) if getattr(self, 'currow', None) else 0.0
        if fn == 'COUNTIF':
            rg, cr = flat(a[0]), a[1]
            return float(sum(1 for x in rg if match_crit(x, cr)))
        if fn == 'RANK':
            x = numify(a[0])
            rg = [numify(v) for v in flat(a[1]) if isinstance(v, (int, float))
                  and not isinstance(v, bool)]
            asc = len(a) > 2 and numify(a[2]) != 0
            return float(1 + sum(1 for v in rg if ((v < x) if asc else (v > x))))
        if fn == 'DATE':
            return date(int(numify(a[0])), int(numify(a[1])), int(numify(a[2])))
        if fn == 'YEAR':
            return float(_asdate(a[0]).year)
        if fn == 'MONTH':
            return float(_asdate(a[0]).month)
        if fn == 'DAY':
            return float(_asdate(a[0]).day)
        if fn == 'WEEKDAY':
            d = _asdate(a[0])
            t = int(numify(a[1])) if len(a) > 1 else 1
            if t == 3:
                return float(d.weekday())
            if t == 2:
                return float(d.weekday() + 1)
            return float((d.weekday() + 1) % 7 + 1)
        if fn == 'TEXT':
            v, f = numify(a[0]), _txt(a[1])
            nd = len(f.split('.')[1]) if '.' in f else 0
            return ('%%.%df' % nd) % v
        if fn == 'SUMPRODUCT_':
            return None
        raise ValueError('미구현 함수 %s' % fn)


class Parser:
    """수식 하나를 파싱·계산한다. 재진입 가능하도록 상태를 인스턴스에 둔다."""

    def __init__(self, bk, toks, sheet):
        self.bk, self.toks, self.sheet, self.i = bk, toks, sheet, 0

    def run(self):
        v = self._cmp()
        assert self.i == len(self.toks), ('잔여 토큰', self.toks, self.i)
        return v

    def _peek(self):
        return self.toks[self.i] if self.i < len(self.toks) else (None, None)

    def _eat(self, val=None):
        t, v = self._peek()
        if val is not None and v != val:
            raise ValueError('기대 %r 실제 %r' % (val, v))
        self.i += 1
        return v

    def _cmp(self):
        v = self._concat()
        while self._peek()[1] in ('=', '<>', '<', '>', '<=', '>='):
            op = self._eat()
            v = binop(op, v, self._concat())
        return v

    def _concat(self):
        v = self._add()
        while self._peek()[1] == '&':
            self._eat()
            v = binop('&', v, self._add())
        return v

    def _add(self):
        v = self._mul()
        while self._peek()[1] in ('+', '-'):
            op = self._eat()
            v = binop(op, v, self._mul())
        return v

    def _mul(self):
        v = self._un()
        while self._peek()[1] in ('*', '/'):
            op = self._eat()
            v = binop(op, v, self._un())
        return v

    def _un(self):
        if self._peek()[1] in ('-', '+'):
            op = self._eat()
            v = self._un()
            return binop('-', 0, v) if op == '-' else v
        return self._pow()

    def _pow(self):
        v = self._atom()
        while self._peek()[1] == '^':
            self._eat()
            v = binop('^', v, self._atom())
        return v

    def _args(self):
        self._eat('(')
        args = []
        if self._peek()[1] == ')':
            self._eat(')')
            return args
        while True:
            args.append(self._cmp())
            if self._peek()[1] == ',':
                self._eat()
                continue
            self._eat(')')
            return args

    def _atom(self):
        t, v = self._peek()
        if v == '(':
            self._eat('(')
            r = self._cmp()
            self._eat(')')
            return r
        if t == 'num':
            self._eat()
            return float(v)
        if t == 'str':
            self._eat()
            return v[1:-1].replace('""', '"')
        if t == 'func':
            self._eat()
            return self.bk.call(v.upper(), self._args())
        if t in ('ref', 'cell'):
            self._eat()
            return self.bk.resolve(v, self.sheet)
        if t == 'name':
            self._eat()
            if v.upper() in ('TRUE', 'FALSE'):
                return v.upper() == 'TRUE'
            tgt = self.bk.names.get(v)
            assert tgt, '이름 없음 %r' % v
            sh, a1 = tgt.split('!', 1)
            return self.bk.rng(sh.strip("'"), a1)
        raise ValueError('예상 못한 토큰 %r' % (v,))


# ── 검증 ──────────────────────────────────────────────────────────
def rowmap(ws, col=1):
    """라벨 -> 행 번호. 시트 구조가 바뀌어도 주소를 손으로 적지 않는다."""
    m = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and v and not v.startswith('='):
            m.setdefault(v, r)
    return m


def counts(wb):
    out, nf, nv = {}, 0, 0
    for ws in wb.worksheets:
        f = v = 0
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or c.value == '':
                    continue
                if isinstance(c.value, str) and c.value.startswith('='):
                    f += 1
                else:
                    v += 1
        nf += f
        nv += v
        out[ws.title] = {'행': ws.max_row, '열': ws.max_column, '수식셀': f, '값셀': v}
    return out, nf, nv


def evaluate_all(bk):
    err = []
    for ws in bk.wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    try:
                        bk.cell(ws.title, c.coordinate)
                    except Exception as e:
                        err.append('%s!%s  %s  → %s' % (ws.title, c.coordinate,
                                                        c.value[:80], e))
    return err


def facts():
    return json.load(open(os.path.join(BASE, 'ledger_facts.json'), encoding='utf-8'))


def main():
    bk = Book(XLSX)
    wb = bk.wb
    fx = facts()
    rep = {'파일': XLSX, '사본': os.path.expanduser(
        '~/Downloads/payhug_검산엑셀/검산_투자자어드민_20260901.xlsx')}
    rep['바이트 동일'] = (os.path.exists(rep['사본'])
                          and open(XLSX, 'rb').read() == open(rep['사본'], 'rb').read())
    rep['시트'], nf, nv = counts(wb)
    rep['수식셀 합계'], rep['값셀 합계'] = nf, nv
    rep['이름정의'] = len(bk.names)

    err = evaluate_all(bk)
    rep['평가 실패 수'] = len(err)
    rep['평가 실패'] = err[:12]

    # 값 셀이 허용된 자리에만 있는가 — 입력 가정값 · 채권 Di · 화면 값
    vcells = {}
    for ws in wb.worksheets:
        n = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None or v == '' or isinstance(v, str):
                    continue
                if isinstance(v, bool):
                    continue
                n += 1
        vcells[ws.title] = n
    rep['숫자 값셀'] = vcells

    # ── 화면대조 ────────────────────────────────────────────────
    ws = wb['화면대조']
    blocks = {'항등식': [], '불변식': [], '교차': [], '화면 정합': [], '가맹점': [],
              '모델 잔차': [], '화면 값 원본': []}
    for r in range(2, ws.max_row + 1):
        g = ws.cell(r, 1).value
        if g not in blocks:
            continue
        lab = ws.cell(r, 2).value
        if g == '화면 값 원본':
            blocks[g].append([lab, bk.cell('화면대조', 'E%d' % r)])
            continue
        row = [lab, bk.cell('화면대조', 'C%d' % r), bk.cell('화면대조', 'E%d' % r),
               bk.cell('화면대조', 'F%d' % r), bk.cell('화면대조', 'G%d' % r)]
        blocks[g].append(row)
    judged = [x for k in ('항등식', '불변식', '교차', '화면 정합', '가맹점')
              for x in blocks[k]]
    rep['화면대조 판정 행'] = len(judged)
    rep['화면대조 차이'] = sum(1 for x in judged if x[4] != '일치')
    rep['화면대조 차이 목록'] = [x for x in judged if x[4] != '일치']
    rep['모델 잔차 행'] = len(blocks['모델 잔차'])
    rep['모델 잔차'] = [[x[0], x[1], x[2], x[3]] for x in blocks['모델 잔차']]
    rep['화면 값 원본 행'] = len(blocks['화면 값 원본'])

    # 모델 잔차 행에 사유가 비어 있으면 안 된다
    nowhy = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == '모델 잔차' and not ws.cell(r, 8).value:
            nowhy.append(ws.cell(r, 2).value)
    rep['사유 빠진 잔차 행'] = nowhy

    # ── 화면 값이 ledger_facts.json 과 같은가 ────────────────────
    sv = dict(blocks['화면 값 원본'])
    chk = [('투자실행액', fx['exec']), ('순현금', fx['cash']), ('투자자산', fx['total']),
           ('W금융일수 raw', float(fx['wRaw'])), ('W금융일수 표기', float(fx['w'])),
           ('Ty수익율(%)', float(fx['ty'])), ('S입금부족율 raw(%)', float(fx['sRaw'])),
           ('S입금부족율 표기(%)', float(fx['s'])), ('할인율(%)', float(fx['rate'])),
           ('하루 평균 투자실행금', fx['dayAvg']), ('채권 건수', fx['receivables']),
           ('미회수 채권 건수', fx['openReceivables']),
           ('채권 Di 최소', fx['diRange'][0]), ('채권 Di 최대', fx['diRange'][1]),
           ('로스터 건수', len(fx['merchants'])),
           ('주간 PSA', fx['weekExec']), ('주간 PSM', fx['weekProfit']),
           ('주간 PSD raw', float(fx['weekWRaw'])), ('주간 PSC', fx['weekPsc']),
           ('주간 ④(%)', float(fx['weekTy'])), ('주간 ⑤(%)', float(fx['weekTyAsset'])),
           ('전 구간 PSA', fx['fullExec']), ('전 구간 PSM', fx['fullProfit']),
           ('전 구간 PSC', fx['fullPsc']), ('전 구간 ④(%)', float(fx['fullTy'])),
           ('전 구간 ⑤(%)', float(fx['fullTyAsset']))]
    bad = [[k, sv.get(k), v] for k, v in chk if abs(numify(sv.get(k)) - v) > 1e-9]
    rep['화면 값 ↔ ledger_facts 불일치'] = bad

    # ── 가중치 대조 ─────────────────────────────────────────────
    ws = wb['가중치 대조']
    wm = rowmap(ws)
    g = lambda a: bk.cell('가중치 대조', a)
    W = {}
    for lab, key in [('W금융일수', 'W'), ('W금융일수 표기', 'W표기'), ('Ty수익율(%)', 'Ty'),
                     ('미회수 잔량 W', '미회수W'), ('미회수 잔량 Ty(%)', '미회수Ty')]:
        r = wm[lab]
        W[key] = {'(가) 금액 실측': g('C%d' % r), '(나) 엑셀 MAU': g('D%d' % r),
                  '(다) 로스터 금액가중': g('E%d' % r), '(가)-(나)': g('F%d' % r),
                  '(다)-(가)': g('E%d' % r) - g('C%d' % r)}
    rep['가중치 대조'] = W
    wm2 = rowmap(ws, 2)
    rep['가중치 판정'] = ws.cell(wm['판정'], 3).value
    rep['워드 인용'] = [ws.cell(wm2[k], 3).value for k in
                        ('w금융일수 산식', 'Ai 정의', 'Di 정의')]
    rep['확인 문항'] = [ws.cell(wm[k], 2).value for k in ('확인 문항 1', '확인 문항 2')]
    v0 = wm['배달앱/전체']
    sens = []
    for r in range(v0 + 1, v0 + 6):
        sens.append([g('A%d' % r), g('F%d' % r), g('G%d' % r), g('H%d' % r),
                     ws.cell(r, 9).value])
    rep['0.35 민감도'] = sens

    # ── 기간집계 · 플랫폼 주요값 ────────────────────────────────
    pm = rowmap(wb['기간집계'])
    rep['기간집계'] = {k: bk.cell('기간집계', 'B%d' % r) for k, r in pm.items()
                       if k not in ('항목',)}

    #   S입금부족율 — 불변식은 ROUND(...,2) 뒤 0.07 = 0.07 만 본다. 반올림이 차를 가린다.
    #   raw 로 다시 재고, 그 차가 `모델 잔차` 표에 등재돼 있는지까지 본다.
    s_row = pm['S입금부족율(%)']
    s_x = bk.cell('기간집계', 'B%d' % s_row)
    s_s = float(fx['sRaw'])
    rep['S입금부족율 raw'] = {'엑셀 기간집계 B%d' % s_row: s_x, 'ledger_facts.sRaw': s_s,
                              '차': s_x - s_s,
                              '잔차율(%)': (s_x - s_s) / s_s * 100 if s_s else None}
    rep['S입금부족율 raw 잔차 등재'] = [x[0] for x in blocks['모델 잔차']
                                       if 'S입금부족율 raw' in (x[0] or '')]
    rep['S입금부족율 표기 비교 잔존'] = [x[0] for k in ('불변식',) for x in blocks[k]
                                        if x[0] == 'S입금부족율(%)']

    #   ⑤ 는 대표가 수식을 새로 써서 다시 주기로 했다. 값은 두되 표식이 살아 있어야 한다.
    r5 = pm['⑤ 투자자산 대비 ty수익율 (정의)']
    r6 = pm['⑥ 투자실행금액 대비 ty수익율 (정의)']
    rep['⑤ 미확정 표식'] = {'행': r5, '화면 반영': wb['기간집계'].cell(r5, 5).value,
                             '출처': wb['기간집계'].cell(r5, 4).value,
                             '값': bk.cell('기간집계', 'B%d' % r5)}
    rep['⑥ 미확정 표식'] = {'행': r6, '화면 반영': wb['기간집계'].cell(r6, 5).value,
                             '출처': wb['기간집계'].cell(r6, 4).value,
                             '값': bk.cell('기간집계', 'B%d' % r6)}

    #   대표 DM 2026-08-31 16:45 — 근사식과 원식을 나란히. 어느 쪽이 정본인지는 판정하지 않는다.
    dm45 = {k: bk.cell('기간집계', 'B%d' % pm[k]) for k in
            ('DM 16:45 (가) 대표 근사식(%)', 'DM 16:45 (나) 원식 PSMR(%)',
             'DM 16:45 (나) - (가) (%p)', 'DM 16:45 (나) / (가) (배)',
             'DM 16:45 1 / (1 - 할인율) (배)', 'DM 16:45 차감합',
             'DM 16:45 차감합 0 일 때 (가) (%)', 'DM 16:45 차감합 0 일 때 (나) (%)')}
    rate_pct = bk.cell('입력', 'C%d' % rowmap(wb['입력'], 2)['할인율(%)'])
    dm45['차감합 0 일 때 두 값의 차(%p)'] = (dm45['DM 16:45 차감합 0 일 때 (나) (%)']
                                            - dm45['DM 16:45 차감합 0 일 때 (가) (%)'])
    rep['대표 DM 16:45 항등식'] = dm45
    rep['대표 DM 16:45 판정'] = {
        '두 값이 셀로 나란히 있다': all(isinstance(dm45[k], float) for k in
                                        ('DM 16:45 (가) 대표 근사식(%)',
                                         'DM 16:45 (나) 원식 PSMR(%)')),
        '차 셀이 두 값의 차와 같다': abs(dm45['DM 16:45 (나) - (가) (%p)']
                                        - (dm45['DM 16:45 (나) 원식 PSMR(%)']
                                           - dm45['DM 16:45 (가) 대표 근사식(%)'])) < 1e-12,
        '차감합 0 일 때 (가) = 할인율': abs(dm45['DM 16:45 차감합 0 일 때 (가) (%)']
                                           - rate_pct) < 1e-12,
        '차감합 0 일 때도 두 값이 갈린다 (근사식이다)':
            abs(dm45['차감합 0 일 때 두 값의 차(%p)']) > 1e-9,
        '갈림 배수가 1/(1-할인율)': abs(
            dm45['DM 16:45 1 / (1 - 할인율) (배)']
            - dm45['DM 16:45 차감합 0 일 때 (나) (%)']
            / dm45['DM 16:45 차감합 0 일 때 (가) (%)']) < 1e-12,
        '판정 낱말을 넣지 않았다': not any(
            isinstance(v, str) and ('성립' in v) for v in dm45.values()),
    }

    #   대표 DM 2026-08-31 16:27 — ty수익률 연 환산 · 일 환산 두 갈래.
    am = pm
    ty2 = {k: bk.cell('기간집계', 'B%d' % am[k]) for k in
           ('1년 회전수 (365 / 표기 W)', 'Ty수익율 — 연 환산(%)', 'Ty수익율 — 일 환산(%)',
            '검산 — 일 환산 x 365 = 연 환산')}
    rep['대표 DM 16:27 연·일 환산'] = ty2
    rep['대표 DM 16:27 판정'] = {
        '연 환산 = 화면 Ty': abs(ty2['Ty수익율 — 연 환산(%)']
                                 - bk.cell('기간집계', 'B%d' % am['Ty수익율(%)'])) < 1e-12,
        '일 환산 자리 있음': isinstance(ty2['Ty수익율 — 일 환산(%)'], float),
        '일 환산 x 365 = 연 환산': abs(ty2['검산 — 일 환산 x 365 = 연 환산']) < 1e-9,
        '1년 회전수 = 365 / 표기 W': abs(
            ty2['1년 회전수 (365 / 표기 W)']
            - 365 / bk.cell('기간집계',
                            'B%d' % am['W금융일수 표기 (스위치 ② 적용)'])) < 1e-9,
    }

    #   대표 DM 인용 — 「읽는 법」 시트에 옮긴 원문이 DM 원본의 부분문자열인가.
    #   대상이 0건이면 FAIL 이다.
    dmraw = open(os.path.join(BASE, 'dm_0831', 'dm_20260831_raw.md'),
                 encoding='utf-8').read()
    dmlines = [l.replace('*', '').strip() for l in dmraw.splitlines()]
    gw = wb['읽는 법']
    quotes = []
    for r in range(2, gw.max_row + 1):
        t = str(gw.cell(r, 2).value or '')
        if not re.match(r'^2026-\d\d-\d\d \d\d:\d\d$', t):
            continue
        q = _runs(gw.cell(r, 3).value)[0] if gw.cell(r, 3).value is not None else ''
        quotes.append([t, q[:70], bool(q) and any(q in l for l in dmlines)])
    rep['대표 DM 인용'] = quotes
    rep['대표 DM 인용 판정'] = {'인용 0건 아님': len(quotes) > 0,
                                '전건이 DM 원문 부분문자열': bool(quotes)
                                and all(x[2] for x in quotes)}

    #   산식 — 대표 정의서 원문 줄마다 대응 셀(또는 왜 비어 있는지)이 적혀 있는가.
    fs = wb['산식']
    rep['산식 원문 줄'] = sum(1 for r in range(2, fs.max_row + 1) if fs.cell(r, 2).value)
    rep['산식 대응 셀 빈 줄'] = [fs.cell(r, 2).value for r in range(2, fs.max_row + 1)
                                if fs.cell(r, 2).value and not fs.cell(r, 3).value]
    rep['플랫폼'] = {wb['플랫폼'].cell(r, 1).value: bk.cell('플랫폼', 'B%d' % r)
                     for r in range(16, 23) if wb['플랫폼'].cell(r, 1).value}

    # ── 항등식 ──────────────────────────────────────────────────
    ws = wb['화면대조']
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == '항등식':
            rep['항등식'] = {'좌변': bk.cell('화면대조', 'C%d' % r),
                             '우변': bk.cell('화면대조', 'E%d' % r),
                             '차': bk.cell('화면대조', 'F%d' % r)}
            break

    # ── 프리셋 · 비중 최대잉여법 · 계약 상태 ─────────────────────
    ws = wb['입력']
    im = rowmap(ws)
    pr = im['프리셋']
    rep['프리셋'] = [[ws.cell(r, 3).value, bk.cell('입력', 'D%d' % r),
                      bk.cell('입력', 'E%d' % r), bk.cell('입력', 'F%d' % r),
                      bk.cell('입력', 'I%d' % r)] for r in range(pr + 1, pr + 7)]
    rep['프리셋 차 합계'] = sum(x[4] for x in rep['프리셋'])
    gm = wb['가맹점']
    n = 0
    while gm.cell(2 + n, 3).value and bk.cell('가맹점', 'C%d' % (2 + n)) != '합계':
        n += 1
    rep['비중 최대잉여법'] = [[bk.cell('가맹점', 'C%d' % (2 + i)),
                               bk.cell('가맹점', 'S%d' % (2 + i)) * 100,
                               bk.cell('가맹점', 'Z%d' % (2 + i)),
                               bk.cell('가맹점', 'AB%d' % (2 + i)),
                               bk.cell('가맹점', 'AC%d' % (2 + i))] for i in range(n)]
    rep['비중 합'] = sum(x[4] for x in rep['비중 최대잉여법'])
    rep['비중 잔차 최대(pp)'] = max(abs(x[4] - x[1]) for x in rep['비중 최대잉여법'])
    rep['서명 대기'] = [bk.cell('가맹점', 'C%d' % (2 + i)) for i in range(n)
                        if bk.cell('가맹점', 'AD%d' % (2 + i)) == '서명 대기']

    print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


def incells(wb):
    """입력 시트의 셀 주소를 라벨에서 뽑는다 — 행이 움직여도 검사기를 손대지 않는다.

    대상을 못 찾으면 KeyError 로 죽는다. 조용히 건너뛰어 통과하는 자리를 두지 않는다.
    """
    ws = wb['입력']
    m1, m2 = rowmap(ws, 1), rowmap(ws, 2)
    return {'①': 'B%d' % m1['① W금융일수 모집단'], '②': 'B%d' % m1['② W 표기 자릿수'],
            '③': 'B%d' % m1['③ 가맹점 수'], '④': 'B%d' % m1['④ 산출 방향'],
            '총투자자산': 'C%d' % m2['총 투자자산'], '유휴비율': 'C%d' % m2['유휴 비율'],
            '투자실행액목표': 'C%d' % m2['투자실행액 목표'], '순현금': 'C%d' % m2['순현금'],
            '구성비0': m1['코드'] + 1, '배달앱비중': 'B%d' % m1['배달앱/전체']}


def switch_test(quiet=False):
    """스위치 4개를 바꿔 넣고 값이 실제로 움직이는지 본다."""
    base = Book(XLSX)
    pm = rowmap(base.wb['기간집계'])
    gm = rowmap(base.wb['가맹점'])
    IC = incells(base.wb)
    pick = [('W raw', '기간집계', 'B%d' % pm['W금융일수 raw (스위치 ① 적용)']),
            ('W 표기', '기간집계', 'B%d' % pm['W금융일수 표기 (스위치 ② 적용)']),
            ('Ty(%)', '기간집계', 'B%d' % pm['Ty수익율(%)']),
            ('투자실행액', '기간집계', 'B%d' % pm['투자실행액']),
            ('투자자산', '기간집계', 'B%d' % pm['투자자산']),
            ('S(%)', '기간집계', 'B%d' % pm['S입금부족율(%)']),
            ('가맹점수', '가맹점', 'B%d' % gm['가맹점 수 (적용)']),
            ('하루선정산액합계', '가맹점', 'B%d' % gm['하루 선정산액 합계 (적용)'])]
    cases = [('기본', {}),
             ('① 만기 도래분만', {IC['①']: '만기 도래분만'}),
             ('① 미회수 잔량만', {IC['①']: '미회수 잔량만'}),
             ('② 표기 1자리', {IC['②']: 1}),
             ('③ 가맹점 4곳', {IC['③']: 4}),
             ('③ 가맹점 1곳', {IC['③']: 1}),
             ('④ 방향 B', {IC['④']: 'B'}),
             ('④ 방향 B + 4곳', {IC['④']: 'B', IC['③']: 4})]
    out = []
    for nm, ov in cases:
        bk = Book(XLSX)
        for k, v in ov.items():
            bk.wb['입력'][k] = v
        row = {'케이스': nm}
        for lab, sh, ad in pick:
            row[lab] = bk.cell(sh, ad)
        # 항등식은 어느 스위치에서도 성립해야 한다
        ws = bk.wb['화면대조']
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == '항등식':
                row['항등식 차'] = bk.cell('화면대조', 'F%d' % r)
                break
        out.append(row)
    moved = {}
    b = out[0]
    for lab in ('W raw', 'W 표기', '가맹점수', '하루선정산액합계'):
        moved[lab] = sorted({round(x[lab], 8) if isinstance(x[lab], float) else x[lab]
                             for x in out})
    rep = {'케이스': out, '스위치가 실제로 움직인 값': moved,
           '항등식 차 (기본 조합)': out[0]['항등식 차'],
           '항등식 차 (스위치 ①②)': [out[i]['항등식 차'] for i in (1, 2, 3)],
           '항등식 차 (스위치 ③④)': [out[i]['항등식 차'] for i in (4, 5, 6, 7)],
           '항등식 주석': '채권 시트의 미회수 건수는 기본 조합에서 푼 정수다. '
                          '스위치 ③④로 가맹점 구성이 바뀌면 좌변이 그만큼 어긋난다.'}
    ok = {'① 모집단': len({round(out[i]['W raw'], 8) for i in (0, 1, 2)}) == 3,
          '② 자릿수': out[0]['W 표기'] != out[3]['W 표기'],
          '③ 가맹점 수': len({out[i]['가맹점수'] for i in (0, 4, 5)}) == 3,
          '④ 방향': abs(out[0]['하루선정산액합계'] - out[6]['하루선정산액합계']) > 1,
          '항등식 기본 0': abs(out[0]['항등식 차']) < 0.5}
    rep['스위치 동작'] = ok
    rep['스위치 4개 모두 동작'] = all(ok.values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


def sens_test(quiet=False):
    """배달앱/전체 입력을 0.30 · 0.40 으로 바꿔 (나) W·Ty 가 움직이는지 본다."""
    out = []
    for v in (0.30, 0.35, 0.40):
        bk = Book(XLSX)
        bk.wb['입력'][incells(bk.wb)['배달앱비중']] = v
        wm = rowmap(bk.wb['가중치 대조'])
        out.append({'배달앱/전체': v,
                    '(나) W': bk.cell('가중치 대조', 'D%d' % wm['W금융일수']),
                    '(나) Ty(%)': bk.cell('가중치 대조', 'D%d' % wm['Ty수익율(%)'])})
    rep = {'민감도': out, '움직임': len({round(x['(나) W'], 8) for x in out}) == 3}
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


def mix_test(quiet=False):
    """구성비 4칸을 MAU 값으로 바꿔 넣으면 (가) 가 (나) 와 같아지는가."""
    bk = Book(XLSX)
    wm = rowmap(bk.wb['가중치 대조'])
    mau = [bk.cell('가중치 대조', 'D%d' % (wm['카드사'] + i)) for i in range(4)]
    bk2 = Book(XLSX)
    c0 = incells(bk2.wb)['구성비0']
    for i, v in enumerate(mau):
        bk2.wb['입력']['C%d' % (c0 + i)] = v
    rep = {'MAU 로 바꾼 뒤 (가) W': bk2.cell('가중치 대조', 'C%d' % wm['W금융일수']),
           '(나) W': bk2.cell('가중치 대조', 'D%d' % wm['W금융일수'])}
    rep['일치'] = abs(rep['MAU 로 바꾼 뒤 (가) W'] - rep['(나) W']) < 1e-9
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


IDLE_PCT = (20, 25, 30)
VARIANT = [(pct, os.path.join(VDIR, '검산_투자자어드민_20260901_유휴%d.xlsx' % pct),
            os.path.expanduser('~/Downloads/payhug_검산엑셀/'
                               '검산_투자자어드민_20260901_유휴%d.xlsx' % pct))
           for pct in IDLE_PCT]


def idle_test(quiet=False):
    """총 투자자산 · 유휴 비율 두 칸에서 투자실행액 목표 · 순현금이 갈라지는가.

    두 칸을 손으로 갈라 넣던 자리다. 이제 한 칸(유휴 비율)만 돌려도 나머지가 따라와야 한다.
    """
    base = Book(XLSX)
    IC = incells(base.wb)
    inp = base.wb['입력']
    raw = {k: inp[IC[k]].value for k in ('총투자자산', '유휴비율', '투자실행액목표', '순현금')}
    isf = lambda k: isinstance(raw[k], str) and raw[k].startswith('=')
    rows = []
    for pct in IDLE_PCT:
        bk = Book(XLSX)
        bk.wb['입력'][IC['유휴비율']] = pct / 100
        t = bk.cell('입력', IC['총투자자산'])
        e = bk.cell('입력', IC['투자실행액목표'])
        c = bk.cell('입력', IC['순현금'])
        rows.append({'유휴(%)': pct, '총 투자자산': t, '투자실행액 목표': e, '순현금': c,
                     '합 - 총 투자자산': e + c - t,
                     '목표 - 총자산x(1-유휴)': e - t * (1 - pct / 100),
                     '순현금 - 총자산x유휴': c - t * pct / 100})
    rep = {'입력 칸': IC, '두 칸 원문': raw, '유휴 비율 세 벌': rows}
    rep['판정'] = {
        '총 투자자산이 입력 값 셀': not isf('총투자자산'),
        '유휴 비율이 입력 값 셀': not isf('유휴비율'),
        '투자실행액 목표가 수식': isf('투자실행액목표'),
        '순현금이 수식': isf('순현금'),
        '세 벌 합 = 총 투자자산': all(abs(x['합 - 총 투자자산']) < 0.5 for x in rows),
        '세 벌 목표 = 총자산 x (1-유휴)': all(abs(x['목표 - 총자산x(1-유휴)']) < 0.5
                                             for x in rows),
        '세 벌 순현금 = 총자산 x 유휴': all(abs(x['순현금 - 총자산x유휴']) < 0.5
                                           for x in rows),
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


def variant_test(quiet=False):
    """세 벌(유휴 20·25·30%)이 실물로 있고, 채권 원장이 그 목표를 원 단위까지 맞추는가.

    유휴 비율만 손으로 바꾸면 금융일수 Di 가 그대로라 투자실행액이 75,000,001 · 70,000,013
    으로 어긋난다. 세 벌은 그 비율로 다시 푼 원장이라 차가 0 이어야 한다.
    대상 파일이 하나라도 없으면 FAIL — 못 찾아서 통과하는 자리를 두지 않는다.
    """
    rep = {'세 벌': [], '없는 파일': [], '사본 다름': []}
    for pct, src, dst in VARIANT:
        if not (os.path.exists(src) and os.path.exists(dst)):
            rep['없는 파일'].append(os.path.basename(src))
            continue
        if open(src, 'rb').read() != open(dst, 'rb').read():
            rep['사본 다름'].append(os.path.basename(src))
        bk = Book(src)
        IC = incells(bk.wb)
        pm = rowmap(bk.wb['기간집계'])
        gm = rowmap(bk.wb['가맹점'])
        t = bk.cell('입력', IC['총투자자산'])
        want_e = bk.cell('입력', IC['투자실행액목표'])
        want_c = bk.cell('입력', IC['순현금'])
        got_e = bk.cell('기간집계', 'B%d' % pm['투자실행액'])
        got_a = bk.cell('기간집계', 'B%d' % pm['투자자산'])
        ws = bk.wb['화면대조']
        judge = {}
        for r in range(2, ws.max_row + 1):
            lab = ws.cell(r, 2).value
            if ws.cell(r, 1).value == '불변식' and lab in ('투자실행액 = 입력 목표',
                                                           '투자자산 = 총 투자자산'):
                judge[lab] = bk.cell('화면대조', 'G%d' % r)
        rep['세 벌'].append({
            '유휴(%)': pct, '파일': os.path.basename(src),
            '총 투자자산': t, '입력 투자실행액 목표': want_e, '입력 순현금': want_c,
            '기간집계 투자실행액': got_e, '기간집계 투자자산': got_a,
            '목표 대비 차(원)': got_e - want_e, '총자산 대비 차(원)': got_a - t,
            '유휴 비율 실측(%)': bk.cell('기간집계', 'B%d' % pm['순현금 비중']) * 100,
            'W raw': bk.cell('기간집계', 'B%d' % pm['W금융일수 raw (스위치 ① 적용)']),
            'W 표기': bk.cell('기간집계', 'B%d' % pm['W금융일수 표기 (스위치 ② 적용)']),
            'Ty(%)': bk.cell('기간집계', 'B%d' % pm['Ty수익율(%)']),
            'S raw(%)': bk.cell('기간집계', 'B%d' % pm['S입금부족율(%)']),
            'PSA': bk.cell('기간집계', 'B%d' % pm['PSA 투자실행금']),
            'PSM': bk.cell('기간집계', 'B%d' % pm['PSM 투자수익 (정의 · 차감 반영)']),
            '② 상환액': bk.cell('기간집계', 'B%d' % pm['② 상환액']),
            'PSD': bk.cell('기간집계', 'B%d' % pm['PSD']),
            '④(%)': bk.cell('기간집계',
                            'B%d' % pm['④ 투자실행금액 대비 ty수익율 (정의)']),
            '⑤(%)': bk.cell('기간집계',
                            'B%d' % pm['⑤ 투자자산 대비 ty수익율 (정의)']),
            '⑤ 화면 반영 표식': bk.wb['기간집계'].cell(
                pm['⑤ 투자자산 대비 ty수익율 (정의)'], 5).value,
            #   기본 벌에 넣은 구멍 셋(평균순현금 칸 · EC 표식 · 갈림 표)이 이 벌에도 있는가.
            '평균순현금 행': pm.get('평균순현금'),
            '평균순현금 값': (bk.wb['기간집계'].cell(pm['평균순현금'], 2).value
                              if '평균순현금' in pm else '칸 없음'),
            '평균순현금 표식': (bk.wb['기간집계'].cell(pm['평균순현금'], 5).value
                                if '평균순현금' in pm else None),
            'EC 열머리': ec_head(bk.wb)[2],
            '갈림 표 머리': bk.wb['읽는 법'].cell(1, 2).value,
            '불변식 판정': judge,
            '가맹점 수': bk.cell('가맹점', 'B%d' % gm['가맹점 수 (적용)']),
        })
    v = rep['세 벌']
    rep['판정'] = {
        '세 벌 전건 있음': len(v) == len(VARIANT) and not rep['없는 파일'],
        '사본 바이트 동일': not rep['사본 다름'],
        '투자실행액이 목표와 원 단위까지 같다': bool(v) and all(
            abs(x['목표 대비 차(원)']) < 0.5 for x in v),
        '투자자산이 총 투자자산과 같다': bool(v) and all(
            abs(x['총자산 대비 차(원)']) < 0.5 for x in v),
        '유휴 비율이 지정값과 같다': bool(v) and all(
            abs(x['유휴 비율 실측(%)'] - x['유휴(%)']) < 0.05 for x in v),
        '불변식 두 줄 전건 일치': bool(v) and all(
            len(x['불변식 판정']) == 2 and set(x['불변식 판정'].values()) == {'일치'}
            for x in v),
        '⑤ 가 미확정 표식': bool(v) and all(x['⑤ 화면 반영 표식'] == '미확정' for x in v),
        '세 벌에 평균순현금 칸이 있다': bool(v) and all(x['평균순현금 행'] for x in v),
        '세 벌 평균순현금이 값 없이 미확정 표식': bool(v) and all(
            x['평균순현금 값'] is None and x['평균순현금 표식'] == '미확정' for x in v),
        '세 벌 EC 열머리에 미확정 표식': bool(v) and all(
            '미확정' in (x['EC 열머리'] or '') for x in v),
        '세 벌 첫 시트 맨 위가 갈림 표': bool(v) and all(
            x['갈림 표 머리'] == '화면 ↔ 엑셀 갈림' for x in v),
        '세 벌이 서로 다른 벌이다': len({x['입력 순현금'] for x in v}) == len(VARIANT),
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


def five_test(quiet=False):
    """⑤ 가 `입력` 시트 한 칸에서만 정의되고, ⑥ 이 그 ⑤ 셀을 실제로 참조하는가.

    산식 칸을 갈아 끼우면 ⑤ 도 ⑥ 도 엑셀 안에서 따라 움직여야 한다.
    라벨을 못 찾으면 KeyError 로 죽는다. 조용히 건너뛰어 통과하는 자리를 두지 않는다.
    """
    base = Book(XLSX)
    wb = base.wb
    im = rowmap(wb['입력'], 2)
    pm = rowmap(wb['기간집계'])
    F5 = 'C%d' % im['⑤ 산식 (미확정 · 대표 재작성 대기)']
    N3 = 'C%d' % im['③ 지시 대상 (미확정)']
    R5, R6 = (pm['⑤ 투자자산 대비 ty수익율 (정의)'],
              pm['⑥ 투자실행금액 대비 ty수익율 (정의)'])
    R4, PSA, PSC = (pm['④ 투자실행금액 대비 ty수익율 (정의)'],
                    pm['PSA 투자실행금'], pm['PSC 순현금 합'])
    f5src = str(wb['입력'][F5].value or '')
    f5v = str(wb['기간집계'].cell(R5, 2).value or '')
    f6v = str(wb['기간집계'].cell(R6, 2).value or '')

    #   ⑤ 를 계산하는 자리 — PSA 와 PSC 를 함께 문 수식 셀. 하나뿐이어야 한다.
    hard = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if not (isinstance(f, str) and f.startswith('=')):
                    continue
                if ('기간집계!B%d' % PSA) in f and ('기간집계!B%d' % PSC) in f:
                    hard.append('%s!%s' % (ws.title, c.coordinate))
    #   산식 설명 문자열로 남은 ⑤ 계산 — `입력` 시트 ⑤ 칸 밖에 있으면 안 된다.
    #   대표 정의서 원문 열(QUOTE_CELLS)은 인용이라 뺀다.
    shape = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if is_quote(ws.title, c.column, c.row):
                    continue
                t, _ = _runs(c.value) if c.value is not None else ('', [])
                if t.startswith('='):
                    continue
                if 'PSA + PSC' in t or 'PSA+PSC' in t:
                    shape.append(['%s!%s' % (ws.title, c.coordinate), t[:70]])

    #   ⑤ 칸을 갈아 끼우면 ⑤ 셀과 ⑥ 셀이 따라 움직이는가.
    def probe(f5val, n3val):
        bk = Book(XLSX)
        bk.wb['입력'][F5] = f5val
        bk.wb['입력'][N3] = n3val
        return bk.cell('기간집계', 'B%d' % R5), bk.cell('기간집계', 'B%d' % R6)
    a5, a6 = probe(2.0, 100.0)
    b5, b6 = probe(4.0, 100.0)
    t4 = base.cell('기간집계', 'B%d' % R4)
    rep = {
        '⑤ 산식 칸': '입력!%s' % F5, '③ 칸': '입력!%s' % N3,
        '⑤ 산식 칸 수식': f5src,
        '기간집계 ⑤ 수식': f5v, '기간집계 ⑥ 수식': f6v,
        '③ 칸 값 (비어 있어야 한다)': wb['입력'][N3].value,
        '기본 벌 ⑤ 값': base.cell('기간집계', 'B%d' % R5),
        '기본 벌 ⑥ 값': base.cell('기간집계', 'B%d' % R6),
        'PSA·PSC 를 함께 문 수식 셀': hard,
        '「PSA + PSC」 문자열이 남은 셀': shape,
        '⑤=2 일 때 (⑤셀, ⑥셀)': [a5, a6], '⑤=4 일 때 (⑤셀, ⑥셀)': [b5, b6],
        '④ 값': t4,
    }
    rep['판정'] = {
        '⑤ 산식 칸이 수식이다': f5src.startswith('='),
        '기간집계 ⑤ 가 그 칸만 참조한다': f5v.replace(' ', '') == '=입력!%s' % F5,
        '⑥ 수식이 ⑤ 셀 주소를 담고 있다': ('B%d' % R5) in f6v,
        '⑥ 수식이 ③ 칸 주소를 담고 있다': ('입력!%s' % N3) in f6v,
        '⑤ 를 계산하는 수식 셀이 입력 칸 하나뿐': hard == ['입력!%s' % F5],
        '⑤ 산식 문자열이 다른 셀에 남지 않았다': not shape,
        '③ 칸이 비어 있다': wb['입력'][N3].value is None,
        '③ 이 비면 ⑥ 은 미확정': base.cell('기간집계', 'B%d' % R6) == '미확정',
        '⑤ 칸을 바꾸면 ⑤ 셀이 따라온다': (a5, b5) == (2.0, 4.0),
        '⑤ 칸을 바꾸면 ⑥ 셀이 따라온다': (isinstance(a6, float) and isinstance(b6, float)
                                          and abs(a6 - b6) > 1e-12),
        '⑥ 이 (④/③) x 365 / ⑤ 그대로': (isinstance(a6, float)
                                          and abs(a6 - t4 / 100.0 * 365 / 2.0) < 1e-9),
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


# ── 기호 표기 ─────────────────────────────────────────────────────
#   기준표를 여기 두지 않는다. `dm_0831/symbol_rule_0831.md` 갈아끼움표에서 읽어 온다.
#   대문자 `D` 는 금융일수, 소문자 `d` 는 오늘 날짜다. `d-1` 은 날짜가 아니라
#   「정산예정일이 어제인 대상정산금채권 집합」이다.
#   생성기는 글자열을 바꾸지 않고 서식만 갈랐다. 평문으로 읽으면 이 고침이 안 보여
#   rich_text=True 로 다시 읽어 글자별 아래첨자 여부로 판정한다.
#   대상이 0건이면 FAIL 이다 — 못 찾아서 통과하는 구조를 두지 않는다.
RULE_MD = os.path.join(BASE, 'dm_0831', 'symbol_rule_0831.md')
#   대표 정의서 원문을 그대로 옮긴 자리. 여기만 옛 표기(`D-1`)를 그대로 둔다.
#   (시트, 열, 시작 행). 머리 행은 인용이 아니라 뺀다.
QUOTE_CELLS = [('산식', 2, 2)]


def is_quote(sheet, col, row):
    return any(sheet == a and col == b and row >= c for a, b, c in QUOTE_CELLS)


def symbol_rule(path=RULE_MD):
    """갈아끼움표를 읽어 (새 표기 쌍, 금지된 옛 표기)를 만든다.

    파일이 없거나 표를 못 읽으면 예외로 죽는다. 조용히 건너뛰어 통과하는 자리를 두지 않는다.
    """
    txt = open(path, encoding='utf-8').read()
    seg = txt.split('## 갈아 끼우는 표', 1)[1].split('\n## ', 1)[0]
    tick = re.compile(r'`([^`]+)`')
    pairs, forbid, bare = [], [], False
    for line in seg.splitlines():
        line = line.strip()
        if not line.startswith('|') or set(line) <= set('|- '):
            continue
        col = [c.strip() for c in line.strip('|').split('|')]
        if len(col) < 2:
            continue
        a, b = tick.findall(col[0]), tick.findall(col[1])
        if not a or not b:
            continue                              # 「그대로」 처럼 새 표기가 없는 줄
        o, n = a[0], b[0]
        if '그대로' in col[1] and '_' not in o:
            continue                              # 「PSD · PSA … 그대로」 처럼 안 바뀌는 줄
        if o == 'D' and n == 'd':
            bare = True                           # 홀로 선 `D` 를 오늘 날짜로 쓰지 않는다
            continue
        if '_' in n:                              # `SB_{d-1}` · `B_{d-1,i}` · `D_i`
            h, t = n.split('_', 1)
            t = t.strip('{}').replace(',', '')
            if (h + t) not in [x[0] for x in pairs]:
                pairs.append((h + t, h, t))
        if '_' in o:
            ho, to = o.split('_', 1)
            po = ho + to.strip('{}').replace(',', '')
            if po != (n.split('_', 1)[0] + n.split('_', 1)[1].strip('{}').replace(',', '')):
                forbid.append(po)
        else:
            for tok in re.split(r'\s*~\s*', o):  # `D-20 ~ D-11`
                tok = tok.strip()
                if tok and tok != n:
                    forbid.append(tok)
    if not pairs:
        raise ValueError('갈아끼움표에서 기호 쌍을 못 읽었다: %s' % path)
    if not forbid:
        raise ValueError('갈아끼움표에서 옛 표기를 못 읽었다: %s' % path)
    return {'pairs': pairs, 'forbid': sorted(set(forbid), key=len, reverse=True),
            'bare_D': bare, 'src': os.path.relpath(path, BASE)}


RULE = symbol_rule()
NOTMAP = {k: (h, t) for k, h, t in RULE['pairs']}
NOTRE = re.compile(r'(?<![A-Za-z0-9])(%s)(?![A-Za-z0-9])'
                   % '|'.join(re.escape(k) for k, _, _ in RULE['pairs']))
#   규칙 문서가 손대지 않은 꼬리(`i` · `pi`)는 워크북에서 직접 찾는다.
#   기준표를 두 벌 두지 않으려고 낱말 목록 대신 꼬리 규칙 하나로 잡는다.
TAILRE = re.compile(r'(?<![A-Za-z0-9])(S?[A-Z]{1,3})(pi|i)(?![A-Za-z0-9])')
OLDRE = re.compile(r'(?<![A-Za-z0-9])(%s)(?![A-Za-z0-9])'
                   % '|'.join(re.escape(x) for x in RULE['forbid']))
BAREDRE = re.compile(r'(?<![A-Za-z0-9])D(?=\s|$)')
#   「만기」를 남겨 둔 자리 — 우리 분석 조어라 원문에 대응 낱말이 없다.
#   그 밖의 「만기」가 새로 생기면 FAIL 이다.
COINED = '만기 도래'
#   채권 열머리에 쓰면 안 되는 약칭 — 원문에 없는 우리 조어.
ABBREV = re.compile(r'(?<![A-Za-z0-9])(Mi|Bi)(?![A-Za-z0-9])')
#   워크북 XML 전체(정의된 이름 · 시트 이름 · 데이터 유효성 · 주석)에서 막는 낱말.
BANNED = re.compile('가중평균만기|만기|[Dd]uration')
#   그 낱말을 품고도 남겨 두는 자리 — 개수까지 못 박는다. 늘어나면 FAIL 이다.
BAN_OK = {COINED: 8, 'platform_duration.py': 5}


def _runs(v):
    """셀 값 -> (평문, 글자별 아래첨자 여부). 리치텍스트가 아니면 전부 False."""
    try:
        from openpyxl.cell.rich_text import CellRichText
    except ImportError:
        CellRichText = ()
    if isinstance(v, CellRichText):
        txt, flag = '', []
        for b in v:
            t = b.text if hasattr(b, 'text') else str(b)
            f = getattr(b, 'font', None)
            sub = f is not None and getattr(f, 'vertAlign', None) == 'subscript'
            txt += t
            flag.extend([sub] * len(t))
        return txt, flag
    s = v if isinstance(v, str) else str(v)
    return s, [False] * len(s)


def notation_spans(txt):
    """기호 자리 -> (시작, 끝, 머리 길이, 평문). 규칙 쌍이 먼저, 남은 자리를 꼬리 규칙이 줍는다."""
    out, taken = [], []
    for m in NOTRE.finditer(txt):
        h, _ = NOTMAP[m.group(1)]
        out.append((m.start(), m.end(), len(h), m.group(1)))
        taken.append((m.start(), m.end()))
    for m in TAILRE.finditer(txt):
        if any(m.start() < e and s0 < m.end() for s0, e in taken):
            continue
        out.append((m.start(), m.end(), len(m.group(1)), m.group(0)))
    return sorted(out)


def notation_bad(txt, flag):
    """규약을 어긴 기호 자리. 머리는 평서, 꼬리는 아래첨자여야 한다."""
    out = []
    for i, j, hl, tok in notation_spans(txt):
        hd, tl = flag[i:i + hl], flag[i + hl:j]
        if any(hd) or not (tl and all(tl)):
            out.append(tok)
    return out


def banned_xml(path=None):
    """워크북 XML 전체에서 금지어를 찾는다. 셀 값만 보면 정의된 이름에 숨은 것을 놓친다.

    openpyxl 은 한글을 숫자 문자 참조로 쓴다. 먼저 되돌린 뒤 본다.
    """
    import zipfile, html as _html
    hit, seen = [], {k: 0 for k in BAN_OK}
    with zipfile.ZipFile(path or XLSX) as z:
        parts = [n for n in z.namelist() if n.endswith(('.xml', '.rels'))]
        for n in parts:
            t = _html.unescape(z.read(n).decode('utf-8', 'replace'))
            for k in BAN_OK:
                seen[k] += t.count(k)
                t = t.replace(k, '')
            for m in BANNED.finditer(t):
                hit.append([n, m.group(0), t[max(0, m.start() - 40):m.end() + 20]])
    return {'검사한 XML 조각': len(parts), '금지어 잔여': hit, '예외 자리 실측': seen,
            '예외 자리 못 박은 수': BAN_OK}


def notation_test(quiet=False):
    """기호 셀이 아래첨자 서식을 달았는가 · 옛 표기가 남았는가 · XML 에 금지어가 있는가."""
    wb = openpyxl.load_workbook(XLSX, rich_text=True)
    cells = subs = 0
    bad, mgi, abbrev, old, bareD, quoted = [], [], [], [], [], 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                txt, flag = _runs(c.value)
                if txt.startswith('='):
                    continue
                if is_quote(ws.title, c.column, c.row):
                    quoted += 1                    # 대표 원문 인용 — 옛 표기 검사에서 뺀다
                else:
                    if OLDRE.search(txt):
                        old.append([ws.title, c.coordinate, txt[:60],
                                    sorted(set(OLDRE.findall(txt)))])
                    #   칸 전체가 `D` 하나면 기호 사전 항목이다. 그 밖의 홀로 선 `D` 만 잡는다.
                    if (RULE['bare_D'] and txt.strip() != 'D'
                            and BAREDRE.search(txt)):
                        bareD.append([ws.title, c.coordinate, txt[:60]])
                if notation_spans(txt):
                    cells += 1
                    subs += sum(1 for x in flag if x)
                    b = notation_bad(txt, flag)
                    if b:
                        bad.append([ws.title, c.coordinate, txt[:60], b])
                if '만기' in txt and COINED not in txt:
                    mgi.append([ws.title, c.coordinate, txt[:60]])
    hdr = [_runs(wb['채권'].cell(1, c).value or '')[0]
           for c in range(1, wb['채권'].max_column + 1)]
    for h in hdr:
        if ABBREV.search(h):
            abbrev.append(h)
    xml = banned_xml()
    #   자기시험 — 같은 글자열에서 아래첨자를 떼면 판정기가 잡아야 한다.
    runs = [('SB', False), ('d-1', True), (' 상환액 · ', False),
            ('A', False), ('i', True), (' x ', False), ('D', False), ('i', True)]
    probe = ''.join(t for t, _ in runs)
    self_ok = notation_bad(probe, [f for t, f in runs for _ in t])
    self_flat = notation_bad(probe, [False] * len(probe))
    #   자기시험 — 옛 표기로 되돌린 글자열을 잡아야 한다.
    self_old = sorted(set(OLDRE.findall('SBD-1 상환액 · 표본 D-20 ~ D-11')))
    rep = {'기준표 출처': RULE['src'],
           '기준표 기호 쌍': [k for k, _, _ in RULE['pairs']],
           '기준표 금지 표기': RULE['forbid'],
           '기호를 담은 셀': cells, '아래첨자 글자': subs,
           '원문 인용 셀 (옛 표기 예외)': quoted,
           '규약 어긴 자리': bad, '옛 표기가 남은 자리': old,
           '홀로 선 대문자 D': bareD,
           '조어 밖 「만기」 셀': mgi, '채권 열머리 조어 약칭': abbrev,
           'XML 금지어': xml,
           '자기시험 — 아래첨자 뗀 벌에서 잡은 자리': self_flat,
           '자기시험 — 제대로 단 벌에서 잡은 자리': self_ok,
           '자기시험 — 옛 표기 벌에서 잡은 자리': self_old}
    rep['판정'] = {
        '기준표 기호 쌍 0건 아님': len(RULE['pairs']) > 0,
        '기준표 금지 표기 0건 아님': len(RULE['forbid']) > 0,
        '대상 0건 아님': cells > 0,
        '아래첨자 글자 0건 아님': subs > 0,
        '규약 어긴 자리 0건': not bad,
        '옛 표기가 남은 자리 0건 (원문 인용 제외)': not old,
        '홀로 선 대문자 D 0건 (원문 인용 제외)': not bareD,
        '원문 인용 셀 0건 아님': quoted > 0,
        '조어(만기 도래) 밖 「만기」 0건': not mgi,
        '채권 열머리에 Mi·Bi 0건': not abbrev,
        'XML 조각 0건 아님': xml['검사한 XML 조각'] > 0,
        'XML 금지어 잔여 0건': not xml['금지어 잔여'],
        'XML 예외 자리가 못 박은 수와 같음': xml['예외 자리 실측'] == BAN_OK,
        '자기시험 판별력 — 아래첨자': len(self_flat) == 3 and not self_ok,
        '자기시험 판별력 — 옛 표기': len(self_old) == 3,
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


# ── 워드 변수 전수 대응 ───────────────────────────────────────────
#   합격 기준 1(대표 2026-08-30 00:47:05 「워드에 있는 것들은 전부 엑셀에 나타나면 되지」)을
#   글자 그대로 잰다. 기준표를 여기 두지 않는다 — 이름은 `ceo_definitions.md` 원문에서 뽑는다.
#   `산식` 시트는 원문 ↔ 셀 색인이지 칸이 아니고, `읽는 법` 은 안내다. 둘을 빼고 찾는다.
DEFS_MD = os.path.join(BASE, 'ceo_definitions.md')
INDEX_SHEETS = ('산식', '읽는 법')
#   값으로 서 있어도 되는 원문 줄 수. 전부 입력 상수 자리다(할인율 · Di · 플랫폼ID · 기준일).
#   개수까지 못 박아 예외가 번지는 것을 막는다. 늘어도 줄어도 FAIL 이다.
VALUE_LINES = 8


def word_core(name):
    """원문의 이름 -> 핵심어. 괄호 안·공백·앞의 Σ 를 떼고 마지막 `의` 뒤 조각을 쓴다."""
    s = re.sub(r'\([^)]*\)', '', name or '')
    s = re.sub(r'\s+', '', s).strip('·,.').lstrip('Σ')
    if '의' in s[:-1]:
        t = s.rsplit('의', 1)[1]
        if len(t) >= 2:
            s = t
    return s


def word_named_vars(path=DEFS_MD):
    """대표 정의서 원문에서 이름 붙은 변수를 뽑는다.

    못 읽으면 예외로 죽는다. 대상이 0건이면 판정에서 FAIL 이다.
    """
    txt = open(path, encoding='utf-8').read()
    lines = [l[2:].strip() for l in txt.splitlines() if l.startswith('- ')]
    if len(lines) < 40:
        raise ValueError('원문 줄을 못 읽었다 (%d줄): %s' % (len(lines), path))
    #   (가) 줄 머리 이름 — `이름: …` 또는 `이름 = …`. 쉼표가 든 줄은 이름이 아니라 서술이다.
    heads = []
    for l in lines:
        h = re.split(r'[=:]', l, 1)[0].strip()
        if h and ',' not in h and len(h) <= 30:
            heads.append(word_core(h))
    #   (나) 줄 안에서만 이름이 불리는 산출 대상 — `아래와 같이 … 산출해` 열거.
    enum = []
    for l in lines:
        m = re.search(r'아래와 같이\s*(.+?)\s*산출해', l)
        if not m:
            continue
        enum = [word_core(t) for t in m.group(1).split(',') if word_core(t)]
    #   (다) 괄호로 이름 붙인 기호 — 대문자 두 자 이상. `(D-1)` 같은 날짜 꼬리는 뺀다.
    syms = sorted({s for s in re.findall(r'\(([A-Za-z][A-Za-z0-9-]*)\)', txt)
                   if len(s) <= 6 and len(re.findall(r'[A-Z]', s)) >= 2})
    return {'줄': len(lines), '줄 머리 이름': sorted(set(heads)),
            '열거 산출 대상': enum, '괄호 기호': syms}


def data_cells(wb):
    """색인·안내 시트를 뺀 시트의 평문 셀 -> [(주소, 원문, 공백 뺀 소문자)]."""
    out = []
    for ws in wb.worksheets:
        if ws.title in INDEX_SHEETS:
            continue
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                t = _runs(c.value)[0]
                if not t or t.startswith('='):
                    continue
                out.append(('%s!%s' % (ws.title, c.coordinate), t,
                            re.sub(r'\s+', '', t).lower()))
    return out


def _cellrefs(wb, text):
    """`산식` C열이 지시한 자리 문자열 -> [(시트, [주소…])].

    괄호 안은 다른 파일·풀이말이라 뺀다. 시트 이름이 앞서지 않은 주소는 버린다.
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    t = re.sub(r'\([^)]*\)', ' ', text or '')
    marks = sorted((m.start(), s) for s in wb.sheetnames
                   for m in re.finditer(re.escape(s), t))
    tok = re.compile(r'(?P<rng>[A-Z]{1,2}\d+:[A-Z]{1,2}\d+)'
                     r'|(?P<cell>[A-Z]{1,2}\d+)'
                     r'|(?P<cols>(?:[A-Z]·)*[A-Z]열)')
    out = []
    for m in tok.finditer(t):
        sh = None
        for pos, s in marks:
            if pos < m.start():
                sh = s
            else:
                break
        if sh is None:
            continue
        g, v = m.lastgroup, m.group()
        if g == 'cell':
            out.append((sh, [v]))
        elif g == 'rng':
            a, b = v.split(':')
            ca, ra = re.match(r'([A-Z]+)(\d+)', a).groups()
            cb, rb = re.match(r'([A-Z]+)(\d+)', b).groups()
            out.append((sh, ['%s%d' % (get_column_letter(ci), ri)
                             for ci in range(column_index_from_string(ca),
                                             column_index_from_string(cb) + 1)
                             for ri in range(int(ra), int(rb) + 1)]))
        else:
            ws = wb[sh]
            out.append((sh, ['%s%d' % (c, ri) for c in v.replace('열', '').split('·')
                             for ri in range(2, ws.max_row + 1)]))
    return out


def word_test(quiet=False):
    """원문이 이름 부른 변수가 전부 칸을 갖는가 · 수식이어야 할 자리가 값으로 있지 않은가."""
    wb = openpyxl.load_workbook(XLSX, rich_text=True)
    V = word_named_vars()
    cells = data_cells(wb)

    def where_ko(k):
        k2 = re.sub(r'\s+', '', k).lower()
        return [a for a, _t, n in cells if k2 in n][:3]

    def where_sym(k):
        rx = re.compile(r'(?<![A-Za-z0-9])%s(?![A-Za-z0-9])' % re.escape(k), re.I)
        return [a for a, t, _n in cells if rx.search(t)][:3]

    enum = {k: where_ko(k) for k in V['열거 산출 대상']}
    syms = {k: where_sym(k) for k in V['괄호 기호']}

    #   `산식` C열이 빈 줄 — 줄 머리 이름은 이 색인으로 셀을 지시받는다.
    fs = wb['산식']
    lines = [(r, _runs(fs.cell(r, 2).value)[0],
              _runs(fs.cell(r, 3).value)[0] if fs.cell(r, 3).value is not None else '')
             for r in range(2, fs.max_row + 1) if fs.cell(r, 2).value]
    blank = [x[1][:34] for x in lines if not x[2]]

    #   지시받은 자리가 수식인가 값인가. 값 셀이 하나라도 있으면 그 줄은 `값` 이다.
    kinds = []
    for r, line, cmap in lines:
        nf = nv = 0
        ex = []
        for sh, ads in _cellrefs(wb, cmap):
            ws = wb[sh]
            for a in ads:
                v = ws[a].value
                if v is None or v == '':
                    continue
                if isinstance(v, str) and v.startswith('='):
                    nf += 1
                else:
                    nv += 1
                    if len(ex) < 2:
                        ex.append('%s!%s' % (sh, a))
        kind = '값' if nv else ('수식' if nf else '자리 지시만')
        kinds.append({'산식 행': r, '원문': line[:34], '지시': cmap[:46],
                      '분류': kind, '수식셀': nf, '값셀': nv, '보기': ex})
    val = [x for x in kinds if x['분류'] == '값']
    vsheets = sorted({e.split('!')[0] for x in val for e in x['보기']})

    rep = {'원문': os.path.relpath(DEFS_MD, BASE), '원문 줄': V['줄'],
           '줄 머리 이름': V['줄 머리 이름'],
           '열거 산출 대상 ↔ 칸': enum, '괄호 기호 ↔ 칸': syms,
           '산식 대응 셀 빈 줄': blank,
           '줄 분류': {k: sum(1 for x in kinds if x['분류'] == k)
                       for k in ('수식', '값', '자리 지시만')},
           '값으로 선 줄': [[x['산식 행'], x['원문'], x['지시'], x['보기']] for x in val],
           '값으로 선 자리의 시트': vsheets,
           '값으로 선 줄 못 박은 수': VALUE_LINES,
           '검사한 평문 셀': len(cells)}
    rep['판정'] = {
        '원문 45줄을 읽었다': V['줄'] == 45,
        '줄 머리 이름 0건 아님': len(V['줄 머리 이름']) > 0,
        '열거 산출 대상 0건 아님': len(enum) > 0,
        '괄호 기호 0건 아님': len(syms) > 0,
        '검사한 평문 셀 0건 아님': len(cells) > 0,
        '열거 산출 대상 전건이 칸을 갖는다': all(bool(v) for v in enum.values()),
        '괄호 기호 전건이 칸을 갖는다': all(bool(v) for v in syms.values()),
        '원문 전건이 산식 색인에서 셀을 지시받는다': not blank,
        '산식 색인이 45줄 전건을 덮는다': len(lines) == V['줄'],
        '값으로 선 줄이 못 박은 수와 같다': len(val) == VALUE_LINES,
        '수식으로 선 줄 0건 아님': rep['줄 분류']['수식'] > 0,
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


# ── 표식 · 옮겨 적기 ──────────────────────────────────────────────
def ec_head(wb):
    """`일별` EC 열 자리를 열머리에서 찾는다. 못 찾으면 죽는다 — 열 주소를 손으로 적지 않는다."""
    ws = wb['일별']
    for c in range(1, ws.max_column + 1):
        t = _runs(ws.cell(1, c).value)[0] if ws.cell(1, c).value is not None else ''
        if re.match(r'^\s*EC(?![A-Za-z0-9])', t):
            return c, get_column_letter(c), t
    raise ValueError('일별 시트에서 EC 열머리를 못 찾았다')


def mark_test(quiet=False):
    """`일별` EC 열이 상수인 까닭이 적혀 있는가 · `읽는 법` 갈림 표가 `화면대조` 와 같은가."""
    bk = Book(XLSX)
    wb = openpyxl.load_workbook(XLSX, rich_text=True)
    ec, L, echead = ec_head(wb)
    ws = wb['일별']
    fml, rows = [], []
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        v = ws.cell(r, ec).value
        if not (isinstance(a, str) and a.startswith('=')):
            break                                   # 합계 행에서 멈춘다
        if isinstance(v, str) and v.startswith('='):
            fml.append(v)
            rows.append('%s%d' % (L, r))
    const = bool(fml) and len(set(fml)) == 1
    #   까닭 표식 — `일별` 시트의 평문 칸. 무엇이 없어서 상수인지가 적혀 있어야 한다.
    why = []
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            t = _runs(c.value)[0]
            if t.startswith('='):
                continue
            if all(k in t for k in ('미확정', 'EC', '원장', '순현금')):
                why.append(['일별!%s' % c.coordinate, t[:90]])

    # ── 읽는 법 갈림 표 ↔ 화면대조
    gw = wb['읽는 법']
    h = None
    for r in range(1, gw.max_row + 1):
        v = gw.cell(r, 2).value
        if v is not None and _runs(v)[0].strip() == '화면 ↔ 엑셀 갈림':
            h = r
            break
    if h is None:
        raise ValueError('읽는 법 시트에서 갈림 표 머리를 못 찾았다')
    cw = wb['화면대조']
    gap, bad = [], []
    r = h + 1
    while gw.cell(r, 2).value is not None:
        lab = _runs(gw.cell(r, 2).value)[0]
        m = re.search(r'(\d+)\s*행', str(gw.cell(r, 4).value or ''))
        if not m:
            bad.append([lab, '화면대조 행 표기 없음'])
            r += 1
            continue
        rr = int(m.group(1))
        src = _runs(cw.cell(rr, 2).value)[0] if cw.cell(rr, 2).value is not None else ''
        pair = [('엑셀 값', 'E%d' % r, 'C%d' % rr), ('화면 값', 'F%d' % r, 'E%d' % rr),
                ('차이', 'G%d' % r, 'F%d' % rr)]
        row = {'항목': lab, '화면대조 행': rr, '구분': cw.cell(rr, 1).value,
               '원본 항목': src}
        for nm, ga, ca in pair:
            a, b = bk.cell('읽는 법', ga), bk.cell('화면대조', ca)
            row[nm] = a
            if abs(numify(a) - numify(b)) > 1e-9:
                bad.append([lab, nm, a, b])
        wa = bk.cell('읽는 법', 'C%d' % r)
        wsrc = bk.cell('화면대조', 'H%d' % rr)
        row['까닭 같음'] = _txt(wa) == _txt(wsrc)
        if not row['까닭 같음']:
            bad.append([lab, '까닭', _txt(wa)[:40], _txt(wsrc)[:40]])
        if src != lab:
            bad.append([lab, '항목 이름', lab, src])
        if cw.cell(rr, 1).value != '모델 잔차':
            bad.append([lab, '화면대조 구분', cw.cell(rr, 1).value])
        gap.append(row)
        r += 1

    rep = {'EC 열': '일별!%s' % L, 'EC 열머리': echead,
           'EC 칸': len(fml), 'EC 수식 가짓수': len(set(fml)),
           'EC 수식': sorted(set(fml))[:3], 'EC 가 상수': const,
           'EC 까닭 표식': why,
           '갈림 표 머리 행': h, '갈림 표 행': gap, '갈림 표 어긋난 자리': bad}
    rep['판정'] = {
        'EC 칸 0건 아님': len(fml) > 0,
        'EC 가 상수면 열머리에 미확정 표식': (not const) or ('미확정' in echead),
        'EC 가 상수면 까닭 표식이 있다': (not const) or bool(why),
        '갈림 표 행 0건 아님': len(gap) > 0,
        '갈림 표가 화면대조와 어긋나지 않는다': not bad,
        '갈림 표 전건이 모델 잔차 행이다': bool(gap) and all(
            x['구분'] == '모델 잔차' for x in gap),
        '갈림 표 전건이 실제로 갈린다': bool(gap) and all(
            abs(numify(x['차이'])) > 1e-9 for x in gap),
    }
    rep['통과'] = all(rep['판정'].values())
    if not quiet:
        print(json.dumps(rep, ensure_ascii=False, indent=1, default=str))
    return rep


if __name__ == '__main__':
    arg = sys.argv[1] if len(sys.argv) > 1 else ''
    if arg == 'notation':
        sys.exit(0 if notation_test()['통과'] else 1)
    if arg == 'idle':
        sys.exit(0 if idle_test()['통과'] else 1)
    if arg == 'variant':
        sys.exit(0 if variant_test()['통과'] else 1)
    if arg == 'five':
        sys.exit(0 if five_test()['통과'] else 1)
    if arg == 'word':
        sys.exit(0 if word_test()['통과'] else 1)
    if arg == 'mark':
        sys.exit(0 if mark_test()['통과'] else 1)
    if arg == 'switch':
        switch_test()
    elif arg == 'sens':
        sens_test()
    elif arg == 'mix':
        mix_test()
    elif arg == 'all':
        r = main()
        s1 = switch_test(True)
        s2 = sens_test(True)
        s3 = mix_test(True)
        s4 = notation_test(True)
        s5 = idle_test(True)
        s6 = variant_test(True)
        s7 = five_test(True)
        s8 = word_test(True)
        s9 = mark_test(True)
        print(json.dumps({'스위치': s1, '배달앱비중 민감도': s2, '구성비 입력 반응': s3,
                          '기호 표기': s4, '총 투자자산·유휴 비율': s5,
                          '유휴 세 벌': s6, '⑤ 단일 원천 · ⑥ 참조': s7,
                          '워드 변수 전수 대응': s8, 'EC 표식 · 갈림 표': s9},
                         ensure_ascii=False, indent=1, default=str))
        ok = (r['평가 실패 수'] == 0 and r['화면대조 차이'] == 0
              and not r['화면 값 ↔ ledger_facts 불일치'] and r['바이트 동일']
              and s1['스위치 4개 모두 동작'] and s2['움직임'] and s3['일치']
              and s4['통과'] and s5['통과'] and s6['통과'] and s7['통과']
              and s8['통과'] and s9['통과']
              and r['S입금부족율 raw 잔차 등재'] and r['S입금부족율 표기 비교 잔존']
              and r['⑤ 미확정 표식']['화면 반영'] == '미확정'
              and '대기' in (r['⑤ 미확정 표식']['출처'] or '')
              and r['⑥ 미확정 표식']['화면 반영'] == '미확정'
              and all(r['대표 DM 16:45 판정'].values())
              and all(r['대표 DM 16:27 판정'].values())
              and all(r['대표 DM 인용 판정'].values())
              and r['산식 원문 줄'] > 0 and not r['산식 대응 셀 빈 줄']
              and not r['사유 빠진 잔차 행'] and abs(r['비중 합'] - 100.0) < 1e-9)
        print(json.dumps({'전체 통과': ok}, ensure_ascii=False))
        sys.exit(0 if ok else 1)
    else:
        main()
